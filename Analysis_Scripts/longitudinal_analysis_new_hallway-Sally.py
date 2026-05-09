"""Longitudinal analysis of behavioral data across multiple mice.
Original Author: Brenna Manuel

TEST COUNTING LOGIC:
To test the zone/event counting logic on a single trial_log file before running the full analysis:
    
    from longitudinal_analysis_new_hallway import test_matching_logic
    test_matching_logic('path/to/your/trial_log.csv')
    
Or run this script and call the function interactively in Python console.
"""

import ast
import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from datetime import datetime
import colorsys

# ── Global Matplotlib defaults ────────────────────────────────────────────────
plt.rcParams["font.family"] = "sans-serif"
plt.rcParams["font.sans-serif"] = ["Arial"]
plt.rcParams["svg.fonttype"] = "none"
plt.rcParams.update({
    "font.size": 8,
    "axes.titlesize": 10,
    "axes.labelsize": 8,
    "xtick.labelsize": 8,
    "ytick.labelsize": 8,
    "legend.fontsize": 7.5,
    "figure.titlesize": 10,
    "lines.linewidth": 0.9,
    "lines.markersize": 3,
    "figure.figsize": (4, 2.5),
})
import sys
import pickle
import hashlib
import warnings
import math
from concurrent.futures import ThreadPoolExecutor as _TPE

# Add Analysis_Scripts to path to import lick detection algorithm
script_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, script_dir)
import lick_detection_algorithm as lda
from scipy.stats import norm, mannwhitneyu
from scipy.signal import butter, filtfilt
from timeline_refactored import (
    safe_literal_eval,
    uniformly_sample_treadmill,
    uniformly_sample_capacitive,
    create_aligned_windows,
)

# Cache directory for KDE values
KDE_CACHE_DIR = os.path.join(script_dir, '.kde_cache')
if not os.path.exists(KDE_CACHE_DIR):
    os.makedirs(KDE_CACHE_DIR)

def get_file_hash(filepath):
    """Generate a hash based on file path and modification time."""
    stat = os.stat(filepath)
    # Use file path, size, and modification time to create unique identifier
    hash_input = f"{filepath}_{stat.st_size}_{stat.st_mtime}".encode('utf-8')
    return hashlib.md5(hash_input).hexdigest()

def get_cached_kde(capacitive_filepath):
    """Retrieve cached KDE value for a capacitive file if it exists."""
    file_hash = get_file_hash(capacitive_filepath)
    cache_file = os.path.join(KDE_CACHE_DIR, f"{file_hash}.pkl")
    
    if os.path.exists(cache_file):
        try:
            with open(cache_file, 'rb') as f:
                cached_data = pickle.load(f)
                return cached_data['kde_value']
        except Exception as e:
            # If cache is corrupted, ignore and recalculate
            print(f"Warning: Cache read failed for {capacitive_filepath}: {e}")
            return None
    return None

def cache_kde_value(capacitive_filepath, kde_value):
    """Cache the KDE value for a capacitive file."""
    file_hash = get_file_hash(capacitive_filepath)
    cache_file = os.path.join(KDE_CACHE_DIR, f"{file_hash}.pkl")
    
    try:
        with open(cache_file, 'wb') as f:
            pickle.dump({'kde_value': kde_value}, f)
    except Exception as e:
        # If caching fails, just continue without caching
        print(f"Warning: Cache write failed for {capacitive_filepath}: {e}")


# ── Session-level result cache ────────────────────────────────────────────────
# Caches all per-session computed metrics so that re-runs load in seconds
# instead of re-reading and re-processing every CSV file.
SESSION_CACHE_DIR = os.path.join(script_dir, '.session_cache')
if not os.path.exists(SESSION_CACHE_DIR):
    os.makedirs(SESSION_CACHE_DIR)

_SESSION_CACHE_VERSION = 5  # bump this to invalidate all cached sessions after code changes

# ── Diagnostic mode ───────────────────────────────────────────────────────────
# Set to True to print per-trial matching, lick-latency, and lick-prop traces
# to the console while sessions are being processed.  Has no effect on results.
_DIAGNOSTIC_MODE = False


def _session_cache_key(paths):
    """Hash a list of file paths by path + size + mtime to create a unique session key."""
    parts = []
    for p in paths:
        p = str(p) if p else ''
        if p and os.path.exists(p):
            st = os.stat(p)
            parts.append(f"{p}|{st.st_size}|{st.st_mtime}")
        else:
            parts.append(f"missing:{p}")
    return hashlib.md5("|".join(parts).encode('utf-8')).hexdigest()


def _load_session_cache(key):
    """Return cached session data dict, or None on miss / version mismatch."""
    cache_file = os.path.join(SESSION_CACHE_DIR, f"{key}.pkl")
    if os.path.exists(cache_file):
        try:
            with open(cache_file, 'rb') as f:
                data = pickle.load(f)
            if data.get('_version') == _SESSION_CACHE_VERSION:
                return data
        except Exception:
            pass
    return None


def _save_session_cache(key, data):
    """Persist session data dict to disk."""
    data['_version'] = _SESSION_CACHE_VERSION
    cache_file = os.path.join(SESSION_CACHE_DIR, f"{key}.pkl")
    try:
        with open(cache_file, 'wb') as f:
            pickle.dump(data, f, protocol=pickle.HIGHEST_PROTOCOL)
    except Exception as e:
        print(f"Warning: session cache write failed: {e}")


def _read_three_csvs(path_treadmill, path_capacitive, path_trial_log):
    """Read three session CSV files in parallel (I/O-bound, uses threads).

    Returns
    -------
    treadmill_df, capacitive_df, trial_log_df : DataFrames or None
    missing : list of label strings for files that could not be read
    """
    results = [None, None, None]
    missing = []
    labels  = ['treadmill', 'capacitive', 'trial_log']
    paths   = [path_treadmill, path_capacitive, path_trial_log]

    def _read(idx):
        try:
            results[idx] = pd.read_csv(paths[idx])
        except Exception:
            pass  # results[idx] stays None

    with _TPE(max_workers=3) as ex:
        list(ex.map(_read, range(3)))

    for i, label in enumerate(labels):
        if results[i] is None:
            missing.append(label)

    return results[0], results[1], results[2], missing


def detect_capacitive_gaps(cap_df: pd.DataFrame,
                            max_gap_s: float = 1.0) -> dict:
    """Detect time gaps in a capacitive recording that exceed max_gap_s.

    The sensor runs at ~50 Hz (20 ms per sample). Any inter-sample interval
    longer than max_gap_s indicates a section of data was deliberately removed
    (e.g. by the baseline-spike trimmer). Knowing where these gaps are is
    required for correct epoch alignment and lick-bout analysis.

    Parameters
    ----------
    cap_df : DataFrame with an 'elapsed_time' column
    max_gap_s : minimum inter-sample interval (seconds) to count as a gap

    Returns
    -------
    dict with keys:
        'has_gaps'   : bool
        'n_gaps'     : int
        'gaps'       : list of (gap_start_s, gap_end_s, gap_duration_s) tuples
        'total_gap_s': float  — total trimmed time
    """
    if 'elapsed_time' not in cap_df.columns:
        return {'has_gaps': False, 'n_gaps': 0, 'gaps': [], 'total_gap_s': 0.0}

    times = pd.to_numeric(cap_df['elapsed_time'], errors='coerce').dropna().values
    if len(times) < 2:
        return {'has_gaps': False, 'n_gaps': 0, 'gaps': [], 'total_gap_s': 0.0}

    diffs = np.diff(times)
    gap_mask = diffs > max_gap_s
    gap_indices = np.where(gap_mask)[0]

    gaps = []
    for idx in gap_indices:
        t_start = float(times[idx])
        t_end   = float(times[idx + 1])
        gaps.append((t_start, t_end, t_end - t_start))

    total_gap_s = sum(g[2] for g in gaps)
    return {
        'has_gaps':    len(gaps) > 0,
        'n_gaps':      len(gaps),
        'gaps':        gaps,
        'total_gap_s': total_gap_s,
    }


def compute_valid_session_duration(cap_df: pd.DataFrame,
                                    max_gap_s: float = 1.0) -> float:
    """Return the true valid data duration of a (possibly trimmed) capacitive file.

    For un-trimmed files this equals elapsed_time.max() - elapsed_time.min().
    For trimmed files (where trim_baseline_spikes.py removed rows) the
    elapsed_time column retains the original timestamps from the full session,
    so elapsed_time.max() would over-count by the total trimmed time.

    This function sums only the inter-sample intervals that are <= max_gap_s
    (normal ~20 ms steps at 50 Hz), giving the total time actually covered by
    valid data — excluding any removed segments.

    Parameters
    ----------
    cap_df    : DataFrame with an 'elapsed_time' column
    max_gap_s : inter-sample threshold above which an interval counts as a gap

    Returns
    -------
    float — valid data duration in seconds (NaN if elapsed_time is absent)
    """
    if 'elapsed_time' not in cap_df.columns:
        return float('nan')
    times = pd.to_numeric(cap_df['elapsed_time'], errors='coerce').dropna().values
    if len(times) < 2:
        return float(times[0]) if len(times) == 1 else 0.0
    diffs = np.diff(times)
    return float(diffs[diffs <= max_gap_s].sum())


def compute_session_distance(treadmill_df):
    """Compute total distance traversed in a session from raw treadmill data.

    Finds the first non-zero value in the 'distance' column, subtracts it from
    all rows to normalise the cumulative odometer to zero at the start of
    motion, then returns both the full normalised series and the total distance
    (final normalised value) for the session.

    Parameters
    ----------
    treadmill_df : pd.DataFrame
        Raw treadmill CSV loaded as a DataFrame (must contain a 'distance' column).

    Returns
    -------
    normalised : pd.Series
        Per-row cumulative distance from the first non-zero sample (same index
        as treadmill_df).  All rows before the first non-zero value are NaN.
    total_distance : float
        Distance traversed from first non-zero sample to end of session (cm).
        NaN if no non-zero distance is found.
    """
    dist = pd.to_numeric(treadmill_df['distance'], errors='coerce')
    nonzero_mask = dist != 0
    if not nonzero_mask.any():
        return pd.Series(np.nan, index=treadmill_df.index), float('nan')

    first_nonzero_val = dist[nonzero_mask].iloc[0]
    first_nonzero_idx = nonzero_mask.idxmax()           # label of first True

    normalised = dist.copy()
    # Rows before first non-zero are meaningless pre-motion artefacts; set NaN
    normalised.loc[:first_nonzero_idx] = np.nan
    normalised = normalised - first_nonzero_val          # shift so first point = 0
    normalised.loc[first_nonzero_idx] = 0.0              # ensure exactly 0

    total_distance = float(normalised.dropna().iloc[-1]) if len(normalised.dropna()) > 0 else float('nan')
    return normalised, total_distance


def generate_colors(n):
    """Generate n distinct colors"""
    colors = []
    for i in range(n):
        hue = i / n
        saturation = 0.7 + np.random.rand() * 0.3  # Random between 0.7-1.0
        value = 0.7 + np.random.rand() * 0.3       # Random between 0.7-1.0
        rgb = colorsys.hsv_to_rgb(hue, saturation, value)
        colors.append(rgb)
    return colors

# ── Canonical condition colours ───────────────────────────────────────────────
_COLOR_0PCT  = "#808080"   # 0% CA starting condition (gray)
_COLOR_2PCT  = "#c0392b"   # 2% CA starting condition (red)
_COLOR_OTHER = "#7f3f98"   # fallback (purple)


def _condition_to_color(label: str) -> str:
    """Map a starting_condition label to its canonical hex colour."""
    lo = str(label).lower()
    if "0%" in lo:
        return _COLOR_0PCT
    if "2%" in lo:
        return _COLOR_2PCT
    return _COLOR_OTHER


def apply_common_plot_style(
    ax,
    start_x_at_zero: bool = False,
    remove_top_right: bool = True,
    remove_x_margins: bool = True,
    remove_y_margins: bool = True,
    ticks_in: bool = True,
    draw_zero_dotted_line: bool = True,
):
    """Apply common styling: remove top/right spines, set tick directions, adjust margins."""
    if remove_top_right:
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)

    if ticks_in:
        ax.tick_params(direction='in', which='both', length=5)

    if remove_x_margins:
        ax.margins(x=0)

    if remove_y_margins:
        ax.margins(y=0)
        ax.autoscale(axis='y', tight=True)

    if start_x_at_zero:
        left, right = ax.get_xlim()
        ax.set_xlim(left=0, right=right)

    if draw_zero_dotted_line:
        try:
            ax.axhline(0, linestyle='-', color='0', linewidth=1.0, alpha=0.8, zorder=1)
        except Exception:
            pass

    return ax


def _auto_integer_step(
    vmin: float,
    vmax: float,
    target_ticks: int = 7,
    allow_sub5: bool = False,
) -> int:
    """Choose a 'nice' integer step so about target_ticks cover the range."""
    if not (np.isfinite(vmin) and np.isfinite(vmax)):
        return 1
    range_int = int(abs(math.ceil(vmax) - math.floor(vmin)))
    if range_int <= 0:
        return 1
    approx = max(1.0, range_int / max(1, target_ticks))
    pow10 = 10 ** int(math.floor(math.log10(approx)))
    multipliers = (1, 2, 2.5, 3, 4, 5) if allow_sub5 else (1, 2, 5)
    for m in multipliers:
        step = int(max(1, math.ceil(m * pow10)))
        if range_int / step <= target_ticks:
            return step
    return int(max(1, 10 * pow10))


def _apply_integer_axis(
    ax,
    *,
    axis: str,
    data_min: float,
    data_max: float,
    step: int,
    clamp_min=None,
    left_pad_steps: int = 0,
    right_pad_steps: int = 1,
) -> None:
    """Apply integer ticks and limits to the chosen axis with one extra step beyond data."""
    step = int(max(1, step))
    base_start = int(math.floor(data_min / step) * step)
    base_end_tick = int(math.ceil(data_max / step) * step)
    tick_start = base_start - left_pad_steps * step
    tick_end = base_end_tick + right_pad_steps * step
    start = tick_start
    if clamp_min is not None and start < clamp_min:
        start = clamp_min
    end = int(data_max) + right_pad_steps * step
    if end <= start:
        end = start + step
    all_ticks = list(range(tick_start, tick_end + 1, step))
    ticks = [t for t in all_ticks if start <= t <= end]
    if axis == 'x':
        ax.set_xlim(start, end)
        ax.set_xticks(ticks)
        ax.xaxis.set_major_formatter(mticker.FormatStrFormatter('%d'))
    elif axis == 'y':
        ax.set_ylim(start, end)
        ax.set_yticks(ticks)
        ax.yaxis.set_major_formatter(mticker.FormatStrFormatter('%.0f'))

# ── Locomotion bout detection constants ──────────────────────────────────────
BOUT_THRESHOLD_CM_S  = 2.0   # cm/s — minimum speed to count as moving
MIN_BOUT_DURATION_S  = 2.0   # s    — minimum continuous time above threshold
MAX_INTER_BOUT_GAP_S = 2.0   # s    — gaps <= this between bouts are merged


def detect_locomotion_bouts(time, speed_cm_s,
                             threshold=BOUT_THRESHOLD_CM_S,
                             min_duration=MIN_BOUT_DURATION_S,
                             max_inter_bout_gap=MAX_INTER_BOUT_GAP_S):
    """
    1. Build contiguous above/below-threshold runs.
    2. Bridge short below-threshold gaps (<= max_inter_bout_gap).
    3. Keep merged above-threshold spans >= min_duration.
    Returns list of (t_start, t_end) tuples in seconds.
    """
    above = speed_cm_s >= threshold
    n = len(above)

    runs = []
    start = 0
    for i in range(1, n):
        if above[i] != above[start]:
            runs.append((bool(above[start]), start, i - 1))
            start = i
    runs.append((bool(above[start]), start, n - 1))

    merged = []
    i = 0
    while i < len(runs):
        if runs[i][0]:
            j = i
            while j + 2 < len(runs):
                gap_dur = time[runs[j + 1][2]] - time[runs[j + 1][1]]
                if (not runs[j + 1][0]) and gap_dur <= max_inter_bout_gap:
                    j += 2
                else:
                    break
            merged.append((True, runs[i][1], runs[j][2]))
            i = j + 1
        else:
            merged.append(runs[i])
            i += 1

    bouts = []
    for is_above, s, e in merged:
        if is_above and (time[e] - time[s]) >= min_duration:
            bouts.append((time[s], time[e]))
    return bouts


def test_matching_logic(trial_log_path):
    """
    Test the zone/event counting logic on a single trial_log CSV file.
    
    Logic:
    - Valid zones (opportunities) = stay_texture_change_time entries (excluding re-entries within 0.05s)
    - Hits = count(reward_event) + count(hits_event)
    - Misses = valid zones - hits
    - Sensitivity = hits / valid zones
    
    Args:
        trial_log_path: Path to a single trial_log.csv file
    
    Example usage:
        test_matching_logic('f:/virtual_foraging_cohort_4/VF11/Session_20/beh/1774635104trial_log.csv')
    """
    print("\n" + "="*80)
    print("TESTING ZONE/EVENT COUNTING LOGIC")
    print("="*80)
    print(f"File: {trial_log_path}")
    
    # Read trial log
    trial_log = pd.read_csv(trial_log_path)
    
    print(f"\nTotal rows in trial_log: {len(trial_log)}")
    
    # Extract ALL stay zone entry times (if column exists; older data may not have it)
    if 'stay_texture_change_time' in trial_log.columns:
        stay_zone_times = pd.to_numeric(trial_log['stay_texture_change_time'], errors='coerce').dropna().values
        print(f"Total stay zone entry times: {len(stay_zone_times)}")
        
        # Collect all re-entry times to exclude them
        re_entry_times_set = set()
        if 'zone_re_entry_time' in trial_log.columns:
            for val in trial_log['zone_re_entry_time']:
                if pd.notna(val) and val != '':
                    try:
                        # Handle both single values and arrays
                        if isinstance(val, str) and val.strip():
                            import ast
                            try:
                                re_entry_list = ast.literal_eval(val)
                                if isinstance(re_entry_list, (list, tuple)):
                                    for t in re_entry_list:
                                        if pd.notna(t):
                                            re_entry_times_set.add(float(t))
                                else:
                                    re_entry_times_set.add(float(re_entry_list))
                            except:
                                pass
                        elif not isinstance(val, str):
                            re_entry_times_set.add(float(val))
                    except (ValueError, TypeError):
                        pass
        
        print(f"Re-entry times found: {len(re_entry_times_set)}")
        if len(re_entry_times_set) > 0:
            print(f"  Sample re-entry times: {sorted(list(re_entry_times_set))[:5]}")
        
        # Filter out stay_zone_times that match re-entries (within 0.05 seconds)
        valid_zone_times = []
        excluded_zones = []
        excluded_details = []  # Store details for reporting
        for zone_time in stay_zone_times:
            is_reentry = False
            for re_entry_time in re_entry_times_set:
                if abs(zone_time - re_entry_time) <= 0.05:
                    is_reentry = True
                    excluded_zones.append(zone_time)
                    excluded_details.append((zone_time, re_entry_time, abs(zone_time - re_entry_time)))
                    break
            if not is_reentry:
                valid_zone_times.append(zone_time)
        
        valid_zone_times = np.array(valid_zone_times)
        print(f"\nZones excluded as re-entries: {len(excluded_zones)}")
        if len(excluded_zones) > 0:
            print(f"  Excluded zone details (first 5):")
            for zone, reentry, diff in sorted(excluded_details)[:5]:
                print(f"    Zone {zone:.2f} ← Re-entry {reentry:.2f} (diff={diff:.3f}s)")
            if len(excluded_zones) > 5:
                print(f"    ... and {len(excluded_zones) - 5} more")
        print(f"Valid zone times (after re-entry filtering): {len(valid_zone_times)}")
        total_opportunities = len(valid_zone_times)
    else:
        print("\nNote: 'stay_texture_change_time' not found. Using texture_history fallback for opportunities.")
        total_opportunities = len(trial_log[trial_log['texture_history'] == 'assets/reward_mean100.jpg'])
        print(f"Reward opportunities (texture_history fallback): {total_opportunities}")
    
    # Collect all reward_event and hits_event times
    reward_event_times = pd.to_numeric(trial_log['reward_event'], errors='coerce').dropna().values
    
    # Check if hits_event column exists (older data may not have it)
    if 'hits_event' in trial_log.columns:
        hits_event_times = pd.to_numeric(trial_log['hits_event'], errors='coerce').dropna().values
    else:
        hits_event_times = np.array([])  # Empty array if column doesn't exist
        print("\nNote: 'hits_event' column not found. Using only 'reward_event' as hits.")
    
    print(f"\nReward events found: {len(reward_event_times)}")
    print(f"Hits events found: {len(hits_event_times)}")
    
    # Simple counting: hits = rewards + hits, misses = opportunities - hits
    if total_opportunities > 0:
        reward_count = len(reward_event_times) + len(hits_event_times)
        misses = total_opportunities - reward_count
        sensitivity = float(reward_count) / float(total_opportunities) if total_opportunities > 0 else 0.0
        
        print(f"\n" + "-"*80)
        print("FINAL RESULTS:")
        print("-"*80)
        print(f"Total valid zones (opportunities): {total_opportunities}")
        print(f"Hits (total successful outcomes): {reward_count}")
        print(f"  - reward_event count: {len(reward_event_times)}")
        print(f"  - hits_event count: {len(hits_event_times)}")
        print(f"Misses: {misses}")
        print(f"Sensitivity: {sensitivity:.3f} ({reward_count}/{total_opportunities})")
        print("="*80 + "\n")
    else:
        print("\nNo valid zones found!")
        print("="*80 + "\n")

# ── Behavioral epoch constants & helpers ────────────────────────────────────
EPOCH_WINDOW_S       = 5            # half-window size (seconds each side of event)
EPOCH_N_SAMPLES      = 501          # canonical time-axis length (10 s / 0.02 s + 1)
EPOCH_CANONICAL_TIME = np.linspace(-EPOCH_WINDOW_S, EPOCH_WINDOW_S, EPOCH_N_SAMPLES)
# Lick-epoch time axis: 250 ms steps over the same ±5 s window (41 points).
# Used with a 500 ms sliding window to give biologically meaningful lick-rate estimates.
LICK_EPOCH_TIME      = np.arange(-EPOCH_WINDOW_S, EPOCH_WINDOW_S + 1e-9, 0.25)


def _extract_reward_zone_entry_times(trial_log_df):
    """Return sorted list of reward zone entry timestamps from a trial log.

    Handles two trial-log formats and excludes re-entries in the new format:

    New hallway format (preferred, detected by 'stay_texture_change_time' column):
        Each row may contain a scalar zone-entry timestamp in
        'stay_texture_change_time'.  Any entry whose time matches a timestamp
        in 'zone_re_entry_time' (within 0.05 s) is excluded as a re-entry.

    Old hallway format (fallback):
        Each row stores lists in 'texture_history' and 'texture_change_time'.
        Every index where texture == 'assets/reward_mean100.jpg' yields one
        zone-entry timestamp.

    Parameters
    ----------
    trial_log_df : pd.DataFrame
        Trial log loaded from a *_trial_log.csv file.

    Returns
    -------
    list[float] : Sorted zone entry timestamps (seconds), re-entries excluded.
    """
    # ── New hallway format ────────────────────────────────────────────────────
    if 'stay_texture_change_time' in trial_log_df.columns:
        raw_times = pd.to_numeric(
            trial_log_df['stay_texture_change_time'], errors='coerce'
        ).dropna().values

        # Build re-entry set (values can be scalars, lists, or string-encoded lists)
        re_entry_times_set = set()
        if 'zone_re_entry_time' in trial_log_df.columns:
            for val in trial_log_df['zone_re_entry_time']:
                if pd.isna(val) or val == '':
                    continue
                try:
                    if isinstance(val, str) and val.strip():
                        parsed = ast.literal_eval(val)
                        if isinstance(parsed, (list, tuple)):
                            for t in parsed:
                                if pd.notna(t):
                                    re_entry_times_set.add(float(t))
                        else:
                            re_entry_times_set.add(float(parsed))
                    elif not isinstance(val, str):
                        re_entry_times_set.add(float(val))
                except (ValueError, TypeError):
                    pass

        # Filter out re-entries (tolerance 0.05 s, matching the source code)
        if re_entry_times_set:
            re_entry_arr = np.array(sorted(re_entry_times_set))
            zone_entry_times = [
                float(t) for t in raw_times
                if np.min(np.abs(re_entry_arr - t)) > 0.05
            ]
        else:
            zone_entry_times = [float(t) for t in raw_times if t > 0]

        return sorted(zone_entry_times)

    # ── Old hallway format (fallback) ─────────────────────────────────────────
    zone_entry_times = []
    for _, log_row in trial_log_df.iterrows():
        texture_hist  = safe_literal_eval(log_row.get('texture_history',  '[]'))
        texture_times = safe_literal_eval(log_row.get('texture_change_time', '[]'))
        for i, texture in enumerate(texture_hist):
            if texture == "assets/reward_mean100.jpg" and i < len(texture_times):
                try:
                    t = float(texture_times[i])
                except (TypeError, ValueError):
                    continue
                if not np.isnan(t) and t > 0:
                    zone_entry_times.append(t)
    return sorted(zone_entry_times)


def _extract_punish_zone_entry_times(trial_log_df):
    """Return sorted list of punishment zone entry timestamps from a trial log.

    Handles two trial-log formats:

    New hallway format (detected by 'stay_punish_texture_change_time' column):
        Each row may contain a scalar zone-entry timestamp.

    Old hallway / scalar-per-row format (fallback, uses 'texture_history' column):
        Supports both list-per-row encoding (where texture_history stores a
        stringified Python list) and scalar-per-row encoding (where each row
        is a single texture encounter).  Any row whose resolved texture equals
        'assets/punish_mean100.jpg' contributes its texture_change_time.

    Parameters
    ----------
    trial_log_df : pd.DataFrame
        Trial log loaded from a *_trial_log.csv file.

    Returns
    -------
    list[float] : Sorted punish zone entry timestamps (seconds).
    """
    # ── New hallway format (dedicated column) ─────────────────────────────────
    if 'stay_punish_texture_change_time' in trial_log_df.columns:
        raw_times = pd.to_numeric(
            trial_log_df['stay_punish_texture_change_time'], errors='coerce'
        ).dropna().values
        return sorted(float(t) for t in raw_times if t > 0)

    # ── Old / scalar-per-row format (fallback via texture_history) ────────────
    if 'texture_history' not in trial_log_df.columns:
        return []

    zone_entry_times = []
    # Determine which timestamp column to use for this trial log
    _time_col = ('texture_change_time'
                 if 'texture_change_time' in trial_log_df.columns
                 else 'stay_texture_change_time')
    for _, log_row in trial_log_df.iterrows():
        raw_hist = log_row.get('texture_history', '')
        raw_time = log_row.get(_time_col, '')
        if pd.isna(raw_hist) or raw_hist == '':
            continue
        # safe_literal_eval returns the string itself when it cannot parse a
        # container, so a scalar-per-row entry returns the texture name directly.
        parsed_hist = safe_literal_eval(raw_hist) if isinstance(raw_hist, str) else raw_hist
        if isinstance(parsed_hist, list):
            # List-per-row encoding (legacy old-hallway format)
            parsed_time = (safe_literal_eval(raw_time)
                           if isinstance(raw_time, str) else raw_time)
            if not isinstance(parsed_time, list):
                parsed_time = [parsed_time]
            for i, texture in enumerate(parsed_hist):
                if texture == 'assets/punish_mean100.jpg' and i < len(parsed_time):
                    try:
                        t = float(parsed_time[i])
                    except (TypeError, ValueError):
                        continue
                    if not np.isnan(t) and t > 0:
                        zone_entry_times.append(t)
        else:
            # Scalar-per-row encoding (each row is one texture encounter)
            if str(parsed_hist).strip() == 'assets/punish_mean100.jpg':
                try:
                    t = float(raw_time)
                except (TypeError, ValueError):
                    continue
                if not np.isnan(t) and t > 0:
                    zone_entry_times.append(t)

    return sorted(zone_entry_times)


def _build_epoch_matrix(time_array, data_array, event_times,
                        window_s=EPOCH_WINDOW_S,
                        canonical_time=EPOCH_CANONICAL_TIME):
    """Extract ±window_s windows around each event and resample to a canonical grid.

    Parameters
    ----------
    time_array     : np.ndarray (1-D) — uniformly sampled time axis (seconds).
    data_array     : np.ndarray (1-D) — signal values, same length as time_array.
    event_times    : sequence of float — event timestamps (seconds).
    window_s       : float — half-window size in seconds.
    canonical_time : np.ndarray (1-D) — target time axis for every window.

    Returns
    -------
    np.ndarray of shape (n_events, len(canonical_time)), dtype float64.
    Values outside the recorded range are NaN (events near session edges).
    Returns None if event_times is empty.
    """
    if len(event_times) == 0:
        return None
    n_pts = len(canonical_time)
    rows  = []
    for t0 in event_times:
        mask     = (time_array >= t0 - window_s) & (time_array <= t0 + window_s)
        seg_time = time_array[mask]
        seg_data = data_array[mask].astype(float)
        if len(seg_time) < 2:
            rows.append(np.full(n_pts, np.nan))
            continue
        rel_time = seg_time - t0
        row = np.interp(canonical_time, rel_time, seg_data, left=np.nan, right=np.nan)
        rows.append(row)
    return np.array(rows, dtype=float)


def _build_lick_epoch_matrix(lick_event_times, event_times,
                              window_s=EPOCH_WINDOW_S,
                              canonical_time=LICK_EPOCH_TIME,
                              bin_s=0.5):
    """Build an epoch matrix of raw lick counts aligned to zone entry times.

    For each zone entry in *event_times* and each point on the canonical time
    axis, counts the lick timestamps that fall within a sliding window of
    *bin_s* seconds centred on that time point.  Returns raw counts (integers
    as float64) — NOT divided by bin width — so the y-axis is in licks per
    bin (licks per 500 ms window by default).

    Parameters
    ----------
    lick_event_times : np.ndarray (1-D) — absolute timestamps of detected lick events.
    event_times      : sequence of float — zone-entry timestamps (absolute seconds).
    window_s         : float — half-window size in seconds (epoch span).
    canonical_time   : np.ndarray (1-D) — target time axis (relative to each event).
    bin_s            : float — sliding window width in seconds (default 0.5 s).

    Returns
    -------
    np.ndarray of shape (n_events, len(canonical_time)), dtype float64, in licks/bin.
    Returns None if lick_event_times or event_times is empty.
    """
    if len(lick_event_times) == 0 or len(event_times) == 0:
        return None
    half = bin_s / 2.0
    rows = []
    for t0 in event_times:
        t_abs = canonical_time + t0          # absolute time axis for this epoch
        rate = np.array(
            [np.sum((lick_event_times >= (ta - half)) & (lick_event_times < (ta + half)))
             for ta in t_abs],
            dtype=float,
        )                                    # raw lick count per bin (no division by bin width)
        rows.append(rate)
    return np.array(rows, dtype=float)


def _filter_event_times_by_gaps(event_times, gap_info, window_s=EPOCH_WINDOW_S):
    """Return only those event times whose ±window_s epoch does not overlap a gap.

    Used to exclude trials from capacitive/lick epoch plots and latency
    calculations when the capacitive file has been trimmed and a data gap
    falls inside the analysis window.

    Parameters
    ----------
    event_times : sequence of float — event timestamps in seconds.
    gap_info    : dict returned by detect_capacitive_gaps.
    window_s    : float — half-window size in seconds (default = EPOCH_WINDOW_S).

    Returns
    -------
    list of float — event times whose epoch is fully within a continuous data segment.
    """
    if not gap_info['has_gaps'] or len(event_times) == 0:
        return list(event_times)
    gaps = gap_info['gaps']  # list of (t_start, t_end, duration)
    return [
        t0 for t0 in event_times
        if not any(g[0] < t0 + window_s and g[1] > t0 - window_s for g in gaps)
    ]


# ── RV-cohort helpers ─────────────────────────────────────────────────────────

def _is_rv_cohort(mouse_name: str) -> bool:
    """Return True if *mouse_name* belongs to the RV cohort.

    RV mice are identified by the naming prefix 'RV' (case-insensitive),
    e.g. RV1, RV2, RV10.  Their reward delivery time relative to zone entry
    is variable (not the fixed 0.65 s used for other cohorts), so downstream
    calculations must use the actual matched delivery time instead.
    """
    return str(mouse_name).upper().startswith('RV')


def _match_rewards_to_zones(trial_log_df, max_window_s: float = 30.0):
    """Match each reward delivery event to its preceding reward zone entry.

    Used for cohorts (e.g. RV) where reward delivery delay is variable rather
    than a fixed offset from zone entry.  For each ``reward_event`` (and
    ``hits_event``) timestamp, the function finds the most recent
    ``stay_texture_change_time`` that preceded it within *max_window_s* and
    pairs them.

    The zone-entry times are extracted with the same logic as
    :func:`_extract_reward_zone_entry_times` (re-entry filtering included), so
    ``test_matching_logic`` remains fully valid as an independent diagnostic.

    Parameters
    ----------
    trial_log_df : pd.DataFrame
        Trial log loaded from a ``*_trial_log.csv`` file.
    max_window_s : float
        Maximum allowed zone-entry → reward-delivery interval (seconds).
        Pairs with a delay outside this range are discarded.

    Returns
    -------
    list of (zone_entry_time: float, reward_delivery_time: float)
        Sorted by zone entry time.  Empty list if required columns are absent
        or no matches are found.
    """
    if 'reward_event' not in trial_log_df.columns:
        return []

    zone_times = np.array(_extract_reward_zone_entry_times(trial_log_df))
    if len(zone_times) == 0:
        return []

    reward_times = pd.to_numeric(
        trial_log_df['reward_event'], errors='coerce').dropna().values
    hits_times = (pd.to_numeric(trial_log_df['hits_event'], errors='coerce').dropna().values
                  if 'hits_event' in trial_log_df.columns else np.array([]))
    all_delivery_times = np.sort(np.concatenate([reward_times, hits_times]))

    if _DIAGNOSTIC_MODE:
        print(f"    [MATCH] _match_rewards_to_zones: {len(zone_times)} zone entries, "
              f"{len(all_delivery_times)} reward delivery events")

    matched = []
    unmatched = []
    for r_t in all_delivery_times:
        # All zone entries that precede this reward within the allowed window
        candidates = zone_times[(zone_times < r_t) & (r_t - zone_times <= max_window_s)]
        if len(candidates) == 0:
            if _DIAGNOSTIC_MODE:
                unmatched.append(r_t)
            continue
        # Pair with the most recent preceding zone entry
        z_t = float(candidates[-1])
        matched.append((z_t, float(r_t)))
        if _DIAGNOSTIC_MODE:
            print(f"      reward@{r_t:.3f}s  ←  zone@{z_t:.3f}s  (delay {r_t - z_t:.3f}s)")

    if _DIAGNOSTIC_MODE and unmatched:
        print(f"      [MATCH] {len(unmatched)} reward(s) had NO preceding zone within {max_window_s}s:")
        for _u in unmatched:
            print(f"        reward@{_u:.3f}s — skipped")
    if _DIAGNOSTIC_MODE:
        print(f"    [MATCH] → {len(matched)} matched pairs")

    return sorted(matched, key=lambda x: x[0])


# ── Plot selection ────────────────────────────────────────────────────────────
_ALL_PLOT_KEYS = {
    'speed', 'sensitivity', 'lick_count', 'reward_count',
    'false_alarms', 'correct_rejections', 'specificity', 'dprime',
    'avg_reward', 'sex_reward', 'avg_sex_speed',
    'distance', 'sex_distance', 'condition_distance',
    'condition_distance_bar', 'total_distance_bar',
    'condition_reward', 'condition_speed', 'condition_lick', 'condition_bar', 'condition_speed_bar',
    'levels', 'level_speed', 'level_speed_condition',
    'avg_lick_rate', 'sex_lick_rate', 'condition_lick_rate', 'condition_lick_bar',
    'level_lick', 'level_lick_condition',
    'level_dist', 'level_dist_condition', 'level_dist_condition_excl_last',
    'bout_count', 'avg_bout_count', 'condition_bout_count', 'condition_bout_count_bar',
    'rewards_per_bout', 'condition_rewards_per_bout', 'condition_rewards_per_bout_bar',
    'weekday_reward_bar', 'weekday_reward_bar_condition',
    'first_lick_latency', 'condition_first_lick_latency', 'condition_first_lick_latency_bar',
    'level_bout', 'level_bout_condition',
    'bout_avg_speed', 'condition_bout_avg_speed', 'condition_bout_avg_speed_bar',
    'bout_avg_dist',  'condition_bout_avg_dist',  'condition_bout_avg_dist_bar',
    'level_bout_avg_speed', 'level_bout_avg_speed_condition',
    'level_bout_avg_dist',  'level_bout_avg_dist_condition',
    'last_level_bar',
    'level_survivor',
    'time_to_level2',
    'lick_after_reward_prop', 'lick_after_reward_prop_bar',
    'epoch_reward_speed', 'epoch_reward_cap',
    'epoch_reward_speed_sess', 'epoch_reward_cap_sess',
    'epoch_reward_speed_early_late', 'epoch_reward_cap_early_late',
    'epoch_reward_speed_early_late_ev', 'epoch_reward_cap_early_late_ev',
    'epoch_reward_speed_early_late_ev_clean', 'epoch_reward_cap_early_late_ev_clean',
    'epoch_reward_speed_sess_clean', 'epoch_reward_cap_sess_clean',
    'epoch_reward_speed_early_late_clean', 'epoch_reward_cap_early_late_clean',
    'epoch_punish_speed', 'epoch_punish_cap',
    'epoch_punish_speed_sess', 'epoch_punish_cap_sess',
    'epoch_punish_speed_sess_clean', 'epoch_punish_cap_sess_clean',
    'sex_speed', 'sex_distance_indiv', 'sex_reward_indiv',
    'epoch_reward_speed_sess_sex', 'epoch_reward_cap_sess_sex',
    'epoch_punish_speed_sess_sex', 'epoch_punish_cap_sess_sex',
    'epoch_reward_speed_pre_post',
    'epoch_reward_speed_diff',
    'epoch_reward_cap_pre_post',
    'epoch_reward_cap_diff',
    'epoch_reward_speed_pre_post_entry',
    'epoch_reward_speed_diff_entry',
    'epoch_reward_speed_pre_post_entry_1s',
    'epoch_reward_speed_diff_entry_1s',
    'epoch_reward_lick_count_sess',
    'epoch_reward_lick_count_sess_clean',
    'epoch_punish_lick_count_sess',
    'epoch_punish_speed_pre_post',
    'epoch_punish_speed_diff',
    'epoch_punish_speed_pre_post_entry',
    'epoch_punish_speed_diff_entry',
    'epoch_punish_cap_pre_post',
    'epoch_punish_cap_diff',
    'epoch_punish_cap_pre_post_entry',
    'epoch_punish_cap_diff_entry',
    'lick_reward_ratio',
    'condition_lick_reward_ratio',
    'condition_lick_reward_ratio_bar',
    'condition_punish_zone_pct_bar',
    'expl_lick_reward_ratio_distfit',
    'expl_speed_histogram',
    'expl_speed_distfit',
    'expl_speed_boxplot',
    'expl_speed_rm_anova_resid',
    'expl_cap_histogram',
    'expl_cap_boxplot',
    'expl_cap_rm_anova_resid',
    'expl_cap_distfit',
    'expl_lick_distfit',
    'expl_lick_boxplot',
    'expl_lick_rm_anova_resid',
    'expl_lick_rate_distfit',
}

_PLOT_LABELS = [
    ('speed',               'Individual: Average speed over time'),
    ('sensitivity',         'Individual: Sensitivity over time'),
    ('lick_count',          'Individual: Lick count over time'),
    ('reward_count',        'Individual: Reward count over time'),
    ('lick_reward_ratio',       'Individual: Lick count / Reward count ratio over time'),
    ('condition_lick_reward_ratio', 'Condition: Lick count / Reward count ratio over time by starting condition'),
    ('condition_lick_reward_ratio_bar', 'Condition: Lick count / Reward count ratio — collapsed bar chart (one avg per mouse)'),
    ('condition_punish_zone_pct_bar',   'Condition: % punishment zones across all sessions — collapsed bar chart (one value per mouse)'),

    ('false_alarms',        'Individual: False alarms over time'),
    ('correct_rejections',  'Individual: Correct rejections over time'),
    ('specificity',         'Individual: Specificity over time'),
    ('dprime',              "Individual: d' over time"),
    ('avg_reward',          'Aggregate: Average reward rate across all mice'),
    ('sex_reward',          'Aggregate: Sex-specific average reward rate'),
    ('avg_sex_speed',      'Aggregate: Sex-specific average speed'),
    ('distance',            'Individual: Total distance per session (m)'),
    ('bout_count',               'Individual: Locomotion bout count per session'),
    ('avg_bout_count',           'Aggregate: Average bout count across all mice'),
    ('condition_bout_count',     'Condition: Bout count over time by condition'),
    ('condition_bout_count_bar', 'Condition: Average bout count — collapsed bar chart'),
    ('rewards_per_bout',          'Individual: Average rewards per locomotion bout per session'),
    ('condition_rewards_per_bout', 'Condition: Average rewards per bout over time by starting condition'),
    ('condition_rewards_per_bout_bar', 'Condition: Average rewards per locomotion bout — collapsed bar chart (one avg per mouse)'),
    ('weekday_reward_bar',           'Weekday: Average reward count by training weekday — all mice pooled'),
    ('weekday_reward_bar_condition', 'Weekday: Average reward count by training weekday — split by starting condition'),
    ('first_lick_latency',               'Individual: Average first-lick latency after reward delivery per session (s)'),
    ('condition_first_lick_latency',     'Condition: Average first-lick latency after reward delivery over time by starting condition'),
    ('condition_first_lick_latency_bar', 'Condition: Average first-lick latency after reward delivery — collapsed bar chart (one avg per mouse)'),
    ('lick_after_reward_prop',     'Condition: Proportion of reward deliveries with licks in 2 s post-delivery window — line plot over time by condition'),
    ('lick_after_reward_prop_bar', 'Condition: Proportion of reward deliveries with licks in 2 s post-delivery window — collapsed bar chart (one avg per mouse)'),
    ('sex_distance',        'Aggregate: Sex-specific average distance per session (m)'),
    ('condition_distance',  'Condition: Distance per session by starting condition (m)'),
    ('condition_distance_bar', 'Condition: Average distance per session — collapsed bar chart (m)'),
    ('total_distance_bar',    'Condition: Total distance per mouse — collapsed bar chart (m)'),
    ('avg_lick_rate',       'Aggregate: Average lick rate across all mice'),
    ('sex_lick_rate',       'Aggregate: Sex-specific average lick rate'),
    ('condition_reward',    'Condition: Reward rate over time (line)'),
    ('condition_speed',     'Condition: Speed over time'),
    ('condition_lick',      'Condition: Lick count over time'),
    ('condition_lick_rate', 'Condition: Lick rate over time (line)'),
    ('condition_bar',       'Condition: Reward rate — collapsed bar chart'),
    ('condition_speed_bar', 'Condition: Average speed — collapsed bar chart'),
    ('condition_lick_bar',  'Condition: Lick rate — collapsed bar chart'),
    ('levels',              'Level: Reward rate by level (requires transitions CSV)'),
    ('level_speed',           'Level: Average speed by level — collapsed (requires transitions CSV)'),
    ('level_speed_condition', 'Level: Average speed by level — by condition (requires transitions CSV)'),
    ('level_lick',            'Level: Average lick rate by level — collapsed (requires transitions CSV)'),
    ('level_lick_condition',  'Level: Average lick rate by level — by condition (requires transitions CSV)'),
    ('level_dist',            'Level: Distance traveled by level — collapsed (requires transitions CSV)'),
    ('level_dist_condition',  'Level: Distance traveled by level — by condition (requires transitions CSV)'),
    ('level_dist_condition_excl_last', 'Level: Distance traveled by level — by condition, last level excluded (requires transitions CSV)'),
    ('level_bout',           'Level: Locomotion bout count by level — collapsed (requires transitions CSV)'),
    ('level_bout_condition', 'Level: Locomotion bout count by level — by condition (requires transitions CSV)'),
    ('level_bout_avg_speed',           'Level: Avg speed per bout by level — collapsed (requires transitions CSV)'),
    ('level_bout_avg_speed_condition', 'Level: Avg speed per bout by level — by condition (requires transitions CSV)'),
    ('level_bout_avg_dist',            'Level: Avg distance per bout by level — collapsed (requires transitions CSV)'),
    ('level_bout_avg_dist_condition',  'Level: Avg distance per bout by level — by condition (requires transitions CSV)'),
    ('last_level_bar',                 'Level: Final (last day) level reached per mouse — collapsed bar chart by condition (requires transitions CSV)'),
    ('level_survivor',                 'Level: Survivor plot — proportion of mice that experienced each level, by condition (requires transitions CSV)'),
    ('time_to_level2',                 'Level: Cumulative time (min) until first level 1→2 transition — bar chart by condition (requires transitions CSV)'),
    ('bout_avg_speed',               'Individual: Average speed per locomotion bout'),
    ('condition_bout_avg_speed',     'Condition: Average speed per bout over time'),
    ('condition_bout_avg_speed_bar', 'Condition: Average speed per bout — collapsed bar chart'),
    ('bout_avg_dist',                'Individual: Average distance per locomotion bout'),
    ('condition_bout_avg_dist',      'Condition: Average distance per bout over time'),
    ('condition_bout_avg_dist_bar',  'Condition: Average distance per bout — collapsed bar chart'),
    ('epoch_reward_speed',      'Epoch: Treadmill speed — event-averaged, aligned to reward zone entry (per-mouse + condition)'),
    ('epoch_reward_cap',        'Epoch: Capacitive value — event-averaged, aligned to reward zone entry (per-mouse + condition)'),
    ('epoch_reward_speed_sess', 'Epoch: Treadmill speed — session-averaged, aligned to reward zone entry (per-mouse + condition)'),
    ('epoch_reward_cap_sess',   'Epoch: Capacitive value — session-averaged, aligned to reward zone entry (per-mouse + condition)'),
    ('epoch_reward_speed_early_late',    'Epoch: Treadmill speed — early vs late sessions, session-averaged, aligned to reward zone entry (per-mouse + condition)'),
    ('epoch_reward_cap_early_late',      'Epoch: Capacitive value — early vs late sessions, session-averaged, aligned to reward zone entry (per-mouse + condition)'),
    ('epoch_reward_speed_early_late_ev', 'Epoch: Treadmill speed — early vs late sessions, event-averaged, aligned to reward zone entry (per-mouse + condition)'),
    ('epoch_reward_cap_early_late_ev',   'Epoch: Capacitive value — early vs late sessions, event-averaged, aligned to reward zone entry (per-mouse + condition)'),
    ('epoch_reward_speed_early_late_ev_clean', 'Epoch: Treadmill speed — early vs late sessions, event-averaged, by condition only (no individual traces)'),
    ('epoch_reward_cap_early_late_ev_clean',   'Epoch: Capacitive value — early vs late sessions, event-averaged, by condition only (no individual traces)'),
    ('epoch_reward_speed_sess_clean',        'Epoch: Treadmill speed — session-averaged, by condition only (no individual traces)'),
    ('epoch_reward_cap_sess_clean',          'Epoch: Capacitive value — session-averaged, by condition only (no individual traces)'),
    ('epoch_reward_speed_early_late_clean',  'Epoch: Treadmill speed — early vs late sessions, session-averaged, by condition only (no individual traces)'),
    ('epoch_reward_cap_early_late_clean',    'Epoch: Capacitive value — early vs late sessions, session-averaged, by condition only (no individual traces)'),
    ('epoch_punish_speed',           'Epoch: Treadmill speed — event-averaged, aligned to punishment zone entry (per-mouse + condition)'),
    ('epoch_punish_cap',             'Epoch: Capacitive value — event-averaged, aligned to punishment zone entry (per-mouse + condition)'),
    ('epoch_punish_speed_sess',      'Epoch: Treadmill speed — session-averaged, aligned to punishment zone entry (per-mouse + condition)'),
    ('epoch_punish_cap_sess',        'Epoch: Capacitive value — session-averaged, aligned to punishment zone entry (per-mouse + condition)'),
    ('epoch_punish_speed_sess_clean','Epoch: Treadmill speed — punishment zone, session-averaged, by condition only (no individual traces)'),
    ('epoch_punish_cap_sess_clean',  'Epoch: Capacitive value — punishment zone, session-averaged, by condition only (no individual traces)'),
    ('sex_speed',          'Individual: Average speed over time — colored by sex'),
    ('sex_distance_indiv', 'Individual: Total distance per session — colored by sex (m)'),
    ('sex_reward_indiv',   'Individual: Reward rate per session — colored by sex (rewards/min)'),
    ('epoch_reward_speed_sess_sex', 'Epoch: Treadmill speed — session-averaged, aligned to reward zone entry (by sex)'),
    ('epoch_reward_cap_sess_sex',   'Epoch: Capacitive value — session-averaged, aligned to reward zone entry (by sex)'),
    ('epoch_punish_speed_sess_sex', 'Epoch: Treadmill speed — punishment zone, session-averaged, by sex'),
    ('epoch_punish_cap_sess_sex',   'Epoch: Capacitive value — punishment zone, session-averaged, by sex'),
    ('epoch_reward_speed_pre_post',  'Epoch: Speed — pre- vs post-reward delivery bar chart, by condition (0–0.65 s vs 0.65–1.3 s)'),
    ('epoch_reward_speed_diff',      'Epoch: Speed — post-minus-pre-reward difference bar chart, by condition (positive = faster after reward)'),
    ('epoch_reward_cap_pre_post',    'Epoch: Capacitive (z-scored) — pre- vs post-reward delivery bar chart, by condition (0–0.65 s vs 0.65–1.3 s)'),
    ('epoch_reward_cap_diff',        'Epoch: Capacitive (z-scored) — post-minus-pre-reward difference bar chart, by condition (Mann-Whitney U)'),
    ('epoch_reward_speed_pre_post_entry', 'Epoch: Reward zone speed — 0.65 s pre- vs 0.65 s post-zone entry bar chart, by condition'),
    ('epoch_reward_speed_diff_entry',     'Epoch: Reward zone speed — post-minus-pre zone entry difference bar chart (0.65 s windows), by condition'),
    ('epoch_reward_speed_pre_post_entry_1s', 'Epoch: Reward zone speed — 1 s pre- vs 1 s post-zone entry bar chart, by condition'),
    ('epoch_reward_speed_diff_entry_1s',     'Epoch: Reward zone speed — post-minus-pre zone entry difference bar chart (1 s windows), by condition'),
    ('epoch_reward_lick_count_sess',       'Epoch: Lick count — session-averaged, aligned to reward zone entry (per-mouse + condition)'),
    ('epoch_reward_lick_count_sess_clean',  'Epoch: Lick count — session-averaged, aligned to reward zone entry, condition average only (no individual traces)'),
    ('epoch_punish_lick_count_sess',      'Epoch: Lick count — session-averaged, aligned to punishment zone entry (per-mouse + condition)'),
    ('epoch_punish_speed_pre_post',  'Epoch: Punishment zone speed — pre vs post 0.65 s cutoff bar chart, by condition (0–0.65 s vs 0.65–1.3 s)'),
    ('epoch_punish_speed_diff',      'Epoch: Punishment zone speed — post-minus-pre 0.65 s cutoff difference bar chart, by condition'),
    ('epoch_punish_speed_pre_post_entry', 'Epoch: Punishment zone speed — 1 s pre- vs 1 s post-zone entry bar chart, by condition'),
    ('epoch_punish_speed_diff_entry',     'Epoch: Punishment zone speed — post-minus-pre zone entry difference bar chart (1 s windows), by condition'),
    ('epoch_punish_cap_pre_post',    'Epoch: Punishment zone capacitive (z-scored) — pre vs post 0.65 s cutoff bar chart, by condition (0–0.65 s vs 0.65–1.3 s)'),
    ('epoch_punish_cap_diff',        'Epoch: Punishment zone capacitive (z-scored) — post-minus-pre 0.65 s cutoff difference bar chart, by condition (Mann-Whitney U)'),
    ('epoch_punish_cap_pre_post_entry', 'Epoch: Punishment zone capacitive (z-scored) — 1 s pre- vs 1 s post-zone entry bar chart, by condition'),
    ('epoch_punish_cap_diff_entry',     'Epoch: Punishment zone capacitive (z-scored) — post-minus-pre zone entry difference bar chart (1 s windows), by condition (Mann-Whitney U)'),
]

_EXPL_PLOT_LABELS = [
    ('expl_lick_reward_ratio_distfit',  'Exploratory: Per-mouse mean lick/reward ratio — KDE + Shapiro-Wilk normality test + Q-Q plot'),
    ('expl_speed_histogram',       'Exploratory: Speed distribution — histogram (all session speeds + per-mouse means, Shapiro-Wilk)'),
    ('expl_speed_distfit',         'Exploratory: Average speed distribution fit — Normal, Log-normal, Gamma (histogram + Q-Q + AIC comparison)'),
    ('expl_speed_boxplot',         'Exploratory: Speed distribution — box and whisker (per-mouse + overall)'),
    ('expl_speed_rm_anova_resid',  'Exploratory: Repeated-measures ANOVA residuals — Q-Q plot, histogram, residuals vs fitted (condition × session)'),
    ('expl_cap_histogram',         'Exploratory: Z-scored mean capacitive sensor value distribution — histogram (all session values + per-mouse means, Shapiro-Wilk)'),
    ('expl_cap_boxplot',           'Exploratory: Z-scored mean capacitive sensor value distribution — box and whisker (per-mouse + overall)'),
    ('expl_cap_rm_anova_resid',    'Exploratory: Capacitive RM ANOVA residuals — Q-Q plot, histogram, residuals vs fitted (condition × session)'),
    ('expl_cap_distfit',           'Exploratory: Capacitive sensor value distribution fit — Normal, Log-normal, Gamma (histogram + Q-Q + AIC comparison)'),
    ('expl_lick_distfit',          'Exploratory: Lick count Poisson vs Negative Binomial distribution fit — rootogram, Q-Q, mean–variance, AIC/χ² comparison'),
    ('expl_lick_boxplot',          'Exploratory: Raw lick count distribution — box and whisker (per-mouse + overall)'),
    ('expl_lick_rm_anova_resid',   'Exploratory: Raw lick count RM ANOVA residuals — Q-Q plot, histogram, residuals vs fitted (DV = log(1+count), condition × session)'),
    ('expl_lick_rate_distfit',     'Exploratory: Average lick rate distribution fit — Normal, Log-normal, Gamma (histogram + Q-Q + AIC comparison)'),
]


def _ask_plot_selection(root, labels=None, title='Select Plots to Generate'):
    """Show a scrollable checkbox dialog and return the frozenset of selected plot keys."""
    dialog = tk.Toplevel(root)
    dialog.title(title)
    dialog.resizable(True, True)
    dialog.grab_set()

    # ── Header (outside the scroll area) ─────────────────────────────────────
    tk.Label(dialog, text='Select which plots to generate:',
             font=('Arial', 11, 'bold')).pack(anchor='w', padx=14, pady=(12, 4))

    # ── Scrollable canvas + inner frame ──────────────────────────────────────
    container = tk.Frame(dialog)
    container.pack(fill='both', expand=True, padx=14, pady=4)

    canvas = tk.Canvas(container, borderwidth=0, highlightthickness=0)
    scrollbar = tk.Scrollbar(container, orient='vertical', command=canvas.yview)
    canvas.configure(yscrollcommand=scrollbar.set)

    scrollbar.pack(side='right', fill='y')
    canvas.pack(side='left', fill='both', expand=True)

    inner = tk.Frame(canvas)
    inner_id = canvas.create_window((0, 0), window=inner, anchor='nw')

    def _on_inner_configure(event):
        canvas.configure(scrollregion=canvas.bbox('all'))

    def _on_canvas_configure(event):
        canvas.itemconfig(inner_id, width=event.width)

    inner.bind('<Configure>', _on_inner_configure)
    canvas.bind('<Configure>', _on_canvas_configure)

    # Mouse-wheel scrolling (cross-platform)
    def _on_mousewheel(event):
        if event.num == 4:          # Linux scroll up
            canvas.yview_scroll(-1, 'units')
        elif event.num == 5:        # Linux scroll down
            canvas.yview_scroll(1, 'units')
        else:                       # Windows / macOS
            canvas.yview_scroll(int(-1 * (event.delta / 120)), 'units')

    canvas.bind('<MouseWheel>', _on_mousewheel)
    canvas.bind('<Button-4>',   _on_mousewheel)
    canvas.bind('<Button-5>',   _on_mousewheel)
    inner.bind('<MouseWheel>',  _on_mousewheel)
    inner.bind('<Button-4>',    _on_mousewheel)
    inner.bind('<Button-5>',    _on_mousewheel)

    # ── Checkboxes ────────────────────────────────────────────────────────────
    if labels is None:
        labels = _PLOT_LABELS
    vars_ = {}
    for key, label in labels:
        var = tk.BooleanVar(value=True)
        vars_[key] = var
        cb = tk.Checkbutton(inner, text=label, variable=var, anchor='w')
        cb.pack(fill='x', padx=8, pady=1)
        cb.bind('<MouseWheel>', _on_mousewheel)
        cb.bind('<Button-4>',   _on_mousewheel)
        cb.bind('<Button-5>',   _on_mousewheel)

    # Cap the dialog height to 80 % of the screen
    dialog.update_idletasks()
    screen_h = dialog.winfo_screenheight()
    max_h    = int(screen_h * 0.80)
    natural_h = inner.winfo_reqheight() + 120   # header + buttons
    dialog_h  = min(natural_h, max_h)
    dialog.geometry(f'520x{dialog_h}')

    # ── Buttons (outside the scroll area) ────────────────────────────────────
    btn_frame = tk.Frame(dialog)
    btn_frame.pack(fill='x', padx=14, pady=(8, 12))

    def _select_all():
        for v in vars_.values():
            v.set(True)

    def _deselect_all():
        for v in vars_.values():
            v.set(False)

    tk.Button(btn_frame, text='Select All',   command=_select_all).pack(side='left',  padx=4)
    tk.Button(btn_frame, text='Deselect All', command=_deselect_all).pack(side='left', padx=4)
    tk.Button(btn_frame, text='OK', command=dialog.destroy,
              default='active').pack(side='right', padx=4)

    root.wait_window(dialog)
    return frozenset(key for key, var in vars_.items() if var.get())


def analyze_levels(data_files, transitions_csv_path, animal_conditions=None, selected_plots=None):
    """Analyze rewards/min for each level across all mice, split by starting condition.

    Uses a transitions CSV (produced by level_sorter.py) to slice each session's
    trial_log into per-level time windows.  Multi-session level visits (where a
    mouse ends a session mid-level and continues the next day) are aggregated
    before computing rpm: total rewards and total active time are summed across
    ALL sessions for a given (animal, level) pair, producing ONE rpm value per
    animal per level.  This prevents time gaps between sessions from penalising
    the rate calculation and ensures partial sessions at either end of a level
    are fully included.

    Boundary rules per session slice:
        start_ts = previous level's transition_ts in this session, or 0
        end_ts   = this level's transition_ts (completed) or capacitive
                   elapsed_time.max() (incomplete last level of session)

    Parameters
    ----------
    animal_conditions : dict[str, str] | None
        Mapping of animal_id -> starting_condition (e.g. {'CAH1': 'SC', 'CAH2': 'VF'}).
        When provided, the bar chart is split into grouped bars by condition.
    """
    # (animal_id, level) -> {'rewards': int, 'duration_min': float, 'condition': str}
    animal_level_accum: dict[tuple, dict] = {}
    # animal_id -> last level seen (updated each session; final value = potentially incomplete level)
    animal_last_level: dict[str, str] = {}
    # (animal_id, session_num) -> session length in seconds (for time-to-level2 calculation)
    animal_session_lengths: dict[tuple, float] = {}

    # Load transitions CSV -------------------------------------------------------
    if not transitions_csv_path:
        print("  [WARN] No transitions CSV provided — level plot will be empty.")
        return plt.figure(figsize=(15, 8)), None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None
    try:
        transitions_df = pd.read_csv(transitions_csv_path)
        transitions_df['date'] = pd.to_datetime(transitions_df['date'])
    except Exception as e:
        print(f"  [ERROR] Cannot read transitions CSV: {e}")
        return plt.figure(figsize=(15, 8)), None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None

    # Build lookup (animal_id, date_normalised) -> {trial_log, capacitive} ------
    session_lookup = {}
    for data_file in data_files:
        animal_id = os.path.basename(data_file).split('_')[0]
        try:
            df = pd.read_csv(data_file)
        except Exception:
            continue
        if 'date' in df.columns:
            df['date'] = pd.to_datetime(df['date'])
        for _, row in df.iterrows():
            date_key = pd.Timestamp(row['date']).normalize()
            session_lookup[(animal_id, date_key)] = {
                'trial_log':  row.get('trial_log'),
                'capacitive': row.get('capacitive'),
                'treadmill':  row.get('treadmill'),
            }

    # Accumulate rewards and time per (animal, level) across all sessions ---------
    for animal_id, animal_group in transitions_df.groupby('animal_id'):
        for session_num, session_group in animal_group.groupby('session_num'):
            session_group = session_group.reset_index(drop=True)
            date_key = pd.Timestamp(session_group.iloc[0]['date']).normalize()

            sess_info = session_lookup.get((animal_id, date_key))
            if sess_info is None:
                continue
            trial_log_path  = sess_info.get('trial_log')
            capacitive_path = sess_info.get('capacitive')
            treadmill_path  = sess_info.get('treadmill')

            if not isinstance(trial_log_path, str) or pd.isna(trial_log_path):
                continue
            try:
                tl = pd.read_csv(trial_log_path)
            except Exception as e:
                print(f"  [WARN] {animal_id} session {session_num}: cannot read trial_log — {e}")
                continue
            if 'reward_event' not in tl.columns or 'texture_change_time' not in tl.columns:
                continue

            # Session end time and lick event detection from capacitive data -----
            lick_event_times = np.array([])
            session_end_time = None
            if isinstance(capacitive_path, str) and not pd.isna(capacitive_path):
                try:
                    cap = pd.read_csv(capacitive_path)
                    if 'elapsed_time' in cap.columns:
                        session_end_time = float(cap['elapsed_time'].max())
                    if 'capacitive_value' in cap.columns:
                        cap_df = cap.copy()
                        cap_df['Time_sec'] = cap_df['elapsed_time']

                        # ── Gap detection ──────────────────────────────────
                        _gap_info = detect_capacitive_gaps(cap_df)
                        if _gap_info['has_gaps']:
                            print(f"  [GAP] {animal_id} session {session_num}: "
                                  f"{_gap_info['n_gaps']} gap(s) detected in capacitive file "
                                  f"(total trimmed: {_gap_info['total_gap_s']:.2f}s) — "
                                  + ", ".join(f"{g[0]:.2f}s–{g[1]:.2f}s" for g in _gap_info['gaps']))

                        kde_val = get_cached_kde(capacitive_path)
                        if kde_val is None:
                            kde_val = lda.compute_KDE(cap_df, 'capacitive_value')
                            cache_kde_value(capacitive_path, kde_val)
                        cap_df = lda.compute_KDE_normalizations(cap_df, 'capacitive_value', kde_val)
                        events_df, _ = lda.detect_events_above_threshold(
                            cap_df, 'capacitive_value', threshold=None)
                        mask_evt = events_df['capacitive_value_event'] == 1
                        lick_event_times = events_df.loc[mask_evt, 'Time_sec'].values.astype(float)
                except Exception:
                    pass
            if session_end_time is None:
                tc_vals = tl['texture_change_time'].dropna()
                session_end_time = float(tc_vals.max()) if len(tc_vals) else None

            # Load treadmill speed data for this session -----------------------
            treadmill_df = None
            if isinstance(treadmill_path, str) and not pd.isna(treadmill_path):
                try:
                    tm = pd.read_csv(treadmill_path)
                    if 'global_time' in tm.columns and 'speed' in tm.columns:
                        load_cols = ['global_time', 'speed']
                        if 'distance' in tm.columns:
                            load_cols.append('distance')
                        treadmill_df = tm[load_cols].dropna(subset=['global_time', 'speed'])
                        # Apply Butterworth low-pass (0.25 Hz, order 3) to full session speed
                        if len(treadmill_df) >= 15:
                            _fs_lvl = 1.0 / treadmill_df['global_time'].diff().median()
                            _b_lvl, _a_lvl = butter(3, 0.25 / (_fs_lvl / 2.0), btype='low')
                            treadmill_df = treadmill_df.copy()
                            treadmill_df['speed'] = filtfilt(_b_lvl, _a_lvl, treadmill_df['speed'].values)
                except Exception:
                    pass

            levels   = session_group['level'].tolist()
            trans_ts = session_group['transition_ts'].tolist()

            # start_ts for each level = end_ts of the previous level (or 0)
            start_times = [0.0]
            for ts in trans_ts[:-1]:
                start_times.append(float(ts) if pd.notna(ts) else float('inf'))

            # end_ts for each level = its own transition_ts, or session end time
            end_times = []
            for ts in trans_ts:
                if pd.notna(ts):
                    end_times.append(float(ts))
                elif session_end_time is not None:
                    end_times.append(session_end_time)
                else:
                    end_times.append(None)

            reward_events = tl['reward_event'].dropna().to_numpy(dtype=float)
            hits_events   = (tl['hits_event'].dropna().to_numpy(dtype=float)
                             if 'hits_event' in tl.columns else np.array([]))

            for i, level in enumerate(levels):
                start_t = start_times[i]
                end_t   = end_times[i]
                if end_t is None or start_t >= end_t:
                    continue
                duration_min = (end_t - start_t) / 60.0
                if duration_min <= 0:
                    continue
                count = int(np.sum((reward_events >= start_t) & (reward_events < end_t)))
                count += int(np.sum((hits_events   >= start_t) & (hits_events   < end_t)))

                condition = (animal_conditions or {}).get(animal_id, 'Unknown')
                key = (animal_id, level)
                if key not in animal_level_accum:
                    animal_level_accum[key] = {'rewards': 0, 'duration_min': 0.0,
                                               'condition': condition,
                                               'speed_sum': 0.0, 'speed_count': 0,
                                               'lick_count': 0, 'dist_sum': 0.0,
                                               'bout_count_lvl': 0,
                                               'bout_spd_sum': 0.0, 'bout_spd_cnt': 0,
                                               'bout_dist_sum': 0.0, 'bout_dist_cnt': 0}
                animal_level_accum[key]['rewards']      += count
                animal_level_accum[key]['duration_min'] += duration_min
                if treadmill_df is not None and len(treadmill_df) > 0:
                    mask = (treadmill_df['global_time'] >= start_t) & (treadmill_df['global_time'] < end_t)
                    lvl_speeds = treadmill_df.loc[mask, 'speed'].values / 10.0
                    lvl_times  = treadmill_df.loc[mask, 'global_time'].values
                    animal_level_accum[key]['speed_sum']   += float(np.sum(lvl_speeds))
                    animal_level_accum[key]['speed_count'] += len(lvl_speeds)
                    if len(lvl_times) >= 2:
                        _bouts = detect_locomotion_bouts(lvl_times, lvl_speeds)
                        animal_level_accum[key]['bout_count_lvl'] += len(_bouts)
                        _lvl_dist_arr = (
                            pd.to_numeric(treadmill_df.loc[mask, 'distance'], errors='coerce').values
                            if 'distance' in treadmill_df.columns else None
                        )
                        for _t0, _t1 in _bouts:
                            _bmask = (lvl_times >= _t0) & (lvl_times <= _t1)
                            if np.any(_bmask):
                                animal_level_accum[key]['bout_spd_sum'] += float(np.mean(lvl_speeds[_bmask]))
                                animal_level_accum[key]['bout_spd_cnt'] += 1
                            if _lvl_dist_arr is not None:
                                _dv = _lvl_dist_arr[_bmask]
                                _dv = _dv[~np.isnan(_dv)]
                                if len(_dv) >= 2:
                                    animal_level_accum[key]['bout_dist_sum'] += float(_dv[-1] - _dv[0])
                                    animal_level_accum[key]['bout_dist_cnt'] += 1
                if len(lick_event_times) > 0:
                    lick_mask = (lick_event_times >= start_t) & (lick_event_times < end_t)
                    animal_level_accum[key]['lick_count'] += int(np.sum(lick_mask))
                if treadmill_df is not None and 'distance' in treadmill_df.columns:
                    dist_col = pd.to_numeric(treadmill_df['distance'], errors='coerce')
                    dist_window = dist_col[
                        (treadmill_df['global_time'] >= start_t) &
                        (treadmill_df['global_time'] < end_t)
                    ].dropna()
                    if len(dist_window) >= 2:
                        animal_level_accum[key]['dist_sum'] += float(
                            dist_window.iloc[-1] - dist_window.iloc[0])
            # Record the last level of this session; final session's value = the incomplete level
            if levels:
                animal_last_level[animal_id] = levels[-1]
            # Record session length for time-to-level2 cumulative sum
            if session_end_time is not None:
                animal_session_lengths[(animal_id, session_num)] = float(session_end_time)

    # Collapse accumulators: one rpm, mean speed, and mean lick rate per (animal, level)
    # condition_level_data[condition][level] = [rpm_animal1, rpm_animal2, ...]
    condition_level_data: dict[str, dict[str, list[float]]] = {}
    condition_level_speed: dict[str, dict[str, list[float]]] = {}
    collapsed_level_speed: dict[str, list[float]] = {}
    condition_level_lick: dict[str, dict[str, list[float]]] = {}
    collapsed_level_lick: dict[str, list[float]] = {}
    condition_level_dist: dict[str, dict[str, list[float]]] = {}
    collapsed_level_dist: dict[str, list[float]] = {}
    condition_level_bout: dict[str, dict[str, list[float]]] = {}
    collapsed_level_bout: dict[str, list[float]] = {}
    condition_level_bout_avg_spd: dict[str, dict[str, list[float]]] = {}
    collapsed_level_bout_avg_spd: dict[str, list[float]] = {}
    condition_level_bout_avg_dist_lvl: dict[str, dict[str, list[float]]] = {}
    collapsed_level_bout_avg_dist_lvl: dict[str, list[float]] = {}
    for (animal_id, level), accum in animal_level_accum.items():
        condition = accum['condition']
        if accum['duration_min'] > 0:
            rpm = accum['rewards'] / accum['duration_min']
            condition_level_data.setdefault(condition, {}).setdefault(level, []).append(rpm)
        if accum['speed_count'] > 0:
            mean_spd = accum['speed_sum'] / accum['speed_count']
            condition_level_speed.setdefault(condition, {}).setdefault(level, []).append(mean_spd)
            collapsed_level_speed.setdefault(level, []).append(mean_spd)
            # bout count shares the same treadmill-data gate as speed
            bc = accum.get('bout_count_lvl', 0)
            condition_level_bout.setdefault(condition, {}).setdefault(level, []).append(float(bc))
            collapsed_level_bout.setdefault(level, []).append(float(bc))
        if accum.get('bout_spd_cnt', 0) > 0:
            avg_spd_bout = accum['bout_spd_sum'] / accum['bout_spd_cnt']
            condition_level_bout_avg_spd.setdefault(condition, {}).setdefault(level, []).append(avg_spd_bout)
            collapsed_level_bout_avg_spd.setdefault(level, []).append(avg_spd_bout)
        if accum.get('bout_dist_cnt', 0) > 0:
            avg_dist_bout_m = accum['bout_dist_sum'] / accum['bout_dist_cnt'] / 1000.0  # mm → m
            condition_level_bout_avg_dist_lvl.setdefault(condition, {}).setdefault(level, []).append(avg_dist_bout_m)
            collapsed_level_bout_avg_dist_lvl.setdefault(level, []).append(avg_dist_bout_m)
        if accum['lick_count'] > 0 and accum['duration_min'] > 0:
            lpm = accum['lick_count'] / accum['duration_min']
            condition_level_lick.setdefault(condition, {}).setdefault(level, []).append(lpm)
            collapsed_level_lick.setdefault(level, []).append(lpm)
        if accum.get('dist_sum', 0.0) > 0:
            dist_m = accum['dist_sum'] / 1000.0  # mm → m
            condition_level_dist.setdefault(condition, {}).setdefault(level, []).append(dist_m)
            collapsed_level_dist.setdefault(level, []).append(dist_m)

    # Per-condition distance excluding each mouse's last (potentially incomplete) level
    condition_level_dist_excl_last: dict[str, dict[str, list[float]]] = {}
    for (animal_id, level), accum in animal_level_accum.items():
        if level == animal_last_level.get(animal_id):
            continue  # skip last level — final session may be incomplete
        if accum.get('dist_sum', 0.0) > 0:
            condition = accum['condition']
            dist_m = accum['dist_sum'] / 1000.0  # mm → m
            condition_level_dist_excl_last.setdefault(condition, {}).setdefault(level, []).append(dist_m)

    # Sort levels ----------------------------------------------------------------
    def sort_key(x):
        if x.startswith('level_'):
            try:
                return (0, int(x.split('_')[1].split('.')[0]))
            except (ValueError, IndexError):
                pass
        return (1, x)

    all_levels = sorted(
        {lv for cond_data in condition_level_data.values() for lv in cond_data},
        key=sort_key,
    )
    conditions_sorted = sorted(condition_level_data.keys())
    n_conditions = len(conditions_sorted)
    n_levels     = len(all_levels)

    cond_color_map = {c: _condition_to_color(c) for c in conditions_sorted}

    # Grouped bar chart ----------------------------------------------------------
    level_reward_fig, ax = plt.subplots(figsize=(max(15, n_levels * 0.8), 8))

    bar_width   = 0.8 / max(n_conditions, 1)
    x_positions = np.arange(n_levels)

    for cond_idx, condition in enumerate(conditions_sorted):
        cond_data = condition_level_data[condition]
        means = []
        sems  = []
        ns    = []
        for level in all_levels:
            vals = cond_data.get(level, [])
            if vals:
                means.append(float(np.mean(vals)))
                sems.append(float(np.std(vals) / np.sqrt(len(vals))))
                ns.append(len(vals))
            else:
                means.append(np.nan)
                sems.append(np.nan)
                ns.append(0)

        offset = (cond_idx - (n_conditions - 1) / 2) * bar_width
        bars = ax.bar(
            x_positions + offset, means,
            width=bar_width, yerr=sems,
            capsize=4,
            color=cond_color_map[condition],
            label=condition,
            error_kw={'elinewidth': 1.2},
        )

        # n= annotations above each bar
        for xi, (m, n_val) in enumerate(zip(means, ns)):
            if n_val > 0 and not np.isnan(m):
                ax.text(
                    x_positions[xi] + offset,
                    m + (sems[xi] if not np.isnan(sems[xi]) else 0),
                    f'n={n_val}',
                    ha='center', va='bottom',
                    fontsize=7,
                    color=cond_color_map[condition],
                )

    ax.set_xticks(x_positions)
    ax.set_xticklabels(
        [lv.replace('level_', 'L').replace('.json', '') for lv in all_levels],
        rotation=45, ha='right',
    )
    ax.set_title('Average Rewards Per Minute by Level — by Starting Condition')
    ax.set_xlabel('Level')
    ax.set_ylabel('Rewards per Minute (Mean ± SEM)')
    ax.set_ylim(bottom=0)
    ax.tick_params(axis='both', direction='in')
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.legend(title='Starting Condition')
    level_reward_fig.tight_layout()

    if selected_plots is None:
        selected_plots = set(_ALL_PLOT_KEYS)

    # ── Collapsed speed by level (all mice) ───────────────────────────────────
    level_speed_collapsed_fig = None
    if 'level_speed' in selected_plots and collapsed_level_speed:
        all_levels_spd = sorted(collapsed_level_speed.keys(), key=sort_key)
        level_speed_collapsed_fig, ax_sc = plt.subplots(
            figsize=(max(15, len(all_levels_spd) * 0.8), 8))
        x_sc = np.arange(len(all_levels_spd))
        sc_means, sc_sems, sc_ns = [], [], []
        for lv in all_levels_spd:
            vals = collapsed_level_speed[lv]
            sc_means.append(float(np.mean(vals)))
            sc_sems.append(float(np.std(vals) / np.sqrt(len(vals))))
            sc_ns.append(len(vals))
        ax_sc.bar(x_sc, sc_means, yerr=sc_sems, capsize=4, color='steelblue',
                  error_kw={'elinewidth': 1.2})
        for xi, (m, n_val) in enumerate(zip(sc_means, sc_ns)):
            if n_val > 0 and not np.isnan(m):
                ax_sc.text(xi, m + (sc_sems[xi] if not np.isnan(sc_sems[xi]) else 0),
                           f'n={n_val}', ha='center', va='bottom', fontsize=7)
        ax_sc.set_xticks(x_sc)
        ax_sc.set_xticklabels(
            [lv.replace('level_', 'L').replace('.json', '') for lv in all_levels_spd],
            rotation=45, ha='right',
        )
        ax_sc.set_title('Average Speed by Level — All Mice')
        ax_sc.set_xlabel('Level')
        ax_sc.set_ylabel('Average Speed cm/s (Mean ± SEM)')
        ax_sc.set_ylim(bottom=0)
        ax_sc.tick_params(axis='both', direction='in')
        ax_sc.spines['top'].set_visible(False)
        ax_sc.spines['right'].set_visible(False)
        level_speed_collapsed_fig.tight_layout()

    # ── Speed by level split by condition ────────────────────────────────────
    level_speed_condition_fig = None
    if 'level_speed_condition' in selected_plots and condition_level_speed:
        all_levels_sc = sorted(
            {lv for cd in condition_level_speed.values() for lv in cd},
            key=sort_key,
        )
        conds_sc = sorted(condition_level_speed.keys())
        n_conds_sc = len(conds_sc)
        level_speed_condition_fig, ax_scc = plt.subplots(
            figsize=(max(15, len(all_levels_sc) * 0.8), 8))
        bw_sc = 0.8 / max(n_conds_sc, 1)
        x_scc = np.arange(len(all_levels_sc))
        for ci, cond in enumerate(conds_sc):
            cdata = condition_level_speed[cond]
            means_c, sems_c, ns_c = [], [], []
            for lv in all_levels_sc:
                vals = cdata.get(lv, [])
                if vals:
                    means_c.append(float(np.mean(vals)))
                    sems_c.append(float(np.std(vals) / np.sqrt(len(vals))))
                    ns_c.append(len(vals))
                else:
                    means_c.append(np.nan)
                    sems_c.append(np.nan)
                    ns_c.append(0)
            offset_c = (ci - (n_conds_sc - 1) / 2) * bw_sc
            ax_scc.bar(
                x_scc + offset_c, means_c,
                width=bw_sc, yerr=sems_c, capsize=4,
                color=cond_color_map.get(cond, 'gray'),
                label=cond,
                error_kw={'elinewidth': 1.2},
            )
            for xi, (m, n_val) in enumerate(zip(means_c, ns_c)):
                if n_val > 0 and not np.isnan(m):
                    ax_scc.text(
                        x_scc[xi] + offset_c,
                        m + (sems_c[xi] if not np.isnan(sems_c[xi]) else 0),
                        f'n={n_val}', ha='center', va='bottom', fontsize=7,
                        color=cond_color_map.get(cond, 'gray'),
                    )
        ax_scc.set_xticks(x_scc)
        ax_scc.set_xticklabels(
            [lv.replace('level_', 'L').replace('.json', '') for lv in all_levels_sc],
            rotation=45, ha='right',
        )
        ax_scc.set_title('Average Speed by Level — by Starting Condition')
        ax_scc.set_xlabel('Level')
        ax_scc.set_ylabel('Average Speed cm/s (Mean ± SEM)')
        ax_scc.set_ylim(bottom=0)
        ax_scc.tick_params(axis='both', direction='in')
        ax_scc.spines['top'].set_visible(False)
        ax_scc.spines['right'].set_visible(False)
        ax_scc.legend(title='Starting Condition')
        level_speed_condition_fig.tight_layout()

    # ── Collapsed lick rate by level (all mice) ───────────────────────────────
    level_lick_collapsed_fig = None
    if 'level_lick' in selected_plots and collapsed_level_lick:
        all_levels_lk = sorted(collapsed_level_lick.keys(), key=sort_key)
        level_lick_collapsed_fig, ax_lk = plt.subplots(
            figsize=(max(15, len(all_levels_lk) * 0.8), 8))
        x_lk = np.arange(len(all_levels_lk))
        lk_means, lk_sems, lk_ns = [], [], []
        for lv in all_levels_lk:
            vals = collapsed_level_lick[lv]
            lk_means.append(float(np.mean(vals)))
            lk_sems.append(float(np.std(vals) / np.sqrt(len(vals))))
            lk_ns.append(len(vals))
        ax_lk.bar(x_lk, lk_means, yerr=lk_sems, capsize=4, color='steelblue',
                  error_kw={'elinewidth': 1.2})
        for xi, (m, n_val) in enumerate(zip(lk_means, lk_ns)):
            if n_val > 0 and not np.isnan(m):
                ax_lk.text(xi, m + (lk_sems[xi] if not np.isnan(lk_sems[xi]) else 0),
                           f'n={n_val}', ha='center', va='bottom', fontsize=7)
        ax_lk.set_xticks(x_lk)
        ax_lk.set_xticklabels(
            [lv.replace('level_', 'L').replace('.json', '') for lv in all_levels_lk],
            rotation=45, ha='right',
        )
        ax_lk.set_title('Average Lick Rate by Level — All Mice')
        ax_lk.set_xlabel('Level')
        ax_lk.set_ylabel('Licks per Minute (Mean ± SEM)')
        ax_lk.set_ylim(bottom=0)
        ax_lk.tick_params(axis='both', direction='in')
        ax_lk.spines['top'].set_visible(False)
        ax_lk.spines['right'].set_visible(False)
        level_lick_collapsed_fig.tight_layout()

    # ── Lick rate by level split by condition ─────────────────────────────────
    level_lick_condition_fig = None
    if 'level_lick_condition' in selected_plots and condition_level_lick:
        all_levels_lkc = sorted(
            {lv for cd in condition_level_lick.values() for lv in cd},
            key=sort_key,
        )
        conds_lkc = sorted(condition_level_lick.keys())
        n_conds_lkc = len(conds_lkc)
        level_lick_condition_fig, ax_lkc = plt.subplots(
            figsize=(max(15, len(all_levels_lkc) * 0.8), 8))
        bw_lkc = 0.8 / max(n_conds_lkc, 1)
        x_lkc = np.arange(len(all_levels_lkc))
        for ci, cond in enumerate(conds_lkc):
            cdata = condition_level_lick[cond]
            means_lkc, sems_lkc, ns_lkc = [], [], []
            for lv in all_levels_lkc:
                vals = cdata.get(lv, [])
                if vals:
                    means_lkc.append(float(np.mean(vals)))
                    sems_lkc.append(float(np.std(vals) / np.sqrt(len(vals))))
                    ns_lkc.append(len(vals))
                else:
                    means_lkc.append(np.nan)
                    sems_lkc.append(np.nan)
                    ns_lkc.append(0)
            offset_lkc = (ci - (n_conds_lkc - 1) / 2) * bw_lkc
            ax_lkc.bar(
                x_lkc + offset_lkc, means_lkc,
                width=bw_lkc, yerr=sems_lkc, capsize=4,
                color=cond_color_map.get(cond, 'gray'),
                label=cond,
                error_kw={'elinewidth': 1.2},
            )
            for xi, (m, n_val) in enumerate(zip(means_lkc, ns_lkc)):
                if n_val > 0 and not np.isnan(m):
                    ax_lkc.text(
                        x_lkc[xi] + offset_lkc,
                        m + (sems_lkc[xi] if not np.isnan(sems_lkc[xi]) else 0),
                        f'n={n_val}', ha='center', va='bottom', fontsize=7,
                        color=cond_color_map.get(cond, 'gray'),
                    )
        ax_lkc.set_xticks(x_lkc)
        ax_lkc.set_xticklabels(
            [lv.replace('level_', 'L').replace('.json', '') for lv in all_levels_lkc],
            rotation=45, ha='right',
        )
        ax_lkc.set_title('Average Lick Rate by Level — by Starting Condition')
        ax_lkc.set_xlabel('Level')
        ax_lkc.set_ylabel('Licks per Minute (Mean ± SEM)')
        ax_lkc.set_ylim(bottom=0)
        ax_lkc.tick_params(axis='both', direction='in')
        ax_lkc.spines['top'].set_visible(False)
        ax_lkc.spines['right'].set_visible(False)
        ax_lkc.legend(title='Starting Condition')
        level_lick_condition_fig.tight_layout()

    # ── Collapsed distance by level (all mice) ───────────────────────────────
    level_dist_collapsed_fig = None
    if 'level_dist' in selected_plots and collapsed_level_dist:
        all_levels_dc = sorted(collapsed_level_dist.keys(), key=sort_key)
        level_dist_collapsed_fig, ax_dc = plt.subplots(
            figsize=(max(15, len(all_levels_dc) * 0.8), 8))
        x_dc = np.arange(len(all_levels_dc))
        dc_means, dc_sems, dc_ns = [], [], []
        for lv in all_levels_dc:
            vals = collapsed_level_dist[lv]
            dc_means.append(float(np.mean(vals)))
            dc_sems.append(float(np.std(vals) / np.sqrt(len(vals))))
            dc_ns.append(len(vals))
        ax_dc.bar(x_dc, dc_means, yerr=dc_sems, capsize=4, color='steelblue',
                  error_kw={'elinewidth': 1.2})
        for xi, (m, n_val) in enumerate(zip(dc_means, dc_ns)):
            if n_val > 0 and not np.isnan(m):
                ax_dc.text(xi, m + (dc_sems[xi] if not np.isnan(dc_sems[xi]) else 0),
                           f'n={n_val}', ha='center', va='bottom', fontsize=7)
        ax_dc.set_xticks(x_dc)
        ax_dc.set_xticklabels(
            [lv.replace('level_', 'L').replace('.json', '') for lv in all_levels_dc],
            rotation=45, ha='right',
        )
        ax_dc.set_title('Distance Traveled by Level \u2014 All Mice')
        ax_dc.set_xlabel('Level')
        ax_dc.set_ylabel('Distance (m, Mean \u00b1 SEM)')
        ax_dc.set_ylim(bottom=0)
        ax_dc.tick_params(axis='both', direction='in')
        ax_dc.spines['top'].set_visible(False)
        ax_dc.spines['right'].set_visible(False)
        level_dist_collapsed_fig.tight_layout()

    # ── Distance by level split by condition ─────────────────────────────────
    level_dist_condition_fig = None
    if 'level_dist_condition' in selected_plots and condition_level_dist:
        all_levels_dcc = sorted(
            {lv for cd in condition_level_dist.values() for lv in cd},
            key=sort_key,
        )
        conds_dcc = sorted(condition_level_dist.keys())
        n_conds_dcc = len(conds_dcc)
        level_dist_condition_fig, ax_dcc = plt.subplots(
            figsize=(max(15, len(all_levels_dcc) * 0.8), 8))
        bw_dcc = 0.8 / max(n_conds_dcc, 1)
        x_dcc = np.arange(len(all_levels_dcc))
        for ci, cond in enumerate(conds_dcc):
            cdata = condition_level_dist[cond]
            means_dcc, sems_dcc, ns_dcc = [], [], []
            for lv in all_levels_dcc:
                vals = cdata.get(lv, [])
                if vals:
                    means_dcc.append(float(np.mean(vals)))
                    sems_dcc.append(float(np.std(vals) / np.sqrt(len(vals))))
                    ns_dcc.append(len(vals))
                else:
                    means_dcc.append(np.nan)
                    sems_dcc.append(np.nan)
                    ns_dcc.append(0)
            offset_dcc = (ci - (n_conds_dcc - 1) / 2) * bw_dcc
            ax_dcc.bar(
                x_dcc + offset_dcc, means_dcc,
                width=bw_dcc, yerr=sems_dcc, capsize=4,
                color=cond_color_map.get(cond, 'gray'),
                label=cond,
                error_kw={'elinewidth': 1.2},
            )
            for xi, (m, n_val) in enumerate(zip(means_dcc, ns_dcc)):
                if n_val > 0 and not np.isnan(m):
                    ax_dcc.text(
                        x_dcc[xi] + offset_dcc,
                        m + (sems_dcc[xi] if not np.isnan(sems_dcc[xi]) else 0),
                        f'n={n_val}', ha='center', va='bottom', fontsize=7,
                        color=cond_color_map.get(cond, 'gray'),
                    )
        ax_dcc.set_xticks(x_dcc)
        ax_dcc.set_xticklabels(
            [lv.replace('level_', 'L').replace('.json', '') for lv in all_levels_dcc],
            rotation=45, ha='right',
        )
        ax_dcc.set_title('Distance Traveled by Level \u2014 by Starting Condition')
        ax_dcc.set_xlabel('Level')
        ax_dcc.set_ylabel('Distance (m, Mean \u00b1 SEM)')
        ax_dcc.set_ylim(bottom=0)
        ax_dcc.tick_params(axis='both', direction='in')
        ax_dcc.spines['top'].set_visible(False)
        ax_dcc.spines['right'].set_visible(False)
        ax_dcc.legend(title='Starting Condition')
        level_dist_condition_fig.tight_layout()

    # ── Distance by level split by condition (each mouse's last level excluded) ─
    level_dist_condition_excl_last_fig = None
    if 'level_dist_condition_excl_last' in selected_plots and condition_level_dist_excl_last:
        all_levels_dcx = sorted(
            {lv for cd in condition_level_dist_excl_last.values() for lv in cd},
            key=sort_key,
        )
        conds_dcx = sorted(condition_level_dist_excl_last.keys())
        n_conds_dcx = len(conds_dcx)
        level_dist_condition_excl_last_fig, ax_dcx = plt.subplots(
            figsize=(max(15, len(all_levels_dcx) * 0.8), 8))
        bw_dcx = 0.8 / max(n_conds_dcx, 1)
        x_dcx = np.arange(len(all_levels_dcx))
        for ci, cond in enumerate(conds_dcx):
            cdata = condition_level_dist_excl_last[cond]
            means_dcx, sems_dcx, ns_dcx = [], [], []
            for lv in all_levels_dcx:
                vals = cdata.get(lv, [])
                if vals:
                    means_dcx.append(float(np.mean(vals)))
                    sems_dcx.append(float(np.std(vals) / np.sqrt(len(vals))))
                    ns_dcx.append(len(vals))
                else:
                    means_dcx.append(np.nan)
                    sems_dcx.append(np.nan)
                    ns_dcx.append(0)
            offset_dcx = (ci - (n_conds_dcx - 1) / 2) * bw_dcx
            ax_dcx.bar(
                x_dcx + offset_dcx, means_dcx,
                width=bw_dcx, yerr=sems_dcx, capsize=4,
                color=cond_color_map.get(cond, 'gray'),
                label=cond,
                error_kw={'elinewidth': 1.2},
            )
            for xi, (m, n_val) in enumerate(zip(means_dcx, ns_dcx)):
                if n_val > 0 and not np.isnan(m):
                    ax_dcx.text(
                        x_dcx[xi] + offset_dcx,
                        m + (sems_dcx[xi] if not np.isnan(sems_dcx[xi]) else 0),
                        f'n={n_val}', ha='center', va='bottom', fontsize=7,
                        color=cond_color_map.get(cond, 'gray'),
                    )
        ax_dcx.set_xticks(x_dcx)
        ax_dcx.set_xticklabels(
            [lv.replace('level_', 'L').replace('.json', '') for lv in all_levels_dcx],
            rotation=45, ha='right',
        )
        ax_dcx.set_title('Distance Traveled by Level \u2014 by Starting Condition (last level excluded)')
        ax_dcx.set_xlabel('Level')
        ax_dcx.set_ylabel('Distance (m, Mean \u00b1 SEM)')
        ax_dcx.set_ylim(bottom=0)
        ax_dcx.tick_params(axis='both', direction='in')
        ax_dcx.spines['top'].set_visible(False)
        ax_dcx.spines['right'].set_visible(False)
        ax_dcx.legend(title='Starting Condition')
        level_dist_condition_excl_last_fig.tight_layout()

    # ── Collapsed bout count by level (all mice) ─────────────────────────────
    level_bout_collapsed_fig = None
    if 'level_bout' in selected_plots and collapsed_level_bout:
        all_levels_bc = sorted(collapsed_level_bout.keys(), key=sort_key)
        level_bout_collapsed_fig, ax_bc = plt.subplots(
            figsize=(max(15, len(all_levels_bc) * 0.8), 8))
        x_bc = np.arange(len(all_levels_bc))
        bc_means, bc_sems, bc_ns = [], [], []
        for lv in all_levels_bc:
            vals = collapsed_level_bout[lv]
            bc_means.append(float(np.mean(vals)))
            bc_sems.append(float(np.std(vals) / np.sqrt(len(vals))) if len(vals) > 1 else 0.0)
            bc_ns.append(len(vals))
        ax_bc.bar(x_bc, bc_means, yerr=bc_sems, capsize=4, color='steelblue',
                  error_kw={'elinewidth': 1.2})
        for xi, (m, n_val) in enumerate(zip(bc_means, bc_ns)):
            if n_val > 0 and not np.isnan(m):
                ax_bc.text(xi, m + (bc_sems[xi] if not np.isnan(bc_sems[xi]) else 0),
                           f'n={n_val}', ha='center', va='bottom', fontsize=7)
        ax_bc.set_xticks(x_bc)
        ax_bc.set_xticklabels(
            [lv.replace('level_', 'L').replace('.json', '') for lv in all_levels_bc],
            rotation=45, ha='right',
        )
        ax_bc.set_title('Locomotion Bout Count by Level \u2014 All Mice')
        ax_bc.set_xlabel('Level')
        ax_bc.set_ylabel('Bout Count (Mean \u00b1 SEM)')
        ax_bc.set_ylim(bottom=0)
        ax_bc.tick_params(axis='both', direction='in')
        ax_bc.spines['top'].set_visible(False)
        ax_bc.spines['right'].set_visible(False)
        level_bout_collapsed_fig.tight_layout()

    # ── Bout count by level split by condition ────────────────────────────────
    level_bout_condition_fig = None
    if 'level_bout_condition' in selected_plots and condition_level_bout:
        all_levels_bcc = sorted(
            {lv for cd in condition_level_bout.values() for lv in cd},
            key=sort_key,
        )
        conds_bcc = sorted(condition_level_bout.keys())
        n_conds_bcc = len(conds_bcc)
        level_bout_condition_fig, ax_bcc = plt.subplots(
            figsize=(max(15, len(all_levels_bcc) * 0.8), 8))
        bw_bcc = 0.8 / max(n_conds_bcc, 1)
        x_bcc = np.arange(len(all_levels_bcc))
        for ci, cond in enumerate(conds_bcc):
            cdata = condition_level_bout[cond]
            means_bcc, sems_bcc, ns_bcc = [], [], []
            for lv in all_levels_bcc:
                vals = cdata.get(lv, [])
                if vals:
                    means_bcc.append(float(np.mean(vals)))
                    sems_bcc.append(float(np.std(vals) / np.sqrt(len(vals))) if len(vals) > 1 else 0.0)
                    ns_bcc.append(len(vals))
                else:
                    means_bcc.append(np.nan)
                    sems_bcc.append(np.nan)
                    ns_bcc.append(0)
            offset_bcc = (ci - (n_conds_bcc - 1) / 2) * bw_bcc
            ax_bcc.bar(
                x_bcc + offset_bcc, means_bcc,
                width=bw_bcc, yerr=sems_bcc, capsize=4,
                color=cond_color_map.get(cond, 'gray'),
                label=cond,
                error_kw={'elinewidth': 1.2},
            )
            for xi, (m, n_val) in enumerate(zip(means_bcc, ns_bcc)):
                if n_val > 0 and not np.isnan(m):
                    ax_bcc.text(
                        x_bcc[xi] + offset_bcc,
                        m + (sems_bcc[xi] if not np.isnan(sems_bcc[xi]) else 0),
                        f'n={n_val}', ha='center', va='bottom', fontsize=7,
                        color=cond_color_map.get(cond, 'gray'),
                    )
        ax_bcc.set_xticks(x_bcc)
        ax_bcc.set_xticklabels(
            [lv.replace('level_', 'L').replace('.json', '') for lv in all_levels_bcc],
            rotation=45, ha='right',
        )
        ax_bcc.set_title('Locomotion Bout Count by Level \u2014 by Starting Condition')
        ax_bcc.set_xlabel('Level')
        ax_bcc.set_ylabel('Bout Count (Mean \u00b1 SEM)')
        ax_bcc.set_ylim(bottom=0)
        ax_bcc.tick_params(axis='both', direction='in')
        ax_bcc.spines['top'].set_visible(False)
        ax_bcc.spines['right'].set_visible(False)
        ax_bcc.legend(title='Starting Condition')
        level_bout_condition_fig.tight_layout()

    # ── Collapsed avg speed per bout by level (all mice) ─────────────────────
    level_bout_avg_speed_collapsed_fig = None
    if 'level_bout_avg_speed' in selected_plots and collapsed_level_bout_avg_spd:
        all_levels_bas = sorted(collapsed_level_bout_avg_spd.keys(), key=sort_key)
        level_bout_avg_speed_collapsed_fig, ax_bas = plt.subplots(
            figsize=(max(15, len(all_levels_bas) * 0.8), 8))
        x_bas = np.arange(len(all_levels_bas))
        bas_means, bas_sems, bas_ns = [], [], []
        for lv in all_levels_bas:
            vals = collapsed_level_bout_avg_spd[lv]
            bas_means.append(float(np.mean(vals)))
            bas_sems.append(float(np.std(vals) / np.sqrt(len(vals))) if len(vals) > 1 else 0.0)
            bas_ns.append(len(vals))
        ax_bas.bar(x_bas, bas_means, yerr=bas_sems, capsize=4, color='steelblue',
                   error_kw={'elinewidth': 1.2})
        for xi, (m, n_val) in enumerate(zip(bas_means, bas_ns)):
            if n_val > 0 and not np.isnan(m):
                ax_bas.text(xi, m + (bas_sems[xi] if not np.isnan(bas_sems[xi]) else 0),
                            f'n={n_val}', ha='center', va='bottom', fontsize=7)
        ax_bas.set_xticks(x_bas)
        ax_bas.set_xticklabels(
            [lv.replace('level_', 'L').replace('.json', '') for lv in all_levels_bas],
            rotation=45, ha='right',
        )
        ax_bas.set_title('Average Speed per Locomotion Bout by Level \u2014 All Mice')
        ax_bas.set_xlabel('Level')
        ax_bas.set_ylabel('Speed per Bout (cm/s, Mean \u00b1 SEM)')
        ax_bas.set_ylim(bottom=0)
        ax_bas.tick_params(axis='both', direction='in')
        ax_bas.spines['top'].set_visible(False)
        ax_bas.spines['right'].set_visible(False)
        level_bout_avg_speed_collapsed_fig.tight_layout()

    # ── Avg speed per bout by level split by condition ────────────────────────
    level_bout_avg_speed_condition_fig = None
    if 'level_bout_avg_speed_condition' in selected_plots and condition_level_bout_avg_spd:
        all_levels_basc = sorted(
            {lv for cd in condition_level_bout_avg_spd.values() for lv in cd},
            key=sort_key,
        )
        conds_basc = sorted(condition_level_bout_avg_spd.keys())
        n_conds_basc = len(conds_basc)
        level_bout_avg_speed_condition_fig, ax_basc = plt.subplots(
            figsize=(max(15, len(all_levels_basc) * 0.8), 8))
        bw_basc = 0.8 / max(n_conds_basc, 1)
        x_basc = np.arange(len(all_levels_basc))
        for ci, cond in enumerate(conds_basc):
            cdata = condition_level_bout_avg_spd[cond]
            means_basc, sems_basc, ns_basc = [], [], []
            for lv in all_levels_basc:
                vals = cdata.get(lv, [])
                if vals:
                    means_basc.append(float(np.mean(vals)))
                    sems_basc.append(float(np.std(vals) / np.sqrt(len(vals))) if len(vals) > 1 else 0.0)
                    ns_basc.append(len(vals))
                else:
                    means_basc.append(np.nan)
                    sems_basc.append(np.nan)
                    ns_basc.append(0)
            offset_basc = (ci - (n_conds_basc - 1) / 2) * bw_basc
            ax_basc.bar(
                x_basc + offset_basc, means_basc,
                width=bw_basc, yerr=sems_basc, capsize=4,
                color=cond_color_map.get(cond, 'gray'),
                label=cond,
                error_kw={'elinewidth': 1.2},
            )
            for xi, (m, n_val) in enumerate(zip(means_basc, ns_basc)):
                if n_val > 0 and not np.isnan(m):
                    ax_basc.text(
                        x_basc[xi] + offset_basc,
                        m + (sems_basc[xi] if not np.isnan(sems_basc[xi]) else 0),
                        f'n={n_val}', ha='center', va='bottom', fontsize=7,
                        color=cond_color_map.get(cond, 'gray'),
                    )
        ax_basc.set_xticks(x_basc)
        ax_basc.set_xticklabels(
            [lv.replace('level_', 'L').replace('.json', '') for lv in all_levels_basc],
            rotation=45, ha='right',
        )
        ax_basc.set_title('Average Speed per Locomotion Bout by Level \u2014 by Starting Condition')
        ax_basc.set_xlabel('Level')
        ax_basc.set_ylabel('Speed per Bout (cm/s, Mean \u00b1 SEM)')
        ax_basc.set_ylim(bottom=0)
        ax_basc.tick_params(axis='both', direction='in')
        ax_basc.spines['top'].set_visible(False)
        ax_basc.spines['right'].set_visible(False)
        ax_basc.legend(title='Starting Condition')
        level_bout_avg_speed_condition_fig.tight_layout()

    # ── Collapsed avg distance per bout by level (all mice) ───────────────────
    level_bout_avg_dist_collapsed_fig = None
    if 'level_bout_avg_dist' in selected_plots and collapsed_level_bout_avg_dist_lvl:
        all_levels_bad = sorted(collapsed_level_bout_avg_dist_lvl.keys(), key=sort_key)
        level_bout_avg_dist_collapsed_fig, ax_bad = plt.subplots(
            figsize=(max(15, len(all_levels_bad) * 0.8), 8))
        x_bad = np.arange(len(all_levels_bad))
        bad_means, bad_sems, bad_ns = [], [], []
        for lv in all_levels_bad:
            vals = collapsed_level_bout_avg_dist_lvl[lv]
            bad_means.append(float(np.mean(vals)))
            bad_sems.append(float(np.std(vals) / np.sqrt(len(vals))) if len(vals) > 1 else 0.0)
            bad_ns.append(len(vals))
        ax_bad.bar(x_bad, bad_means, yerr=bad_sems, capsize=4, color='steelblue',
                   error_kw={'elinewidth': 1.2})
        for xi, (m, n_val) in enumerate(zip(bad_means, bad_ns)):
            if n_val > 0 and not np.isnan(m):
                ax_bad.text(xi, m + (bad_sems[xi] if not np.isnan(bad_sems[xi]) else 0),
                            f'n={n_val}', ha='center', va='bottom', fontsize=7)
        ax_bad.set_xticks(x_bad)
        ax_bad.set_xticklabels(
            [lv.replace('level_', 'L').replace('.json', '') for lv in all_levels_bad],
            rotation=45, ha='right',
        )
        ax_bad.set_title('Average Distance per Locomotion Bout by Level \u2014 All Mice')
        ax_bad.set_xlabel('Level')
        ax_bad.set_ylabel('Distance per Bout (m, Mean \u00b1 SEM)')
        ax_bad.set_ylim(bottom=0)
        ax_bad.tick_params(axis='both', direction='in')
        ax_bad.spines['top'].set_visible(False)
        ax_bad.spines['right'].set_visible(False)
        level_bout_avg_dist_collapsed_fig.tight_layout()

    # ── Avg distance per bout by level split by condition ─────────────────────
    level_bout_avg_dist_condition_fig = None
    if 'level_bout_avg_dist_condition' in selected_plots and condition_level_bout_avg_dist_lvl:
        all_levels_badc = sorted(
            {lv for cd in condition_level_bout_avg_dist_lvl.values() for lv in cd},
            key=sort_key,
        )
        conds_badc = sorted(condition_level_bout_avg_dist_lvl.keys())
        n_conds_badc = len(conds_badc)
        level_bout_avg_dist_condition_fig, ax_badc = plt.subplots(
            figsize=(max(15, len(all_levels_badc) * 0.8), 8))
        bw_badc = 0.8 / max(n_conds_badc, 1)
        x_badc = np.arange(len(all_levels_badc))
        for ci, cond in enumerate(conds_badc):
            cdata = condition_level_bout_avg_dist_lvl[cond]
            means_badc, sems_badc, ns_badc = [], [], []
            for lv in all_levels_badc:
                vals = cdata.get(lv, [])
                if vals:
                    means_badc.append(float(np.mean(vals)))
                    sems_badc.append(float(np.std(vals) / np.sqrt(len(vals))) if len(vals) > 1 else 0.0)
                    ns_badc.append(len(vals))
                else:
                    means_badc.append(np.nan)
                    sems_badc.append(np.nan)
                    ns_badc.append(0)
            offset_badc = (ci - (n_conds_badc - 1) / 2) * bw_badc
            ax_badc.bar(
                x_badc + offset_badc, means_badc,
                width=bw_badc, yerr=sems_badc, capsize=4,
                color=cond_color_map.get(cond, 'gray'),
                label=cond,
                error_kw={'elinewidth': 1.2},
            )
            for xi, (m, n_val) in enumerate(zip(means_badc, ns_badc)):
                if n_val > 0 and not np.isnan(m):
                    ax_badc.text(
                        x_badc[xi] + offset_badc,
                        m + (sems_badc[xi] if not np.isnan(sems_badc[xi]) else 0),
                        f'n={n_val}', ha='center', va='bottom', fontsize=7,
                        color=cond_color_map.get(cond, 'gray'),
                    )
        ax_badc.set_xticks(x_badc)
        ax_badc.set_xticklabels(
            [lv.replace('level_', 'L').replace('.json', '') for lv in all_levels_badc],
            rotation=45, ha='right',
        )
        ax_badc.set_title('Average Distance per Locomotion Bout by Level \u2014 by Starting Condition')
        ax_badc.set_xlabel('Level')
        ax_badc.set_ylabel('Distance per Bout (m, Mean \u00b1 SEM)')
        ax_badc.set_ylim(bottom=0)
        ax_badc.tick_params(axis='both', direction='in')
        ax_badc.spines['top'].set_visible(False)
        ax_badc.spines['right'].set_visible(False)
        ax_badc.legend(title='Starting Condition')
        level_bout_avg_dist_condition_fig.tight_layout()

    # ── Final level bar chart (one value per mouse, grouped by condition) ─────
    last_level_bar_fig = None
    if 'last_level_bar' in selected_plots and animal_last_level:
        # Build a mapping from animal_id → condition using the accumulator
        _animal_cond_map = {aid: accum['condition']
                            for (aid, _), accum in animal_level_accum.items()}

        # Convert last level string to a numeric level number (e.g. 'level_5' → 5)
        def _level_to_num(lv_str):
            if isinstance(lv_str, str) and lv_str.startswith('level_'):
                try:
                    return float(lv_str.split('_')[1].split('.')[0])
                except (ValueError, IndexError):
                    pass
            return float('nan')

        # Group per condition: list of (animal_id, numeric_level)
        _last_lvl_by_cond: dict[str, list] = {}
        for _aid, _last_lv in animal_last_level.items():
            _cond = _animal_cond_map.get(_aid)
            if _cond is None:
                continue
            _lv_num = _level_to_num(_last_lv)
            if not np.isnan(_lv_num):
                _last_lvl_by_cond.setdefault(_cond, []).append((_aid, _lv_num))

        if _last_lvl_by_cond:
            last_level_bar_fig, ax_llb = plt.subplots(figsize=(8, 6))
            _conds_llb = sorted(_last_lvl_by_cond.keys())
            _x_pos_llb = np.arange(len(_conds_llb))
            _rng_llb = np.random.default_rng(seed=42)

            for ci, cond in enumerate(_conds_llb):
                entries = _last_lvl_by_cond[cond]
                mouse_vals = [v for _, v in entries]
                _mean_llb = float(np.mean(mouse_vals))
                _sem_llb  = (float(np.std(mouse_vals, ddof=1) / np.sqrt(len(mouse_vals)))
                             if len(mouse_vals) > 1 else 0.0)
                color = cond_color_map.get(cond, 'steelblue')
                ax_llb.bar(ci, _mean_llb, width=0.5, color=color, alpha=0.8,
                           yerr=_sem_llb, capsize=7,
                           error_kw={'elinewidth': 1.5, 'capthick': 1.5})
                jitter = (_rng_llb.random(len(mouse_vals)) - 0.5) * 0.22
                for j, (_aid_llb, _lv_val) in enumerate(entries):
                    ax_llb.plot(ci + jitter[j], _lv_val, 'o',
                                color='white', markeredgecolor=color,
                                markeredgewidth=1.8, markersize=7, zorder=3)

            ax_llb.set_xticks(_x_pos_llb)
            ax_llb.set_xticklabels(_conds_llb)
            ax_llb.set_xlabel('Starting Condition')
            ax_llb.set_ylabel('Final Level (Mean \u00b1 SEM)')
            ax_llb.set_ylim(bottom=0)
            ax_llb.tick_params(axis='both', direction='in')
            ax_llb.spines['top'].set_visible(False)
            ax_llb.spines['right'].set_visible(False)

            # ── Mann-Whitney U test: pairwise significance brackets ──────────
            from scipy.stats import mannwhitneyu as _mwu_llb
            import itertools as _it_llb
            _llb_results = []
            for (_lc1, _lc2) in _it_llb.combinations(_conds_llb, 2):
                _lv1 = np.array([v for _, v in _last_lvl_by_cond[_lc1]])
                _lv2 = np.array([v for _, v in _last_lvl_by_cond[_lc2]])
                if len(_lv1) >= 2 and len(_lv2) >= 2:
                    _ll_stat, _ll_p = _mwu_llb(_lv1, _lv2, alternative='two-sided')
                    _llb_results.append((_lc1, _lc2, float(_ll_stat), float(_ll_p)))

            if _llb_results:
                _llb_y_max = max(
                    max(max(v for _, v in _last_lvl_by_cond[c]) for c in _conds_llb),
                    float(ax_llb.get_ylim()[1]),
                )
                _llb_step = _llb_y_max * 0.12
                for _bk_idx, (_lc1, _lc2, _ll_stat, _ll_p) in enumerate(_llb_results):
                    _bx1 = _conds_llb.index(_lc1)
                    _bx2 = _conds_llb.index(_lc2)
                    _bk_y = _llb_y_max + _llb_step * (_bk_idx + 1)
                    ax_llb.plot(
                        [_bx1, _bx1, _bx2, _bx2],
                        [_bk_y - _llb_step * 0.2, _bk_y,
                         _bk_y, _bk_y - _llb_step * 0.2],
                        color='black', linewidth=1.2,
                    )
                    if _ll_p < 0.001:
                        _bsig = '***'
                    elif _ll_p < 0.01:
                        _bsig = '**'
                    elif _ll_p < 0.05:
                        _bsig = '*'
                    else:
                        _bsig = f'ns  p={_ll_p:.3f}'
                    ax_llb.text(
                        (_bx1 + _bx2) / 2.0,
                        _bk_y + _llb_step * 0.05,
                        _bsig, ha='center', va='bottom', fontsize=9,
                    )
                ax_llb.set_ylim(
                    bottom=0,
                    top=_llb_y_max + _llb_step * (len(_llb_results) + 1.8),
                )

            ax_llb.set_title('Final Level Reached by Starting Condition\n(last training day; Mann-Whitney U test)')
            last_level_bar_fig.tight_layout()

    # ── Level survivor plot (proportion of mice that experienced each level) ──
    level_survivor_fig = None
    if 'level_survivor' in selected_plots and animal_level_accum:
        # All levels seen across all animals (sorted)
        all_lvls_surv = sorted(
            {lv for (_, lv) in animal_level_accum},
            key=sort_key,
        )
        # Per-condition: set of animal IDs and set of (animal_id, level) pairs
        _surv_cond_mice: dict[str, set] = {}
        _surv_cond_lvl_mice: dict[str, dict[str, set]] = {}
        for (aid, lv), accum in animal_level_accum.items():
            cond = accum['condition']
            _surv_cond_mice.setdefault(cond, set()).add(aid)
            _surv_cond_lvl_mice.setdefault(cond, {}).setdefault(lv, set()).add(aid)

        if _surv_cond_mice:
            level_survivor_fig, ax_surv = plt.subplots(figsize=(max(10, len(all_lvls_surv) * 0.6), 6))
            _surv_conds = sorted(_surv_cond_mice.keys())
            _surv_color_map = {c: _condition_to_color(c) for c in _surv_conds}

            for cond in _surv_conds:
                n_total = len(_surv_cond_mice[cond])
                proportions = []
                for lv in all_lvls_surv:
                    n_with_lv = len(_surv_cond_lvl_mice[cond].get(lv, set()))
                    proportions.append(n_with_lv / n_total if n_total > 0 else 0.0)
                x_idx = np.arange(len(all_lvls_surv))
                color = _surv_color_map.get(cond, 'steelblue')
                ax_surv.step(x_idx, proportions, where='post',
                             color=color, linewidth=2, label=f'{cond} (n={n_total})')

            def _lv_num_label(lv):
                if isinstance(lv, str) and lv.startswith('level_'):
                    try:
                        return str(int(lv.split('_')[1].split('.')[0]))
                    except (ValueError, IndexError):
                        pass
                return str(lv)

            lv_labels = [_lv_num_label(lv) for lv in all_lvls_surv]
            ax_surv.set_xticks(np.arange(len(all_lvls_surv)))
            ax_surv.set_xticklabels(lv_labels)
            ax_surv.set_xlim(-0.5, len(all_lvls_surv) - 0.5)
            ax_surv.set_xlabel('Level')
            ax_surv.set_ylabel('Proportion of Mice')
            ax_surv.set_ylim(0, 1.05)
            ax_surv.set_title('Level Attainment Survivor Plot\n(proportion of mice that experienced each level, by condition)')
            ax_surv.legend(title='Starting Condition')
            ax_surv.tick_params(axis='both', direction='in')
            ax_surv.spines['top'].set_visible(False)
            ax_surv.spines['right'].set_visible(False)
            level_survivor_fig.tight_layout()

    # ── Time to first level 1 → level 2 transition (cumulative minutes) ──────
    time_to_level2_fig = None
    if 'time_to_level2' in selected_plots:
        # For each animal find the first session where level_1 has a non-NaN transition_ts.
        # Cumulative time = sum of all prior session lengths + the transition_ts within that session.
        _is_l1 = lambda lv: (
            isinstance(lv, str) and lv.startswith('level_') and
            lv.split('_')[1].split('.')[0] == '1'
        )
        animal_time_to_l2: dict[str, float] = {}
        animal_time_to_l2_cond: dict[str, str] = {}

        for _aid, _ag in transitions_df.groupby('animal_id'):
            _sessions = sorted(_ag['session_num'].unique())
            _cum_s = 0.0
            _found = False
            for _sn in _sessions:
                _sg = _ag[_ag['session_num'] == _sn]
                # Check whether level_1 completes in this session
                for _, _row in _sg.iterrows():
                    if _is_l1(_row['level']) and pd.notna(_row['transition_ts']):
                        _total_min = (_cum_s + float(_row['transition_ts'])) / 60.0
                        animal_time_to_l2[_aid] = _total_min
                        animal_time_to_l2_cond[_aid] = (
                            (animal_conditions or {}).get(_aid, 'Unknown')
                        )
                        _found = True
                        break
                if _found:
                    break
                # Add this session's length to the cumulative total
                _cum_s += animal_session_lengths.get((_aid, _sn), 0.0)

        if animal_time_to_l2:
            _t2_by_cond: dict[str, list] = {}
            for _aid, _t in animal_time_to_l2.items():
                _cond = animal_time_to_l2_cond.get(_aid, 'Unknown')
                _t2_by_cond.setdefault(_cond, []).append(_t)

            _t2_conds = sorted(_t2_by_cond.keys())
            time_to_level2_fig, _ax_t2 = plt.subplots(
                figsize=(max(5, len(_t2_conds) * 1.8 + 1.5), 6))
            _t2_x = np.arange(len(_t2_conds))
            _rng = np.random.default_rng(42)

            for _ci, _cond in enumerate(_t2_conds):
                _vals = np.array(_t2_by_cond[_cond])
                _mean = float(np.mean(_vals))
                _sem  = (float(np.std(_vals, ddof=1) / np.sqrt(len(_vals)))
                         if len(_vals) > 1 else 0.0)
                _color = cond_color_map.get(_cond, 'steelblue')
                _ax_t2.bar(_t2_x[_ci], _mean, width=0.5, yerr=_sem, capsize=5,
                           color=_color, label=f'{_cond} (n={len(_vals)})',
                           error_kw={'elinewidth': 1.5}, zorder=2)
                # Individual data points with jitter
                _jit = _rng.uniform(-0.10, 0.10, len(_vals))
                _ax_t2.scatter(np.full(len(_vals), _t2_x[_ci]) + _jit, _vals,
                               color=_color, edgecolors='black', s=40, zorder=5,
                               linewidths=0.7)
                _ax_t2.text(_t2_x[_ci], _mean + _sem + max(_mean * 0.02, 0.5),
                            f'n={len(_vals)}', ha='center', va='bottom', fontsize=9)

            # Mann-Whitney U bracket if exactly 2 conditions
            if len(_t2_conds) == 2:
                _v1 = np.array(_t2_by_cond[_t2_conds[0]])
                _v2 = np.array(_t2_by_cond[_t2_conds[1]])
                if len(_v1) >= 2 and len(_v2) >= 2:
                    try:
                        _, _pv = mannwhitneyu(_v1, _v2, alternative='two-sided')
                        _all_vals = np.concatenate([_v1, _v2])
                        _ymax_bk = float(np.max(_all_vals)) * 1.15
                        _stars = ('****' if _pv < 0.0001 else
                                  '***'  if _pv < 0.001  else
                                  '**'   if _pv < 0.01   else
                                  '*'    if _pv < 0.05   else 'ns')
                        _ax_t2.plot([0, 1], [_ymax_bk, _ymax_bk], color='black',
                                    linewidth=1.2, zorder=6)
                        _ax_t2.text(0.5, _ymax_bk * 1.01, f'{_stars}\np={_pv:.3f}',
                                    ha='center', va='bottom', fontsize=9)
                        _ax_t2.set_ylim(bottom=0, top=_ymax_bk * 1.15)
                    except Exception:
                        pass

            _ax_t2.set_xticks(_t2_x)
            _ax_t2.set_xticklabels(_t2_conds)
            _ax_t2.set_xlabel('Starting Condition')
            _ax_t2.set_ylabel('Cumulative time to first level change (min)')
            _ax_t2.set_title(
                'Time to First Level 1\u21922 Transition\n'
                '(cumulative: prior session durations + within-session transition time)')
            _ax_t2.set_ylim(bottom=0)
            _ax_t2.tick_params(axis='both', direction='in')
            _ax_t2.spines['top'].set_visible(False)
            _ax_t2.spines['right'].set_visible(False)
            _ax_t2.legend(title='Starting Condition')
            time_to_level2_fig.tight_layout()

    # ── Package level data for the descriptive stats report ──────────────────
    _level_stats = {
        'condition_level_reward_rate':   condition_level_data,
        'condition_level_speed':         condition_level_speed,
        'condition_level_lick':          condition_level_lick,
        'condition_level_dist':          condition_level_dist,
        'condition_level_bout':          condition_level_bout,
        'condition_level_bout_avg_spd':  condition_level_bout_avg_spd,
        'condition_level_bout_avg_dist': condition_level_bout_avg_dist_lvl,
    } if condition_level_data else None

    return level_reward_fig, level_speed_collapsed_fig, level_speed_condition_fig, level_lick_collapsed_fig, level_lick_condition_fig, level_dist_collapsed_fig, level_dist_condition_fig, level_dist_condition_excl_last_fig, level_bout_collapsed_fig, level_bout_condition_fig, level_bout_avg_speed_collapsed_fig, level_bout_avg_speed_condition_fig, level_bout_avg_dist_collapsed_fig, level_bout_avg_dist_condition_fig, last_level_bar_fig, level_survivor_fig, time_to_level2_fig, _level_stats


# ── Descriptive statistics report ────────────────────────────────────────────

def generate_descriptive_stats_report(all_results, level_stats_data=None, output_dir=None):
    """Generate a descriptive statistics Excel workbook (.xlsx) with one sheet
    per metric sliced across training sessions (time) and, if level data are
    available, per level.  Metrics are grouped by starting condition; statistics
    reported per group are: mean, SEM, SD, N, min, max, median.

    Parameters
    ----------
    all_results : list[dict]
        Session-level result list produced by analyze_mouse_data.
    level_stats_data : dict | None
        Dict produced by analyze_levels containing condition_level_* dicts.
        Pass None to skip level sheets.
    output_dir : str | None
        Directory to save the report.  Defaults to current working directory.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from scipy.stats import t as t_dist
    except ImportError:
        print("[WARN] openpyxl not found — descriptive stats report skipped. "
              "Install it with:  pip install openpyxl")
        return

    STAT_COLS = ['mean', 'sem', 'sd', 'n', 'min', 'max', 'median', 'ci95_lo', 'ci95_hi']
    HEADER_FONT   = Font(bold=True, color='FFFFFF')
    HEADER_FILL_A = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
    ROW_FILLS = [
        PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid'),
        PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
    ]

    # ── Helpers ──────────────────────────────────────────────────────────────

    def _clean(v):
        if v is None:
            return False
        if isinstance(v, float) and np.isnan(v):
            return False
        return True

    def _desc_stats(vals):
        a = np.array([v for v in vals if _clean(v)], dtype=float)
        if len(a) == 0:
            return dict(mean=np.nan, sem=np.nan, sd=np.nan, n=0,
                        min=np.nan, max=np.nan, median=np.nan,
                        ci95_lo=np.nan, ci95_hi=np.nan)
        _mean = float(np.mean(a))
        _sd   = float(np.std(a, ddof=1)) if len(a) > 1 else np.nan
        _sem  = float(_sd / np.sqrt(len(a))) if len(a) > 1 else np.nan
        if len(a) > 1:
            _t = t_dist.ppf(0.975, df=len(a) - 1)
            _ci_lo = _mean - _t * _sem
            _ci_hi = _mean + _t * _sem
        else:
            _ci_lo = _ci_hi = np.nan
        return dict(
            mean=_mean,
            sem=_sem,
            sd=_sd,
            n=int(len(a)),
            min=float(np.min(a)),
            max=float(np.max(a)),
            median=float(np.median(a)),
            ci95_lo=_ci_lo,
            ci95_hi=_ci_hi,
        )

    def _sort_level(lv):
        if lv.startswith('level_'):
            try:
                return (0, int(lv.split('_')[1].split('.')[0]))
            except (ValueError, IndexError):
                pass
        return (1, lv)

    def _write_sheet(writer, sheet_name, index_label, rows):
        """Write a stats sheet to the Excel workbook.

        rows : list[dict]
            Each dict has {index_label: display_value, cond_name: [float_list], ...}.
        """
        all_conds = sorted({k for row in rows for k in row if k != index_label})
        col_names = [index_label]
        for cond in all_conds:
            for stat in STAT_COLS:
                col_names.append(f'{cond}_{stat}')

        records = []
        for row in rows:
            rec = {index_label: row[index_label]}
            for cond in all_conds:
                st = _desc_stats(row.get(cond, []))
                for stat in STAT_COLS:
                    rec[f'{cond}_{stat}'] = st[stat]
            records.append(rec)

        df_sheet = pd.DataFrame(records, columns=col_names)
        df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)

        ws = writer.sheets[sheet_name]
        ws.freeze_panes = 'B2'

        # Style header
        for cell in ws[1]:
            cell.font  = HEADER_FONT
            cell.fill  = HEADER_FILL_A
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        # Alternating condition shading
        for ci, cond in enumerate(all_conds):
            fill = ROW_FILLS[ci % len(ROW_FILLS)]
            col_start = 2 + ci * len(STAT_COLS)
            col_end   = col_start + len(STAT_COLS) - 1
            for row_idx in range(2, len(records) + 2):
                for col_idx in range(col_start, col_end + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

        # Auto-fit column widths (capped 10–25)
        for col_cells in ws.columns:
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0) for c in col_cells
            )
            ws.column_dimensions[
                get_column_letter(col_cells[0].column)
            ].width = max(10, min(max_len + 2, 25))

    # ── Time-based rows ───────────────────────────────────────────────────────

    conditions_sorted = sorted({r['starting_condition'] for r in all_results})
    max_sessions = max(len(r['df']) for r in all_results)

    def _extract(df_r, sess_idx, col):
        if sess_idx >= len(df_r):
            return np.nan
        v = df_r.iat[sess_idx, df_r.columns.get_loc(col)]
        return float(v) if pd.notna(v) else np.nan

    def _session_rows(col_or_fn):
        """Build rows for a time-based sheet.
        col_or_fn: column name (str) or callable(df_r, sess_idx) -> float.
        """
        rows = []
        for sess_idx in range(max_sessions):
            row = {'Session': sess_idx + 1}
            for cond in conditions_sorted:
                vals = []
                for r in all_results:
                    if r['starting_condition'] != cond:
                        continue
                    df_r = r['df'].reset_index(drop=True)
                    if callable(col_or_fn):
                        v = col_or_fn(df_r, sess_idx)
                    else:
                        v = _extract(df_r, sess_idx, col_or_fn)
                    if _clean(v):
                        vals.append(float(v))
                row[cond] = vals
            rows.append(row)
        return rows

    def _reward_rate(df_r, i):
        h = _extract(df_r, i, 'hits')
        s = _extract(df_r, i, 'session_length')
        return h / s if _clean(h) and _clean(s) and s > 0 else np.nan

    def _lick_rate(df_r, i):
        lc = _extract(df_r, i, 'lick_count')
        s  = _extract(df_r, i, 'session_length')
        return lc / s if _clean(lc) and _clean(s) and s > 0 else np.nan

    def _dist_m(df_r, i):
        v = _extract(df_r, i, 'total_distance')
        return v / 1000.0 if _clean(v) else np.nan

    def _bout_dist_m(df_r, i):
        v = _extract(df_r, i, 'avg_dist_per_bout')
        return v / 1000.0 if _clean(v) else np.nan

    time_sheets = [
        ('Time - Speed',          _session_rows('average_speed')),
        ('Time - Reward Count',   _session_rows('hits')),
        ('Time - Reward Rate',    _session_rows(_reward_rate)),
        ('Time - Lick Rate',      _session_rows(_lick_rate)),
        ('Time - Distance (m)',   _session_rows(_dist_m)),
        ('Time - Bout Count',     _session_rows('bout_count')),
        ('Time - Bout Avg Speed', _session_rows('avg_speed_per_bout')),
        ('Time - Bout Avg Dist',  _session_rows(_bout_dist_m)),
    ]

    # ── Level-based rows ──────────────────────────────────────────────────────

    level_sheets = []
    if level_stats_data is not None:

        def _level_rows(cond_level_dict):
            if not cond_level_dict:
                return []
            lvls = sorted(
                {lv for cd in cond_level_dict.values() for lv in cd},
                key=_sort_level,
            )
            rows = []
            for lv in lvls:
                row = {'Level': lv.replace('level_', 'L').replace('.json', '')}
                for cond in sorted(cond_level_dict.keys()):
                    row[cond] = cond_level_dict[cond].get(lv, [])
                rows.append(row)
            return rows

        level_sheets = [
            ('Level - Reward Rate',    _level_rows(level_stats_data['condition_level_reward_rate'])),
            ('Level - Speed',          _level_rows(level_stats_data['condition_level_speed'])),
            ('Level - Lick Rate',      _level_rows(level_stats_data['condition_level_lick'])),
            ('Level - Distance (m)',   _level_rows(level_stats_data['condition_level_dist'])),
            ('Level - Bout Count',     _level_rows(level_stats_data['condition_level_bout'])),
            ('Level - Bout Avg Speed', _level_rows(level_stats_data['condition_level_bout_avg_spd'])),
            ('Level - Bout Avg Dist',  _level_rows(level_stats_data['condition_level_bout_avg_dist'])),
        ]

    # ── Write workbook ────────────────────────────────────────────────────────

    if output_dir is None:
        output_dir = os.getcwd()
    os.makedirs(output_dir, exist_ok=True)
    timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_path = os.path.join(output_dir, f'descriptive_stats_{timestamp_str}.xlsx')

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        for sheet_name, rows in time_sheets:
            if rows:
                _write_sheet(writer, sheet_name, 'Session', rows)
        for sheet_name, rows in level_sheets:
            if rows:
                _write_sheet(writer, sheet_name, 'Level', rows)

    print(f"\nDescriptive stats report saved to:\n  {out_path}")


def print_session_weekday_alignment(all_results):
    """Print a verification table mapping each session to its assigned weekday.

    Sessions are numbered sequentially (0-based) per mouse and assigned to
    the fixed 4-day training cycle: Monday → Tuesday → Thursday → Friday,
    repeating from the start of each new week.

    A ``<-- week boundary`` marker is printed at the start of every new
    week (i.e. every 4th session after the first) so you can visually
    confirm that the cycle resets correctly.

    Parameters
    ----------
    all_results : list[dict]
        As returned by ``analyze_mouse_data``.  Each dict must contain
        keys ``'mouse'``, ``'starting_condition'``, and ``'df'`` (the
        per-mouse results DataFrame, which must include a ``'weekday'``
        column — added automatically during data collection).
    """
    DOW_CYCLE = ['Monday', 'Tuesday', 'Thursday', 'Friday']
    _sep = '=' * 66
    _inner = '-' * 66
    print('\n' + _sep)
    print('  SESSION  ↔  WEEKDAY  ALIGNMENT  CHECK')
    print(_sep)
    for result in all_results:
        mouse     = result['mouse']
        condition = result['starting_condition']
        df_r      = result['df']
        print(f'\n  Mouse : {mouse}   (starting condition: {condition})')
        print(f'  {"Sess #":>6}  {"Date":<14}  {"Weekday":<12}  {"Week #":>6}')
        print('  ' + _inner)
        for sess_i, (_, row) in enumerate(df_r.iterrows()):
            date_str = str(row['date'])[:10]
            wd       = row.get('weekday', DOW_CYCLE[sess_i % 4])
            week_num = sess_i // 4 + 1
            marker   = '  <-- new week' if sess_i % 4 == 0 and sess_i > 0 else ''
            print(f'  {sess_i + 1:>6}  {date_str:<14}  {wd:<12}  {week_num:>6}{marker}')
    print('\n' + _sep + '\n')


def _plot_epoch_panels(all_results, signal_key, ylabel, title_prefix,
                      condition_color_map, window_s=EPOCH_WINDOW_S,
                      hierarchy='event', use_sd=False, show_individual_traces=True,
                      reward_delivery_vline=True, group_key='starting_condition',
                      group_color_map=None, group_label='By Starting Condition',
                      canonical_time=None):
    """Create per-mouse and condition-averaged epoch figures for a given signal.

    Parameters
    ----------
    all_results         : list[dict] — one entry per mouse from analyze_mouse_data.
    signal_key          : str — key in each result dict holding the epoch matrix.
                          For hierarchy='event'  : shape (n_events   × EPOCH_N_SAMPLES).
                          For hierarchy='session': shape (n_sessions  × EPOCH_N_SAMPLES),
                          where each row is already the per-session mean trace.
    ylabel              : str — y-axis label.
    title_prefix        : str — base title string.
    condition_color_map : dict — maps condition name → matplotlib colour.
    window_s            : float — half-window size (seconds) for x-axis limits.
    hierarchy           : 'event' or 'session'.
                          'event'   — rows are individual zone-entry epochs; mean ± SEM
                                      weights every zone entry equally.
                          'session' — rows are per-session mean traces; mean ± SEM
                                      weights every session equally.  Individual session
                                      traces are shown as thin background lines in the
                                      per-mouse panel.

    Returns
    -------
    fig_per_mouse : matplotlib.figure.Figure  (one subplot per mouse)
    fig_cond      : matplotlib.figure.Figure  (condition-averaged overlay)
    """
    canonical_time = EPOCH_CANONICAL_TIME if canonical_time is None else canonical_time
    unit_label     = 'sessions' if hierarchy == 'session' else 'events'
    hier_label     = '(session-averaged)' if hierarchy == 'session' else '(event-averaged)'
    _active_color_map = group_color_map if group_color_map is not None else condition_color_map

    # ── Figure 1: per-mouse ───────────────────────────────────────────────────
    n_mice = len(all_results)
    ncols  = min(3, n_mice)
    nrows  = (n_mice + ncols - 1) // ncols
    fig_per_mouse, axs = plt.subplots(nrows, ncols,
                                      figsize=(5 * ncols, 4 * nrows),
                                      sharex=True, sharey=True,
                                      squeeze=False)
    axs_flat = axs.flatten()
    _yvals_per = []

    for i, result in enumerate(all_results):
        matrix     = result.get(signal_key)
        ax         = axs_flat[i]
        mouse_name = result['mouse']
        condition  = result.get(group_key, 'unknown')
        color      = _active_color_map.get(condition, 'steelblue')

        if matrix is not None and matrix.shape[0] > 0:
            with warnings.catch_warnings():
                warnings.simplefilter('ignore', RuntimeWarning)
                mean_trace = np.nanmean(matrix, axis=0)
                if use_sd:
                    err_trace = np.nanstd(matrix, axis=0, ddof=1)
                else:
                    n_valid   = np.sum(~np.isnan(matrix), axis=0)
                    err_trace = np.where(n_valid > 1,
                                         np.nanstd(matrix, axis=0, ddof=1) / np.sqrt(n_valid),
                                         0.0)
            _yvals_per.append(float(np.nanmax(mean_trace + err_trace)))
            _yvals_per.append(float(np.nanmin(mean_trace - err_trace)))
            ax.plot(canonical_time, mean_trace, color=color, linewidth=1.8,
                    label=f'n={matrix.shape[0]} {unit_label}')
            ax.fill_between(canonical_time,
                            mean_trace - err_trace,
                            mean_trace + err_trace,
                            color=color, alpha=0.25)
        else:
            ax.text(0.5, 0.5, 'No data', ha='center', va='center',
                    transform=ax.transAxes, fontsize=9)

        ax.axvline(0, color='red', linestyle='--', linewidth=1.2)
        if reward_delivery_vline:
            ax.axvline(0.65, color='black', linestyle='--', linewidth=1.0)
        ax.set_title(mouse_name, fontsize=10)
        ax.set_xlim(-window_s, window_s)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.legend(fontsize=8, loc='upper right')

    for j in range(n_mice, len(axs_flat)):
        axs_flat[j].set_visible(False)

    for ax in axs_flat[:n_mice]:
        ax.set_xlabel('Time from reward zone entry (s)', fontsize=9)
        ax.set_ylabel(ylabel, fontsize=9)

    # Shared y-axis: fitted to mean ± SEM of averaged data (5 % headroom).
    # With sharey=True one set_ylim call propagates everywhere.
    if _yvals_per:
        _ymax_per = float(np.nanmax(_yvals_per))
        _ymin_per = float(np.nanmin(_yvals_per))
    else:
        _ymax_per, _ymin_per = 1.0, 0.0
    _bot_per = _ymin_per * 1.05 if _ymin_per < 0 else 0.0
    axs_flat[0].set_ylim(_bot_per, _ymax_per * 1.05)

    fig_per_mouse.suptitle(f'{title_prefix} {hier_label} — Per Mouse', fontsize=13)
    fig_per_mouse.tight_layout()

    # ── Figure 2: condition-averaged ─────────────────────────────────────────
    fig_cond, ax_cond = plt.subplots(figsize=(9, 5))

    condition_mouse_means  = {}
    condition_all_matrices = {}
    _yvals_cond = []
    for result in all_results:
        matrix    = result.get(signal_key)
        condition = result.get(group_key, 'unknown')
        if matrix is not None and matrix.shape[0] > 0:
            with warnings.catch_warnings():
                warnings.simplefilter('ignore', RuntimeWarning)
                mouse_mean = np.nanmean(matrix, axis=0)
            condition_mouse_means.setdefault(condition, []).append(mouse_mean)
            if use_sd:
                condition_all_matrices.setdefault(condition, []).append(matrix)

    for condition in sorted(condition_mouse_means.keys()):
        color       = _active_color_map.get(condition, 'steelblue')
        mouse_means = np.array(condition_mouse_means[condition])  # (n_mice_in_cond, 501)
        n_mice_cond = mouse_means.shape[0]

        # Thin per-mouse lines for within-condition spread (optional)
        if show_individual_traces:
            for mm in mouse_means:
                ax_cond.plot(canonical_time, mm, color=color, linewidth=0.8, alpha=0.4)
                _yvals_cond.append(float(np.nanmax(mm)))
                _yvals_cond.append(float(np.nanmin(mm)))

        with warnings.catch_warnings():
            warnings.simplefilter('ignore', RuntimeWarning)
            cond_mean = np.nanmean(mouse_means, axis=0)
            if use_sd:
                pooled   = np.vstack(condition_all_matrices[condition])
                cond_err = np.nanstd(pooled, axis=0, ddof=1)
            else:
                cond_err = (np.nanstd(mouse_means, axis=0, ddof=1) / np.sqrt(n_mice_cond)
                            if n_mice_cond > 1 else np.zeros_like(cond_mean))
        _yvals_cond.append(float(np.nanmax(cond_mean + cond_err)))
        _yvals_cond.append(float(np.nanmin(cond_mean - cond_err)))

        ax_cond.plot(canonical_time, cond_mean, color=color, linewidth=2.2,
                     label=f'{condition} (n={n_mice_cond} mice)')
        ax_cond.fill_between(canonical_time,
                             cond_mean - cond_err,
                             cond_mean + cond_err,
                             color=color, alpha=0.20)

    ax_cond.axvline(0, color='red', linestyle='--', linewidth=1.5, label='Zone entry (t=0)')
    if reward_delivery_vline:
        ax_cond.axvline(0.65, color='black', linestyle='--', linewidth=1.0, label='Reward delivery (t=0.65 s)')
    ax_cond.set_xlabel('Time from reward zone entry (s)')
    ax_cond.set_ylabel(ylabel)
    ax_cond.set_title(f'{title_prefix} {hier_label} — {group_label}')
    ax_cond.set_xlim(-window_s, window_s)
    # Fit y-axis to mean ± SEM of averaged data (5 % headroom)
    if _yvals_cond:
        _ymax_cond = float(np.nanmax(_yvals_cond))
        _ymin_cond = float(np.nanmin(_yvals_cond))
    else:
        _ymax_cond, _ymin_cond = 1.0, 0.0
    _bot_cond = _ymin_cond * 1.05 if _ymin_cond < 0 else 0.0
    ax_cond.set_ylim(_bot_cond, _ymax_cond * 1.05)
    ax_cond.legend()
    ax_cond.spines['top'].set_visible(False)
    ax_cond.spines['right'].set_visible(False)
    fig_cond.tight_layout()

    return fig_per_mouse, fig_cond


def _plot_epoch_early_late_panels(all_results, signal_key, ylabel, title_prefix,
                                   condition_color_map, window_s=EPOCH_WINDOW_S,
                                   indices_key=None, row_unit='sessions', use_sd=False,
                                   show_individual_traces=True, reward_delivery_vline=True):
    """Early vs late session epoch panels — each half in its own independent figure.

    Uses a 2-level hierarchy: individual row traces (thin background)
    → group mean ± SEM (bold foreground).  The early/late boundary is the same
    global split point for every mouse so that a mouse with missing sessions is
    always split at the same session index as its complete cohort-mates.

    Parameters
    ----------
    indices_key : str | None
        Key in each result dict holding the per-row session index array.
        Defaults to ``signal_key.replace('_means', '_indices')``.
    row_unit : str
        Label string used in the legend (e.g. 'sessions' or 'events').

    Returns
    -------
    fig_pm_early, fig_pm_late, fig_cond_early, fig_cond_late
    """
    canonical_time = EPOCH_CANONICAL_TIME
    if indices_key is None:
        indices_key = signal_key.replace('_means', '_indices')
    n_mice         = len(all_results)

    # ── global split point ────────────────────────────────────────────────────
    _all_idx = [result.get(indices_key) for result in all_results
                if result.get(indices_key) is not None and len(result.get(indices_key)) > 0]
    if _all_idx:
        global_n    = int(max(idx.max() for idx in _all_idx)) + 1
        global_half = global_n // 2
    else:
        global_n    = 0
        global_half = 0

    def _make_half_figs(half_tag):
        """Build per-mouse and condition figures for one half (early or late)."""
        is_early   = (half_tag == 'early')
        half_label = (f'Early Sessions (1\u2013{global_half})'
                      if is_early else
                      f'Late Sessions ({global_half + 1}\u2013{global_n})')

        # ── per-mouse figure (ceil(n_mice/2) rows × 2 cols) ──────────────────
        n_cols = 2
        n_rows = math.ceil(n_mice / n_cols)
        fig_pm, axs = plt.subplots(n_rows, n_cols,
                                   figsize=(12, 4 * n_rows),
                                   sharex=True, sharey=True,
                                   squeeze=False)
        # hide any unused axes in the last row
        for _empty in range(n_mice, n_rows * n_cols):
            axs[_empty // n_cols, _empty % n_cols].set_visible(False)
        _yvals_pm = []

        for i, result in enumerate(all_results):
            matrix     = result.get(signal_key)
            indices    = result.get(indices_key)
            mouse_name = result['mouse']
            condition  = result['starting_condition']
            color      = condition_color_map.get(condition, 'steelblue')
            ax         = axs[i // n_cols, i % n_cols]

            ax.axvline(0, color='red', linestyle='--', linewidth=1.2)
            if reward_delivery_vline:
                ax.axvline(0.65, color='black', linestyle='--', linewidth=1.0)
            ax.set_xlim(-window_s, window_s)
            ax.set_ylabel(ylabel, fontsize=9)
            ax.set_title(f'{mouse_name} — {half_label}', fontsize=10)
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            # x-label on bottom row only
            if i // n_cols == n_rows - 1:
                ax.set_xlabel('Time from reward zone entry (s)', fontsize=9)

            if matrix is not None and indices is not None and len(indices) > 0:
                mask = (indices < global_half) if is_early else (indices >= global_half)
                sub  = matrix[mask]

                if sub.shape[0] == 0:
                    ax.text(0.5, 0.5, 'No data', ha='center', va='center',
                            transform=ax.transAxes, fontsize=9)
                else:
                    with warnings.catch_warnings():
                        warnings.simplefilter('ignore', RuntimeWarning)
                        mean_trace = np.nanmean(sub, axis=0)
                        if use_sd:
                            err_trace = np.nanstd(sub, axis=0, ddof=1)
                        else:
                            n_valid   = np.sum(~np.isnan(sub), axis=0)
                            err_trace = np.where(n_valid > 1,
                                                  np.nanstd(sub, axis=0, ddof=1) / np.sqrt(n_valid),
                                                  0.0)
                    _yvals_pm.append(float(np.nanmax(mean_trace + err_trace)))
                    _yvals_pm.append(float(np.nanmin(mean_trace - err_trace)))
                    ax.plot(canonical_time, mean_trace, color=color, linewidth=1.8,
                            label=f'n={sub.shape[0]} {row_unit}')
                    ax.fill_between(canonical_time,
                                    mean_trace - err_trace,
                                    mean_trace + err_trace,
                                    color=color, alpha=0.25)
            else:
                ax.text(0.5, 0.5, 'No data', ha='center', va='center',
                        transform=ax.transAxes, fontsize=9)

            ax.legend(fontsize=8, loc='upper right')

        if _yvals_pm:
            _ymax_pm = float(np.nanmax(_yvals_pm))
            _ymin_pm = float(np.nanmin(_yvals_pm))
        else:
            _ymax_pm, _ymin_pm = 1.0, 0.0
        _bot_pm = _ymin_pm * 1.05 if _ymin_pm < 0 else 0.0
        axs[0, 0].set_ylim(_bot_pm, _ymax_pm * 1.05)
        fig_pm.suptitle(f'{title_prefix} — {half_label} (Per Mouse)', fontsize=13)
        fig_pm.tight_layout()

        # ── condition figure ──────────────────────────────────────────────────
        fig_cond, ax_cond = plt.subplots(figsize=(10, 5))

        cond_data: dict = {}
        cond_raw:  dict = {}
        _yvals_cond_half = []
        for result in all_results:
            matrix    = result.get(signal_key)
            indices   = result.get(indices_key)
            condition = result['starting_condition']
            if matrix is None or indices is None or len(indices) == 0:
                continue
            mask = (indices < global_half) if is_early else (indices >= global_half)
            sub  = matrix[mask]
            if sub.shape[0] == 0:
                continue
            with warnings.catch_warnings():
                warnings.simplefilter('ignore', RuntimeWarning)
                cond_data.setdefault(condition, []).append(np.nanmean(sub, axis=0))
            if use_sd:
                cond_raw.setdefault(condition, []).append(sub)

        for condition in sorted(cond_data.keys()):
            color       = condition_color_map.get(condition, 'steelblue')
            mouse_means = np.array(cond_data[condition])
            n_m         = mouse_means.shape[0]
            if show_individual_traces:
                for mm in mouse_means:
                    ax_cond.plot(canonical_time, mm, color=color,
                                 linewidth=0.6, alpha=0.3)
                    _yvals_cond_half.append(float(np.nanmax(mm)))
                    _yvals_cond_half.append(float(np.nanmin(mm)))
            with warnings.catch_warnings():
                warnings.simplefilter('ignore', RuntimeWarning)
                cond_mean = np.nanmean(mouse_means, axis=0)
                if use_sd:
                    pooled   = np.vstack(cond_raw[condition])
                    cond_err = np.nanstd(pooled, axis=0, ddof=1)
                else:
                    cond_err = (np.nanstd(mouse_means, axis=0, ddof=1) / np.sqrt(n_m)
                                if n_m > 1 else np.zeros_like(cond_mean))
            _yvals_cond_half.append(float(np.nanmax(cond_mean + cond_err)))
            _yvals_cond_half.append(float(np.nanmin(cond_mean - cond_err)))
            ax_cond.plot(canonical_time, cond_mean, color=color, linewidth=2.2,
                         label=f'{condition} (n={n_m} mice)')
            ax_cond.fill_between(canonical_time,
                                 cond_mean - cond_err,
                                 cond_mean + cond_err,
                                 color=color, alpha=0.15)

        ax_cond.axvline(0, color='red', linestyle='--', linewidth=1.5, label='Zone entry (t=0)')
        if reward_delivery_vline:
            ax_cond.axvline(0.65, color='black', linestyle='--', linewidth=1.0, label='Reward delivery (t=0.65 s)')
        ax_cond.set_xlabel('Time from reward zone entry (s)')
        ax_cond.set_ylabel(ylabel)
        ax_cond.set_title(f'{title_prefix} — {half_label} (By Condition)')
        ax_cond.set_xlim(-window_s, window_s)
        if _yvals_cond_half:
            _ymax_cond = float(np.nanmax(_yvals_cond_half))
            _ymin_cond = float(np.nanmin(_yvals_cond_half))
        else:
            _ymax_cond, _ymin_cond = 1.0, 0.0
        _bot_cond_half = _ymin_cond * 1.05 if _ymin_cond < 0 else 0.0
        ax_cond.set_ylim(_bot_cond_half, _ymax_cond * 1.05)
        ax_cond.legend()
        ax_cond.spines['top'].set_visible(False)
        ax_cond.spines['right'].set_visible(False)
        fig_cond.tight_layout()

        return fig_pm, fig_cond

    fig_pm_early, fig_cond_early = _make_half_figs('early')
    fig_pm_late,  fig_cond_late  = _make_half_figs('late')
    return fig_pm_early, fig_pm_late, fig_cond_early, fig_cond_late


def analyze_mouse_data(data_files, markers, starting_conditions, transitions_csv_path=None, selected_plots=None, save_lick_plots=False, output_dir=None):
    # Create dictionaries to map mouse names to markers and starting conditions
    markers = {os.path.basename(file).split("_")[0]: marker for file, marker in zip(data_files, markers)}
    conditions = {os.path.basename(file).split("_")[0]: condition for file, condition in zip(data_files, starting_conditions)}
    if selected_plots is None:
        selected_plots = set(_ALL_PLOT_KEYS)

    # Create output directory for lick detection plots if needed
    if save_lick_plots and output_dir:
        lick_plots_dir = os.path.join(output_dir, 'lick_detection_plots')
        os.makedirs(lick_plots_dir, exist_ok=True)
        print(f"\nSaving lick detection plots to: {lick_plots_dir}")
    
    # Create color mapping based on starting conditions
    # Sort so that the condition→color assignment is deterministic across runs
    unique_conditions = sorted(set(starting_conditions))
    condition_color_map = {cond: _condition_to_color(cond) for cond in unique_conditions}
    
    speed_fig             = plt.figure(figsize=(12, 6)) if 'speed'              in selected_plots else None
    sensitivity_fig       = plt.figure(figsize=(12, 6)) if 'sensitivity'        in selected_plots else None
    lick_fig              = plt.figure(figsize=(12, 6)) if 'lick_count'         in selected_plots else None
    reward_fig            = plt.figure(figsize=(12, 6)) if 'reward_count'       in selected_plots else None
    lick_reward_ratio_fig = plt.figure(figsize=(12, 6)) if 'lick_reward_ratio'  in selected_plots else None
    avg_reward_fig        = plt.figure(figsize=(12, 6)) if 'avg_reward'         in selected_plots else None
    sex_reward_fig        = plt.figure(figsize=(12, 6)) if 'sex_reward'         in selected_plots else None
    avg_sex_speed_fig     = plt.figure(figsize=(12, 6)) if 'avg_sex_speed'      in selected_plots else None
    avg_lick_rate_fig     = plt.figure(figsize=(12, 6)) if 'avg_lick_rate'      in selected_plots else None
    sex_lick_rate_fig     = plt.figure(figsize=(12, 6)) if 'sex_lick_rate'      in selected_plots else None
    false_alarm_fig       = plt.figure(figsize=(12, 6)) if 'false_alarms'       in selected_plots else None
    correct_rejection_fig = plt.figure(figsize=(12, 6)) if 'correct_rejections' in selected_plots else None
    specificity_fig       = plt.figure(figsize=(12, 6)) if 'specificity'        in selected_plots else None
    dprime_fig            = plt.figure(figsize=(12, 6)) if 'dprime'             in selected_plots else None
    distance_fig          = plt.figure(figsize=(12, 6)) if 'distance'           in selected_plots else None
    bout_count_fig        = plt.figure(figsize=(12, 6)) if 'bout_count'         in selected_plots else None
    avg_bout_count_fig    = plt.figure(figsize=(12, 6)) if 'avg_bout_count'     in selected_plots else None
    rewards_per_bout_fig  = plt.figure(figsize=(12, 6)) if 'rewards_per_bout'   in selected_plots else None
    first_lick_latency_fig= plt.figure(figsize=(12, 6)) if 'first_lick_latency' in selected_plots else None
    bout_avg_speed_fig    = plt.figure(figsize=(12, 6)) if 'bout_avg_speed'     in selected_plots else None
    bout_avg_dist_fig     = plt.figure(figsize=(12, 6)) if 'bout_avg_dist'      in selected_plots else None
    sex_speed_fig         = plt.figure(figsize=(12, 6)) if 'sex_speed'          in selected_plots else None
    sex_distance_indiv_fig= plt.figure(figsize=(12, 6)) if 'sex_distance_indiv' in selected_plots else None
    sex_reward_indiv_fig  = plt.figure(figsize=(12, 6)) if 'sex_reward_indiv'   in selected_plots else None
    colors = generate_colors(len(data_files))  # Generate colors based on number of mice
    
    all_results = []
    
    for idx, data_file in enumerate(data_files):
        # Read the combined data file
        df = pd.read_csv(data_file, index_col='timestamp')

        # Detect cohort from the file name prefix (e.g. "RV1_data.csv" → "RV1" → RV cohort)
        _mouse_name_prefix = os.path.basename(data_file).split("_")[0]
        _mouse_is_rv = _is_rv_cohort(_mouse_name_prefix)
        if _mouse_is_rv:
            print(f"[RV COHORT] {_mouse_name_prefix}: reward delivery times will be determined "
                  f"by zone-to-reward matching (variable delay, not fixed 0.65 s)")

        print(f"Reading data from: {data_file}")
        
        # Initialize lists to store results
        dates = []
        speeds = []
        total_distances = []  # List for total session distances (cm)
        bout_counts = []  # List for locomotion bout counts per session
        rewards_per_bout_list = []  # List for avg rewards per locomotion bout per session
        avg_speeds_per_bout = []  # List for avg speed per locomotion bout (cm/s)
        avg_dists_per_bout  = []  # List for avg distance per locomotion bout (mm)
        hits = []  # List for reward events
        hits_gap_aware = []  # Reward count excluding events inside capacitive gaps (for lick/reward ratio only)
        misses_list = []  # List for misses (texture changes minus hits)
        sensitivities = []  # List for sensitivity values
        lick_counts = []  # List for daily lick counts
        session_lengths = []  # List for session lengths in minutes
        avg_cap_values = []  # List for per-session mean raw capacitive sensor value
        false_alarms_list = []  # List for false alarm counts
        correct_rejections_list = []  # List for correct rejection counts
        # Per-mouse accumulators for punishment zone percentage
        _mouse_punish_count = 0
        _mouse_total_zone_count = 0
        specificities_list = []  # List for specificity values
        dprimes_list = []  # List for d-prime values
        session_file_errors = {}  # {date_str: [list of missing file labels]}
        speed_epoch_windows_all       = []  # list of 2-D arrays (n_events × EPOCH_N_SAMPLES) per session
        cap_epoch_windows_all         = []  # same for capacitive
        speed_epoch_session_means_all   = []  # list of 1-D arrays — one per-session mean trace
        cap_epoch_session_means_all     = []  # same for capacitive
        speed_epoch_session_indices_all = []  # 0-based session position in df for each speed mean
        cap_epoch_session_indices_all   = []  # same for capacitive
        speed_epoch_event_indices_all   = []  # session index repeated for every event row in speed_epoch_matrix
        cap_epoch_event_indices_all     = []  # same for capacitive
        punish_speed_epoch_windows_all         = []  # list of 2-D arrays per session (punishment zone)
        punish_cap_epoch_windows_all           = []  # same for capacitive signal (punishment zone)
        punish_speed_epoch_session_means_all   = []  # per-session mean traces (punishment zone)
        punish_cap_epoch_session_means_all     = []  # same for capacitive (punishment zone)
        punish_speed_epoch_session_indices_all = []  # 0-based session index for each punish speed mean
        punish_cap_epoch_session_indices_all   = []  # same for capacitive (punishment zone)
        punish_speed_epoch_event_indices_all   = []  # session index repeated per event row (punishment zone)
        punish_cap_epoch_event_indices_all     = []  # same for capacitive (punishment zone)
        lick_epoch_session_means_all           = []  # per-session mean lick-count trace (reward zone)
        punish_lick_epoch_session_means_all    = []  # per-session mean lick-count trace (punishment zone)
        avg_lick_latency_list                  = []  # per-session avg first-lick latency after reward delivery (s)
        lick_after_reward_prop_list            = []  # per-session proportion of reward deliveries with ≥1 lick in 2 s post-delivery window

        # Process each date's data
        for _sess_idx, (timestamp, row) in enumerate(df.iterrows()):
            date_str = datetime.fromtimestamp(int(timestamp)).strftime('%Y-%m-%d')
            missing_files = []
            try:
                # ── Session-level cache check ─────────────────────────────────────
                _s_paths = [
                    str(row.get('treadmill',  '') or ''),
                    str(row.get('capacitive', '') or ''),
                    str(row.get('trial_log',  '') or ''),
                ]
                _s_key   = _session_cache_key(_s_paths)
                _s_cache = _load_session_cache(_s_key)

                if _s_cache is not None:
                    # ── Cache HIT: restore pre-computed session data ───────────────
                    print(f"  [CACHE] {date_str}: loading pre-computed session data")
                    _c = _s_cache
                    missing_files = _c.get('missing_files', [])

                    if missing_files == ['treadmill', 'capacitive', 'trial_log']:
                        session_file_errors[date_str] = missing_files
                        print(f"  [WARN] {date_str}: all files missing — skipping (cached)")
                        continue
                    if missing_files:
                        session_file_errors[date_str] = missing_files

                    # Restore gap info (needed for gap printing below)
                    _cap_gap_info = _c['cap_gap_info']
                    if _cap_gap_info['has_gaps']:
                        print(f"  [GAP] {date_str}: {_cap_gap_info['n_gaps']} gap(s) in "
                              f"capacitive file (cached) — total trimmed: {_cap_gap_info['total_gap_s']:.2f}s")
                    else:
                        print(f"  [GAP] {date_str}: no gaps (cached)")

                    # Restore scalars
                    date                    = _c['date']
                    avg_speed               = _c['avg_speed']
                    total_distance          = _c['total_distance']
                    bout_count              = _c['bout_count']
                    rewards_per_bout_val    = _c['rewards_per_bout_val']
                    avg_speed_per_bout      = _c['avg_speed_per_bout']
                    avg_dist_per_bout       = _c['avg_dist_per_bout']
                    reward_count            = _c['reward_count']
                    _gap_aware_reward_count = _c['gap_aware_reward_count']
                    misses                  = _c['misses']
                    sensitivity             = _c['sensitivity']
                    lick_count              = _c['lick_count']
                    session_length_minutes  = _c['session_length_minutes']
                    avg_cap_session         = _c['avg_cap_session']
                    false_alarm_count       = _c['false_alarm_count']
                    correct_rejection_count = _c['correct_rejection_count']
                    specificity             = _c['specificity']
                    dprime                  = _c['dprime']
                    _first_lick_lat         = _c['first_lick_lat']
                    _lick_after_rew_prop    = _c['lick_after_rew_prop']

                    # Accumulate punishment zone percentage deltas
                    _mouse_punish_count      += _c.get('punish_count_delta', 0)
                    _mouse_total_zone_count  += _c.get('total_zone_count_delta', 0)

                    # Append scalars to per-mouse result lists
                    dates.append(date)
                    speeds.append(avg_speed)
                    total_distances.append(total_distance)
                    bout_counts.append(bout_count)
                    rewards_per_bout_list.append(rewards_per_bout_val)
                    avg_speeds_per_bout.append(avg_speed_per_bout)
                    avg_dists_per_bout.append(avg_dist_per_bout)
                    hits.append(reward_count)
                    hits_gap_aware.append(_gap_aware_reward_count)
                    misses_list.append(misses)
                    sensitivities.append(sensitivity)
                    lick_counts.append(lick_count)
                    session_lengths.append(session_length_minutes)
                    avg_cap_values.append(avg_cap_session)
                    false_alarms_list.append(false_alarm_count)
                    correct_rejections_list.append(correct_rejection_count)
                    specificities_list.append(specificity)
                    dprimes_list.append(dprime)
                    avg_lick_latency_list.append(_first_lick_lat)
                    lick_after_reward_prop_list.append(_lick_after_rew_prop)

                    # Restore epoch matrices and append to accumulator lists
                    _sp_mat_c = _c.get('sp_epoch_mat')
                    if _sp_mat_c is not None:
                        speed_epoch_windows_all.append(_sp_mat_c)
                        with warnings.catch_warnings():
                            warnings.simplefilter('ignore', RuntimeWarning)
                            speed_epoch_session_means_all.append(np.nanmean(_sp_mat_c, axis=0))
                        speed_epoch_session_indices_all.append(_sess_idx)
                        speed_epoch_event_indices_all.extend([_sess_idx] * _sp_mat_c.shape[0])

                    _cp_mat_c = _c.get('cp_epoch_mat')
                    if _cp_mat_c is not None:
                        cap_epoch_windows_all.append(_cp_mat_c)
                        with warnings.catch_warnings():
                            warnings.simplefilter('ignore', RuntimeWarning)
                            cap_epoch_session_means_all.append(np.nanmean(_cp_mat_c, axis=0))
                        cap_epoch_session_indices_all.append(_sess_idx)
                        cap_epoch_event_indices_all.extend([_sess_idx] * _cp_mat_c.shape[0])

                    _lk_mean_c = _c.get('lick_epoch_mean')
                    if _lk_mean_c is not None:
                        lick_epoch_session_means_all.append(_lk_mean_c)

                    _psp_mat_c = _c.get('punish_sp_epoch_mat')
                    if _psp_mat_c is not None:
                        punish_speed_epoch_windows_all.append(_psp_mat_c)
                        with warnings.catch_warnings():
                            warnings.simplefilter('ignore', RuntimeWarning)
                            punish_speed_epoch_session_means_all.append(np.nanmean(_psp_mat_c, axis=0))
                        punish_speed_epoch_session_indices_all.append(_sess_idx)
                        punish_speed_epoch_event_indices_all.extend([_sess_idx] * _psp_mat_c.shape[0])

                    _pcp_mat_c = _c.get('punish_cp_epoch_mat')
                    if _pcp_mat_c is not None:
                        punish_cap_epoch_windows_all.append(_pcp_mat_c)
                        with warnings.catch_warnings():
                            warnings.simplefilter('ignore', RuntimeWarning)
                            punish_cap_epoch_session_means_all.append(np.nanmean(_pcp_mat_c, axis=0))
                        punish_cap_epoch_session_indices_all.append(_sess_idx)
                        punish_cap_epoch_event_indices_all.extend([_sess_idx] * _pcp_mat_c.shape[0])

                    _plk_mean_c = _c.get('punish_lick_epoch_mean')
                    if _plk_mean_c is not None:
                        punish_lick_epoch_session_means_all.append(_plk_mean_c)

                    continue  # skip all normal processing for this session
                # ── End cache HIT ─────────────────────────────────────────────────

                # ── Cache MISS: read CSVs in parallel, process normally ────────────
                treadmill_data, capacitive_data, trial_log, missing_files = _read_three_csvs(
                    _s_paths[0], _s_paths[1], _s_paths[2])

                # ── Gap detection on freshly-loaded capacitive data ───────────────
                if capacitive_data is not None:
                    _cap_gap_info = detect_capacitive_gaps(capacitive_data)
                    if _cap_gap_info['has_gaps']:
                        print(f"  [GAP] {date_str}: {_cap_gap_info['n_gaps']} gap(s) detected in "
                              f"capacitive file — total trimmed time: {_cap_gap_info['total_gap_s']:.2f}s")
                        for _gi, _g in enumerate(_cap_gap_info['gaps']):
                            print(f"    gap {_gi+1}: {_g[0]:.3f}s → {_g[1]:.3f}s "
                                  f"(duration: {_g[2]:.3f}s)")
                    else:
                        print(f"  [GAP] {date_str}: no gaps detected — file is continuous")
                else:
                    _cap_gap_info = {'has_gaps': False, 'n_gaps': 0, 'gaps': [], 'total_gap_s': 0.0}

                if len(missing_files) == 3:
                    # All three files missing — nothing to process for this date
                    session_file_errors[date_str] = missing_files
                    print(f"  [WARN] {date_str}: all files missing — skipping session entirely")
                    continue

                if missing_files:
                    session_file_errors[date_str] = missing_files
                    print(f"  [WARN] {date_str}: missing files: {', '.join(missing_files)} — partial session kept")

                # ── Treadmill-derived metrics ─────────────────────────────
                if treadmill_data is not None:
                    _speed_raw = treadmill_data['speed']
                    # Butterworth low-pass (0.25 Hz, order 3) applied to raw speed
                    _fs = 1.0 / treadmill_data['global_time'].diff().median()
                    _b, _a = butter(3, 0.25 / (_fs / 2.0), btype='low')
                    _speed_filt = filtfilt(_b, _a, _speed_raw)
                    avg_speed = float(np.mean(_speed_filt)) / 10.0
                    _, total_distance = compute_session_distance(treadmill_data)
                    # Bout detection on filtered speed (mm/s → cm/s)
                    _time_arr = treadmill_data['global_time'].values
                    _speed_cm = _speed_filt / 10.0
                    _bouts = detect_locomotion_bouts(_time_arr, _speed_cm)
                    _session_bouts = _bouts  # saved for rewards_per_bout computation below
                    bout_count = len(_bouts)
                    # Per-bout averages (speed in cm/s; dist in mm)
                    if _bouts:
                        _bout_speed_means = []
                        _bout_dist_means  = []
                        for _t0, _t1 in _bouts:
                            _bmask = (_time_arr >= _t0) & (_time_arr <= _t1)
                            if np.any(_bmask):
                                _bout_speed_means.append(float(np.mean(_speed_cm[_bmask])))
                            if 'distance' in treadmill_data.columns:
                                _dist_vals = pd.to_numeric(
                                    treadmill_data['distance'], errors='coerce').values[_bmask]
                                _dist_vals = _dist_vals[~np.isnan(_dist_vals)]
                                if len(_dist_vals) >= 2:
                                    _bout_dist_means.append(float(_dist_vals[-1] - _dist_vals[0]))
                        avg_speed_per_bout = (float(np.mean(_bout_speed_means))
                                             if _bout_speed_means else float('nan'))
                        avg_dist_per_bout  = (float(np.mean(_bout_dist_means))
                                             if _bout_dist_means  else float('nan'))
                    else:
                        avg_speed_per_bout = float('nan')
                        avg_dist_per_bout  = float('nan')
                else:
                    avg_speed = float('nan')
                    total_distance = float('nan')
                    bout_count = float('nan')
                    avg_speed_per_bout = float('nan')
                    avg_dist_per_bout  = float('nan')
                    _session_bouts = []

                # ── Capacitive-derived metrics ────────────────────────────
                if capacitive_data is not None:
                    session_length_minutes = float(capacitive_data['elapsed_time'].max()) / 60.0

                    cap_df = capacitive_data.copy()
                    cap_df['Time_sec'] = cap_df['elapsed_time']

                    # Compute KDE normalization (with caching)
                    capacitive_filepath = row['capacitive']
                    kde_value = get_cached_kde(capacitive_filepath)
                    if kde_value is None:
                        kde_value = lda.compute_KDE(cap_df, 'capacitive_value')
                        cache_kde_value(capacitive_filepath, kde_value)

                    cap_df = lda.compute_KDE_normalizations(cap_df, 'capacitive_value', kde_value)
                    events_df, threshold_used = lda.detect_events_above_threshold(cap_df, 'capacitive_value', threshold=None)
                    lick_count = events_df['capacitive_value_event'].sum()
                    # Store timestamps of detected lick events for epoch analysis
                    _sess_lick_times = events_df.loc[
                        events_df['capacitive_value_event'] == 1, 'Time_sec'
                    ].values.astype(float)

                    if save_lick_plots and output_dir:
                        mouse_name_plot = os.path.basename(data_file).split('_')[0]
                        plot_filename = f"{mouse_name_plot}_{date_str}_lick_detection.png"
                        plot_path = os.path.join(lick_plots_dir, plot_filename)
                        fig = lda.plot_summary(
                            cap_df, events_df,
                            column='capacitive_value',
                            kde_value=kde_value,
                            threshold=threshold_used,
                            title=f"{mouse_name_plot} - {date_str} - {lick_count} licks detected",
                            show=False
                        )
                        fig.savefig(plot_path, dpi=150, bbox_inches='tight')
                        plt.close(fig)
                    # Per-session mean of the raw capacitive sensor signal
                    _cap_raw = pd.to_numeric(
                        capacitive_data['capacitive_value'], errors='coerce').dropna()
                    avg_cap_session = float(np.nanmean(_cap_raw)) if len(_cap_raw) > 0 else float('nan')
                else:
                    # Fall back to treadmill global_time span if available
                    if treadmill_data is not None and 'global_time' in treadmill_data.columns:
                        _tm_times = pd.to_numeric(treadmill_data['global_time'], errors='coerce').dropna()
                        session_length_minutes = float(_tm_times.max()) / 60.0 if len(_tm_times) > 0 else float('nan')
                    else:
                        session_length_minutes = float('nan')
                    lick_count = float('nan')
                    avg_cap_session = float('nan')

                # ── Trial-log-derived metrics ─────────────────────────────
                if trial_log is not None:
                    # Extract ALL stay zone entry times (if column exists; older data may not have it)
                    if 'stay_texture_change_time' in trial_log.columns:
                        stay_zone_times = pd.to_numeric(trial_log['stay_texture_change_time'], errors='coerce').dropna().values

                        # Collect all re-entry times to exclude them
                        re_entry_times_set = set()
                        if 'zone_re_entry_time' in trial_log.columns:
                            for val in trial_log['zone_re_entry_time']:
                                if pd.notna(val) and val != '':
                                    try:
                                        # Handle both single values and arrays
                                        if isinstance(val, str) and val.strip():
                                            import ast
                                            try:
                                                re_entry_list = ast.literal_eval(val)
                                                if isinstance(re_entry_list, (list, tuple)):
                                                    for t in re_entry_list:
                                                        if pd.notna(t):
                                                            re_entry_times_set.add(float(t))
                                                else:
                                                    re_entry_times_set.add(float(re_entry_list))
                                            except:
                                                pass
                                        elif not isinstance(val, str):
                                            re_entry_times_set.add(float(val))
                                    except (ValueError, TypeError):
                                        pass

                        # Filter out stay_zone_times that match re-entries (within 0.05 seconds)
                        valid_zone_times = []
                        for zone_time in stay_zone_times:
                            is_reentry = False
                            for re_entry_time in re_entry_times_set:
                                if abs(zone_time - re_entry_time) <= 0.05:
                                    is_reentry = True
                                    break
                            if not is_reentry:
                                valid_zone_times.append(zone_time)

                        valid_zone_times = np.array(valid_zone_times)
                        total_opportunities = len(valid_zone_times)
                    else:
                        # Fallback for older data without stay_texture_change_time
                        total_opportunities = len(trial_log[trial_log['texture_history'] == 'assets/reward_mean100.jpg'])

                    # Collect all reward_event and hits_event times
                    reward_event_times = pd.to_numeric(trial_log['reward_event'], errors='coerce').dropna().values

                    # Check if hits_event column exists (older data may not have it)
                    if 'hits_event' in trial_log.columns:
                        hits_event_times = pd.to_numeric(trial_log['hits_event'], errors='coerce').dropna().values
                    else:
                        hits_event_times = np.array([])  # Empty array if column doesn't exist

                    # Simple counting: hits = rewards + hits, misses = opportunities - hits
                    if total_opportunities > 0:
                        reward_count = len(reward_event_times) + len(hits_event_times)
                        misses = total_opportunities - reward_count
                        sensitivity = float(reward_count) / float(total_opportunities) if total_opportunities > 0 else 0.0
                    else:
                        reward_count = 0
                        misses = 0
                        sensitivity = float('nan')

                    # Compute puff opportunities and false alarms
                    puff_event_times = pd.to_numeric(trial_log['puff_event'], errors='coerce').dropna().values

                    if 'go_texture_change_time' in trial_log.columns:
                        go_zone_times = np.sort(
                            pd.to_numeric(trial_log['go_texture_change_time'], errors='coerce').dropna().values
                        )
                        puff_opportunities = len(go_zone_times)

                        # Count false alarms: each go zone that had at least one puff event is one false alarm.
                        false_alarm_count = 0
                        for z_idx, zone_start in enumerate(go_zone_times):
                            zone_end = go_zone_times[z_idx + 1] if z_idx + 1 < len(go_zone_times) else np.inf
                            zone_puffs = puff_event_times[
                                (puff_event_times >= zone_start) & (puff_event_times < zone_end)
                            ]
                            if len(zone_puffs) > 0:
                                false_alarm_count += 1
                    else:
                        # Older format: use texture_change_time from punish rows as zone start times
                        punish_rows = trial_log[trial_log['texture_history'] == 'assets/punish_mean100.jpg']
                        puff_opportunities = len(punish_rows)
                        punish_zone_times = np.sort(
                            pd.to_numeric(punish_rows['texture_change_time'], errors='coerce').dropna().values
                        )

                        false_alarm_count = 0
                        for z_idx, zone_start in enumerate(punish_zone_times):
                            zone_end = punish_zone_times[z_idx + 1] if z_idx + 1 < len(punish_zone_times) else np.inf
                            zone_puffs = puff_event_times[
                                (puff_event_times >= zone_start) & (puff_event_times < zone_end)
                            ]
                            if len(zone_puffs) > 0:
                                false_alarm_count += 1

                    # Compute correct rejections and specificity
                    if puff_opportunities > 0:
                        correct_rejection_count = puff_opportunities - false_alarm_count
                        specificity = float(correct_rejection_count) / float(puff_opportunities)
                    else:
                        false_alarm_count = 0
                        correct_rejection_count = 0
                        specificity = float('nan')

                    # Compute d-prime (signal detection theory)
                    # Log-linear correction handles edge cases of hit/FA rate = 0 or 1
                    if total_opportunities > 0 and puff_opportunities > 0:
                        hr_corrected = (reward_count + 0.5) / (total_opportunities + 1.0)
                        fa_corrected = (false_alarm_count + 0.5) / (puff_opportunities + 1.0)
                        dprime = float(norm.ppf(hr_corrected) - norm.ppf(fa_corrected))
                    else:
                        dprime = float('nan')
                else:
                    reward_count = float('nan')
                    misses = float('nan')
                    sensitivity = float('nan')
                    false_alarm_count = float('nan')
                    correct_rejection_count = float('nan')
                    specificity = float('nan')
                    dprime = float('nan')

                # ── Gap-aware reward count (lick/reward ratio denominator only) ──────
                # Exclude reward/hits events whose timestamp falls inside a capacitive
                # gap window.  All other reward_count uses remain unfiltered.
                if _cap_gap_info['has_gaps'] and not (
                    isinstance(reward_count, float) and np.isnan(reward_count)
                ):
                    _gaps = _cap_gap_info['gaps']
                    _rew_in_gap = np.array([
                        any(g[0] <= t <= g[1] for g in _gaps)
                        for t in reward_event_times
                    ]) if len(reward_event_times) > 0 else np.array([], dtype=bool)
                    _hits_in_gap = np.array([
                        any(g[0] <= t <= g[1] for g in _gaps)
                        for t in hits_event_times
                    ]) if len(hits_event_times) > 0 else np.array([], dtype=bool)
                    _gap_aware_reward_count = int(
                        np.sum(~_rew_in_gap) + np.sum(~_hits_in_gap)
                    )
                else:
                    # No gaps or no trial log: use the same value as reward_count
                    _gap_aware_reward_count = reward_count

                # Convert Unix timestamp to datetime and store results
                date = datetime.fromtimestamp(int(timestamp))

                # ── Rewards per locomotion bout ───────────────────────────
                # Assign each reward event to a bout; compute mean rewards/bout
                if _session_bouts and trial_log is not None:
                    try:
                        _all_reward_ts = np.concatenate([
                            pd.to_numeric(trial_log['reward_event'], errors='coerce').dropna().values,
                            (pd.to_numeric(trial_log['hits_event'], errors='coerce').dropna().values
                             if 'hits_event' in trial_log.columns else np.array([])),
                        ])
                        _rewards_in_bout = 0
                        for _bt0, _bt1 in _session_bouts:
                            _rewards_in_bout += int(np.sum(
                                (_all_reward_ts >= _bt0) & (_all_reward_ts <= _bt1)
                            ))
                        rewards_per_bout_val = _rewards_in_bout / len(_session_bouts)
                    except Exception:
                        rewards_per_bout_val = float('nan')
                else:
                    rewards_per_bout_val = float('nan')

                # ── Punishment zone percentage (accumulated across all sessions) ──
                _sess_punish_count_delta      = 0
                _sess_total_zone_count_delta  = 0
                if trial_log is not None and 'texture_history' in trial_log.columns:
                    _th = trial_log['texture_history'].dropna()
                    _sess_total_zone_count_delta  = len(_th)
                    _sess_punish_count_delta       = int((_th == 'assets/punish_mean100.jpg').sum())
                    _mouse_total_zone_count += _sess_total_zone_count_delta
                    _mouse_punish_count     += _sess_punish_count_delta

                # ── First-lick latency after reward delivery ──────────────────────
                # For each reward delivery, find the first lick that follows it and
                # record the latency.  Average across all trials with a valid first lick.
                #
                # Standard cohorts : reward delivered 0.65 s after zone entry.
                # RV cohort        : actual delivery time determined by zone-to-reward
                #                    matching (_match_rewards_to_zones).
                _first_lick_lat = float('nan')
                _sess_lick_arr = _sess_lick_times if '_sess_lick_times' in dir() else np.array([])
                if trial_log is not None and len(_sess_lick_arr) > 0:
                    try:
                        # Build (zone_entry_time, reward_delivery_time) pairs
                        if _mouse_is_rv:
                            _fll_pairs = _match_rewards_to_zones(trial_log)
                        else:
                            _fll_pairs = [
                                (_zt, _zt + 0.65)
                                for _zt in _extract_reward_zone_entry_times(trial_log)
                            ]
                        if _DIAGNOSTIC_MODE:
                            _diag_method = 'zone-to-reward match' if _mouse_is_rv else 'zone+0.65s'
                            print(f"  [FLL] {_mouse_name_prefix} | {date_str} | "
                                  f"{len(_fll_pairs)} pairs via {_diag_method}")
                        _fll_latencies = []
                        for _fll_zt, _fll_rd in _fll_pairs:
                            # Drop trial if the reward delivery itself lands inside a gap.
                            if _cap_gap_info['has_gaps']:
                                _delivery_in_gap = [
                                    g for g in _cap_gap_info['gaps']
                                    if g[0] <= _fll_rd <= g[1]
                                ]
                                if _delivery_in_gap:
                                    if _DIAGNOSTIC_MODE:
                                        _dg = _delivery_in_gap[0]
                                        print(f"    trial zone@{_fll_zt:.3f}s delivery@{_fll_rd:.3f}s "
                                              f"→ SKIP (delivery inside gap {_dg[0]:.3f}–{_dg[1]:.3f}s)")
                                    continue
                            _fll_candidates = _sess_lick_arr[_sess_lick_arr > _fll_rd]
                            if len(_fll_candidates) == 0:
                                if _DIAGNOSTIC_MODE:
                                    print(f"    trial zone@{_fll_zt:.3f}s delivery@{_fll_rd:.3f}s "
                                          f"→ SKIP (no lick after delivery)")
                                continue
                            _first_lick = _fll_candidates[0]
                            # Drop trial if a gap interrupts the window between
                            # reward delivery and the first detected lick.
                            if _cap_gap_info['has_gaps']:
                                _blocking_gaps = [
                                    g for g in _cap_gap_info['gaps']
                                    if g[0] > _fll_rd and g[0] < _first_lick
                                ]
                                if _blocking_gaps:
                                    if _DIAGNOSTIC_MODE:
                                        print(f"    trial zone@{_fll_zt:.3f}s delivery@{_fll_rd:.3f}s "
                                              f"→ SKIP (gap {_blocking_gaps[0][0]:.3f}–{_blocking_gaps[0][1]:.3f}s "
                                              f"before first lick@{_first_lick:.3f}s)")
                                    continue
                            _lat = float(_first_lick - _fll_rd)
                            if _DIAGNOSTIC_MODE:
                                print(f"    trial zone@{_fll_zt:.3f}s delivery@{_fll_rd:.3f}s "
                                      f"first_lick@{_first_lick:.3f}s → latency {_lat:.3f}s")
                            _fll_latencies.append(_lat)
                        if _fll_latencies:
                            _first_lick_lat = float(np.mean(_fll_latencies))
                        if _DIAGNOSTIC_MODE:
                            print(f"  [FLL] session mean latency: "
                                  f"{_first_lick_lat:.3f}s  ({len(_fll_latencies)}/{len(_fll_pairs)} trials valid)")
                    except Exception:
                        _first_lick_lat = float('nan')
                avg_lick_latency_list.append(_first_lick_lat)

                # ── Proportion of reward deliveries with licks (2 s post-delivery window) ──
                # Window: [delivery_time, delivery_time + 2 s]. Binary per delivery:
                # 1 if any lick detected in window, 0 otherwise. NaN if data unavailable.
                #
                # Standard cohorts : delivery_time = zone_entry + 0.65 s.
                # RV cohort        : delivery_time determined by zone-to-reward matching.
                _lick_after_rew_prop = float('nan')
                if trial_log is not None and capacitive_data is not None:
                    try:
                        # Build list of actual reward delivery times for this session
                        if _mouse_is_rv:
                            _lar_delivery_times = [
                                rd for _, rd in _match_rewards_to_zones(trial_log)
                            ]
                        else:
                            _lar_delivery_times = [
                                _zt + 0.65
                                for _zt in _extract_reward_zone_entry_times(trial_log)
                            ]
                        if _DIAGNOSTIC_MODE:
                            _diag_method = 'zone-to-reward match' if _mouse_is_rv else 'zone+0.65s'
                            print(f"  [LAR] {_mouse_name_prefix} | {date_str} | "
                                  f"{len(_lar_delivery_times)} deliveries via {_diag_method}")
                        if _lar_delivery_times:
                            _lar_binary = []
                            for _lar_rd in _lar_delivery_times:
                                _lar_start = _lar_rd
                                _lar_end   = _lar_rd + 2.0  # 2 s post-delivery detection window
                                # Drop trial if a gap falls anywhere in the detection window.
                                if _cap_gap_info['has_gaps']:
                                    _lar_blocking = [
                                        g for g in _cap_gap_info['gaps']
                                        if g[0] < _lar_end and g[1] > _lar_start
                                    ]
                                    if _lar_blocking:
                                        if _DIAGNOSTIC_MODE:
                                            print(f"    delivery@{_lar_rd:.3f}s window [{_lar_start:.3f},{_lar_end:.3f}] "
                                                  f"→ SKIP (gap {_lar_blocking[0][0]:.3f}–{_lar_blocking[0][1]:.3f}s)")
                                        continue
                                _hit = int(np.any(
                                    (_sess_lick_arr >= _lar_start) & (_sess_lick_arr < _lar_end)
                                ))
                                if _DIAGNOSTIC_MODE:
                                    _licks_in_win = _sess_lick_arr[
                                        (_sess_lick_arr >= _lar_start) & (_sess_lick_arr < _lar_end)
                                    ]
                                    _hit_str = f"HIT  licks={list(np.round(_licks_in_win, 3))}" if _hit else "MISS"
                                    print(f"    delivery@{_lar_rd:.3f}s window [{_lar_start:.3f},{_lar_end:.3f}] → {_hit_str}")
                                _lar_binary.append(_hit)
                            _lick_after_rew_prop = (float(np.mean(_lar_binary))
                                                    if _lar_binary else float('nan'))
                        if _DIAGNOSTIC_MODE:
                            _n_valid = len(_lar_binary) if _lar_delivery_times else 0
                            print(f"  [LAR] session proportion: "
                                  f"{_lick_after_rew_prop:.3f}  ({_n_valid}/{len(_lar_delivery_times)} trials valid)")
                    except Exception:
                        _lick_after_rew_prop = float('nan')
                lick_after_reward_prop_list.append(_lick_after_rew_prop)

                dates.append(date)
                speeds.append(avg_speed)
                total_distances.append(total_distance)
                bout_counts.append(bout_count)
                rewards_per_bout_list.append(rewards_per_bout_val)
                avg_speeds_per_bout.append(avg_speed_per_bout)
                avg_dists_per_bout.append(avg_dist_per_bout)
                hits.append(reward_count)
                hits_gap_aware.append(_gap_aware_reward_count)
                misses_list.append(misses)
                sensitivities.append(sensitivity)
                lick_counts.append(lick_count)
                session_lengths.append(session_length_minutes)
                avg_cap_values.append(avg_cap_session)
                false_alarms_list.append(false_alarm_count)
                correct_rejections_list.append(correct_rejection_count)
                specificities_list.append(specificity)
                dprimes_list.append(dprime)

                # ── Pre-sample continuous signals once (shared for both epoch blocks) ─
                # Avoids reading and interpolating the same large arrays twice per session.
                _sess_sp_time = _sess_sp_val = None
                _sess_cp_time = _sess_cp_val_z = None
                if treadmill_data is not None:
                    _sess_sp_time, _sess_sp_val = uniformly_sample_treadmill(treadmill_data)
                if capacitive_data is not None:
                    _sess_cp_time, _sess_cp_val_raw = uniformly_sample_capacitive(capacitive_data)
                    _cp_mu  = np.nanmean(_sess_cp_val_raw)
                    _cp_sig = np.nanstd(_sess_cp_val_raw, ddof=1)
                    if _cp_sig > 0:
                        _sess_cp_val_z = (_sess_cp_val_raw - _cp_mu) / _cp_sig
                    else:
                        _sess_cp_val_z = _sess_cp_val_raw - _cp_mu

                # Record list lengths before epoch extraction for reliable cache harvesting
                _n_sp_before  = len(speed_epoch_windows_all)
                _n_cp_before  = len(cap_epoch_windows_all)
                _n_lk_before  = len(lick_epoch_session_means_all)
                _n_psp_before = len(punish_speed_epoch_windows_all)
                _n_pcp_before = len(punish_cap_epoch_windows_all)
                _n_plk_before = len(punish_lick_epoch_session_means_all)

                # ── Behavioral epoch extraction (reward zone entry) ──────────────────
                if trial_log is not None and (treadmill_data is not None or capacitive_data is not None):
                    try:
                        _zone_times = _extract_reward_zone_entry_times(trial_log)
                        # Filtered list for capacitive/lick epochs: drop trials whose
                        # ±EPOCH_WINDOW_S window is interrupted by a trimmed gap.
                        _zone_times_cap = _filter_event_times_by_gaps(_zone_times, _cap_gap_info)
                        if _zone_times:
                            if treadmill_data is not None and _sess_sp_time is not None:
                                # Use pre-sampled (unfiltered) speed — computed once above.
                                # Speed epochs are NOT gap-filtered (treadmill data is untrimmed).
                                _sp_mat = _build_epoch_matrix(_sess_sp_time, _sess_sp_val, _zone_times)
                                if _sp_mat is not None:
                                    speed_epoch_windows_all.append(_sp_mat)
                                    with warnings.catch_warnings():
                                        warnings.simplefilter('ignore', RuntimeWarning)
                                        speed_epoch_session_means_all.append(
                                            np.nanmean(_sp_mat, axis=0))
                                    speed_epoch_session_indices_all.append(_sess_idx)
                                    speed_epoch_event_indices_all.extend(
                                        [_sess_idx] * _sp_mat.shape[0])
                            if capacitive_data is not None and _sess_cp_time is not None:
                                # Use pre-sampled, z-scored capacitive signal — computed once above.
                                _cp_mat = _build_epoch_matrix(_sess_cp_time, _sess_cp_val_z, _zone_times_cap)
                                if _cp_mat is not None:
                                    cap_epoch_windows_all.append(_cp_mat)
                                    with warnings.catch_warnings():
                                        warnings.simplefilter('ignore', RuntimeWarning)
                                        cap_epoch_session_means_all.append(
                                            np.nanmean(_cp_mat, axis=0))
                                    cap_epoch_session_indices_all.append(_sess_idx)
                                    cap_epoch_event_indices_all.extend(
                                        [_sess_idx] * _cp_mat.shape[0])
                            # ── Lick count epoch (reward zone) ───────────────────────────
                            if capacitive_data is not None and '_sess_lick_times' in dir():
                                _lk_mat = _build_lick_epoch_matrix(_sess_lick_times, _zone_times_cap)
                                if _lk_mat is not None:
                                    with warnings.catch_warnings():
                                        warnings.simplefilter('ignore', RuntimeWarning)
                                        lick_epoch_session_means_all.append(
                                            np.nanmean(_lk_mat, axis=0))
                    except Exception as _epoch_err:
                        print(f"  [WARN] {date_str}: epoch extraction failed — {_epoch_err}")

                # ── Behavioral epoch extraction (punishment zone entry) ──────────────
                if trial_log is not None and (treadmill_data is not None or capacitive_data is not None):
                    try:
                        _punish_times = _extract_punish_zone_entry_times(trial_log)
                        # Filtered list for capacitive/lick epochs only.
                        _punish_times_cap = _filter_event_times_by_gaps(_punish_times, _cap_gap_info)
                        if _punish_times:
                            if treadmill_data is not None and _sess_sp_time is not None:
                                # Use pre-sampled speed — already computed above.
                                # Speed epochs are NOT gap-filtered (treadmill data is untrimmed).
                                _sp_mat = _build_epoch_matrix(_sess_sp_time, _sess_sp_val, _punish_times)
                                if _sp_mat is not None:
                                    punish_speed_epoch_windows_all.append(_sp_mat)
                                    with warnings.catch_warnings():
                                        warnings.simplefilter('ignore', RuntimeWarning)
                                        punish_speed_epoch_session_means_all.append(
                                            np.nanmean(_sp_mat, axis=0))
                                    punish_speed_epoch_session_indices_all.append(_sess_idx)
                                    punish_speed_epoch_event_indices_all.extend(
                                        [_sess_idx] * _sp_mat.shape[0])
                            if capacitive_data is not None and _sess_cp_time is not None:
                                # Use pre-sampled, z-scored capacitive signal — computed once above.
                                _cp_mat = _build_epoch_matrix(_sess_cp_time, _sess_cp_val_z, _punish_times_cap)
                                if _cp_mat is not None:
                                    punish_cap_epoch_windows_all.append(_cp_mat)
                                    with warnings.catch_warnings():
                                        warnings.simplefilter('ignore', RuntimeWarning)
                                        punish_cap_epoch_session_means_all.append(
                                            np.nanmean(_cp_mat, axis=0))
                                    punish_cap_epoch_session_indices_all.append(_sess_idx)
                                    punish_cap_epoch_event_indices_all.extend(
                                        [_sess_idx] * _cp_mat.shape[0])
                            # ── Lick count epoch (punishment zone) ──────────────────────
                            if capacitive_data is not None and '_sess_lick_times' in dir():
                                _plk_mat = _build_lick_epoch_matrix(_sess_lick_times, _punish_times_cap)
                                if _plk_mat is not None:
                                    with warnings.catch_warnings():
                                        warnings.simplefilter('ignore', RuntimeWarning)
                                        punish_lick_epoch_session_means_all.append(
                                            np.nanmean(_plk_mat, axis=0))
                    except Exception as _punish_epoch_err:
                        print(f"  [WARN] {date_str}: punish epoch extraction failed — {_punish_epoch_err}")

                # ── Save computed session data to cache ────────────────────────────────────────
                try:
                    # Harvest epoch matrices added during this session using pre-recorded lengths
                    _cache_sp_mat   = speed_epoch_windows_all[-1]             if len(speed_epoch_windows_all)             > _n_sp_before  else None
                    _cache_cp_mat   = cap_epoch_windows_all[-1]               if len(cap_epoch_windows_all)               > _n_cp_before  else None
                    _cache_lk_mean  = lick_epoch_session_means_all[-1]        if len(lick_epoch_session_means_all)        > _n_lk_before  else None
                    _cache_psp_mat  = punish_speed_epoch_windows_all[-1]      if len(punish_speed_epoch_windows_all)      > _n_psp_before else None
                    _cache_pcp_mat  = punish_cap_epoch_windows_all[-1]        if len(punish_cap_epoch_windows_all)        > _n_pcp_before else None
                    _cache_plk_mean = punish_lick_epoch_session_means_all[-1] if len(punish_lick_epoch_session_means_all) > _n_plk_before else None
                    _save_session_cache(_s_key, {
                        'missing_files':             missing_files,
                        'cap_gap_info':              _cap_gap_info,
                        'date':                      date,
                        'avg_speed':                 avg_speed,
                        'total_distance':            total_distance,
                        'bout_count':                bout_count,
                        'rewards_per_bout_val':      rewards_per_bout_val,
                        'avg_speed_per_bout':        avg_speed_per_bout,
                        'avg_dist_per_bout':         avg_dist_per_bout,
                        'reward_count':              reward_count,
                        'gap_aware_reward_count':    _gap_aware_reward_count,
                        'misses':                    misses,
                        'sensitivity':               sensitivity,
                        'lick_count':                lick_count,
                        'session_length_minutes':    session_length_minutes,
                        'avg_cap_session':           avg_cap_session,
                        'false_alarm_count':         false_alarm_count,
                        'correct_rejection_count':   correct_rejection_count,
                        'specificity':               specificity,
                        'dprime':                    dprime,
                        'first_lick_lat':            _first_lick_lat,
                        'lick_after_rew_prop':       _lick_after_rew_prop,
                        'punish_count_delta':        _sess_punish_count_delta,
                        'total_zone_count_delta':    _sess_total_zone_count_delta,
                        'sp_epoch_mat':              _cache_sp_mat,
                        'cp_epoch_mat':              _cache_cp_mat,
                        'lick_epoch_mean':           _cache_lk_mean,
                        'punish_sp_epoch_mat':       _cache_psp_mat,
                        'punish_cp_epoch_mat':       _cache_pcp_mat,
                        'punish_lick_epoch_mean':    _cache_plk_mean,
                    })
                except Exception as _cache_err:
                    print(f"  [WARN] {date_str}: session cache save failed — {_cache_err}")

            except Exception as e:
                print(f"Error processing date {timestamp}: {str(e)}")
                continue
        
        # Create results DataFrame
        results_df = pd.DataFrame({
            'date': dates,
            'average_speed': speeds,
            'total_distance': total_distances,
            'bout_count': bout_counts,
            'rewards_per_bout': rewards_per_bout_list,
            'avg_speed_per_bout': avg_speeds_per_bout,
            'avg_dist_per_bout':  avg_dists_per_bout,
            'hits': hits,
            'misses': misses_list,
            'sensitivity': sensitivities,
            'lick_count': lick_counts,
            'session_length': session_lengths,
            'avg_cap': avg_cap_values,
            'false_alarms': false_alarms_list,
            'correct_rejections': correct_rejections_list,
            'specificity': specificities_list,
            'dprime': dprimes_list,
            'avg_lick_latency': avg_lick_latency_list,
            'lick_after_reward_prop': lick_after_reward_prop_list,
            'hits_gap_aware': hits_gap_aware,
        })
        
        # Sort and remove duplicates
        results_df = results_df.drop_duplicates(subset=['date'])
        results_df = results_df.sort_values('date').reset_index(drop=True)

        # Assign each session a weekday in the fixed 4-day training cycle.
        # The schedule is always: Mon → Tue → Thu → Fri → Mon → ...
        _DOW_CYCLE = ['Monday', 'Tuesday', 'Thursday', 'Friday']
        results_df['weekday'] = [_DOW_CYCLE[i % 4] for i in range(len(results_df))]
        
        # Remove the first date as requested for hits, misses, and sensitivity analysis
        results_df.loc[1:, 'hits'] = results_df.loc[1:, 'hits']  # Keep only hits after first date
        results_df.loc[1:, 'misses'] = results_df.loc[1:, 'misses']  # Keep only misses after first date
        results_df.loc[1:, 'sensitivity'] = results_df.loc[1:, 'sensitivity']  # Keep only sensitivity after first date
        
        # Get mouse name
        mouse_name = os.path.basename(data_file).split("_")[0]

        # Stack per-session epoch window matrices into one array per signal
        speed_epoch_matrix        = (np.vstack(speed_epoch_windows_all)
                                     if speed_epoch_windows_all       else None)
        cap_epoch_matrix          = (np.vstack(cap_epoch_windows_all)
                                     if cap_epoch_windows_all         else None)
        speed_epoch_session_means   = (np.vstack(speed_epoch_session_means_all)
                                       if speed_epoch_session_means_all else None)
        cap_epoch_session_means     = (np.vstack(cap_epoch_session_means_all)
                                       if cap_epoch_session_means_all   else None)
        speed_epoch_session_indices = (np.array(speed_epoch_session_indices_all, dtype=int)
                                       if speed_epoch_session_indices_all else None)
        cap_epoch_session_indices   = (np.array(cap_epoch_session_indices_all, dtype=int)
                                       if cap_epoch_session_indices_all   else None)
        speed_epoch_event_indices   = (np.array(speed_epoch_event_indices_all, dtype=int)
                                       if speed_epoch_event_indices_all   else None)
        cap_epoch_event_indices     = (np.array(cap_epoch_event_indices_all, dtype=int)
                                       if cap_epoch_event_indices_all     else None)
        punish_speed_epoch_matrix          = (np.vstack(punish_speed_epoch_windows_all)
                                              if punish_speed_epoch_windows_all         else None)
        punish_cap_epoch_matrix            = (np.vstack(punish_cap_epoch_windows_all)
                                              if punish_cap_epoch_windows_all           else None)
        punish_speed_epoch_session_means   = (np.vstack(punish_speed_epoch_session_means_all)
                                              if punish_speed_epoch_session_means_all   else None)
        punish_cap_epoch_session_means     = (np.vstack(punish_cap_epoch_session_means_all)
                                              if punish_cap_epoch_session_means_all     else None)
        punish_speed_epoch_session_indices = (np.array(punish_speed_epoch_session_indices_all, dtype=int)
                                              if punish_speed_epoch_session_indices_all else None)
        punish_cap_epoch_session_indices   = (np.array(punish_cap_epoch_session_indices_all, dtype=int)
                                              if punish_cap_epoch_session_indices_all   else None)
        punish_speed_epoch_event_indices   = (np.array(punish_speed_epoch_event_indices_all, dtype=int)
                                              if punish_speed_epoch_event_indices_all   else None)
        punish_cap_epoch_event_indices     = (np.array(punish_cap_epoch_event_indices_all, dtype=int)
                                              if punish_cap_epoch_event_indices_all     else None)
        lick_epoch_session_means           = (np.vstack(lick_epoch_session_means_all)
                                              if lick_epoch_session_means_all           else None)
        punish_lick_epoch_session_means    = (np.vstack(punish_lick_epoch_session_means_all)
                                              if punish_lick_epoch_session_means_all    else None)

        # Store results for this mouse
        _pct_punish_zones = (
            float(_mouse_punish_count) / float(_mouse_total_zone_count) * 100.0
            if _mouse_total_zone_count > 0 else float('nan')
        )
        all_results.append({
            'mouse': mouse_name,
            'dates': dates,
            'speeds': speeds,
            'total_distances': total_distances,
            'bout_counts': bout_counts,
            'avg_speeds_per_bout': avg_speeds_per_bout,
            'avg_dists_per_bout':  avg_dists_per_bout,
            'hits': hits,
            'false_alarms': false_alarms_list,
            'correct_rejections': correct_rejections_list,
            'dprimes': dprimes_list,
            'session_lengths': session_lengths,
            'starting_condition': conditions[mouse_name],
            'sex': 'male' if markers[mouse_name] == 's' else 'female',
            'df': results_df,
            'session_file_errors': session_file_errors,
            'speed_epoch_matrix':        speed_epoch_matrix,
            'cap_epoch_matrix':          cap_epoch_matrix,
            'speed_epoch_session_means':   speed_epoch_session_means,
            'cap_epoch_session_means':     cap_epoch_session_means,
            'speed_epoch_session_indices': speed_epoch_session_indices,
            'cap_epoch_session_indices':   cap_epoch_session_indices,
            'speed_epoch_event_indices':   speed_epoch_event_indices,
            'cap_epoch_event_indices':     cap_epoch_event_indices,
            'punish_speed_epoch_matrix':          punish_speed_epoch_matrix,
            'punish_cap_epoch_matrix':            punish_cap_epoch_matrix,
            'punish_speed_epoch_session_means':   punish_speed_epoch_session_means,
            'punish_cap_epoch_session_means':     punish_cap_epoch_session_means,
            'punish_speed_epoch_session_indices': punish_speed_epoch_session_indices,
            'punish_cap_epoch_session_indices':   punish_cap_epoch_session_indices,
            'punish_speed_epoch_event_indices':   punish_speed_epoch_event_indices,
            'punish_cap_epoch_event_indices':     punish_cap_epoch_event_indices,
            'lick_epoch_session_means':           lick_epoch_session_means,
            'punish_lick_epoch_session_means':    punish_lick_epoch_session_means,
            'pct_punish_zones':                   _pct_punish_zones,
            'avg_lick_latency_list':              avg_lick_latency_list,
        })

    # ── Date-aligned per-mouse plots ─────────────────────────────────────────
    # Plotting is deferred to here, after the data-collection loop, so that
    # every mouse is placed on a shared date-based x-axis.  Missing sessions
    # for a given mouse leave a gap rather than compressing the axis leftward.
    all_dates_flat = [d for result in all_results for d in result['df']['date']]
    global_start = min(all_dates_flat).date()
    global_end   = max(all_dates_flat).date()
    total_days   = (global_end - global_start).days + 1
    # max_sessions: number of unique training days (no calendar gaps)
    max_sessions = max(len(result['df']) for result in all_results)

    for result in all_results:
        mouse_name      = result['mouse']
        df_r            = result['df']
        condition_color = condition_color_map[conditions[mouse_name]]
        day_numbers     = list(range(len(df_r)))  # sequential session index, no calendar gaps

        if speed_fig is not None:
            plt.figure(speed_fig.number)
            plt.plot(day_numbers, df_r['average_speed'],
                f'{markers[mouse_name]}-', color=condition_color, markersize=8, label=mouse_name)

        if sensitivity_fig is not None:
            plt.figure(sensitivity_fig.number)
            plt.plot(day_numbers, df_r['sensitivity'],
                f'{markers[mouse_name]}-', color=condition_color, markersize=8, label=mouse_name)

        if lick_fig is not None:
            plt.figure(lick_fig.number)
            plt.plot(day_numbers, df_r['lick_count'],
                f'{markers[mouse_name]}-', color=condition_color, markersize=8, label=mouse_name)

        if reward_fig is not None:
            plt.figure(reward_fig.number)
            plt.plot(day_numbers, df_r['hits'],
                f'{markers[mouse_name]}-', color=condition_color, markersize=8, label=mouse_name)

        if lick_reward_ratio_fig is not None:
            plt.figure(lick_reward_ratio_fig.number)
            _ratio = np.where(
                (pd.to_numeric(df_r['hits_gap_aware'], errors='coerce').fillna(0) > 0),
                pd.to_numeric(df_r['lick_count'], errors='coerce').values /
                pd.to_numeric(df_r['hits_gap_aware'], errors='coerce').replace(0, np.nan).values,
                np.nan,
            )
            plt.plot(day_numbers, _ratio,
                f'{markers[mouse_name]}-', color=condition_color, markersize=8, label=mouse_name)

        if false_alarm_fig is not None:
            plt.figure(false_alarm_fig.number)
            plt.plot(day_numbers, df_r['false_alarms'],
                f'{markers[mouse_name]}-', color=condition_color, markersize=8, label=mouse_name)

        if correct_rejection_fig is not None:
            plt.figure(correct_rejection_fig.number)
            plt.plot(day_numbers, df_r['correct_rejections'],
                f'{markers[mouse_name]}-', color=condition_color, markersize=8, label=mouse_name)

        if specificity_fig is not None:
            plt.figure(specificity_fig.number)
            plt.plot(day_numbers, df_r['specificity'],
                f'{markers[mouse_name]}-', color=condition_color, markersize=8, label=mouse_name)

        if dprime_fig is not None:
            plt.figure(dprime_fig.number)
            plt.plot(day_numbers, df_r['dprime'],
                f'{markers[mouse_name]}-', color=condition_color, markersize=8, label=mouse_name)

        if distance_fig is not None:
            plt.figure(distance_fig.number)
            # convert mm → m
            dist_m = df_r['total_distance'] / 1000.0
            plt.plot(day_numbers, dist_m,
                f'{markers[mouse_name]}-', color=condition_color, markersize=8, label=mouse_name)

        if bout_count_fig is not None:
            plt.figure(bout_count_fig.number)
            plt.plot(day_numbers, df_r['bout_count'],
                f'{markers[mouse_name]}-', color=condition_color, markersize=8, label=mouse_name)

        if rewards_per_bout_fig is not None:
            plt.figure(rewards_per_bout_fig.number)
            plt.plot(day_numbers, pd.to_numeric(df_r['rewards_per_bout'], errors='coerce'),
                f'{markers[mouse_name]}-', color=condition_color, markersize=8, label=mouse_name)

        if first_lick_latency_fig is not None:
            plt.figure(first_lick_latency_fig.number)
            plt.plot(day_numbers, pd.to_numeric(df_r['avg_lick_latency'], errors='coerce'),
                f'{markers[mouse_name]}-', color=condition_color, markersize=8, label=mouse_name)

        if bout_avg_speed_fig is not None:
            plt.figure(bout_avg_speed_fig.number)
            plt.plot(day_numbers, df_r['avg_speed_per_bout'],
                f'{markers[mouse_name]}-', color=condition_color, markersize=8, label=mouse_name)

        if bout_avg_dist_fig is not None:
            plt.figure(bout_avg_dist_fig.number)
            # convert mm → m
            plt.plot(day_numbers, df_r['avg_dist_per_bout'] / 1000.0,
                f'{markers[mouse_name]}-', color=condition_color, markersize=8, label=mouse_name)

        # ── Sex-colored individual trace plots ────────────────────────────────
        _sex_color = 'green' if markers[mouse_name] == 's' else 'purple'

        if sex_speed_fig is not None:
            plt.figure(sex_speed_fig.number)
            plt.plot(day_numbers, df_r['average_speed'],
                f'{markers[mouse_name]}-', color=_sex_color, markersize=8, label=mouse_name)

        if sex_distance_indiv_fig is not None:
            plt.figure(sex_distance_indiv_fig.number)
            _dist_m = df_r['total_distance'] / 1000.0
            plt.plot(day_numbers, _dist_m,
                f'{markers[mouse_name]}-', color=_sex_color, markersize=8, label=mouse_name)

        if sex_reward_indiv_fig is not None:
            plt.figure(sex_reward_indiv_fig.number)
            _rpm_vals = df_r.apply(
                lambda _row: _row['hits'] / _row['session_length']
                if pd.notna(_row['session_length']) and _row['session_length'] > 0
                    and pd.notna(_row['hits'])
                else float('nan'), axis=1
            )
            plt.plot(day_numbers, _rpm_vals.values,
                f'{markers[mouse_name]}-', color=_sex_color, markersize=8, label=mouse_name)

    # Tick spacing for individual mouse plots
    max_day = max_sessions
    if max_day <= 10:
        major_spacing = 1
        minor_spacing = 1
    elif max_day <= 20:
        major_spacing = 2
        minor_spacing = 1
    elif max_day <= 50:
        major_spacing = 5
        minor_spacing = 1
    elif max_day <= 100:
        major_spacing = 10
        minor_spacing = 2
    else:
        major_spacing = 20
        minor_spacing = 5

    # Configure speed plot
    if speed_fig is not None:
        plt.figure(speed_fig.number)
        plt.title('Average Speed Over Time')
        plt.xlabel('Training Day')
        plt.ylabel('Average Speed (cm/s)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_day - 0.5)
        ax.xaxis.set_major_locator(plt.MultipleLocator(major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))

    # Configure sensitivity plot
    if sensitivity_fig is not None:
        plt.figure(sensitivity_fig.number)
        plt.title('Sensitivity Over Time')
        plt.xlabel('Day')
        plt.ylabel('Sensitivity (Hits / Total Trials)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(-0.05, 1.05)  # Sensitivity is between 0 and 1
        ax.set_xlim(left=0, right=max_day - 0.5)
        ax.xaxis.set_major_locator(plt.MultipleLocator(major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))

    # Configure lick count plot
    if lick_fig is not None:
        plt.figure(lick_fig.number)
        plt.title('Lick Counts Over Time')
        plt.xlabel('Training Day')
        plt.ylabel('Number of Licks')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)  # Lick counts cannot be negative
        ax.set_xlim(left=0, right=max_day - 0.5)
        ax.xaxis.set_major_locator(plt.MultipleLocator(major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))

    # Configure reward count plot
    if reward_fig is not None:
        plt.figure(reward_fig.number)
        plt.title('Total Reward Count Over Time')
        plt.xlabel('Training Day')
        plt.ylabel('Number of Rewards')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)  # Reward counts cannot be negative
        ax.set_xlim(left=0, right=max_day - 0.5)
        ax.xaxis.set_major_locator(plt.MultipleLocator(major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))

    # Configure lick / reward ratio plot
    if lick_reward_ratio_fig is not None:
        plt.figure(lick_reward_ratio_fig.number)
        plt.title('Lick Count / Reward Count Ratio Over Time')
        plt.xlabel('Training Day')
        plt.ylabel('Licks per Reward')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_day - 0.5)
        ax.xaxis.set_major_locator(plt.MultipleLocator(major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))

    # Configure false alarm plot
    if false_alarm_fig is not None:
        plt.figure(false_alarm_fig.number)
        plt.title('False Alarms Over Time')
        plt.xlabel('Training Day')
        plt.ylabel('Number of False Alarms')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_day - 0.5)
        ax.xaxis.set_major_locator(plt.MultipleLocator(major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))

    # Configure correct rejection plot
    if correct_rejection_fig is not None:
        plt.figure(correct_rejection_fig.number)
        plt.title('Correct Rejections Over Time')
        plt.xlabel('Training Day')
        plt.ylabel('Number of Correct Rejections')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_day - 0.5)
        ax.xaxis.set_major_locator(plt.MultipleLocator(major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))

    # Configure specificity plot
    if specificity_fig is not None:
        plt.figure(specificity_fig.number)
        plt.title('Specificity Over Time')
        plt.xlabel('Training Day')
        plt.ylabel('Specificity (Correct Rejections / Puff Opportunities)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(-0.05, 1.05)
        ax.set_xlim(left=0, right=max_day - 0.5)
        ax.xaxis.set_major_locator(plt.MultipleLocator(major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))

    # Configure d-prime plot
    if dprime_fig is not None:
        plt.figure(dprime_fig.number)
        plt.title("d' Over Time")
        plt.xlabel('Training Day')
        plt.ylabel("d' (Signal Detection)")
        plt.axhline(y=0, color='gray', linestyle='--', linewidth=0.8, alpha=0.6)
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_xlim(left=0, right=max_day - 0.5)
        ax.xaxis.set_major_locator(plt.MultipleLocator(major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))

    # Configure distance plot
    if distance_fig is not None:
        plt.figure(distance_fig.number)
        plt.title('Total Distance Per Session')
        plt.xlabel('Training Day')
        plt.ylabel('Distance (m)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_day - 0.5)
        ax.xaxis.set_major_locator(plt.MultipleLocator(major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))

    if bout_count_fig is not None:
        plt.figure(bout_count_fig.number)
        plt.title('Locomotion Bout Count Per Session')
        plt.xlabel('Training Day')
        plt.ylabel('Bout Count')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_day - 0.5)
        ax.xaxis.set_major_locator(plt.MultipleLocator(major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))

    if rewards_per_bout_fig is not None:
        plt.figure(rewards_per_bout_fig.number)
        plt.title('Average Rewards per Locomotion Bout Per Session')
        plt.xlabel('Training Day')
        plt.ylabel('Rewards per Bout')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_day - 0.5)
        ax.xaxis.set_major_locator(plt.MultipleLocator(major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))

    if first_lick_latency_fig is not None:
        plt.figure(first_lick_latency_fig.number)
        plt.title('Average First-Lick Latency After Reward Delivery Per Session')
        plt.xlabel('Training Day')
        plt.ylabel('First-Lick Latency (s)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_day - 0.5)
        ax.xaxis.set_major_locator(plt.MultipleLocator(major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))
        plt.legend()

    if bout_avg_speed_fig is not None:
        plt.figure(bout_avg_speed_fig.number)
        plt.title('Average Speed per Locomotion Bout')
        plt.xlabel('Training Day')
        plt.ylabel('Speed per Bout (cm/s)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_day - 0.5)
        ax.xaxis.set_major_locator(plt.MultipleLocator(major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))

    if bout_avg_dist_fig is not None:
        plt.figure(bout_avg_dist_fig.number)
        plt.title('Average Distance per Locomotion Bout')
        plt.xlabel('Training Day')
        plt.ylabel('Distance per Bout (m)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_day - 0.5)
        ax.xaxis.set_major_locator(plt.MultipleLocator(major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))

    # Configure sex-colored speed plot
    if sex_speed_fig is not None:
        plt.figure(sex_speed_fig.number)
        plt.title('Average Speed Over Time — By Sex')
        plt.xlabel('Training Day')
        plt.ylabel('Average Speed (cm/s)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_day - 0.5)
        ax.xaxis.set_major_locator(plt.MultipleLocator(major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))

    # Configure sex-colored distance plot
    if sex_distance_indiv_fig is not None:
        plt.figure(sex_distance_indiv_fig.number)
        plt.title('Total Distance Per Session — By Sex')
        plt.xlabel('Training Day')
        plt.ylabel('Distance (m)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_day - 0.5)
        ax.xaxis.set_major_locator(plt.MultipleLocator(major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))

    # Configure sex-colored reward rate plot
    if sex_reward_indiv_fig is not None:
        plt.figure(sex_reward_indiv_fig.number)
        plt.title('Reward Rate Per Session — By Sex')
        plt.xlabel('Training Day')
        plt.ylabel('Rewards per Minute')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_day - 0.5)
        ax.xaxis.set_major_locator(plt.MultipleLocator(major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))

    # Calculate average rewards/minute and SEM across mice
    max_days = max_sessions
    
    # Dynamic tick spacing based on data range for aggregate plots
    if max_days <= 10:
        agg_major_spacing = 1
        agg_minor_spacing = 1
    elif max_days <= 20:
        agg_major_spacing = 2
        agg_minor_spacing = 1
    elif max_days <= 50:
        agg_major_spacing = 5
        agg_minor_spacing = 1
    elif max_days <= 100:
        agg_major_spacing = 10
        agg_minor_spacing = 2
    else:
        agg_major_spacing = 20
        agg_minor_spacing = 5

    # Build session-indexed per-mouse rewards-per-minute arrays (no calendar gaps)
    all_rewards_per_min = np.full((len(data_files), max_sessions), np.nan)
    male_rewards_per_min   = []
    female_rewards_per_min = []

    for i, result in enumerate(all_results):
        df_r    = result['df']
        rpm_row = np.full(max_sessions, np.nan)
        for session_idx, (_, row) in enumerate(df_r.iterrows()):
            if pd.notna(row['session_length']) and row['session_length'] > 0 and pd.notna(row['hits']):
                rpm_row[session_idx] = row['hits'] / row['session_length']
        all_rewards_per_min[i] = rpm_row

        mouse_name = result['mouse']
        if markers[mouse_name] == 's':  # Male
            male_rewards_per_min.append(rpm_row)
        else:  # Female (marker 'o')
            female_rewards_per_min.append(rpm_row)

    if male_rewards_per_min:
        male_rewards_per_min = np.array(male_rewards_per_min)
    if female_rewards_per_min:
        female_rewards_per_min = np.array(female_rewards_per_min)
    
    # Calculate mean and SEM across mice for each day (only over mice that have data on that day)
    with warnings.catch_warnings():
        warnings.simplefilter('ignore', RuntimeWarning)
        n_all = np.sum(~np.isnan(all_rewards_per_min), axis=0)
        mean_rewards_per_min = np.where(n_all > 0, np.nanmean(all_rewards_per_min, axis=0), np.nan)
        sem_rewards_per_min = np.where(n_all > 1,
                                       np.nanstd(all_rewards_per_min, axis=0) / np.sqrt(n_all),
                                       0)
    
    # Plot average rewards/minute with SEM
    if avg_reward_fig is not None:
        plt.figure(avg_reward_fig.number)
        day_numbers = np.arange(0, max_days)
        plt.plot(day_numbers, mean_rewards_per_min, '-', color='black', linewidth=2, label='Mean')
        plt.fill_between(day_numbers, mean_rewards_per_min - sem_rewards_per_min, mean_rewards_per_min + sem_rewards_per_min, 
                         color='gray', alpha=0.3, label='SEM')
        
        # Configure average rewards plot
        plt.title('Average Rewards Per Minute Across Mice')
        plt.xlabel('Day')
        plt.ylabel('Rewards per Minute (Mean ± SEM)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_days - 1)
        ax.xaxis.set_major_locator(plt.MultipleLocator(5))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(1))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))
        plt.legend()
    
    # Plot sex-specific average rewards/minute with SEM
    if sex_reward_fig is not None:
        plt.figure(sex_reward_fig.number)
        day_numbers = np.arange(0, max_days)
        
        # Plot male data if available
        if len(male_rewards_per_min) > 0:
            # Check if we have any non-NaN values
            valid_male_data = np.any(~np.isnan(male_rewards_per_min))
            if valid_male_data:
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore', RuntimeWarning)
                    n_male = np.sum(~np.isnan(male_rewards_per_min), axis=0)
                    mean_male = np.where(n_male > 0, np.nanmean(male_rewards_per_min, axis=0), np.nan)
                    sem_male = np.where(n_male > 1,
                                        np.nanstd(male_rewards_per_min, axis=0) / np.sqrt(n_male),
                                        0)
                plt.plot(day_numbers, mean_male, '-', color='green', linewidth=2, label=f'Male (n={len(male_rewards_per_min)})')
                plt.fill_between(day_numbers, mean_male - sem_male, mean_male + sem_male,
                                 color='green', alpha=0.2)

        # Plot female data if available
        if len(female_rewards_per_min) > 0:
            # Check if we have any non-NaN values
            valid_female_data = np.any(~np.isnan(female_rewards_per_min))
            if valid_female_data:
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore', RuntimeWarning)
                    n_female = np.sum(~np.isnan(female_rewards_per_min), axis=0)
                    mean_female = np.where(n_female > 0, np.nanmean(female_rewards_per_min, axis=0), np.nan)
                    sem_female = np.where(n_female > 1,
                                          np.nanstd(female_rewards_per_min, axis=0) / np.sqrt(n_female),
                                          0)
                plt.plot(day_numbers, mean_female, '-', color='purple', linewidth=2, label=f'Female (n={len(female_rewards_per_min)})')
                plt.fill_between(day_numbers, mean_female - sem_female, mean_female + sem_female,
                                 color='purple', alpha=0.2)

        # Configure sex-specific rewards plot
        plt.title('Sex-Specific Average Rewards Per Minute')
        plt.xlabel('Training Day')
        plt.ylabel('Average Rewards per Minute')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_days - 1)
        ax.xaxis.set_major_locator(plt.MultipleLocator(agg_major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(agg_minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))
        plt.legend()

    # Build session-indexed per-mouse speed arrays split by sex (no calendar gaps)
    all_speeds_agg    = np.full((len(data_files), max_sessions), np.nan)
    male_speeds_agg   = []
    female_speeds_agg = []

    for i, result in enumerate(all_results):
        df_r     = result['df']
        spd_row  = np.full(max_sessions, np.nan)
        for session_idx, (_, row) in enumerate(df_r.iterrows()):
            if pd.notna(row['average_speed']):
                spd_row[session_idx] = row['average_speed']
        all_speeds_agg[i] = spd_row
        mouse_name = result['mouse']
        if markers[mouse_name] == 's':
            male_speeds_agg.append(spd_row)
        else:
            female_speeds_agg.append(spd_row)

    if male_speeds_agg:
        male_speeds_agg = np.array(male_speeds_agg)
    if female_speeds_agg:
        female_speeds_agg = np.array(female_speeds_agg)

    # Plot sex-specific average speed with SEM
    if avg_sex_speed_fig is not None:
        plt.figure(avg_sex_speed_fig.number)
        day_numbers = np.arange(0, max_days)

        if len(male_speeds_agg) > 0:
            if np.any(~np.isnan(male_speeds_agg)):
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore', RuntimeWarning)
                    n_male_spd = np.sum(~np.isnan(male_speeds_agg), axis=0)
                    mean_male_spd = np.where(n_male_spd > 0, np.nanmean(male_speeds_agg, axis=0), np.nan)
                    sem_male_spd  = np.where(n_male_spd > 1,
                                             np.nanstd(male_speeds_agg, axis=0) / np.sqrt(n_male_spd),
                                             0)
                plt.plot(day_numbers, mean_male_spd, '-', color='green', linewidth=2,
                         label=f'Male (n={len(male_speeds_agg)})')
                plt.fill_between(day_numbers, mean_male_spd - sem_male_spd,
                                 mean_male_spd + sem_male_spd, color='green', alpha=0.2)

        if len(female_speeds_agg) > 0:
            if np.any(~np.isnan(female_speeds_agg)):
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore', RuntimeWarning)
                    n_female_spd = np.sum(~np.isnan(female_speeds_agg), axis=0)
                    mean_female_spd = np.where(n_female_spd > 0, np.nanmean(female_speeds_agg, axis=0), np.nan)
                    sem_female_spd  = np.where(n_female_spd > 1,
                                               np.nanstd(female_speeds_agg, axis=0) / np.sqrt(n_female_spd),
                                               0)
                plt.plot(day_numbers, mean_female_spd, '-', color='purple', linewidth=2,
                         label=f'Female (n={len(female_speeds_agg)})')
                plt.fill_between(day_numbers, mean_female_spd - sem_female_spd,
                                 mean_female_spd + sem_female_spd, color='purple', alpha=0.2)

        plt.title('Sex-Specific Average Speed')
        plt.xlabel('Training Day')
        plt.ylabel('Average Speed (cm/s) (Mean \u00b1 SEM)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_days - 1)
        ax.xaxis.set_major_locator(plt.MultipleLocator(agg_major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(agg_minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))
        plt.legend()

    # Build session-indexed per-mouse licks-per-minute arrays (no calendar gaps)
    all_licks_per_min = np.full((len(data_files), max_sessions), np.nan)
    male_licks_per_min   = []
    female_licks_per_min = []

    for i, result in enumerate(all_results):
        df_r    = result['df']
        lpm_row = np.full(max_sessions, np.nan)
        for session_idx, (_, row) in enumerate(df_r.iterrows()):
            if pd.notna(row['session_length']) and row['session_length'] > 0 and pd.notna(row['lick_count']):
                lpm_row[session_idx] = row['lick_count'] / row['session_length']
        all_licks_per_min[i] = lpm_row
        mouse_name = result['mouse']
        if markers[mouse_name] == 's':
            male_licks_per_min.append(lpm_row)
        else:
            female_licks_per_min.append(lpm_row)

    if male_licks_per_min:
        male_licks_per_min = np.array(male_licks_per_min)
    if female_licks_per_min:
        female_licks_per_min = np.array(female_licks_per_min)

    with warnings.catch_warnings():
        warnings.simplefilter('ignore', RuntimeWarning)
        n_all_lpm = np.sum(~np.isnan(all_licks_per_min), axis=0)
        mean_licks_per_min = np.where(n_all_lpm > 0, np.nanmean(all_licks_per_min, axis=0), np.nan)
        sem_licks_per_min  = np.where(n_all_lpm > 1,
                                      np.nanstd(all_licks_per_min, axis=0) / np.sqrt(n_all_lpm),
                                      0)

    # Plot average licks/minute with SEM
    if avg_lick_rate_fig is not None:
        plt.figure(avg_lick_rate_fig.number)
        day_numbers = np.arange(0, max_days)
        plt.plot(day_numbers, mean_licks_per_min, '-', color='black', linewidth=2, label='Mean')
        plt.fill_between(day_numbers,
                         mean_licks_per_min - sem_licks_per_min,
                         mean_licks_per_min + sem_licks_per_min,
                         color='gray', alpha=0.3, label='SEM')
        plt.title('Average Lick Rate Across Mice')
        plt.xlabel('Day')
        plt.ylabel('Licks per Minute (Mean \u00b1 SEM)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_days - 1)
        ax.xaxis.set_major_locator(plt.MultipleLocator(5))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(1))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))
        plt.legend()

    # Plot sex-specific average licks/minute with SEM
    if sex_lick_rate_fig is not None:
        plt.figure(sex_lick_rate_fig.number)
        day_numbers = np.arange(0, max_days)
        if len(male_licks_per_min) > 0:
            if np.any(~np.isnan(male_licks_per_min)):
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore', RuntimeWarning)
                    n_male_lpm = np.sum(~np.isnan(male_licks_per_min), axis=0)
                    mean_male_lpm = np.where(n_male_lpm > 0, np.nanmean(male_licks_per_min, axis=0), np.nan)
                    sem_male_lpm  = np.where(n_male_lpm > 1,
                                             np.nanstd(male_licks_per_min, axis=0) / np.sqrt(n_male_lpm),
                                             0)
                plt.plot(day_numbers, mean_male_lpm, '-', color='green', linewidth=2,
                         label=f'Male (n={len(male_licks_per_min)})')
                plt.fill_between(day_numbers, mean_male_lpm - sem_male_lpm,
                                 mean_male_lpm + sem_male_lpm, color='green', alpha=0.2)
        if len(female_licks_per_min) > 0:
            if np.any(~np.isnan(female_licks_per_min)):
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore', RuntimeWarning)
                    n_female_lpm = np.sum(~np.isnan(female_licks_per_min), axis=0)
                    mean_female_lpm = np.where(n_female_lpm > 0, np.nanmean(female_licks_per_min, axis=0), np.nan)
                    sem_female_lpm  = np.where(n_female_lpm > 1,
                                               np.nanstd(female_licks_per_min, axis=0) / np.sqrt(n_female_lpm),
                                               0)
                plt.plot(day_numbers, mean_female_lpm, '-', color='purple', linewidth=2,
                         label=f'Female (n={len(female_licks_per_min)})')
                plt.fill_between(day_numbers, mean_female_lpm - sem_female_lpm,
                                 mean_female_lpm + sem_female_lpm, color='purple', alpha=0.2)
        plt.title('Sex-Specific Average Lick Rate')
        plt.xlabel('Training Day')
        plt.ylabel('Licks per Minute (Mean \u00b1 SEM)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_days - 1)
        ax.xaxis.set_major_locator(plt.MultipleLocator(agg_major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(agg_minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))
        plt.legend()

    # Build session-indexed per-mouse bout count arrays
    all_bout_counts = np.full((len(data_files), max_sessions), np.nan)

    for i, result in enumerate(all_results):
        df_r = result['df']
        bc_row = np.full(max_sessions, np.nan)
        for session_idx, (_, row) in enumerate(df_r.iterrows()):
            if pd.notna(row['bout_count']):
                bc_row[session_idx] = row['bout_count']
        all_bout_counts[i] = bc_row

    with warnings.catch_warnings():
        warnings.simplefilter('ignore', RuntimeWarning)
        n_all_bc = np.sum(~np.isnan(all_bout_counts), axis=0)
        mean_bout_counts = np.where(n_all_bc > 0, np.nanmean(all_bout_counts, axis=0), np.nan)
        sem_bout_counts  = np.where(n_all_bc > 1,
                                    np.nanstd(all_bout_counts, axis=0) / np.sqrt(n_all_bc),
                                    0)

    if avg_bout_count_fig is not None:
        plt.figure(avg_bout_count_fig.number)
        day_numbers = np.arange(0, max_days)
        plt.plot(day_numbers, mean_bout_counts, '-', color='black', linewidth=2, label='Mean')
        plt.fill_between(day_numbers,
                         mean_bout_counts - sem_bout_counts,
                         mean_bout_counts + sem_bout_counts,
                         color='gray', alpha=0.3, label='SEM')
        plt.title('Average Locomotion Bout Count Across Mice')
        plt.xlabel('Training Day')
        plt.ylabel('Bout Count (Mean \u00b1 SEM)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_days - 1)
        ax.xaxis.set_major_locator(plt.MultipleLocator(agg_major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(agg_minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))
        plt.legend()

    # Build session-indexed per-mouse distance arrays (mm → m)
    all_distances_m = np.full((len(data_files), max_sessions), np.nan)
    male_distances_m   = []
    female_distances_m = []

    for i, result in enumerate(all_results):
        df_r = result['df']
        dist_row = np.full(max_sessions, np.nan)
        for session_idx, (_, row) in enumerate(df_r.iterrows()):
            if pd.notna(row['total_distance']):
                dist_row[session_idx] = row['total_distance'] / 1000.0  # mm → m
        all_distances_m[i] = dist_row
        mouse_name = result['mouse']
        if markers[mouse_name] == 's':
            male_distances_m.append(dist_row)
        else:
            female_distances_m.append(dist_row)

    if male_distances_m:
        male_distances_m = np.array(male_distances_m)
    if female_distances_m:
        female_distances_m = np.array(female_distances_m)

    # Sex-specific distance plot
    sex_distance_fig = plt.figure(figsize=(12, 6)) if 'sex_distance' in selected_plots else None
    if sex_distance_fig is not None:
        plt.figure(sex_distance_fig.number)
        day_numbers = np.arange(0, max_days)
        if len(male_distances_m) > 0 and np.any(~np.isnan(male_distances_m)):
            with warnings.catch_warnings():
                warnings.simplefilter('ignore', RuntimeWarning)
                n_m = np.sum(~np.isnan(male_distances_m), axis=0)
                mean_m = np.where(n_m > 0, np.nanmean(male_distances_m, axis=0), np.nan)
                sem_m  = np.where(n_m > 1, np.nanstd(male_distances_m, axis=0) / np.sqrt(n_m), 0)
            plt.plot(day_numbers, mean_m, '-', color='green', linewidth=2,
                     label=f'Male (n={len(male_distances_m)})')
            plt.fill_between(day_numbers, mean_m - sem_m, mean_m + sem_m, color='green', alpha=0.2)
        if len(female_distances_m) > 0 and np.any(~np.isnan(female_distances_m)):
            with warnings.catch_warnings():
                warnings.simplefilter('ignore', RuntimeWarning)
                n_f = np.sum(~np.isnan(female_distances_m), axis=0)
                mean_f = np.where(n_f > 0, np.nanmean(female_distances_m, axis=0), np.nan)
                sem_f  = np.where(n_f > 1, np.nanstd(female_distances_m, axis=0) / np.sqrt(n_f), 0)
            plt.plot(day_numbers, mean_f, '-', color='purple', linewidth=2,
                     label=f'Female (n={len(female_distances_m)})')
            plt.fill_between(day_numbers, mean_f - sem_f, mean_f + sem_f, color='purple', alpha=0.2)
        plt.title('Sex-Specific Average Distance Per Session')
        plt.xlabel('Training Day')
        plt.ylabel('Distance (m, Mean \u00b1 SEM)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_days - 1)
        ax.xaxis.set_major_locator(plt.MultipleLocator(agg_major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(agg_minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))
        plt.legend()

    # Condition-based distance plot
    condition_distance_fig = plt.figure(figsize=(12, 6)) if 'condition_distance' in selected_plots else None
    if condition_distance_fig is not None:
        condition_dist_groups = {}
        for result in all_results:
            condition = result['starting_condition']
            if condition not in condition_dist_groups:
                condition_dist_groups[condition] = []
            df_r = result['df']
            dist_array = np.full(max_sessions, np.nan)
            for session_idx, (_, row) in enumerate(df_r.iterrows()):
                if pd.notna(row['total_distance']):
                    dist_array[session_idx] = row['total_distance'] / 1000.0  # mm → m
            condition_dist_groups[condition].append(dist_array)

        day_numbers = np.arange(0, max_sessions)
        for condition, dist_list in condition_dist_groups.items():
            color = condition_color_map[condition]
            padded = np.array(dist_list)
            with warnings.catch_warnings():
                warnings.simplefilter('ignore', RuntimeWarning)
                n_mice = np.sum(~np.isnan(padded), axis=0)
                mean_d = np.where(n_mice > 0, np.nanmean(padded, axis=0), np.nan)
                sem_d  = np.where(n_mice > 1, np.nanstd(padded, axis=0) / np.sqrt(n_mice), 0)
            plt.plot(day_numbers, mean_d, '-', color=color, linewidth=2,
                     label=f'{condition} (n={len(dist_list)})')
            plt.fill_between(day_numbers, mean_d - sem_d, mean_d + sem_d, color=color, alpha=0.2)
        plt.title('Average Distance Per Session by Starting Condition')
        plt.xlabel('Training Day')
        plt.ylabel('Distance (m, Mean \u00b1 SEM)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_days - 1)
        ax.xaxis.set_major_locator(plt.MultipleLocator(agg_major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(agg_minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))
        plt.legend()

    # Create a new figure for condition-based analysis
    condition_reward_fig = plt.figure(figsize=(12, 6)) if 'condition_reward' in selected_plots else None
    if condition_reward_fig is not None:
        # Group mice by starting condition (session-indexed arrays, no calendar gaps)
        condition_groups = {}
        for result in all_results:
            condition = result['starting_condition']
            if condition not in condition_groups:
                condition_groups[condition] = []
            df_r = result['df']
            rpm_array = np.full(max_sessions, np.nan)
            for session_idx, (_, row) in enumerate(df_r.iterrows()):
                if pd.notna(row['session_length']) and row['session_length'] > 0 and pd.notna(row['hits']):
                    rpm_array[session_idx] = row['hits'] / row['session_length']
            condition_groups[condition].append(rpm_array)

        # Plot each condition's data
        day_numbers = np.arange(0, max_sessions)
        for condition, rewards_list in condition_groups.items():
            color = condition_color_map[condition]
            padded_rewards = np.array(rewards_list)

            # Calculate mean and SEM (only over mice that have data on that day)
            with warnings.catch_warnings():
                warnings.simplefilter('ignore', RuntimeWarning)
                n_mice = np.sum(~np.isnan(padded_rewards), axis=0)
                mean_rewards = np.where(n_mice > 0, np.nanmean(padded_rewards, axis=0), np.nan)
                sem_rewards = np.where(n_mice > 1,
                                       np.nanstd(padded_rewards, axis=0) / np.sqrt(n_mice),
                                       0)

            # Plot the data
            plt.plot(day_numbers, mean_rewards, '-', color=color, linewidth=2,
                    label=f'{condition} (n={len(rewards_list)})')
            plt.fill_between(day_numbers, mean_rewards - sem_rewards, mean_rewards + sem_rewards,
                            color=color, alpha=0.2)
        
        # Configure condition-based rewards plot
        plt.title('Average Rewards Per Minute by Starting Condition')
        plt.xlabel('Training Day')
        plt.ylabel('Average Rewards per Minute')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_days - 1)
        ax.xaxis.set_major_locator(plt.MultipleLocator(agg_major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(agg_minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))
        plt.legend()

    # Create a new figure for condition-based speed analysis
    condition_speed_fig = plt.figure(figsize=(12, 6)) if 'condition_speed' in selected_plots else None
    if condition_speed_fig is not None:
        # Group mice by starting condition for speed (session-indexed arrays, no calendar gaps)
        condition_speed_groups = {}
        for result in all_results:
            condition = result['starting_condition']
            if condition not in condition_speed_groups:
                condition_speed_groups[condition] = []
            df_r = result['df']
            speed_array = np.full(max_sessions, np.nan)
            for session_idx, (_, row) in enumerate(df_r.iterrows()):
                speed_array[session_idx] = row['average_speed']
            condition_speed_groups[condition].append(speed_array)

        # Plot each condition's speed data
        day_numbers = np.arange(0, max_sessions)
        for condition, speed_list in condition_speed_groups.items():
            color = condition_color_map[condition]
            padded_speeds = np.array(speed_list)

            # Calculate mean and SEM (only over mice that have data on that day)
            with warnings.catch_warnings():
                warnings.simplefilter('ignore', RuntimeWarning)
                n_mice = np.sum(~np.isnan(padded_speeds), axis=0)
                mean_speeds = np.where(n_mice > 0, np.nanmean(padded_speeds, axis=0), np.nan)
                sem_speeds = np.where(n_mice > 1,
                                      np.nanstd(padded_speeds, axis=0) / np.sqrt(n_mice),
                                      0)

            # Plot the data
            plt.plot(day_numbers, mean_speeds, '-', color=color, linewidth=2,
                    label=f'{condition} (n={len(speed_list)})')
            plt.fill_between(day_numbers, mean_speeds - sem_speeds, mean_speeds + sem_speeds,
                            color=color, alpha=0.2)
        
        # Configure condition-based speed plot
        plt.title('Average Speed by Starting Condition')
        plt.xlabel('Training Day')
        plt.ylabel('Average Speed (cm/s)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_days - 1)
        ax.xaxis.set_major_locator(plt.MultipleLocator(agg_major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(agg_minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))
        plt.legend()

    # Create a new figure for condition-based bout count analysis
    condition_bout_count_fig = plt.figure(figsize=(12, 6)) if 'condition_bout_count' in selected_plots else None
    if condition_bout_count_fig is not None:
        condition_bout_count_groups = {}
        for result in all_results:
            condition = result['starting_condition']
            if condition not in condition_bout_count_groups:
                condition_bout_count_groups[condition] = []
            df_r = result['df']
            bc_array = np.full(max_sessions, np.nan)
            for session_idx, (_, row) in enumerate(df_r.iterrows()):
                if pd.notna(row['bout_count']):
                    bc_array[session_idx] = row['bout_count']
            condition_bout_count_groups[condition].append(bc_array)

        day_numbers = np.arange(0, max_sessions)
        for condition, bc_list in condition_bout_count_groups.items():
            color = condition_color_map[condition]
            padded_bcs = np.array(bc_list)
            with warnings.catch_warnings():
                warnings.simplefilter('ignore', RuntimeWarning)
                n_mice = np.sum(~np.isnan(padded_bcs), axis=0)
                mean_bcs = np.where(n_mice > 0, np.nanmean(padded_bcs, axis=0), np.nan)
                sem_bcs  = np.where(n_mice > 1,
                                    np.nanstd(padded_bcs, axis=0) / np.sqrt(n_mice),
                                    0)
            plt.plot(day_numbers, mean_bcs, '-', color=color, linewidth=2,
                     label=f'{condition} (n={len(bc_list)})')
            plt.fill_between(day_numbers, mean_bcs - sem_bcs, mean_bcs + sem_bcs,
                             color=color, alpha=0.2)

        plt.title('Average Bout Count by Starting Condition')
        plt.xlabel('Training Day')
        plt.ylabel('Bout Count (Mean \u00b1 SEM)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_days - 1)
        ax.xaxis.set_major_locator(plt.MultipleLocator(agg_major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(agg_minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))
        plt.legend()

    # Condition-based average rewards per locomotion bout
    condition_rewards_per_bout_fig = plt.figure(figsize=(12, 6)) if 'condition_rewards_per_bout' in selected_plots else None
    if condition_rewards_per_bout_fig is not None:
        condition_rpb_groups = {}
        for result in all_results:
            condition = result['starting_condition']
            if condition not in condition_rpb_groups:
                condition_rpb_groups[condition] = []
            df_r = result['df']
            rpb_array = np.full(max_sessions, np.nan)
            for session_idx, (_, row) in enumerate(df_r.iterrows()):
                val = pd.to_numeric(row.get('rewards_per_bout', np.nan), errors='coerce')
                if pd.notna(val):
                    rpb_array[session_idx] = val
            condition_rpb_groups[condition].append(rpb_array)

        day_numbers = np.arange(0, max_sessions)
        for condition, rpb_list in condition_rpb_groups.items():
            color = condition_color_map[condition]
            padded_rpb = np.array(rpb_list)
            with warnings.catch_warnings():
                warnings.simplefilter('ignore', RuntimeWarning)
                n_mice = np.sum(~np.isnan(padded_rpb), axis=0)
                mean_rpb = np.where(n_mice > 0, np.nanmean(padded_rpb, axis=0), np.nan)
                sem_rpb  = np.where(n_mice > 1,
                                    np.nanstd(padded_rpb, axis=0) / np.sqrt(n_mice),
                                    0)
            plt.plot(day_numbers, mean_rpb, '-', color=color, linewidth=2,
                     label=f'{condition} (n={len(rpb_list)})')
            plt.fill_between(day_numbers, mean_rpb - sem_rpb, mean_rpb + sem_rpb,
                             color=color, alpha=0.2)

        plt.title('Average Rewards per Locomotion Bout by Starting Condition')
        plt.xlabel('Training Day')
        plt.ylabel('Rewards per Bout (Mean \u00b1 SEM)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_days - 1)
        ax.xaxis.set_major_locator(plt.MultipleLocator(agg_major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(agg_minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))
        plt.legend()

    # ── Rewards per bout bar chart (one avg per mouse, collapsed across sessions) ─
    condition_rewards_per_bout_bar_fig = None
    if 'condition_rewards_per_bout_bar' in selected_plots:
        condition_rewards_per_bout_bar_fig, ax_rpbbar = plt.subplots(figsize=(8, 6))

        condition_mouse_rpb: dict[str, list] = {}
        for result in all_results:
            condition = result['starting_condition']
            df_r = result['df']
            session_vals = pd.to_numeric(df_r['rewards_per_bout'], errors='coerce').dropna().values
            if len(session_vals) > 0:
                if condition not in condition_mouse_rpb:
                    condition_mouse_rpb[condition] = []
                condition_mouse_rpb[condition].append((result['mouse'], float(np.mean(session_vals))))

        conditions_sorted_rpbbar = sorted(condition_mouse_rpb.keys())
        x_pos_rpbbar = np.arange(len(conditions_sorted_rpbbar))

        rng_rpbbar = np.random.default_rng(seed=42)
        for ci, condition in enumerate(conditions_sorted_rpbbar):
            entries = condition_mouse_rpb[condition]
            mouse_vals = [v for _, v in entries]
            mean_rpb = float(np.mean(mouse_vals))
            sem_rpb  = float(np.std(mouse_vals, ddof=1) / np.sqrt(len(mouse_vals))) if len(mouse_vals) > 1 else 0.0
            color = condition_color_map[condition]
            ax_rpbbar.bar(ci, mean_rpb, width=0.5, color=color, alpha=0.8,
                          yerr=sem_rpb, capsize=7, error_kw={'elinewidth': 1.5, 'capthick': 1.5})
            jitter = (rng_rpbbar.random(len(mouse_vals)) - 0.5) * 0.22
            for j, (mouse_name_rpb, rpb_val) in enumerate(entries):
                ax_rpbbar.plot(ci + jitter[j], rpb_val, 'o',
                               color='white', markeredgecolor=color,
                               markeredgewidth=1.8, markersize=7, zorder=3)

        ax_rpbbar.set_xticks(x_pos_rpbbar)
        ax_rpbbar.set_xticklabels(conditions_sorted_rpbbar)
        ax_rpbbar.set_xlabel('Starting Condition')
        ax_rpbbar.set_ylabel('Rewards per Locomotion Bout (Mean \u00b1 SEM)')
        ax_rpbbar.set_ylim(bottom=0)
        ax_rpbbar.tick_params(axis='both', direction='in')
        ax_rpbbar.spines['top'].set_visible(False)
        ax_rpbbar.spines['right'].set_visible(False)

        # ── Mann-Whitney U test: pairwise significance brackets ─────────────
        from scipy.stats import mannwhitneyu as _mwu_rpbbar
        import itertools as _it_rpbbar
        _rpbbar_results = []
        for (_rc1, _rc2) in _it_rpbbar.combinations(conditions_sorted_rpbbar, 2):
            _rv1 = np.array([v for _, v in condition_mouse_rpb[_rc1]])
            _rv2 = np.array([v for _, v in condition_mouse_rpb[_rc2]])
            if len(_rv1) >= 2 and len(_rv2) >= 2:
                _rpb_stat, _rpb_p = _mwu_rpbbar(_rv1, _rv2, alternative='two-sided')
                _rpbbar_results.append((_rc1, _rc2, float(_rpb_stat), float(_rpb_p)))

        if _rpbbar_results:
            _rpbbar_y_max = max(
                max(max(v for _, v in condition_mouse_rpb[c]) for c in conditions_sorted_rpbbar),
                float(ax_rpbbar.get_ylim()[1]),
            )
            _rpbbar_step = _rpbbar_y_max * 0.12
            for _bk_idx, (_rc1, _rc2, _rpb_stat, _rpb_p) in enumerate(_rpbbar_results):
                _bx1 = conditions_sorted_rpbbar.index(_rc1)
                _bx2 = conditions_sorted_rpbbar.index(_rc2)
                _bk_y = _rpbbar_y_max + _rpbbar_step * (_bk_idx + 1)
                ax_rpbbar.plot(
                    [_bx1, _bx1, _bx2, _bx2],
                    [_bk_y - _rpbbar_step * 0.2, _bk_y,
                     _bk_y, _bk_y - _rpbbar_step * 0.2],
                    color='black', linewidth=1.2,
                )
                if _rpb_p < 0.001:
                    _bsig = '***'
                elif _rpb_p < 0.01:
                    _bsig = '**'
                elif _rpb_p < 0.05:
                    _bsig = '*'
                else:
                    _bsig = f'ns  p={_rpb_p:.3f}'
                ax_rpbbar.text(
                    (_bx1 + _bx2) / 2.0,
                    _bk_y + _rpbbar_step * 0.05,
                    _bsig, ha='center', va='bottom', fontsize=9,
                )
            ax_rpbbar.set_ylim(
                bottom=0,
                top=_rpbbar_y_max + _rpbbar_step * (len(_rpbbar_results) + 1.8),
            )

        ax_rpbbar.set_title('Average Rewards per Locomotion Bout by Starting Condition\n(collapsed across all sessions; Mann-Whitney U test)')
        condition_rewards_per_bout_bar_fig.tight_layout()

    # ── Condition-based first-lick latency line plot ──────────────────────────
    condition_first_lick_latency_fig = plt.figure(figsize=(12, 6)) if 'condition_first_lick_latency' in selected_plots else None
    if condition_first_lick_latency_fig is not None:
        _cfll_groups = {}
        for result in all_results:
            condition = result['starting_condition']
            if condition not in _cfll_groups:
                _cfll_groups[condition] = []
            df_r = result['df']
            _fll_arr = np.full(max_sessions, np.nan)
            for session_idx, (_, row) in enumerate(df_r.iterrows()):
                val = pd.to_numeric(row.get('avg_lick_latency', np.nan), errors='coerce')
                if pd.notna(val):
                    _fll_arr[session_idx] = val
            _cfll_groups[condition].append(_fll_arr)

        day_numbers = np.arange(0, max_sessions)
        for condition, _fll_list in _cfll_groups.items():
            color = condition_color_map[condition]
            padded_fll = np.array(_fll_list)
            with warnings.catch_warnings():
                warnings.simplefilter('ignore', RuntimeWarning)
                n_mice = np.sum(~np.isnan(padded_fll), axis=0)
                mean_fll = np.where(n_mice > 0, np.nanmean(padded_fll, axis=0), np.nan)
                sem_fll  = np.where(n_mice > 1,
                                    np.nanstd(padded_fll, axis=0) / np.sqrt(n_mice),
                                    0)
            plt.plot(day_numbers, mean_fll, '-', color=color, linewidth=2,
                     label=f'{condition} (n={len(_fll_list)})')
            plt.fill_between(day_numbers, mean_fll - sem_fll, mean_fll + sem_fll,
                             color=color, alpha=0.2)

        plt.title('Average First-Lick Latency After Reward Delivery by Starting Condition')
        plt.xlabel('Training Day')
        plt.ylabel('First-Lick Latency (s, Mean \u00b1 SEM)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_days - 1)
        ax.xaxis.set_major_locator(plt.MultipleLocator(agg_major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(agg_minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))
        plt.legend()

    # ── First-lick latency bar chart (one avg per mouse, collapsed across sessions) ─
    condition_first_lick_latency_bar_fig = None
    if 'condition_first_lick_latency_bar' in selected_plots:
        condition_first_lick_latency_bar_fig, ax_fllbar = plt.subplots(figsize=(8, 6))

        _condition_mouse_fll: dict[str, list] = {}
        for result in all_results:
            condition = result['starting_condition']
            df_r = result['df']
            session_vals = pd.to_numeric(df_r['avg_lick_latency'], errors='coerce').dropna().values
            if len(session_vals) > 0:
                if condition not in _condition_mouse_fll:
                    _condition_mouse_fll[condition] = []
                _condition_mouse_fll[condition].append((result['mouse'], float(np.mean(session_vals))))

        _conditions_sorted_fllbar = sorted(_condition_mouse_fll.keys())
        _x_pos_fllbar = np.arange(len(_conditions_sorted_fllbar))

        _rng_fllbar = np.random.default_rng(seed=42)
        for ci, condition in enumerate(_conditions_sorted_fllbar):
            entries = _condition_mouse_fll[condition]
            mouse_vals = [v for _, v in entries]
            mean_fll = float(np.mean(mouse_vals))
            sem_fll  = float(np.std(mouse_vals, ddof=1) / np.sqrt(len(mouse_vals))) if len(mouse_vals) > 1 else 0.0
            color = condition_color_map[condition]
            ax_fllbar.bar(ci, mean_fll, width=0.5, color=color, alpha=0.8,
                          yerr=sem_fll, capsize=7, error_kw={'elinewidth': 1.5, 'capthick': 1.5})
            jitter = (_rng_fllbar.random(len(mouse_vals)) - 0.5) * 0.22
            for j, (mouse_name_fll, fll_val) in enumerate(entries):
                ax_fllbar.plot(ci + jitter[j], fll_val, 'o',
                               color='white', markeredgecolor=color,
                               markeredgewidth=1.8, markersize=7, zorder=3)

        ax_fllbar.set_xticks(_x_pos_fllbar)
        ax_fllbar.set_xticklabels(_conditions_sorted_fllbar)
        ax_fllbar.set_xlabel('Starting Condition')
        ax_fllbar.set_ylabel('First-Lick Latency (s, Mean \u00b1 SEM)')
        ax_fllbar.set_ylim(bottom=0)
        ax_fllbar.tick_params(axis='both', direction='in')
        ax_fllbar.spines['top'].set_visible(False)
        ax_fllbar.spines['right'].set_visible(False)

        # ── Mann-Whitney U test: pairwise significance brackets ─────────────
        from scipy.stats import mannwhitneyu as _mwu_fllbar
        import itertools as _it_fllbar
        _fllbar_results = []
        for (_fc1, _fc2) in _it_fllbar.combinations(_conditions_sorted_fllbar, 2):
            _fv1 = np.array([v for _, v in _condition_mouse_fll[_fc1]])
            _fv2 = np.array([v for _, v in _condition_mouse_fll[_fc2]])
            if len(_fv1) >= 2 and len(_fv2) >= 2:
                _fll_stat, _fll_p = _mwu_fllbar(_fv1, _fv2, alternative='two-sided')
                _fllbar_results.append((_fc1, _fc2, float(_fll_stat), float(_fll_p)))

        if _fllbar_results:
            _fllbar_y_max = max(
                max(max(v for _, v in _condition_mouse_fll[c]) for c in _conditions_sorted_fllbar),
                float(ax_fllbar.get_ylim()[1]),
            )
            _fllbar_step = _fllbar_y_max * 0.12
            for _bk_idx, (_fc1, _fc2, _fll_stat, _fll_p) in enumerate(_fllbar_results):
                _bx1 = _conditions_sorted_fllbar.index(_fc1)
                _bx2 = _conditions_sorted_fllbar.index(_fc2)
                _bk_y = _fllbar_y_max + _fllbar_step * (_bk_idx + 1)
                ax_fllbar.plot(
                    [_bx1, _bx1, _bx2, _bx2],
                    [_bk_y - _fllbar_step * 0.2, _bk_y,
                     _bk_y, _bk_y - _fllbar_step * 0.2],
                    color='black', linewidth=1.2,
                )
                if _fll_p < 0.001:
                    _bsig = '***'
                elif _fll_p < 0.01:
                    _bsig = '**'
                elif _fll_p < 0.05:
                    _bsig = '*'
                else:
                    _bsig = f'ns  p={_fll_p:.3f}'
                ax_fllbar.text(
                    (_bx1 + _bx2) / 2.0,
                    _bk_y + _fllbar_step * 0.05,
                    _bsig, ha='center', va='bottom', fontsize=9,
                )
            ax_fllbar.set_ylim(
                bottom=0,
                top=_fllbar_y_max + _fllbar_step * (len(_fllbar_results) + 1.8),
            )

        ax_fllbar.set_title('Average First-Lick Latency After Reward Delivery by Starting Condition\n(collapsed across all sessions; Mann-Whitney U test)')
        condition_first_lick_latency_bar_fig.tight_layout()

    # ── Condition-based lick-after-reward proportion line plot ────────────────
    condition_lick_after_reward_prop_fig = plt.figure(figsize=(12, 6)) if 'lick_after_reward_prop' in selected_plots else None
    if condition_lick_after_reward_prop_fig is not None:
        _clarp_groups = {}
        for result in all_results:
            condition = result['starting_condition']
            if condition not in _clarp_groups:
                _clarp_groups[condition] = []
            df_r = result['df']
            _larp_arr = np.full(max_sessions, np.nan)
            for session_idx, (_, row) in enumerate(df_r.iterrows()):
                val = pd.to_numeric(row.get('lick_after_reward_prop', np.nan), errors='coerce')
                if pd.notna(val):
                    _larp_arr[session_idx] = val
            _clarp_groups[condition].append(_larp_arr)

        day_numbers = np.arange(0, max_sessions)
        for condition, _larp_list in _clarp_groups.items():
            color = condition_color_map[condition]
            padded_larp = np.array(_larp_list)
            with warnings.catch_warnings():
                warnings.simplefilter('ignore', RuntimeWarning)
                n_mice = np.sum(~np.isnan(padded_larp), axis=0)
                mean_larp = np.where(n_mice > 0, np.nanmean(padded_larp, axis=0), np.nan)
                sem_larp  = np.where(n_mice > 1,
                                     np.nanstd(padded_larp, axis=0) / np.sqrt(n_mice),
                                     0)
            plt.plot(day_numbers, mean_larp, '-', color=color, linewidth=2,
                     label=f'{condition} (n={len(_larp_list)})')
            plt.fill_between(day_numbers, mean_larp - sem_larp, mean_larp + sem_larp,
                             color=color, alpha=0.2)

        plt.title('Proportion of Reward Deliveries with Licks by Starting Condition\n(2 s window starting 0.65 s after reward zone entry)')
        plt.xlabel('Training Day')
        plt.ylabel('Proportion with Licks (Mean \u00b1 SEM)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(0, 1.05)
        ax.set_xlim(left=0, right=max_days - 1)
        ax.xaxis.set_major_locator(plt.MultipleLocator(agg_major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(agg_minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))
        plt.legend()

    # ── Lick-after-reward proportion bar chart (one avg per mouse) ────────────
    condition_lick_after_reward_prop_bar_fig = None
    if 'lick_after_reward_prop_bar' in selected_plots:
        condition_lick_after_reward_prop_bar_fig, ax_larpbar = plt.subplots(figsize=(8, 6))

        _condition_mouse_larp: dict[str, list] = {}
        for result in all_results:
            condition = result['starting_condition']
            df_r = result['df']
            session_vals = pd.to_numeric(df_r['lick_after_reward_prop'], errors='coerce').dropna().values
            if len(session_vals) > 0:
                if condition not in _condition_mouse_larp:
                    _condition_mouse_larp[condition] = []
                _condition_mouse_larp[condition].append((result['mouse'], float(np.mean(session_vals))))

        _conditions_sorted_larp = sorted(_condition_mouse_larp.keys())
        _x_pos_larp = np.arange(len(_conditions_sorted_larp))

        _rng_larp = np.random.default_rng(seed=42)
        for ci, condition in enumerate(_conditions_sorted_larp):
            entries = _condition_mouse_larp[condition]
            mouse_vals = [v for _, v in entries]
            mean_larp_b = float(np.mean(mouse_vals))
            sem_larp_b  = float(np.std(mouse_vals, ddof=1) / np.sqrt(len(mouse_vals))) if len(mouse_vals) > 1 else 0.0
            color = condition_color_map[condition]
            ax_larpbar.bar(ci, mean_larp_b, width=0.5, color=color, alpha=0.8,
                           yerr=sem_larp_b, capsize=7, error_kw={'elinewidth': 1.5, 'capthick': 1.5})
            jitter = (_rng_larp.random(len(mouse_vals)) - 0.5) * 0.22
            for j, (mouse_name_larp, larp_val) in enumerate(entries):
                ax_larpbar.plot(ci + jitter[j], larp_val, 'o',
                                color='white', markeredgecolor=color,
                                markeredgewidth=1.8, markersize=7, zorder=3)

        ax_larpbar.set_xticks(_x_pos_larp)
        ax_larpbar.set_xticklabels(_conditions_sorted_larp)
        ax_larpbar.set_xlabel('Starting Condition')
        ax_larpbar.set_ylabel('Proportion of Deliveries with Licks (Mean \u00b1 SEM)')
        ax_larpbar.set_ylim(0, 1.05)
        ax_larpbar.tick_params(axis='both', direction='in')
        ax_larpbar.spines['top'].set_visible(False)
        ax_larpbar.spines['right'].set_visible(False)

        # ── Mann-Whitney U test: pairwise significance brackets ─────────────
        from scipy.stats import mannwhitneyu as _mwu_larp
        import itertools as _it_larp
        _larp_results = []
        for (_lc1, _lc2) in _it_larp.combinations(_conditions_sorted_larp, 2):
            _lv1 = np.array([v for _, v in _condition_mouse_larp[_lc1]])
            _lv2 = np.array([v for _, v in _condition_mouse_larp[_lc2]])
            if len(_lv1) >= 2 and len(_lv2) >= 2:
                _larp_stat, _larp_p = _mwu_larp(_lv1, _lv2, alternative='two-sided')
                _larp_results.append((_lc1, _lc2, float(_larp_stat), float(_larp_p)))

        if _larp_results:
            _larp_y_max = 1.0  # proportion is bounded by [0, 1]; anchor brackets at 1.0
            _larp_step = _larp_y_max * 0.12
            for _bk_idx, (_lc1, _lc2, _larp_stat, _larp_p) in enumerate(_larp_results):
                _bx1 = _conditions_sorted_larp.index(_lc1)
                _bx2 = _conditions_sorted_larp.index(_lc2)
                _bk_y = _larp_y_max + _larp_step * (_bk_idx + 1)
                ax_larpbar.plot(
                    [_bx1, _bx1, _bx2, _bx2],
                    [_bk_y - _larp_step * 0.2, _bk_y,
                     _bk_y, _bk_y - _larp_step * 0.2],
                    color='black', linewidth=1.2,
                )
                if _larp_p < 0.001:
                    _lsig = '***'
                elif _larp_p < 0.01:
                    _lsig = '**'
                elif _larp_p < 0.05:
                    _lsig = '*'
                else:
                    _lsig = f'ns  p={_larp_p:.3f}'
                ax_larpbar.text(
                    (_bx1 + _bx2) / 2.0,
                    _bk_y + _larp_step * 0.05,
                    _lsig, ha='center', va='bottom', fontsize=9,
                )
            ax_larpbar.set_ylim(
                bottom=0,
                top=_larp_y_max + _larp_step * (len(_larp_results) + 1.8),
            )

        ax_larpbar.set_title('Proportion of Reward Deliveries with Licks by Starting Condition\n(2 s post-delivery window, collapsed across all sessions; Mann-Whitney U test)')
        condition_lick_after_reward_prop_bar_fig.tight_layout()

    # ── Weekday reward bar charts ─────────────────────────────────────────────
    # For each mouse average its reward count (hits) across all sessions that
    # share the same weekday label (Mon/Tue/Thu/Fri).  One value per weekday
    # per mouse.
    _DOW_ORDER = ['Monday', 'Tuesday', 'Thursday', 'Friday']

    # Build the shared data structure used by both weekday plots
    # weekday_mouse_data: weekday -> list of (mouse_name, condition, mean_reward)
    _weekday_mouse_data: dict[str, list] = {d: [] for d in _DOW_ORDER}
    for result in all_results:
        df_r      = result['df']
        mouse     = result['mouse']
        condition = result['starting_condition']
        if 'weekday' not in df_r.columns:
            continue
        for wd in _DOW_ORDER:
            _wd_hits = pd.to_numeric(
                df_r.loc[df_r['weekday'] == wd, 'hits'], errors='coerce'
            ).dropna().values
            if len(_wd_hits) > 0:
                _weekday_mouse_data[wd].append((mouse, condition, float(np.mean(_wd_hits))))

    # ── Plot 1: all mice pooled ───────────────────────────────────────────────
    weekday_reward_bar_fig = None
    if 'weekday_reward_bar' in selected_plots:
        weekday_reward_bar_fig, ax_wdbar = plt.subplots(figsize=(8, 6))
        _x_wd = np.arange(len(_DOW_ORDER))
        _rng_wdbar = np.random.default_rng(seed=42)
        for ci, wd in enumerate(_DOW_ORDER):
            _vals = [v for _, _, v in _weekday_mouse_data[wd]]
            if not _vals:
                continue
            _mean_wd = float(np.mean(_vals))
            _sem_wd  = float(np.std(_vals, ddof=1) / np.sqrt(len(_vals))) if len(_vals) > 1 else 0.0
            ax_wdbar.bar(ci, _mean_wd, width=0.5, color='steelblue', alpha=0.8,
                         yerr=_sem_wd, capsize=7,
                         error_kw={'elinewidth': 1.5, 'capthick': 1.5})
            _jit = (_rng_wdbar.random(len(_vals)) - 0.5) * 0.22
            for j, val in enumerate(_vals):
                ax_wdbar.plot(ci + _jit[j], val, 'o',
                              color='white', markeredgecolor='steelblue',
                              markeredgewidth=1.8, markersize=7, zorder=3)
        ax_wdbar.set_xticks(_x_wd)
        ax_wdbar.set_xticklabels(_DOW_ORDER)
        ax_wdbar.set_xlabel('Training Weekday')
        ax_wdbar.set_ylabel('Reward Count (Mean \u00b1 SEM)')
        ax_wdbar.set_ylim(bottom=0)
        ax_wdbar.tick_params(axis='both', direction='in')
        ax_wdbar.spines['top'].set_visible(False)
        ax_wdbar.spines['right'].set_visible(False)
        ax_wdbar.set_title('Average Reward Count by Training Weekday\n(all mice pooled, Mon\u2192Tue\u2192Thu\u2192Fri cycle)')
        weekday_reward_bar_fig.tight_layout()

    # ── Plot 2: split by starting condition ──────────────────────────────────
    weekday_reward_bar_condition_fig = None
    if 'weekday_reward_bar_condition' in selected_plots:
        weekday_reward_bar_condition_fig, ax_wdcbar = plt.subplots(figsize=(10, 6))

        # condition -> weekday -> list of per-mouse mean reward counts
        _wd_cond_data: dict[str, dict[str, list]] = {}
        for result in all_results:
            cond = result['starting_condition']
            if cond not in _wd_cond_data:
                _wd_cond_data[cond] = {d: [] for d in _DOW_ORDER}
            df_r = result['df']
            if 'weekday' not in df_r.columns:
                continue
            for wd in _DOW_ORDER:
                _wd_hits = pd.to_numeric(
                    df_r.loc[df_r['weekday'] == wd, 'hits'], errors='coerce'
                ).dropna().values
                if len(_wd_hits) > 0:
                    _wd_cond_data[cond][wd].append(float(np.mean(_wd_hits)))

        _conds_wd = sorted(_wd_cond_data.keys())
        _n_conds  = len(_conds_wd)
        _wd_x     = np.arange(len(_DOW_ORDER))
        _bar_w    = 0.7 / max(_n_conds, 1)
        _rng_wdcbar = np.random.default_rng(seed=42)

        for ci, cond in enumerate(_conds_wd):
            color  = condition_color_map[cond]
            _offset = (_bar_w * ci) - (_bar_w * (_n_conds - 1) / 2.0)
            for wi, wd in enumerate(_DOW_ORDER):
                _vals = _wd_cond_data[cond][wd]
                if not _vals:
                    continue
                _mean_wd = float(np.mean(_vals))
                _sem_wd  = float(np.std(_vals, ddof=1) / np.sqrt(len(_vals))) if len(_vals) > 1 else 0.0
                _bar_x   = _wd_x[wi] + _offset
                ax_wdcbar.bar(_bar_x, _mean_wd, width=_bar_w * 0.9, color=color, alpha=0.8,
                              yerr=_sem_wd, capsize=5,
                              error_kw={'elinewidth': 1.5, 'capthick': 1.5},
                              label=cond if wi == 0 else '_nolegend_')
                _jit = (_rng_wdcbar.random(len(_vals)) - 0.5) * (_bar_w * 0.35)
                for j, val in enumerate(_vals):
                    ax_wdcbar.plot(_bar_x + _jit[j], val, 'o',
                                   color='white', markeredgecolor=color,
                                   markeredgewidth=1.8, markersize=6, zorder=3)

        ax_wdcbar.set_xticks(_wd_x)
        ax_wdcbar.set_xticklabels(_DOW_ORDER)
        ax_wdcbar.set_xlabel('Training Weekday')
        ax_wdcbar.set_ylabel('Reward Count (Mean \u00b1 SEM)')
        ax_wdcbar.set_ylim(bottom=0)
        ax_wdcbar.tick_params(axis='both', direction='in')
        ax_wdcbar.spines['top'].set_visible(False)
        ax_wdcbar.spines['right'].set_visible(False)
        ax_wdcbar.legend(title='Starting Condition', frameon=False)
        ax_wdcbar.set_title('Average Reward Count by Training Weekday \u2014 by Starting Condition\n(Mon\u2192Tue\u2192Thu\u2192Fri cycle)')
        weekday_reward_bar_condition_fig.tight_layout()

    # Create a new figure for condition-based average speed per bout analysis
    condition_bout_avg_speed_fig = plt.figure(figsize=(12, 6)) if 'condition_bout_avg_speed' in selected_plots else None
    if condition_bout_avg_speed_fig is not None:
        condition_bas_groups = {}
        for result in all_results:
            condition = result['starting_condition']
            if condition not in condition_bas_groups:
                condition_bas_groups[condition] = []
            df_r = result['df']
            bas_array = np.full(max_sessions, np.nan)
            for session_idx, (_, row) in enumerate(df_r.iterrows()):
                if pd.notna(row['avg_speed_per_bout']):
                    bas_array[session_idx] = row['avg_speed_per_bout']
            condition_bas_groups[condition].append(bas_array)

        day_numbers = np.arange(0, max_sessions)
        for condition, bas_list in condition_bas_groups.items():
            color = condition_color_map[condition]
            padded = np.array(bas_list)
            with warnings.catch_warnings():
                warnings.simplefilter('ignore', RuntimeWarning)
                n_mice = np.sum(~np.isnan(padded), axis=0)
                mean_bas = np.where(n_mice > 0, np.nanmean(padded, axis=0), np.nan)
                sem_bas  = np.where(n_mice > 1, np.nanstd(padded, axis=0) / np.sqrt(n_mice), 0)
            plt.plot(day_numbers, mean_bas, '-', color=color, linewidth=2,
                     label=f'{condition} (n={len(bas_list)})')
            plt.fill_between(day_numbers, mean_bas - sem_bas, mean_bas + sem_bas,
                             color=color, alpha=0.2)

        plt.title('Average Speed per Locomotion Bout by Starting Condition')
        plt.xlabel('Training Day')
        plt.ylabel('Speed per Bout (cm/s, Mean \u00b1 SEM)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_days - 1)
        ax.xaxis.set_major_locator(plt.MultipleLocator(agg_major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(agg_minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))
        plt.legend()

    # Create a new figure for condition-based average distance per bout analysis
    condition_bout_avg_dist_fig = plt.figure(figsize=(12, 6)) if 'condition_bout_avg_dist' in selected_plots else None
    if condition_bout_avg_dist_fig is not None:
        condition_bad_groups = {}
        for result in all_results:
            condition = result['starting_condition']
            if condition not in condition_bad_groups:
                condition_bad_groups[condition] = []
            df_r = result['df']
            bad_array = np.full(max_sessions, np.nan)
            for session_idx, (_, row) in enumerate(df_r.iterrows()):
                if pd.notna(row['avg_dist_per_bout']):
                    bad_array[session_idx] = row['avg_dist_per_bout'] / 1000.0  # mm → m
            condition_bad_groups[condition].append(bad_array)

        day_numbers = np.arange(0, max_sessions)
        for condition, bad_list in condition_bad_groups.items():
            color = condition_color_map[condition]
            padded = np.array(bad_list)
            with warnings.catch_warnings():
                warnings.simplefilter('ignore', RuntimeWarning)
                n_mice = np.sum(~np.isnan(padded), axis=0)
                mean_bad = np.where(n_mice > 0, np.nanmean(padded, axis=0), np.nan)
                sem_bad  = np.where(n_mice > 1, np.nanstd(padded, axis=0) / np.sqrt(n_mice), 0)
            plt.plot(day_numbers, mean_bad, '-', color=color, linewidth=2,
                     label=f'{condition} (n={len(bad_list)})')
            plt.fill_between(day_numbers, mean_bad - sem_bad, mean_bad + sem_bad,
                             color=color, alpha=0.2)

        plt.title('Average Distance per Locomotion Bout by Starting Condition')
        plt.xlabel('Training Day')
        plt.ylabel('Distance per Bout (m, Mean \u00b1 SEM)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_days - 1)
        ax.xaxis.set_major_locator(plt.MultipleLocator(agg_major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(agg_minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))
        plt.legend()

    # Create a new figure for condition-based lick count analysis
    condition_lick_fig = plt.figure(figsize=(12, 6)) if 'condition_lick' in selected_plots else None
    if condition_lick_fig is not None:
        # Group mice by starting condition for lick count (session-indexed arrays, no calendar gaps)
        condition_lick_groups = {}
        for result in all_results:
            condition = result['starting_condition']
            if condition not in condition_lick_groups:
                condition_lick_groups[condition] = []
            df_r = result['df']
            lick_array = np.full(max_sessions, np.nan)
            for session_idx, (_, row) in enumerate(df_r.iterrows()):
                lick_array[session_idx] = row['lick_count']
            condition_lick_groups[condition].append(lick_array)

        # Plot each condition's lick count data
        day_numbers = np.arange(0, max_sessions)
        for condition, lick_list in condition_lick_groups.items():
            color = condition_color_map[condition]
            padded_licks = np.array(lick_list)

            # Calculate mean and SEM (only over mice that have data on that day)
            with warnings.catch_warnings():
                warnings.simplefilter('ignore', RuntimeWarning)
                n_mice = np.sum(~np.isnan(padded_licks), axis=0)
                mean_licks = np.where(n_mice > 0, np.nanmean(padded_licks, axis=0), np.nan)
                sem_licks = np.where(n_mice > 1,
                                     np.nanstd(padded_licks, axis=0) / np.sqrt(n_mice),
                                     0)

            # Plot the data
            plt.plot(day_numbers, mean_licks, '-', color=color, linewidth=2,
                    label=f'{condition} (n={len(lick_list)})')
            plt.fill_between(day_numbers, mean_licks - sem_licks, mean_licks + sem_licks,
                            color=color, alpha=0.2)

        # Configure condition-based lick count plot
        plt.title('Average Lick Count by Starting Condition')
        plt.xlabel('Training Day')
        plt.ylabel('Number of Licks (Mean \u00b1 SEM)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_days - 1)
        ax.xaxis.set_major_locator(plt.MultipleLocator(agg_major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(agg_minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))
        plt.legend()

    # Create a new figure for condition-based lick rate analysis
    condition_lick_rate_fig = plt.figure(figsize=(12, 6)) if 'condition_lick_rate' in selected_plots else None
    if condition_lick_rate_fig is not None:
        condition_lick_rate_groups = {}
        for result in all_results:
            condition = result['starting_condition']
            if condition not in condition_lick_rate_groups:
                condition_lick_rate_groups[condition] = []
            df_r = result['df']
            lpm_array = np.full(max_sessions, np.nan)
            for session_idx, (_, row) in enumerate(df_r.iterrows()):
                if pd.notna(row['session_length']) and row['session_length'] > 0 and pd.notna(row['lick_count']):
                    lpm_array[session_idx] = row['lick_count'] / row['session_length']
            condition_lick_rate_groups[condition].append(lpm_array)

        day_numbers = np.arange(0, max_sessions)
        for condition, lpm_list in condition_lick_rate_groups.items():
            color = condition_color_map[condition]
            padded_lpms = np.array(lpm_list)
            with warnings.catch_warnings():
                warnings.simplefilter('ignore', RuntimeWarning)
                n_mice = np.sum(~np.isnan(padded_lpms), axis=0)
                mean_lpms = np.where(n_mice > 0, np.nanmean(padded_lpms, axis=0), np.nan)
                sem_lpms  = np.where(n_mice > 1,
                                     np.nanstd(padded_lpms, axis=0) / np.sqrt(n_mice),
                                     0)
            plt.plot(day_numbers, mean_lpms, '-', color=color, linewidth=2,
                     label=f'{condition} (n={len(lpm_list)})')
            plt.fill_between(day_numbers, mean_lpms - sem_lpms, mean_lpms + sem_lpms,
                             color=color, alpha=0.2)

        plt.title('Average Lick Rate by Starting Condition')
        plt.xlabel('Training Day')
        plt.ylabel('Licks per Minute (Mean \u00b1 SEM)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_days - 1)
        ax.xaxis.set_major_locator(plt.MultipleLocator(agg_major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(agg_minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))
        plt.legend()

    # Create a new figure for condition-based lick/reward ratio analysis
    condition_lick_reward_ratio_fig = plt.figure(figsize=(12, 6)) if 'condition_lick_reward_ratio' in selected_plots else None
    if condition_lick_reward_ratio_fig is not None:
        condition_ratio_groups = {}
        for result in all_results:
            condition = result['starting_condition']
            if condition not in condition_ratio_groups:
                condition_ratio_groups[condition] = []
            df_r = result['df']
            ratio_array = np.full(max_sessions, np.nan)
            for session_idx, (_, row) in enumerate(df_r.iterrows()):
                hits_val = pd.to_numeric(row['hits_gap_aware'], errors='coerce')
                lick_val = pd.to_numeric(row['lick_count'], errors='coerce')
                if pd.notna(hits_val) and hits_val > 0 and pd.notna(lick_val):
                    ratio_array[session_idx] = lick_val / hits_val
            condition_ratio_groups[condition].append(ratio_array)

        day_numbers = np.arange(0, max_sessions)
        for condition, ratio_list in condition_ratio_groups.items():
            color = condition_color_map[condition]
            padded_ratios = np.array(ratio_list)
            with warnings.catch_warnings():
                warnings.simplefilter('ignore', RuntimeWarning)
                n_mice = np.sum(~np.isnan(padded_ratios), axis=0)
                mean_ratios = np.where(n_mice > 0, np.nanmean(padded_ratios, axis=0), np.nan)
                sem_ratios  = np.where(n_mice > 1,
                                       np.nanstd(padded_ratios, axis=0) / np.sqrt(n_mice),
                                       0)
            plt.plot(day_numbers, mean_ratios, '-', color=color, linewidth=2,
                     label=f'{condition} (n={len(ratio_list)})')
            plt.fill_between(day_numbers, mean_ratios - sem_ratios, mean_ratios + sem_ratios,
                             color=color, alpha=0.2)

        plt.title('Lick Count / Reward Count Ratio by Starting Condition')
        plt.xlabel('Training Day')
        plt.ylabel('Licks per Reward (Mean \u00b1 SEM)')
        plt.grid(False)
        ax = plt.gca()
        ax.tick_params(axis='both', direction='in')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(bottom=0)
        ax.set_xlim(left=0, right=max_days - 1)
        ax.xaxis.set_major_locator(plt.MultipleLocator(agg_major_spacing))
        ax.xaxis.set_minor_locator(plt.MultipleLocator(agg_minor_spacing))
        ax.tick_params(axis='x', which='minor', direction='in')
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x)}'))
        plt.legend()

    # Create collapsed lick/reward ratio bar plot (one average per mouse, collapsed across all sessions)
    condition_lick_reward_ratio_bar_fig = None
    if 'condition_lick_reward_ratio_bar' in selected_plots:
        condition_lick_reward_ratio_bar_fig, ax_lrrbar = plt.subplots(figsize=(8, 6))

        condition_mouse_lrr: dict[str, list] = {}
        for result in all_results:
            condition = result['starting_condition']
            df_r = result['df']
            session_ratios = []
            for _, row in df_r.iterrows():
                hits_val = pd.to_numeric(row.get('hits_gap_aware', np.nan), errors='coerce')
                lick_val = pd.to_numeric(row.get('lick_count', np.nan), errors='coerce')
                if pd.notna(hits_val) and hits_val > 0 and pd.notna(lick_val):
                    session_ratios.append(lick_val / hits_val)
            if session_ratios:
                if condition not in condition_mouse_lrr:
                    condition_mouse_lrr[condition] = []
                condition_mouse_lrr[condition].append((result['mouse'], float(np.mean(session_ratios))))

        conditions_sorted_lrrbar = sorted(condition_mouse_lrr.keys())
        x_pos_lrrbar = np.arange(len(conditions_sorted_lrrbar))

        rng_lrrbar = np.random.default_rng(seed=42)
        for ci, condition in enumerate(conditions_sorted_lrrbar):
            entries = condition_mouse_lrr[condition]
            mouse_lrrs = [v for _, v in entries]
            mean_lrr = float(np.mean(mouse_lrrs))
            sem_lrr  = float(np.std(mouse_lrrs, ddof=1) / np.sqrt(len(mouse_lrrs))) if len(mouse_lrrs) > 1 else 0.0
            color = condition_color_map[condition]
            ax_lrrbar.bar(ci, mean_lrr, width=0.5, color=color, alpha=0.8,
                          yerr=sem_lrr, capsize=7, error_kw={'elinewidth': 1.5, 'capthick': 1.5})
            jitter = (rng_lrrbar.random(len(mouse_lrrs)) - 0.5) * 0.22
            for j, (mouse_name_lrrbar, lrr_val) in enumerate(entries):
                ax_lrrbar.plot(ci + jitter[j], lrr_val, 'o',
                               color='white', markeredgecolor=color,
                               markeredgewidth=1.8, markersize=7, zorder=3)

        ax_lrrbar.set_xticks(x_pos_lrrbar)
        ax_lrrbar.set_xticklabels(conditions_sorted_lrrbar)
        ax_lrrbar.set_xlabel('Starting Condition')
        ax_lrrbar.set_ylabel('Licks per Reward (Mean \u00b1 SEM)')
        ax_lrrbar.set_ylim(bottom=0)
        ax_lrrbar.tick_params(axis='both', direction='in')
        ax_lrrbar.spines['top'].set_visible(False)
        ax_lrrbar.spines['right'].set_visible(False)

        # ── Mann-Whitney U test: pairwise significance brackets ─────────────
        from scipy.stats import mannwhitneyu as _mwu_lrrbar
        import itertools as _it_lrrbar
        _lrrbar_results = []
        for (_bc1, _bc2) in _it_lrrbar.combinations(conditions_sorted_lrrbar, 2):
            _v1 = np.array([v for _, v in condition_mouse_lrr[_bc1]])
            _v2 = np.array([v for _, v in condition_mouse_lrr[_bc2]])
            if len(_v1) >= 2 and len(_v2) >= 2:
                _bt_stat, _bt_p = _mwu_lrrbar(_v1, _v2, alternative='two-sided')
                _lrrbar_results.append((_bc1, _bc2, float(_bt_stat), float(_bt_p)))

        if _lrrbar_results:
            # determine y ceiling from current bars + individual points
            _lrrbar_y_max = max(
                max(max(v for _, v in condition_mouse_lrr[c]) for c in conditions_sorted_lrrbar),
                float(ax_lrrbar.get_ylim()[1]),
            )
            _lrrbar_step = _lrrbar_y_max * 0.12
            for _bk_idx, (_bc1, _bc2, _bt_stat, _bt_p) in enumerate(_lrrbar_results):
                _bx1 = conditions_sorted_lrrbar.index(_bc1)
                _bx2 = conditions_sorted_lrrbar.index(_bc2)
                _bk_y = _lrrbar_y_max + _lrrbar_step * (_bk_idx + 1)
                ax_lrrbar.plot(
                    [_bx1, _bx1, _bx2, _bx2],
                    [_bk_y - _lrrbar_step * 0.2, _bk_y,
                     _bk_y, _bk_y - _lrrbar_step * 0.2],
                    color='black', linewidth=1.2,
                )
                if _bt_p < 0.001:
                    _bsig = '***'
                elif _bt_p < 0.01:
                    _bsig = '**'
                elif _bt_p < 0.05:
                    _bsig = '*'
                else:
                    _bsig = f'ns  p={_bt_p:.3f}'
                ax_lrrbar.text(
                    (_bx1 + _bx2) / 2.0,
                    _bk_y + _lrrbar_step * 0.05,
                    _bsig, ha='center', va='bottom', fontsize=9,
                )
            ax_lrrbar.set_ylim(
                bottom=0,
                top=_lrrbar_y_max + _lrrbar_step * (len(_lrrbar_results) + 1.8),
            )

        ax_lrrbar.set_title('Average Lick Count / Reward Count Ratio by Starting Condition\n(collapsed across all sessions; Mann-Whitney U test)')
        condition_lick_reward_ratio_bar_fig.tight_layout()

    # ── Punishment zone percentage bar chart (one value per mouse) ────────────
    condition_punish_zone_pct_bar_fig = None
    if 'condition_punish_zone_pct_bar' in selected_plots:
        condition_punish_zone_pct_bar_fig, ax_pzbar = plt.subplots(figsize=(8, 6))

        condition_mouse_pzpct: dict[str, list] = {}
        for result in all_results:
            condition = result['starting_condition']
            val = result.get('pct_punish_zones', float('nan'))
            if pd.notna(val):
                if condition not in condition_mouse_pzpct:
                    condition_mouse_pzpct[condition] = []
                condition_mouse_pzpct[condition].append((result['mouse'], val))

        conditions_sorted_pzbar = sorted(condition_mouse_pzpct.keys())
        x_pos_pzbar = np.arange(len(conditions_sorted_pzbar))

        rng_pzbar = np.random.default_rng(seed=42)
        for ci, condition in enumerate(conditions_sorted_pzbar):
            entries = condition_mouse_pzpct[condition]
            mouse_vals = [v for _, v in entries]
            mean_pz = float(np.mean(mouse_vals))
            sem_pz  = float(np.std(mouse_vals, ddof=1) / np.sqrt(len(mouse_vals))) if len(mouse_vals) > 1 else 0.0
            color = condition_color_map[condition]
            ax_pzbar.bar(ci, mean_pz, width=0.5, color=color, alpha=0.8,
                         yerr=sem_pz, capsize=7, error_kw={'elinewidth': 1.5, 'capthick': 1.5})
            jitter = (rng_pzbar.random(len(mouse_vals)) - 0.5) * 0.22
            for j, (mouse_name_pz, pz_val) in enumerate(entries):
                ax_pzbar.plot(ci + jitter[j], pz_val, 'o',
                              color='white', markeredgecolor=color,
                              markeredgewidth=1.8, markersize=7, zorder=3)

        ax_pzbar.set_xticks(x_pos_pzbar)
        ax_pzbar.set_xticklabels(conditions_sorted_pzbar)
        ax_pzbar.set_xlabel('Starting Condition')
        ax_pzbar.set_ylabel('Punishment Zones (% of all zones, Mean \u00b1 SEM)')
        ax_pzbar.set_ylim(bottom=0)
        ax_pzbar.tick_params(axis='both', direction='in')
        ax_pzbar.spines['top'].set_visible(False)
        ax_pzbar.spines['right'].set_visible(False)

        # ── Mann-Whitney U test: pairwise significance brackets ─────────────
        from scipy.stats import mannwhitneyu as _mwu_pzbar
        import itertools as _it_pzbar
        _pzbar_results = []
        for (_pc1, _pc2) in _it_pzbar.combinations(conditions_sorted_pzbar, 2):
            _v1 = np.array([v for _, v in condition_mouse_pzpct[_pc1]])
            _v2 = np.array([v for _, v in condition_mouse_pzpct[_pc2]])
            if len(_v1) >= 2 and len(_v2) >= 2:
                _pz_stat, _pz_p = _mwu_pzbar(_v1, _v2, alternative='two-sided')
                _pzbar_results.append((_pc1, _pc2, float(_pz_stat), float(_pz_p)))

        if _pzbar_results:
            _pzbar_y_max = max(
                max(max(v for _, v in condition_mouse_pzpct[c]) for c in conditions_sorted_pzbar),
                float(ax_pzbar.get_ylim()[1]),
            )
            _pzbar_step = _pzbar_y_max * 0.12
            for _bk_idx, (_pc1, _pc2, _pz_stat, _pz_p) in enumerate(_pzbar_results):
                _bx1 = conditions_sorted_pzbar.index(_pc1)
                _bx2 = conditions_sorted_pzbar.index(_pc2)
                _bk_y = _pzbar_y_max + _pzbar_step * (_bk_idx + 1)
                ax_pzbar.plot(
                    [_bx1, _bx1, _bx2, _bx2],
                    [_bk_y - _pzbar_step * 0.2, _bk_y,
                     _bk_y, _bk_y - _pzbar_step * 0.2],
                    color='black', linewidth=1.2,
                )
                if _pz_p < 0.001:
                    _pzsig = '***'
                elif _pz_p < 0.01:
                    _pzsig = '**'
                elif _pz_p < 0.05:
                    _pzsig = '*'
                else:
                    _pzsig = f'ns  p={_pz_p:.3f}'
                ax_pzbar.text(
                    (_bx1 + _bx2) / 2.0,
                    _bk_y + _pzbar_step * 0.05,
                    _pzsig, ha='center', va='bottom', fontsize=9,
                )
            ax_pzbar.set_ylim(
                bottom=0,
                top=_pzbar_y_max + _pzbar_step * (len(_pzbar_results) + 1.8),
            )

        ax_pzbar.set_title('% Punishment Zones Across All Sessions by Starting Condition\n(one value per mouse; Mann-Whitney U test)')
        condition_punish_zone_pct_bar_fig.tight_layout()

    # Create collapsed condition bar plot (one average rpm per mouse, collapsed across all days)
    condition_bar_fig = None
    if 'condition_bar' in selected_plots:
        condition_bar_fig, ax_bar = plt.subplots(figsize=(8, 6))

        condition_mouse_rpms: dict[str, list] = {}
        for result in all_results:
            condition = result['starting_condition']
            df_r = result['df']
            session_rpms = []
            for _, row in df_r.iterrows():
                if pd.notna(row['session_length']) and row['session_length'] > 0 and pd.notna(row['hits']):
                    session_rpms.append(row['hits'] / row['session_length'])
            if session_rpms:
                if condition not in condition_mouse_rpms:
                    condition_mouse_rpms[condition] = []
                condition_mouse_rpms[condition].append((result['mouse'], float(np.mean(session_rpms))))

        conditions_sorted_bar = sorted(condition_mouse_rpms.keys())
        x_pos_bar = np.arange(len(conditions_sorted_bar))

        rng_bar = np.random.default_rng(seed=42)
        for ci, condition in enumerate(conditions_sorted_bar):
            entries = condition_mouse_rpms[condition]
            mouse_rpms = [v for _, v in entries]
            mean_rpm = float(np.mean(mouse_rpms))
            sem_rpm  = float(np.std(mouse_rpms, ddof=1) / np.sqrt(len(mouse_rpms))) if len(mouse_rpms) > 1 else 0.0
            color = condition_color_map[condition]
            ax_bar.bar(ci, mean_rpm, width=0.5, color=color, alpha=0.8,
                       yerr=sem_rpm, capsize=7, error_kw={'elinewidth': 1.5, 'capthick': 1.5})
            jitter = (rng_bar.random(len(mouse_rpms)) - 0.5) * 0.22
            for j, (mouse_name_bar, rpm_val) in enumerate(entries):
                ax_bar.plot(ci + jitter[j], rpm_val, 'o',
                            color='white', markeredgecolor=color,
                            markeredgewidth=1.8, markersize=7, zorder=3)

        ax_bar.set_xticks(x_pos_bar)
        ax_bar.set_xticklabels(conditions_sorted_bar)
        ax_bar.set_title('Average Reward Rate by Starting Condition\n(collapsed across all sessions)')
        ax_bar.set_xlabel('Starting Condition')
        ax_bar.set_ylabel('Rewards per Minute (Mean \u00b1 SEM)')
        ax_bar.set_ylim(bottom=0)
        ax_bar.tick_params(axis='both', direction='in')
        ax_bar.spines['top'].set_visible(False)
        ax_bar.spines['right'].set_visible(False)
        condition_bar_fig.tight_layout()

    # Create collapsed condition bar plot for speed
    condition_speed_bar_fig = None
    if 'condition_speed_bar' in selected_plots:
        condition_speed_bar_fig, ax_sbar = plt.subplots(figsize=(8, 6))

        condition_mouse_speeds: dict[str, list] = {}
        for result in all_results:
            condition = result['starting_condition']
            df_r = result['df']
            session_speeds = pd.to_numeric(df_r['average_speed'], errors='coerce').dropna().tolist()
            if session_speeds:
                if condition not in condition_mouse_speeds:
                    condition_mouse_speeds[condition] = []
                condition_mouse_speeds[condition].append((result['mouse'], float(np.mean(session_speeds))))

        conditions_sorted_sbar = sorted(condition_mouse_speeds.keys())
        x_pos_sbar = np.arange(len(conditions_sorted_sbar))

        rng_sbar = np.random.default_rng(seed=42)
        for ci, condition in enumerate(conditions_sorted_sbar):
            entries = condition_mouse_speeds[condition]
            mouse_speeds = [v for _, v in entries]
            mean_spd = float(np.mean(mouse_speeds))
            sem_spd  = float(np.std(mouse_speeds, ddof=1) / np.sqrt(len(mouse_speeds))) if len(mouse_speeds) > 1 else 0.0
            color = condition_color_map[condition]
            ax_sbar.bar(ci, mean_spd, width=0.5, color=color, alpha=0.8,
                        yerr=sem_spd, capsize=7, error_kw={'elinewidth': 1.5, 'capthick': 1.5})
            jitter = (rng_sbar.random(len(mouse_speeds)) - 0.5) * 0.22
            for j, (mouse_name_sbar, spd_val) in enumerate(entries):
                ax_sbar.plot(ci + jitter[j], spd_val, 'o',
                             color='white', markeredgecolor=color,
                             markeredgewidth=1.8, markersize=7, zorder=3)

        ax_sbar.set_xticks(x_pos_sbar)
        ax_sbar.set_xticklabels(conditions_sorted_sbar)
        ax_sbar.set_title('Average Speed by Starting Condition\n(collapsed across all sessions)')
        ax_sbar.set_xlabel('Starting Condition')
        ax_sbar.set_ylabel('Average Speed cm/s (Mean \u00b1 SEM)')
        ax_sbar.set_ylim(bottom=0)
        ax_sbar.tick_params(axis='both', direction='in')
        ax_sbar.spines['top'].set_visible(False)
        ax_sbar.spines['right'].set_visible(False)
        condition_speed_bar_fig.tight_layout()

    # ── Exploratory: speed histogram ──────────────────────────────────────────
    expl_speed_histogram_fig = None
    if 'expl_speed_histogram' in selected_plots:
        from scipy.stats import shapiro
        from scipy.stats import gaussian_kde as _gkde

        # Collect data (log-transformed: zero/negative values excluded)
        all_session_speeds = []
        per_mouse_means    = []
        for _r in all_results:
            _ss = pd.to_numeric(_r['df']['average_speed'], errors='coerce').dropna()
            _ss = _ss[_ss > 0].tolist()
            all_session_speeds.extend(_ss)
            if _ss:
                per_mouse_means.append(float(np.mean(_ss)))

        all_session_speeds = np.log(np.array(all_session_speeds, dtype=float))
        per_mouse_means    = np.log(np.array(per_mouse_means,    dtype=float))

        expl_speed_histogram_fig, (ax_h1, ax_h2) = plt.subplots(1, 2, figsize=(14, 5))

        for ax_h, data, panel_title in [
            (ax_h1, all_session_speeds,
             f'All session speeds (n={len(all_session_speeds)} sessions)'),
            (ax_h2, per_mouse_means,
             f'Per-mouse mean speed (n={len(per_mouse_means)} mice)'),
        ]:
            ax_h.hist(data, bins='auto', color='steelblue', alpha=0.65,
                      edgecolor='white', linewidth=0.5, density=True)
            # KDE overlay
            if len(data) >= 3:
                _kde_x = np.linspace(data.min(), data.max(), 300)
                try:
                    _kde = _gkde(data)
                    ax_h.plot(_kde_x, _kde(_kde_x), color='navy', linewidth=2, label='KDE')
                except Exception:
                    pass
            # Shapiro-Wilk (only valid for n >= 3)
            if len(data) >= 3:
                _sw_stat, _sw_p = shapiro(data)
                _sw_text = (f'Shapiro-Wilk: W={_sw_stat:.3f}, p={_sw_p:.4f}\n'
                            f'{"Normal (p>0.05)" if _sw_p > 0.05 else "Non-normal (p\u22640.05)"}')
                ax_h.text(0.97, 0.97, _sw_text, transform=ax_h.transAxes,
                          fontsize=8, va='top', ha='right',
                          bbox=dict(boxstyle='round,pad=0.3', facecolor='lightyellow',
                                    edgecolor='gray', alpha=0.8))
            ax_h.set_title(panel_title)
            ax_h.set_xlabel('log(Average speed)  [log cm/s]')
            ax_h.set_ylabel('Density')
            ax_h.tick_params(axis='both', direction='in')
            ax_h.spines['top'].set_visible(False)
            ax_h.spines['right'].set_visible(False)

        expl_speed_histogram_fig.suptitle(
            'Log-Transformed Speed Distribution — Exploratory\n'
            'Left: one value per session per mouse (raw repeated-measures data)  |  '
            'Right: one mean per mouse',
            fontsize=10,
        )
        expl_speed_histogram_fig.tight_layout()

    # ── Exploratory: average speed distribution fit (per-mouse means by cohort) ─
    # One mean per mouse; grouped by starting_condition.
    # Goal: assess normality per cohort to decide whether a t-test is appropriate.
    expl_speed_distfit_fig = None
    if 'expl_speed_distfit' in selected_plots:
        try:
            from scipy.stats import (norm as _norm_sd, probplot as _probplot_sd,
                                     shapiro as _sw_sd, beta as _beta_sd)
            from matplotlib.gridspec import GridSpec as _GridSpec_sd
            import warnings as _warnings_sd

            # ── One mean per mouse, grouped by starting_condition ─────────────
            _sd_groups = {}
            for _r in all_results:
                _sv = pd.to_numeric(
                    _r['df']['average_speed'], errors='coerce'
                ).dropna().values
                _sv = _sv[_sv > 0]
                if len(_sv) > 0:
                    _cond_sd = _r.get('starting_condition', 'Unknown')
                    _sd_groups.setdefault(_cond_sd, []).append(float(np.mean(_sv)))

            _sd_cond_names = sorted(_sd_groups.keys())
            _n_cg_sd = len(_sd_cond_names)
            if _n_cg_sd == 0:
                raise ValueError('No valid speed data found')

            # ── Shapiro-Wilk per cohort ────────────────────────────────────────
            _sd_sw = {}
            for _cn in _sd_cond_names:
                _arr = np.array(_sd_groups[_cn])
                if len(_arr) >= 3:
                    _w, _p = _sw_sd(_arr)
                    _sd_sw[_cn] = (float(_w), float(_p), len(_arr))
                else:
                    _sd_sw[_cn] = (np.nan, np.nan, len(_arr))

            # ── Figure layout ─────────────────────────────────────────────────
            # Row 0 (full width): overlaid histograms + Normal fit + rug marks
            # Rows 1.._n_qq_rows: per-cohort Normal Q-Q plots (2-column grid)
            # Last row (full width): summary normality table
            _n_qq_rows_sd = (_n_cg_sd + 1) // 2
            _n_rows_sd = 1 + 1 + _n_qq_rows_sd + 1  # +1 row for box plot
            _coh_pal_sd = [plt.cm.tab10(i / max(_n_cg_sd - 1, 1))
                           for i in range(_n_cg_sd)]

            expl_speed_distfit_fig = plt.figure(figsize=(13, 4.5 * _n_rows_sd))
            _gs_sd = _GridSpec_sd(_n_rows_sd, 2, figure=expl_speed_distfit_fig,
                                  hspace=0.60, wspace=0.38)
            ax_sdhist = expl_speed_distfit_fig.add_subplot(_gs_sd[0, :])

            # [row 0] Overlaid histograms + Normal fit + rug marks
            _sd_rng = np.random.default_rng(42)
            for _ci, _cn in enumerate(_sd_cond_names):
                _arr = np.array(_sd_groups[_cn])
                _col_sd = _coh_pal_sd[_ci]
                _sw_lbl = (f'SW p={_sd_sw[_cn][1]:.3f}'
                           if not np.isnan(_sd_sw[_cn][1]) else 'SW: n<3')
                _nbins_sd = max(4, len(_arr) // 2 + 1)
                ax_sdhist.hist(_arr, bins=_nbins_sd, alpha=0.40, color=_col_sd,
                               density=True, edgecolor='white', linewidth=0.5,
                               label=f'{_cn}  (n={len(_arr)}, {_sw_lbl})')
                if len(_arr) >= 3:
                    _mn_sd = float(np.mean(_arr))
                    _sdv_sd = float(np.std(_arr, ddof=1))
                    _xfit_sd = np.linspace(_mn_sd - 4 * _sdv_sd,
                                           _mn_sd + 4 * _sdv_sd, 300)
                    ax_sdhist.plot(_xfit_sd, _norm_sd.pdf(_xfit_sd, _mn_sd, _sdv_sd),
                                   color=_col_sd, linewidth=2.0, linestyle='--')
                _jit_sd = _sd_rng.uniform(-0.003, 0.003, len(_arr))
                ax_sdhist.plot(_arr, _jit_sd, '|', color=_col_sd,
                               markersize=14, markeredgewidth=2.5, alpha=0.75)

            ax_sdhist.set_xlabel('Per-mouse mean average speed (cm/s)', fontsize=10)
            ax_sdhist.set_ylabel('Density', fontsize=10)
            ax_sdhist.set_title(
                'Per-mouse mean average speed — distribution by cohort\n'
                '(each point = one mouse; dashed lines = Normal fit; '
                'tick marks = individual mice)',
                fontsize=10,
            )
            ax_sdhist.legend(fontsize=8)
            ax_sdhist.spines['top'].set_visible(False)
            ax_sdhist.spines['right'].set_visible(False)
            ax_sdhist.tick_params(axis='both', direction='in')

            # [row 1] Box-and-whisker with 1.5×IQR outlier highlighting
            ax_sdbox = expl_speed_distfit_fig.add_subplot(_gs_sd[1, :])
            _bp_data_sd = [np.array(_sd_groups[_cn]) for _cn in _sd_cond_names]
            _bplot_sd = ax_sdbox.boxplot(
                _bp_data_sd, labels=_sd_cond_names, patch_artist=True,
                showfliers=False, widths=0.50,
                boxprops=dict(linewidth=1.4),
                whiskerprops=dict(linewidth=1.2, linestyle='--'),
                capprops=dict(linewidth=1.4),
                medianprops=dict(color='black', linewidth=2.0),
            )
            for _bi, _bpatch in enumerate(_bplot_sd['boxes']):
                _bc = _coh_pal_sd[_bi]
                _bpatch.set_facecolor([*_bc[:3], 0.35])
                _bpatch.set_edgecolor(_bc)
            _rng_box_sd = np.random.default_rng(1)
            _outlier_legend_added_sd = False
            for _bi, _cn in enumerate(_sd_cond_names):
                _arr_bx = np.array(_sd_groups[_cn])
                _q1_sd  = float(np.percentile(_arr_bx, 25))
                _q3_sd  = float(np.percentile(_arr_bx, 75))
                _iqr_sd = _q3_sd - _q1_sd
                _lo_sd  = _q1_sd - 1.5 * _iqr_sd
                _hi_sd  = _q3_sd + 1.5 * _iqr_sd
                _out_sd = (_arr_bx < _lo_sd) | (_arr_bx > _hi_sd)
                _jit_sd = _rng_box_sd.uniform(-0.15, 0.15, len(_arr_bx))
                ax_sdbox.scatter(
                    np.full(int(np.sum(~_out_sd)), _bi + 1) + _jit_sd[~_out_sd],
                    _arr_bx[~_out_sd], color=_coh_pal_sd[_bi],
                    s=40, alpha=0.70, zorder=3,
                )
                if _out_sd.any():
                    ax_sdbox.scatter(
                        np.full(int(np.sum(_out_sd)), _bi + 1) + _jit_sd[_out_sd],
                        _arr_bx[_out_sd], color='red', s=80, alpha=0.90,
                        zorder=4, edgecolors='darkred', linewidths=1.5, marker='D',
                        label='Outlier (1.5\u00d7IQR)' if not _outlier_legend_added_sd else '',
                    )
                    _outlier_legend_added_sd = True
                    for _ov in _arr_bx[_out_sd]:
                        ax_sdbox.annotate(
                            f'{_ov:.3g}', xy=(_bi + 1, _ov),
                            xytext=(8, 0), textcoords='offset points',
                            fontsize=7, color='darkred', va='center',
                        )
            ax_sdbox.set_xlabel('Cohort', fontsize=10)
            ax_sdbox.set_ylabel('Per-mouse mean (cm/s)', fontsize=10)
            ax_sdbox.set_title(
                'Box-and-whisker \u2014 per-mouse mean average speed by cohort\n'
                '(whiskers extend to 1.5\u00d7IQR; \u25c6 red diamonds = outliers beyond fence)',
                fontsize=10,
            )
            _handles_sd, _labels_sd = ax_sdbox.get_legend_handles_labels()
            if _handles_sd:
                ax_sdbox.legend(fontsize=8)
            ax_sdbox.spines['top'].set_visible(False)
            ax_sdbox.spines['right'].set_visible(False)
            ax_sdbox.tick_params(axis='both', direction='in')

            # [rows 2+] Per-cohort Normal Q-Q with 95% CI
            for _ci, _cn in enumerate(_sd_cond_names):
                _arr = np.array(_sd_groups[_cn])
                _row_qq = 2 + _ci // 2
                _col_qq = _ci % 2
                _ax_qq = expl_speed_distfit_fig.add_subplot(_gs_sd[_row_qq, _col_qq])
                _col_sd = _coh_pal_sd[_ci]
                if len(_arr) >= 3:
                    (_osm_sd, _osr_sd), (_sl_sd, _int_sd, _) = _probplot_sd(
                        _arr, dist='norm',
                    )
                    _n_qq_sd = len(_arr)
                    with _warnings_sd.catch_warnings():
                        _warnings_sd.simplefilter('ignore')
                        _ci_lo_sd = np.array([
                            _norm_sd.ppf(_beta_sd.ppf(0.025, _i + 1, _n_qq_sd - _i))
                            for _i in range(_n_qq_sd)
                        ])
                        _ci_hi_sd = np.array([
                            _norm_sd.ppf(_beta_sd.ppf(0.975, _i + 1, _n_qq_sd - _i))
                            for _i in range(_n_qq_sd)
                        ])
                    _ax_qq.fill_between(
                        _osm_sd,
                        _sl_sd * _ci_lo_sd + _int_sd,
                        _sl_sd * _ci_hi_sd + _int_sd,
                        color=_col_sd, alpha=0.18, label='95% CI',
                    )
                    _ax_qq.plot(_osm_sd, _osr_sd, 'o', color=_col_sd,
                                markersize=7, alpha=0.85, label='Mouse mean')
                    _ax_qq.plot(
                        [_osm_sd[0], _osm_sd[-1]],
                        [_sl_sd * _osm_sd[0] + _int_sd,
                         _sl_sd * _osm_sd[-1] + _int_sd],
                        'k-', linewidth=1.4, label='Reference line',
                    )
                    _w_sd, _p_sd = _sd_sw[_cn][:2]
                    _verdict_sd = ('Normal (p>0.05) — t-test OK \u2713'
                                   if _p_sd > 0.05
                                   else 'Non-normal (p\u22640.05) — consider Mann-Whitney')
                    _ax_qq.text(
                        0.04, 0.96,
                        f'SW: W={_w_sd:.4f}, p={_p_sd:.4f}\n{_verdict_sd}',
                        transform=_ax_qq.transAxes, fontsize=8, va='top',
                        bbox=dict(boxstyle='round,pad=0.3',
                                  facecolor='#e6ffe6' if _p_sd > 0.05 else '#ffe6e6',
                                  edgecolor=_col_sd, alpha=0.90),
                    )
                else:
                    _ax_qq.text(
                        0.5, 0.5,
                        f'n={len(_arr)} — need \u22653\nfor Shapiro-Wilk / Q-Q',
                        transform=_ax_qq.transAxes,
                        ha='center', va='center', fontsize=10,
                    )
                _ax_qq.set_title(f'{_cn} — Normal Q-Q  (n={len(_arr)} mice)', fontsize=9)
                _ax_qq.set_xlabel('Theoretical quantiles (Normal)', fontsize=8)
                _ax_qq.set_ylabel('Observed quantiles', fontsize=8)
                _ax_qq.legend(fontsize=7)
                _ax_qq.spines['top'].set_visible(False)
                _ax_qq.spines['right'].set_visible(False)
                _ax_qq.tick_params(axis='both', direction='in')

            # Hide unused Q-Q cell when odd cohort count
            if _n_cg_sd % 2 == 1:
                _ax_empty_sd = expl_speed_distfit_fig.add_subplot(
                    _gs_sd[2 + (_n_cg_sd - 1) // 2, 1])
                _ax_empty_sd.axis('off')

            # [last row] Summary table
            ax_sdtbl = expl_speed_distfit_fig.add_subplot(_gs_sd[_n_rows_sd - 1, :])
            ax_sdtbl.axis('off')
            _tbl_cols_sd = ['Cohort', 'N mice', 'Mean (cm/s)', 'SD', 'SW W', 'SW p',
                            'T-test normality OK?']
            _tbl_rows_sd = []
            for _cn in _sd_cond_names:
                _arr = np.array(_sd_groups[_cn])
                _w_sd, _p_sd, _nn_sd = _sd_sw[_cn]
                _mn_sd = float(np.mean(_arr))
                _sdv_str = (f'{np.std(_arr, ddof=1):.4g}'
                            if _nn_sd >= 2 else 'n/a')
                _ok_sd = ('Yes' if (not np.isnan(_p_sd) and _p_sd > 0.05)
                          else ('Insufficient n (need \u22653)'
                                if np.isnan(_p_sd)
                                else 'No \u2014 consider Mann-Whitney U'))
                _tbl_rows_sd.append([
                    _cn, str(_nn_sd), f'{_mn_sd:.4g}', _sdv_str,
                    f'{_w_sd:.4f}' if not np.isnan(_w_sd) else 'n/a',
                    f'{_p_sd:.4f}' if not np.isnan(_p_sd) else 'n/a',
                    _ok_sd,
                ])
            _tbl_sd = ax_sdtbl.table(
                cellText=_tbl_rows_sd, colLabels=_tbl_cols_sd,
                cellLoc='center', loc='center', bbox=[0, 0, 1, 1],
            )
            _tbl_sd.auto_set_font_size(False)
            _tbl_sd.set_fontsize(9)
            for (_ri, _ci2), _cell in _tbl_sd.get_celld().items():
                if _ri == 0:
                    _cell.set_facecolor('#d0d8e8')
                    _cell.set_text_props(fontweight='bold')
                elif _ci2 == 6 and _ri > 0:
                    _txt2 = _tbl_rows_sd[_ri - 1][6]
                    _cell.set_facecolor(
                        '#e6ffe6' if 'Yes' in _txt2
                        else '#ffe6e6' if 'No' in _txt2
                        else '#fffff0'
                    )
            ax_sdtbl.set_title(
                'Normality summary — one mean per mouse per cohort '
                '(Shapiro-Wilk, \u03b1=0.05)',
                fontsize=9, pad=6,
            )

            expl_speed_distfit_fig.suptitle(
                'Average Speed — Per-Mouse Means by Cohort\n'
                '(N=1 value per mouse = mean across all sessions; '
                'goal: assess normality for between-cohort t-test)',
                fontsize=11, y=1.01,
            )
            expl_speed_distfit_fig.tight_layout()

            print('\n\u2500\u2500 Average Speed \u2014 Per-Mouse Means by Cohort \u2500\u2500')
            for _cn in _sd_cond_names:
                _arr = np.array(_sd_groups[_cn])
                _w_sd, _p_sd, _nn_sd = _sd_sw[_cn]
                _sdv_str = f'{np.std(_arr, ddof=1):.4g}' if _nn_sd >= 2 else 'n/a'
                _w_str   = f'{_w_sd:.4f}' if not np.isnan(_w_sd) else 'n/a'
                _p_str   = f'{_p_sd:.4f}' if not np.isnan(_p_sd) else 'n/a'
                print(f'  {_cn}: n={_nn_sd}, mean={np.mean(_arr):.4g}, '
                      f'SD={_sdv_str}, SW W={_w_str}, p={_p_str}')

        except Exception as _e:
            import traceback as _tb_sd
            print(f'[expl_speed_distfit] Error: {_e}')
            _tb_sd.print_exc()
            expl_speed_distfit_fig = None

    # ── Exploratory: speed box-and-whisker ────────────────────────────────────
    expl_speed_boxplot_fig = None
    if 'expl_speed_boxplot' in selected_plots:
        # Collect per-mouse session speeds
        _mouse_names_bx  = []
        _mouse_speeds_bx = []
        _all_speeds_bx   = []
        for _r in all_results:
            _ss = pd.to_numeric(_r['df']['average_speed'], errors='coerce').dropna()
            _ss = _ss[_ss > 0].tolist()
            if _ss:
                _mouse_names_bx.append(_r['mouse'])
                _mouse_speeds_bx.append(np.log(_ss).tolist())
                _all_speeds_bx.extend(np.log(_ss).tolist())

        n_mice_bx = len(_mouse_names_bx)
        expl_speed_boxplot_fig, (ax_bx1, ax_bx2) = plt.subplots(
            1, 2, figsize=(max(10, n_mice_bx * 0.8 + 3), 6),
            gridspec_kw={'width_ratios': [max(3, n_mice_bx), 1]},
        )

        # Left: one box per mouse
        ax_bx1.boxplot(_mouse_speeds_bx, labels=_mouse_names_bx,
                       patch_artist=True,
                       boxprops=dict(facecolor='steelblue', alpha=0.6),
                       medianprops=dict(color='navy', linewidth=2),
                       whiskerprops=dict(color='steelblue'),
                       capprops=dict(color='steelblue'),
                       flierprops=dict(marker='o', markerfacecolor='steelblue',
                                       markersize=4, alpha=0.5, linestyle='none'))
        ax_bx1.set_title('Log-speed distribution per mouse\n(each session = one data point)')
        ax_bx1.set_xlabel('Mouse')
        ax_bx1.set_ylabel('log(Average session speed)  [log cm/s]')
        ax_bx1.tick_params(axis='x', rotation=45)
        ax_bx1.tick_params(axis='both', direction='in')
        ax_bx1.spines['top'].set_visible(False)
        ax_bx1.spines['right'].set_visible(False)

        # Right: overall single box (all sessions pooled)
        ax_bx2.boxplot([_all_speeds_bx], labels=['All mice'],
                       patch_artist=True,
                       boxprops=dict(facecolor='coral', alpha=0.6),
                       medianprops=dict(color='darkred', linewidth=2),
                       whiskerprops=dict(color='coral'),
                       capprops=dict(color='coral'),
                       flierprops=dict(marker='o', markerfacecolor='coral',
                                       markersize=4, alpha=0.5, linestyle='none'))
        ax_bx2.set_title(f'Overall\n(n={len(_all_speeds_bx)} sessions)')
        ax_bx2.set_ylabel('log(Average session speed)  [log cm/s]')
        ax_bx2.tick_params(axis='both', direction='in')
        ax_bx2.spines['top'].set_visible(False)
        ax_bx2.spines['right'].set_visible(False)

        expl_speed_boxplot_fig.suptitle('Log-Transformed Speed Box-and-Whisker — Exploratory', fontsize=11)
        expl_speed_boxplot_fig.tight_layout()

    # ── Exploratory: RM ANOVA residual diagnostics ────────────────────────────
    expl_speed_rm_anova_resid_fig = None
    if 'expl_speed_rm_anova_resid' in selected_plots:
        import warnings as _warnings
        try:
            import statsmodels.formula.api as _smf
            from statsmodels.stats.diagnostic import het_breuschpagan as _bp_test
            from scipy.stats import shapiro as _shapiro, levene as _levene
            from scipy.stats import probplot as _probplot

            # ── Build long-format DataFrame (log-transformed speed) ───────────
            _rows = []
            for _r in all_results:
                _df_r = _r['df'].copy()
                _df_r = _df_r.reset_index(drop=True)
                _df_r['session_num'] = np.arange(1, len(_df_r) + 1, dtype=float)
                _df_r['mouse']       = _r['mouse']
                _df_r['condition']   = _r['starting_condition']
                _raw_spd = pd.to_numeric(_df_r['average_speed'], errors='coerce')
                _df_r['speed']       = np.where(_raw_spd > 0, np.log(_raw_spd), np.nan)
                _rows.append(_df_r[['mouse', 'condition', 'session_num', 'speed']])
            _df_long = pd.concat(_rows, ignore_index=True).dropna(subset=['speed'])

            # ── Fit OLS with mouse as fixed-effect blocking factor ────────────
            # Including C(mouse) as a blocking factor gives within-subject
            # residuals equivalent to RM ANOVA residuals.
            # DV is log(speed) — appropriate for a lognormal response variable.
            _formula = 'speed ~ C(condition) + session_num + C(condition):session_num + C(mouse)'
            with _warnings.catch_warnings():
                _warnings.simplefilter('ignore')
                _ols_result = _smf.ols(_formula, data=_df_long).fit()

            _resid  = _ols_result.resid.values
            _fitted = _ols_result.fittedvalues.values
            _conds  = _df_long['condition'].values
            _resid_mean = float(np.mean(_resid))

            # ── Additional error-term statistics ──────────────────────────────
            # Var(ε): use model MSE (= SS_resid / df_resid) for unbiased estimate
            _resid_var  = float(_ols_result.mse_resid)          # σ² = MSE
            _resid_sd   = float(np.sqrt(_resid_var))            # σ  = RMSE
            _r2         = float(_ols_result.rsquared)
            _r2_adj     = float(_ols_result.rsquared_adj)
            # Variance of response (log-speed) — total variance (needed by PRESS below)
            _y_mean     = float(_df_long['speed'].mean())
            _y_var      = float(_df_long['speed'].var(ddof=1))
            _y_sd       = float(np.sqrt(_y_var))
            # Predicted R²: uses PRESS statistic (leave-one-out cross-validation)
            # PRESS = Σ (εᵢ / (1 − hᵢᵢ))²  where hᵢᵢ = leverage (hat matrix diagonal)
            # Predicted R² = 1 − PRESS / SS_total
            try:
                _influence   = _ols_result.get_influence()
                _hat         = _influence.hat_matrix_diag          # hᵢᵢ for each obs
                _press_resid = _resid / (1.0 - np.clip(_hat, None, 0.9999))
                _press       = float(np.sum(_press_resid ** 2))
                _ss_total    = float(np.sum((_df_long['speed'].values - _y_mean) ** 2))
                _r2_pred     = float(1.0 - _press / _ss_total) if _ss_total > 0 else np.nan
            except Exception:
                _press, _r2_pred = np.nan, np.nan
            # Var(ε) / Var(Y) = unexplained fraction (= 1 − R²)
            _unexplained = 1.0 - _r2
            # Signal-to-noise: mean fitted value / residual SD
            _snr        = float(np.mean(_fitted)) / _resid_sd if _resid_sd > 0 else np.nan

            # ── Shapiro-Wilk on residuals ─────────────────────────────────────
            _sw_stat, _sw_p = _shapiro(_resid) if len(_resid) >= 3 else (np.nan, np.nan)

            # ── Levene's test across condition groups ─────────────────────────
            _cond_groups = [_resid[_conds == c] for c in np.unique(_conds)
                            if np.sum(_conds == c) >= 2]
            if len(_cond_groups) >= 2:
                _lev_stat, _lev_p = _levene(*_cond_groups)
            else:
                _lev_stat, _lev_p = np.nan, np.nan

            # ── Also try pingouin for the ANOVA table (optional) ──────────────
            _pg_table_text = ''
            try:
                import pingouin as _pg
                _pg_result = _pg.mixed_anova(
                    data=_df_long, dv='speed', within='session_num',
                    between='condition', subject='mouse',
                )  # dv='speed' is log(speed) at this point
                # Format key rows
                _pg_lines = ['Mixed ANOVA (pingouin):']
                for _, _row in _pg_result.iterrows():
                    _src   = _row.get('Source', '')
                    _f     = _row.get('F', np.nan)
                    _pval  = _row.get('p-unc', np.nan)
                    _eta   = _row.get('np2', np.nan)
                    _pg_lines.append(
                        f"  {_src:<28} F={_f:.3f}  p={_pval:.4f}  \u03b7\u00b2={_eta:.3f}"
                    )
                _pg_table_text = '\n'.join(_pg_lines)
            except Exception:
                _pg_table_text = 'pingouin not available — ANOVA table omitted'

            # ── Cook's D and leverage (already have _hat from PRESS block) ──────
            try:
                _cooks_d    = _influence.cooks_distance[0]          # shape (n,)
                _lev_diag   = _hat.copy()                            # hᵢᵢ
                _n_params   = len(_ols_result.params)
                # Conventional thresholds
                _cooks_thresh = 4.0 / max(len(_resid) - _n_params, 1)  # 4/(n-p)
                _lev_thresh   = 2.0 * _n_params / max(len(_resid), 1)  # 2p/n
                _high_cooks   = np.where(_cooks_d > _cooks_thresh)[0]
                _high_lev     = np.where(_lev_diag > _lev_thresh)[0]
                _obs_index    = np.arange(len(_resid))
                _mice_labels  = _df_long['mouse'].values
                _sess_nums_lbl = _df_long['session_num'].values.astype(int)
                _obs_labels   = np.array([f'{m}\nS{s}' for m, s in
                                          zip(_mice_labels, _sess_nums_lbl)])
            except Exception:
                _cooks_d = _lev_diag = None
                _cooks_thresh = _lev_thresh = np.nan
                _high_cooks = _high_lev = np.array([], dtype=int)
                _obs_index = np.arange(len(_resid))
                _mice_labels = _df_long['mouse'].values
                _sess_nums_lbl = _df_long['session_num'].values.astype(int)
                _obs_labels   = np.array([f'{m}\nS{s}' for m, s in
                                          zip(_mice_labels, _sess_nums_lbl)])

            # ── Build figure: 3 rows × 2 cols ─────────────────────────────────
            expl_speed_rm_anova_resid_fig, _axes = plt.subplots(3, 2, figsize=(13, 15))
            (ax_qq, ax_rh), (ax_rf, ax_rs), (ax_cd, ax_lv) = _axes

            # ── Q-Q plot ──────────────────────────────────────────────────────
            (_qq_osm, _qq_osr), (_qq_slope, _qq_intercept, _qq_r) = _probplot(_resid, dist='norm')
            ax_qq.plot(_qq_osm, _qq_osr, 'o', color='steelblue',
                       markersize=4, alpha=0.7, label='Residuals')
            ax_qq.plot(
                [_qq_osm[0], _qq_osm[-1]],
                [_qq_slope * _qq_osm[0] + _qq_intercept,
                 _qq_slope * _qq_osm[-1] + _qq_intercept],
                'r-', linewidth=1.5, label='Normal line',
            )
            ax_qq.set_title('Normal Q-Q Plot of Residuals')
            ax_qq.set_xlabel('Theoretical quantiles')
            ax_qq.set_ylabel('Sample quantiles')
            ax_qq.tick_params(axis='both', direction='in')
            ax_qq.spines['top'].set_visible(False)
            ax_qq.spines['right'].set_visible(False)
            ax_qq.legend(fontsize=8)
            _sw_label = (f'Shapiro-Wilk: W={_sw_stat:.3f}, p={_sw_p:.4f}\n'
                         f'{"Normal (p>0.05)" if _sw_p > 0.05 else "Non-normal (p\u22640.05)"}')
            ax_qq.text(0.03, 0.97, _sw_label, transform=ax_qq.transAxes,
                       fontsize=8, va='top', ha='left',
                       bbox=dict(boxstyle='round,pad=0.3', facecolor='lightyellow',
                                 edgecolor='gray', alpha=0.8))
            # Formula explanation box
            _formula_label = (
                'Error terms: \u03b5\u1d62 = y\u1d62 \u2212 \u0177\u1d62\n'
                'y\u1d62 = log(avg speed, session i)\n'
                '\u0177\u1d62 = model fitted value\n'
                'Model: log(speed) ~ condition\n'
                '       + session + condition\u00d7session\n'
                '       + mouse  (blocking factor)'
            )
            ax_qq.text(0.97, 0.03, _formula_label, transform=ax_qq.transAxes,
                       fontsize=7.5, va='bottom', ha='right', family='monospace',
                       bbox=dict(boxstyle='round,pad=0.35', facecolor='#f5f5f5',
                                 edgecolor='gray', alpha=0.88))

            # ── Residual histogram ────────────────────────────────────────────
            from scipy.stats import gaussian_kde as _gkde2
            ax_rh.hist(_resid, bins='auto', color='steelblue', alpha=0.65,
                       edgecolor='white', linewidth=0.5, density=True)
            if len(_resid) >= 3:
                _kde_x2 = np.linspace(_resid.min(), _resid.max(), 300)
                try:
                    _kde2 = _gkde2(_resid)
                    ax_rh.plot(_kde_x2, _kde2(_kde_x2), color='navy',
                               linewidth=2, label='KDE')
                except Exception:
                    pass
            ax_rh.axvline(0, color='red', linewidth=1.5, linestyle='--', label='Zero')
            ax_rh.axvline(_resid_mean, color='darkorange', linewidth=1.5,
                          linestyle=':', label=f'Mean={_resid_mean:.3f}')
            ax_rh.set_title('Histogram of Residuals')
            ax_rh.set_xlabel('Residual \u03b5\u1d62  (log cm/s)')
            ax_rh.set_ylabel('Density')
            ax_rh.tick_params(axis='both', direction='in')
            ax_rh.spines['top'].set_visible(False)
            ax_rh.spines['right'].set_visible(False)
            ax_rh.legend(fontsize=8)
            # Extended error-term stats box
            _r2_pred_str = f'{_r2_pred:.4f}' if not np.isnan(_r2_pred) else 'n/a'
            _stats_box = (
                f'E[\u03b5]         = {_resid_mean:+.5f}  (\u22480 by OLS)\n'
                f'Var(\u03b5)  = MSE = {_resid_var:.5f}\n'
                f'SD(\u03b5)   = RMSE = {_resid_sd:.5f}\n'
                f'\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\n'
                f'Var(Y)         = {_y_var:.5f}\n'
                f'Var(\u03b5)/Var(Y) = {_unexplained:.4f}  (= 1\u2212R\u00b2)\n'
                f'R\u00b2             = {_r2:.4f}\n'
                f'R\u00b2 adj         = {_r2_adj:.4f}\n'
                f'R\u00b2 pred (PRESS) = {_r2_pred_str}\n'
                f'SNR (fitted/SD\u03b5) = {_snr:.3f}'
            )
            ax_rh.text(0.98, 0.98, _stats_box, transform=ax_rh.transAxes,
                       fontsize=7.5, va='top', ha='right', family='monospace',
                       bbox=dict(boxstyle='round,pad=0.35', facecolor='#f0f4ff',
                                 edgecolor='steelblue', alpha=0.92))

            # ── Residuals vs Fitted ───────────────────────────────────────────
            _uniq_conds = np.unique(_conds)
            _cond_color_rm = {c: _condition_to_color(c) for c in _uniq_conds}
            for _c in _uniq_conds:
                _mask = _conds == _c
                ax_rf.scatter(_fitted[_mask], _resid[_mask],
                              color=_cond_color_rm[_c], alpha=0.5,
                              s=18, label=str(_c))
            ax_rf.axhline(0, color='red', linewidth=1.5, linestyle='--')
            # Lowess smoothing overlay
            try:
                from statsmodels.nonparametric.smoothers_lowess import lowess as _lowess
                _lw = _lowess(_resid, _fitted, frac=0.4)
                ax_rf.plot(_lw[:, 0], _lw[:, 1], color='black',
                           linewidth=1.5, linestyle='-', label='Lowess')
            except Exception:
                pass
            ax_rf.set_title('Residuals vs Fitted Values')
            ax_rf.set_xlabel('Fitted values (log cm/s)')
            ax_rf.set_ylabel('Residuals (log cm/s)')
            ax_rf.tick_params(axis='both', direction='in')
            ax_rf.spines['top'].set_visible(False)
            ax_rf.spines['right'].set_visible(False)
            ax_rf.legend(fontsize=8, title='Condition')

            # ── Residuals vs session number ───────────────────────────────────
            _sess_nums = _df_long['session_num'].values
            ax_rs.scatter(_sess_nums, _resid,
                          color='steelblue', alpha=0.4, s=18)
            ax_rs.axhline(0, color='red', linewidth=1.5, linestyle='--')
            try:
                _lw2 = _lowess(_resid, _sess_nums, frac=0.4)
                ax_rs.plot(_lw2[:, 0], _lw2[:, 1], color='black',
                           linewidth=1.5, linestyle='-', label='Lowess')
                ax_rs.legend(fontsize=8)
            except Exception:
                pass
            ax_rs.set_title('Residuals vs Session Number\n(check for time-trend)')
            ax_rs.set_xlabel('Session number')
            ax_rs.set_ylabel('Residuals (log cm/s)')
            ax_rs.tick_params(axis='both', direction='in')
            ax_rs.spines['top'].set_visible(False)
            ax_rs.spines['right'].set_visible(False)

            # ── Cook's D bar chart ────────────────────────────────────────────
            if _cooks_d is not None:
                _cd_colors = np.where(_cooks_d > _cooks_thresh, 'crimson', 'steelblue')
                ax_cd.bar(_obs_index, _cooks_d, color=_cd_colors, alpha=0.75, width=0.8)
                ax_cd.axhline(_cooks_thresh, color='crimson', linewidth=1.5,
                              linestyle='--',
                              label=f'Threshold 4/(n\u2212p)={_cooks_thresh:.3f}')
                # Label flagged observations with mouse ID + session
                for _idx in _high_cooks:
                    ax_cd.text(_idx, _cooks_d[_idx] * 1.04, _obs_labels[_idx],
                               fontsize=6, ha='center', va='bottom', color='crimson',
                               linespacing=1.2)
                ax_cd.set_title(
                    f"Cook's D per Observation\n"
                    f"({len(_high_cooks)} flagged > threshold, shown in red)"
                )
                ax_cd.set_xlabel('Observation index')
                ax_cd.set_ylabel("Cook's D")
                ax_cd.tick_params(axis='both', direction='in')
                ax_cd.spines['top'].set_visible(False)
                ax_cd.spines['right'].set_visible(False)
                ax_cd.legend(fontsize=8)
            else:
                ax_cd.text(0.5, 0.5, "Cook's D unavailable",
                           transform=ax_cd.transAxes, ha='center', va='center')
                ax_cd.axis('off')

            # ── Leverage vs Cook's D (influence plot) ─────────────────────────
            if _cooks_d is not None:
                # Colour by condition
                for _c in _uniq_conds:
                    _mask = _conds == _c
                    ax_lv.scatter(_lev_diag[_mask], _cooks_d[_mask],
                                  color=_cond_color_rm[_c], alpha=0.6,
                                  s=20, label=str(_c))
                # Threshold lines
                ax_lv.axvline(_lev_thresh, color='darkorange', linewidth=1.5,
                              linestyle='--',
                              label=f'Leverage 2p/n={_lev_thresh:.3f}')
                ax_lv.axhline(_cooks_thresh, color='crimson', linewidth=1.5,
                              linestyle='--',
                              label=f"Cook's D 4/(n\u2212p)={_cooks_thresh:.3f}")
                # Label points in the high-influence quadrant (both thresholds exceeded)
                _both_flag = np.where((_cooks_d > _cooks_thresh) &
                                      (_lev_diag > _lev_thresh))[0]
                for _idx in _both_flag:
                    ax_lv.annotate(_obs_labels[_idx],
                                   (_lev_diag[_idx], _cooks_d[_idx]),
                                   textcoords='offset points', xytext=(4, 4),
                                   fontsize=6, color='crimson')
                ax_lv.set_title(
                    f'Leverage vs Cook\u2019s D (Influence Plot)\n'
                    f'High-leverage: {len(_high_lev)} obs  |  '
                    f'High-influence: {len(_high_cooks)} obs  |  '
                    f'Both: {len(_both_flag)} obs'
                )
                ax_lv.set_xlabel('Leverage hᵢᵢ')
                ax_lv.set_ylabel("Cook's D")
                ax_lv.tick_params(axis='both', direction='in')
                ax_lv.spines['top'].set_visible(False)
                ax_lv.spines['right'].set_visible(False)
                ax_lv.legend(fontsize=7, title='Condition', ncol=2)
            else:
                ax_lv.text(0.5, 0.5, 'Leverage unavailable',
                           transform=ax_lv.transAxes, ha='center', va='center')
                ax_lv.axis('off')

            # ── Suptitle: summary stats ───────────────────────────────────────
            _n_mice   = _df_long['mouse'].nunique()
            _n_obs    = len(_df_long)
            _lev_str  = (f'Levene (across conditions): F={_lev_stat:.3f}, p={_lev_p:.4f}'
                         if not np.isnan(_lev_p)
                         else 'Levene: insufficient groups')
            _summary  = (
                f'RM ANOVA residual diagnostics  |  '
                f'n={_n_mice} mice, {_n_obs} obs  |  '
                f'E[\u03b5]={_resid_mean:+.4f} (\u22480)  '
                f'Var(\u03b5)={_resid_var:.4f}  SD(\u03b5)={_resid_sd:.4f}  '
                f'Var(\u03b5)/Var(Y)={_unexplained:.3f} (=1\u2212R\u00b2)  '
                f'R\u00b2={_r2:.3f}  |  {_lev_str}'
            )
            expl_speed_rm_anova_resid_fig.suptitle(_summary, fontsize=8.5, y=1.01, wrap=True)

            # ── ANOVA table as figure text ────────────────────────────────────
            if _pg_table_text:
                expl_speed_rm_anova_resid_fig.text(
                    0.01, -0.02, _pg_table_text,
                    fontsize=7.5, family='monospace', va='top',
                    transform=expl_speed_rm_anova_resid_fig.transFigure,
                    bbox=dict(boxstyle='round,pad=0.4', facecolor='#f5f5f5',
                              edgecolor='gray', alpha=0.9),
                )

            expl_speed_rm_anova_resid_fig.tight_layout()
            print('\n── RM ANOVA Residual Diagnostics ──')
            print(f'  Formula:          \u03b5\u1d62 = y\u1d62 \u2212 \u0177\u1d62  (observed \u2212 model fitted value)')
            print(f'  Model DV:         log(average speed)  [log cm/s]')
            print(f'  N mice:           {_n_mice}')
            print(f'  N observations:   {_n_obs}')
            print(f'  E[\u03b5]  (mean):    {_resid_mean:+.6f}  (\u22480 by OLS construction)')
            print(f'  Var(\u03b5) (MSE):    {_resid_var:.6f}')
            print(f'  SD(\u03b5)  (RMSE):   {_resid_sd:.6f}')
            print(f'  Var(Y):           {_y_var:.6f}')
            print(f'  Var(\u03b5)/Var(Y):   {_unexplained:.4f}  (= 1 \u2212 R\u00b2 = unexplained variance fraction)')
            print(f'  R\u00b2:              {_r2:.4f}')
            print(f'  R\u00b2 adj:          {_r2_adj:.4f}')
            print(f'  R\u00b2 pred (PRESS): {_r2_pred_str}  (leave-one-out; gap vs R\u00b2 indicates overfitting)')
            print(f'  SNR (\u0177\u0305/SD\u03b5):    {_snr:.4f}')
            print(f'  Shapiro-Wilk:     W={_sw_stat:.4f}, p={_sw_p:.6f}')
            print(f'  Levene test:      F={_lev_stat:.4f}, p={_lev_p:.6f}')
            if _cooks_d is not None:
                print(f'  Cook\'s D thresh:  {_cooks_thresh:.4f}  (4/(n\u2212p))')
                print(f'  Leverage thresh:  {_lev_thresh:.4f}  (2p/n)')
                _flag_strs = [f'{m}(S{s})' for m, s in
                              zip(_mice_labels[_high_cooks], _sess_nums_lbl[_high_cooks])]
                print(f'  High-influence obs (Cook\'s D > thresh): {len(_high_cooks)}'
                      + (f'  [{", ".join(_flag_strs)}]' if len(_high_cooks) else ''))
                _lev_strs = [f'{m}(S{s})' for m, s in
                             zip(_mice_labels[_high_lev], _sess_nums_lbl[_high_lev])]
                print(f'  High-leverage obs (h\u1d62\u1d62 > thresh):      {len(_high_lev)}'
                      + (f'  [{", ".join(_lev_strs)}]' if len(_high_lev) else ''))
            if _pg_table_text:
                print(_pg_table_text)

        except ImportError as _e:
            print(f'[expl_speed_rm_anova_resid] Missing dependency: {_e}')
            print('  Install statsmodels: conda install statsmodels')
        except Exception as _e:
            print(f'[expl_speed_rm_anova_resid] Error: {_e}')

    # ── Exploratory: z-scored mean capacitive sensor value histogram ──────────
    # DV = z-scored mean raw capacitive sensor value (avg_cap), globally
    # z-scored across all sessions. Parallels the log(speed) exploratory histogram.
    expl_cap_histogram_fig = None
    if 'expl_cap_histogram' in selected_plots:
        from scipy.stats import shapiro as _shapiro_cap
        from scipy.stats import gaussian_kde as _gkde_cap

        # Collect per-session mean raw capacitive sensor values (avg_cap column)
        _cap_all_raw = []
        _cap_per_mouse_raw = []
        for _r in all_results:
            _df_r = _r['df']
            _cv = pd.to_numeric(_df_r.get('avg_cap', pd.Series(dtype=float)), errors='coerce')
            _cv_valid = _cv.dropna().values
            _cap_all_raw.extend(_cv_valid.tolist())
            if len(_cv_valid) > 0:
                _cap_per_mouse_raw.append(float(np.mean(_cv_valid)))

        # Global z-score
        _cap_all_raw = np.array(_cap_all_raw, dtype=float)
        _cap_per_m   = np.array(_cap_per_mouse_raw, dtype=float)
        _glb_mu, _glb_sd = float(np.mean(_cap_all_raw)), float(np.std(_cap_all_raw, ddof=1))
        if _glb_sd > 0:
            _cap_all_z = (_cap_all_raw - _glb_mu) / _glb_sd
            _cap_per_z = (_cap_per_m   - _glb_mu) / _glb_sd
        else:
            _cap_all_z = _cap_all_raw - _glb_mu
            _cap_per_z = _cap_per_m  - _glb_mu

        expl_cap_histogram_fig, (ax_ch1, ax_ch2) = plt.subplots(1, 2, figsize=(14, 5))
        for ax_ch, data_z, panel_title in [
            (ax_ch1, _cap_all_z,
             f'All session values (n={len(_cap_all_z)} sessions)'),
            (ax_ch2, _cap_per_z,
             f'Per-mouse mean (n={len(_cap_per_z)} mice)'),
        ]:
            ax_ch.hist(data_z, bins='auto', color='teal', alpha=0.65,
                       edgecolor='white', linewidth=0.5, density=True)
            if len(data_z) >= 3:
                _kde_xc = np.linspace(data_z.min(), data_z.max(), 300)
                try:
                    _kdec = _gkde_cap(data_z)
                    ax_ch.plot(_kde_xc, _kdec(_kde_xc), color='darkslategray',
                               linewidth=2, label='KDE')
                except Exception:
                    pass
            if len(data_z) >= 3:
                _swc_stat, _swc_p = _shapiro_cap(data_z)
                _swc_text = (f'Shapiro-Wilk: W={_swc_stat:.3f}, p={_swc_p:.4f}\n'
                             f'{"Normal (p>0.05)" if _swc_p > 0.05 else "Non-normal (p\u22640.05)"}')
                ax_ch.text(0.97, 0.97, _swc_text, transform=ax_ch.transAxes,
                           fontsize=8, va='top', ha='right',
                           bbox=dict(boxstyle='round,pad=0.3', facecolor='lightyellow',
                                     edgecolor='gray', alpha=0.8))
            ax_ch.set_title(panel_title)
            ax_ch.set_xlabel('Z-scored mean capacitive value  [z(raw cap units)]')
            ax_ch.set_ylabel('Density')
            ax_ch.tick_params(axis='both', direction='in')
            ax_ch.spines['top'].set_visible(False)
            ax_ch.spines['right'].set_visible(False)

        expl_cap_histogram_fig.suptitle(
            'Z-Scored Mean Capacitive Sensor Value Distribution — Exploratory\n'
            'Left: one value per session per mouse (raw repeated-measures data)  |  '
            'Right: one mean per mouse\n'
            f'Global z-score: \u03bc={_glb_mu:.4f} (raw cap units), \u03c3={_glb_sd:.4f}',
            fontsize=10,
        )
        expl_cap_histogram_fig.tight_layout()

    # ── Exploratory: z-scored mean capacitive value box-and-whisker ──────────
    expl_cap_boxplot_fig = None
    if 'expl_cap_boxplot' in selected_plots:
        _cap_bx_names = []
        _cap_bx_vals  = []
        _cap_bx_all   = []
        # Recompute global z-score parameters if histogram block was skipped
        _all_cv_tmp = []
        for _r in all_results:
            _df_r = _r['df']
            _cv = pd.to_numeric(_df_r.get('avg_cap', pd.Series(dtype=float)), errors='coerce')
            _all_cv_tmp.extend(_cv.dropna().values.tolist())
        _all_cv_tmp = np.array(_all_cv_tmp, dtype=float)
        _bx_mu  = float(np.mean(_all_cv_tmp)) if len(_all_cv_tmp) else 0.0
        _bx_sd  = float(np.std(_all_cv_tmp, ddof=1)) if len(_all_cv_tmp) > 1 else 1.0
        if _bx_sd == 0:
            _bx_sd = 1.0

        for _r in all_results:
            _df_r = _r['df']
            _cv = pd.to_numeric(_df_r.get('avg_cap', pd.Series(dtype=float)), errors='coerce')
            _cv_valid = _cv.dropna().values
            if len(_cv_valid) > 0:
                _z = (_cv_valid - _bx_mu) / _bx_sd
                _cap_bx_names.append(_r['mouse'])
                _cap_bx_vals.append(_z.tolist())
                _cap_bx_all.extend(_z.tolist())

        n_mice_cbx = len(_cap_bx_names)
        expl_cap_boxplot_fig, (ax_cbx1, ax_cbx2) = plt.subplots(
            1, 2, figsize=(max(10, n_mice_cbx * 0.8 + 3), 6),
            gridspec_kw={'width_ratios': [max(3, n_mice_cbx), 1]},
        )
        ax_cbx1.boxplot(_cap_bx_vals, labels=_cap_bx_names,
                        patch_artist=True,
                        boxprops=dict(facecolor='teal', alpha=0.6),
                        medianprops=dict(color='darkslategray', linewidth=2),
                        whiskerprops=dict(color='teal'),
                        capprops=dict(color='teal'),
                        flierprops=dict(marker='o', markerfacecolor='teal',
                                        markersize=4, alpha=0.5, linestyle='none'))
        ax_cbx1.set_title('Z-scored mean capacitive value per mouse\n(each session = one data point)')
        ax_cbx1.set_xlabel('Mouse')
        ax_cbx1.set_ylabel('Z-scored mean capacitive value  [z(raw cap)]')
        ax_cbx1.tick_params(axis='x', rotation=45)
        ax_cbx1.tick_params(axis='both', direction='in')
        ax_cbx1.spines['top'].set_visible(False)
        ax_cbx1.spines['right'].set_visible(False)

        ax_cbx2.boxplot([_cap_bx_all], labels=['All mice'],
                        patch_artist=True,
                        boxprops=dict(facecolor='mediumaquamarine', alpha=0.6),
                        medianprops=dict(color='darkgreen', linewidth=2),
                        whiskerprops=dict(color='mediumaquamarine'),
                        capprops=dict(color='mediumaquamarine'),
                        flierprops=dict(marker='o', markerfacecolor='mediumaquamarine',
                                        markersize=4, alpha=0.5, linestyle='none'))
        ax_cbx2.set_title(f'Overall\n(n={len(_cap_bx_all)} sessions)')
        ax_cbx2.set_ylabel('Z-scored mean capacitive value  [z(raw cap)]')
        ax_cbx2.tick_params(axis='both', direction='in')
        ax_cbx2.spines['top'].set_visible(False)
        ax_cbx2.spines['right'].set_visible(False)

        expl_cap_boxplot_fig.suptitle('Z-Scored Mean Capacitive Sensor Value Box-and-Whisker — Exploratory', fontsize=11)
        expl_cap_boxplot_fig.tight_layout()

    # ── Exploratory: capacitive RM ANOVA residual diagnostics ─────────────────
    expl_cap_rm_anova_resid_fig = None
    if 'expl_cap_rm_anova_resid' in selected_plots:
        import warnings as _warnings_cap
        try:
            import statsmodels.formula.api as _smf_cap
            from scipy.stats import shapiro as _shapiro_c, levene as _levene_c
            from scipy.stats import probplot as _probplot_c

            # ── Build long-format DataFrame (z-scored log-transformed mean cap value) ──
            # Log transform first (lognormal fit), then z-score.
            # First pass: collect log(avg_cap) values to compute global z-score params
            _c_all_cv = []
            for _r in all_results:
                _df_r = _r['df']
                _cv = pd.to_numeric(_df_r.get('avg_cap', pd.Series(dtype=float)), errors='coerce')
                _cv_pos = _cv.dropna()
                _cv_pos = _cv_pos[_cv_pos > 0]
                _c_all_cv.extend(np.log(_cv_pos.values).tolist())
            _c_all_cv = np.array(_c_all_cv, dtype=float)
            _c_mu = float(np.mean(_c_all_cv)) if len(_c_all_cv) else 0.0
            _c_sd = float(np.std(_c_all_cv, ddof=1)) if len(_c_all_cv) > 1 else 1.0
            if _c_sd == 0:
                _c_sd = 1.0

            # Second pass: build long-format df with z-scored log(avg_cap)
            _c_rows = []
            for _r in all_results:
                _df_r = _r['df'].copy()
                _df_r = _df_r.reset_index(drop=True)
                _df_r['session_num'] = np.arange(1, len(_df_r) + 1, dtype=float)
                _df_r['mouse']       = _r['mouse']
                _df_r['condition']   = _r['starting_condition']
                _cv = pd.to_numeric(_df_r.get('avg_cap', pd.Series(dtype=float)), errors='coerce')
                _df_r['z_avg_cap'] = np.where(
                    _cv.notna() & (_cv > 0),
                    (np.log(_cv.clip(lower=1e-9)) - _c_mu) / _c_sd,
                    np.nan,
                )
                _c_rows.append(_df_r[['mouse', 'condition', 'session_num', 'z_avg_cap']])
            _df_c_long = pd.concat(_c_rows, ignore_index=True).dropna(subset=['z_avg_cap'])

            # ── Fit OLS with mouse as fixed-effect blocking factor ────────────
            _c_formula = 'z_avg_cap ~ C(condition) + session_num + C(condition):session_num + C(mouse)'
            with _warnings_cap.catch_warnings():
                _warnings_cap.simplefilter('ignore')
                _c_ols = _smf_cap.ols(_c_formula, data=_df_c_long).fit()

            _c_resid  = _c_ols.resid.values
            _c_fitted = _c_ols.fittedvalues.values
            _c_conds  = _df_c_long['condition'].values
            _c_resid_mean = float(np.mean(_c_resid))
            _c_resid_var  = float(_c_ols.mse_resid)
            _c_resid_sd   = float(np.sqrt(_c_resid_var))
            _c_r2         = float(_c_ols.rsquared)
            _c_r2_adj     = float(_c_ols.rsquared_adj)
            _c_y_var      = float(_df_c_long['z_avg_cap'].var(ddof=1))
            _c_y_mean     = float(_df_c_long['z_avg_cap'].mean())
            _c_unexplained = 1.0 - _c_r2
            _c_snr         = float(np.mean(_c_fitted)) / _c_resid_sd if _c_resid_sd > 0 else np.nan

            # PRESS / predicted R²
            try:
                _c_influence = _c_ols.get_influence()
                _c_hat       = _c_influence.hat_matrix_diag
                _c_press_r   = _c_resid / (1.0 - np.clip(_c_hat, None, 0.9999))
                _c_press     = float(np.sum(_c_press_r ** 2))
                _c_ss_total  = float(np.sum((_df_c_long['z_avg_cap'].values - _c_y_mean) ** 2))
                _c_r2_pred   = float(1.0 - _c_press / _c_ss_total) if _c_ss_total > 0 else np.nan
            except Exception:
                _c_press, _c_r2_pred, _c_hat = np.nan, np.nan, None
                _c_influence = None

            # ── Shapiro-Wilk on residuals ─────────────────────────────────────
            _c_sw_stat, _c_sw_p = _shapiro_c(_c_resid) if len(_c_resid) >= 3 else (np.nan, np.nan)

            # ── Levene's test across condition groups ─────────────────────────
            _c_cond_groups = [_c_resid[_c_conds == c] for c in np.unique(_c_conds)
                              if np.sum(_c_conds == c) >= 2]
            if len(_c_cond_groups) >= 2:
                _c_lev_stat, _c_lev_p = _levene_c(*_c_cond_groups)
            else:
                _c_lev_stat, _c_lev_p = np.nan, np.nan

            # ── pingouin mixed ANOVA table (optional) ─────────────────────────
            _c_pg_text = ''
            try:
                import pingouin as _pg_c
                _c_pg_result = _pg_c.mixed_anova(
                    data=_df_c_long, dv='z_avg_cap', within='session_num',
                    between='condition', subject='mouse',
                )
                _c_pg_lines = ['Mixed ANOVA (pingouin):']
                for _, _row in _c_pg_result.iterrows():
                    _src  = _row.get('Source', '')
                    _f    = _row.get('F', np.nan)
                    _pval = _row.get('p-unc', np.nan)
                    _eta  = _row.get('np2', np.nan)
                    _c_pg_lines.append(
                        f"  {_src:<28} F={_f:.3f}  p={_pval:.4f}  \u03b7\u00b2={_eta:.3f}"
                    )
                _c_pg_text = '\n'.join(_c_pg_lines)
            except Exception:
                _c_pg_text = 'pingouin not available \u2014 ANOVA table omitted'

            # ── Cook's D and leverage ─────────────────────────────────────────
            _c_mice_labels   = _df_c_long['mouse'].values
            _c_sess_nums_lbl = _df_c_long['session_num'].values.astype(int)
            _c_obs_labels    = np.array([f'{m}\nS{s}' for m, s in
                                         zip(_c_mice_labels, _c_sess_nums_lbl)])
            _c_obs_index     = np.arange(len(_c_resid))
            try:
                _c_cooks_d  = _c_influence.cooks_distance[0]
                _c_lev_diag = _c_hat.copy()
                _c_n_params = len(_c_ols.params)
                _c_cooks_thresh = 4.0 / max(len(_c_resid) - _c_n_params, 1)
                _c_lev_thresh   = 2.0 * _c_n_params / max(len(_c_resid), 1)
                _c_high_cooks   = np.where(_c_cooks_d > _c_cooks_thresh)[0]
                _c_high_lev     = np.where(_c_lev_diag > _c_lev_thresh)[0]
            except Exception:
                _c_cooks_d = _c_lev_diag = None
                _c_cooks_thresh = _c_lev_thresh = np.nan
                _c_high_cooks = _c_high_lev = np.array([], dtype=int)

            # ── Build figure: 3 rows × 2 cols ────────────────────────────────
            expl_cap_rm_anova_resid_fig, _c_axes = plt.subplots(3, 2, figsize=(13, 15))
            (ax_cqq, ax_crh), (ax_crf, ax_crs), (ax_ccd, ax_clv) = _c_axes

            # ── Q-Q plot ──────────────────────────────────────────────────────
            (_c_qq_osm, _c_qq_osr), (_c_qq_slope, _c_qq_int, _) = _probplot_c(_c_resid, dist='norm')
            ax_cqq.plot(_c_qq_osm, _c_qq_osr, 'o', color='teal',
                        markersize=4, alpha=0.7, label='Residuals')
            ax_cqq.plot(
                [_c_qq_osm[0], _c_qq_osm[-1]],
                [_c_qq_slope * _c_qq_osm[0] + _c_qq_int,
                 _c_qq_slope * _c_qq_osm[-1] + _c_qq_int],
                'r-', linewidth=1.5, label='Normal line',
            )
            ax_cqq.set_title('Normal Q-Q Plot of Residuals')
            ax_cqq.set_xlabel('Theoretical quantiles')
            ax_cqq.set_ylabel('Sample quantiles')
            ax_cqq.tick_params(axis='both', direction='in')
            ax_cqq.spines['top'].set_visible(False)
            ax_cqq.spines['right'].set_visible(False)
            ax_cqq.legend(fontsize=8)
            _c_sw_label = (f'Shapiro-Wilk: W={_c_sw_stat:.3f}, p={_c_sw_p:.4f}\n'
                           f'{"Normal (p>0.05)" if _c_sw_p > 0.05 else "Non-normal (p\u22640.05)"}')
            ax_cqq.text(0.03, 0.97, _c_sw_label, transform=ax_cqq.transAxes,
                        fontsize=8, va='top', ha='left',
                        bbox=dict(boxstyle='round,pad=0.3', facecolor='lightyellow',
                                  edgecolor='gray', alpha=0.8))
            _c_formula_label = (
                'Error terms: \u03b5\u1d62 = y\u1d62 \u2212 \u0177\u1d62\n'
                'y\u1d62 = z-score(log(mean cap value), session i)\n'
                '\u0177\u1d62 = model fitted value\n'
                'Model: z(log(avg_cap)) ~ condition\n'
                '       + session + condition\u00d7session\n'
                '       + mouse  (blocking factor)'
            )
            ax_cqq.text(0.97, 0.03, _c_formula_label, transform=ax_cqq.transAxes,
                        fontsize=7.5, va='bottom', ha='right', family='monospace',
                        bbox=dict(boxstyle='round,pad=0.35', facecolor='#f5f5f5',
                                  edgecolor='gray', alpha=0.88))

            # ── Residual histogram ────────────────────────────────────────────
            from scipy.stats import gaussian_kde as _gkde_cr
            ax_crh.hist(_c_resid, bins='auto', color='teal', alpha=0.65,
                        edgecolor='white', linewidth=0.5, density=True)
            if len(_c_resid) >= 3:
                _c_kde_x = np.linspace(_c_resid.min(), _c_resid.max(), 300)
                try:
                    _c_kde2 = _gkde_cr(_c_resid)
                    ax_crh.plot(_c_kde_x, _c_kde2(_c_kde_x), color='darkslategray',
                                linewidth=2, label='KDE')
                except Exception:
                    pass
            ax_crh.axvline(0, color='red', linewidth=1.5, linestyle='--', label='Zero')
            ax_crh.axvline(_c_resid_mean, color='darkorange', linewidth=1.5,
                           linestyle=':', label=f'Mean={_c_resid_mean:.3f}')
            ax_crh.set_title('Histogram of Residuals')
            ax_crh.set_xlabel('Residual \u03b5\u1d62  [z-score units — z(mean cap value)]')
            ax_crh.set_ylabel('Density')
            ax_crh.tick_params(axis='both', direction='in')
            ax_crh.spines['top'].set_visible(False)
            ax_crh.spines['right'].set_visible(False)
            ax_crh.legend(fontsize=8)
            _c_r2_pred_str = f'{_c_r2_pred:.4f}' if not np.isnan(_c_r2_pred) else 'n/a'
            _c_stats_box = (
                f'E[\u03b5]         = {_c_resid_mean:+.5f}  (\u22480 by OLS)\n'
                f'Var(\u03b5)  = MSE = {_c_resid_var:.5f}\n'
                f'SD(\u03b5)   = RMSE = {_c_resid_sd:.5f}\n'
                f'\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\n'
                f'Var(Y)         = {_c_y_var:.5f}\n'
                f'Var(\u03b5)/Var(Y) = {_c_unexplained:.4f}  (= 1\u2212R\u00b2)\n'
                f'R\u00b2             = {_c_r2:.4f}\n'
                f'R\u00b2 adj         = {_c_r2_adj:.4f}\n'
                f'R\u00b2 pred (PRESS) = {_c_r2_pred_str}\n'
                f'SNR (fitted/SD\u03b5) = {_c_snr:.3f}'
            )
            ax_crh.text(0.98, 0.98, _c_stats_box, transform=ax_crh.transAxes,
                        fontsize=7.5, va='top', ha='right', family='monospace',
                        bbox=dict(boxstyle='round,pad=0.35', facecolor='#f0f4ff',
                                  edgecolor='steelblue', alpha=0.92))

            # ── Residuals vs Fitted ───────────────────────────────────────────
            _c_uniq_conds = np.unique(_c_conds)
            _c_cond_color = {c: _condition_to_color(c) for c in _c_uniq_conds}
            for _cc in _c_uniq_conds:
                _cmask = _c_conds == _cc
                ax_crf.scatter(_c_fitted[_cmask], _c_resid[_cmask],
                               color=_c_cond_color[_cc], alpha=0.5, s=18, label=str(_cc))
            ax_crf.axhline(0, color='red', linewidth=1.5, linestyle='--')
            try:
                from statsmodels.nonparametric.smoothers_lowess import lowess as _c_lowess
                _c_lw = _c_lowess(_c_resid, _c_fitted, frac=0.4)
                ax_crf.plot(_c_lw[:, 0], _c_lw[:, 1], color='black',
                            linewidth=1.5, linestyle='-', label='Lowess')
            except Exception:
                pass
            ax_crf.set_title('Residuals vs Fitted Values')
            ax_crf.set_xlabel('Fitted values [z(log(mean cap))]')
            ax_crf.set_ylabel('Residuals [z(log(mean cap))]')
            ax_crf.tick_params(axis='both', direction='in')
            ax_crf.spines['top'].set_visible(False)
            ax_crf.spines['right'].set_visible(False)
            ax_crf.legend(fontsize=8, title='Condition')

            # ── Residuals vs session number ───────────────────────────────────
            _c_sess_nums = _df_c_long['session_num'].values
            ax_crs.scatter(_c_sess_nums, _c_resid,
                           color='teal', alpha=0.4, s=18)
            ax_crs.axhline(0, color='red', linewidth=1.5, linestyle='--')
            try:
                _c_lw2 = _c_lowess(_c_resid, _c_sess_nums, frac=0.4)
                ax_crs.plot(_c_lw2[:, 0], _c_lw2[:, 1], color='black',
                            linewidth=1.5, linestyle='-', label='Lowess')
                ax_crs.legend(fontsize=8)
            except Exception:
                pass
            ax_crs.set_title('Residuals vs Session Number\n(check for time-trend)')
            ax_crs.set_xlabel('Session number')
            ax_crs.set_ylabel('Residuals [z(log(mean cap))]')
            ax_crs.tick_params(axis='both', direction='in')
            ax_crs.spines['top'].set_visible(False)
            ax_crs.spines['right'].set_visible(False)

            # ── Cook's D bar chart ────────────────────────────────────────────
            if _c_cooks_d is not None:
                _c_cd_colors = np.where(_c_cooks_d > _c_cooks_thresh, 'crimson', 'teal')
                ax_ccd.bar(_c_obs_index, _c_cooks_d, color=_c_cd_colors, alpha=0.75, width=0.8)
                ax_ccd.axhline(_c_cooks_thresh, color='crimson', linewidth=1.5,
                               linestyle='--',
                               label=f'Threshold 4/(n\u2212p)={_c_cooks_thresh:.3f}')
                for _idx in _c_high_cooks:
                    ax_ccd.text(_idx, _c_cooks_d[_idx] * 1.04, _c_obs_labels[_idx],
                                fontsize=6, ha='center', va='bottom', color='crimson',
                                linespacing=1.2)
                ax_ccd.set_title(
                    f"Cook's D per Observation\n"
                    f"({len(_c_high_cooks)} flagged > threshold, shown in red)"
                )
                ax_ccd.set_xlabel('Observation index')
                ax_ccd.set_ylabel("Cook's D")
                ax_ccd.tick_params(axis='both', direction='in')
                ax_ccd.spines['top'].set_visible(False)
                ax_ccd.spines['right'].set_visible(False)
                ax_ccd.legend(fontsize=8)
            else:
                ax_ccd.text(0.5, 0.5, "Cook's D unavailable",
                            transform=ax_ccd.transAxes, ha='center', va='center')
                ax_ccd.axis('off')

            # ── Leverage vs Cook's D (influence plot) ─────────────────────────
            if _c_cooks_d is not None:
                for _cc in _c_uniq_conds:
                    _cmask = _c_conds == _cc
                    ax_clv.scatter(_c_lev_diag[_cmask], _c_cooks_d[_cmask],
                                   color=_c_cond_color[_cc], alpha=0.6,
                                   s=20, label=str(_cc))
                ax_clv.axvline(_c_lev_thresh, color='darkorange', linewidth=1.5,
                               linestyle='--',
                               label=f'Leverage 2p/n={_c_lev_thresh:.3f}')
                ax_clv.axhline(_c_cooks_thresh, color='crimson', linewidth=1.5,
                               linestyle='--',
                               label=f"Cook's D 4/(n\u2212p)={_c_cooks_thresh:.3f}")
                _c_both_flag = np.where((_c_cooks_d > _c_cooks_thresh) &
                                        (_c_lev_diag > _c_lev_thresh))[0]
                for _idx in _c_both_flag:
                    ax_clv.annotate(_c_obs_labels[_idx],
                                    (_c_lev_diag[_idx], _c_cooks_d[_idx]),
                                    textcoords='offset points', xytext=(4, 4),
                                    fontsize=6, color='crimson')
                ax_clv.set_title(
                    f'Leverage vs Cook\u2019s D (Influence Plot)\n'
                    f'High-leverage: {len(_c_high_lev)} obs  |  '
                    f'High-influence: {len(_c_high_cooks)} obs  |  '
                    f'Both: {len(_c_both_flag)} obs'
                )
                ax_clv.set_xlabel('Leverage h\u1d62\u1d62')
                ax_clv.set_ylabel("Cook's D")
                ax_clv.tick_params(axis='both', direction='in')
                ax_clv.spines['top'].set_visible(False)
                ax_clv.spines['right'].set_visible(False)
                ax_clv.legend(fontsize=7, title='Condition', ncol=2)
            else:
                ax_clv.text(0.5, 0.5, 'Leverage unavailable',
                            transform=ax_clv.transAxes, ha='center', va='center')
                ax_clv.axis('off')

            # ── Suptitle: summary stats ───────────────────────────────────────
            _c_n_mice   = _df_c_long['mouse'].nunique()
            _c_n_obs    = len(_df_c_long)
            _c_lev_str  = (f'Levene (across conditions): F={_c_lev_stat:.3f}, p={_c_lev_p:.4f}'
                           if not np.isnan(_c_lev_p)
                           else 'Levene: insufficient groups')
            _c_summary  = (
                f'Capacitive RM ANOVA residual diagnostics (DV = z(log(mean cap value)))  |  '
                f'n={_c_n_mice} mice, {_c_n_obs} obs  |  '
                f'E[\u03b5]={_c_resid_mean:+.4f} (\u22480)  '
                f'Var(\u03b5)={_c_resid_var:.4f}  SD(\u03b5)={_c_resid_sd:.4f}  '
                f'Var(\u03b5)/Var(Y)={_c_unexplained:.3f} (=1\u2212R\u00b2)  '
                f'R\u00b2={_c_r2:.3f}  |  {_c_lev_str}'
            )
            expl_cap_rm_anova_resid_fig.suptitle(_c_summary, fontsize=8.5, y=1.01, wrap=True)

            if _c_pg_text:
                expl_cap_rm_anova_resid_fig.text(
                    0.01, -0.02, _c_pg_text,
                    fontsize=7.5, family='monospace', va='top',
                    transform=expl_cap_rm_anova_resid_fig.transFigure,
                    bbox=dict(boxstyle='round,pad=0.4', facecolor='#f5f5f5',
                              edgecolor='gray', alpha=0.9),
                )

            expl_cap_rm_anova_resid_fig.tight_layout()
            print('\n\u2500\u2500 Capacitive RM ANOVA Residual Diagnostics \u2500\u2500')
            print(f'  Formula:          \u03b5\u1d62 = y\u1d62 \u2212 \u0177\u1d62  (observed \u2212 model fitted value)')
            print(f'  Model DV:         z-score(log(mean capacitive value))  [z(log(cap units))]')
            print(f'  Global z-score:   \u03bc={_c_mu:.4f} log(cap units), \u03c3={_c_sd:.4f}  (computed on log scale)')
            print(f'  N mice:           {_c_n_mice}')
            print(f'  N observations:   {_c_n_obs}')
            print(f'  E[\u03b5]  (mean):    {_c_resid_mean:+.6f}  (\u22480 by OLS construction)')
            print(f'  Var(\u03b5) (MSE):    {_c_resid_var:.6f}')
            print(f'  SD(\u03b5)  (RMSE):   {_c_resid_sd:.6f}')
            print(f'  Var(Y):           {_c_y_var:.6f}')
            print(f'  Var(\u03b5)/Var(Y):   {_c_unexplained:.4f}  (= 1 \u2212 R\u00b2 = unexplained variance fraction)')
            print(f'  R\u00b2:              {_c_r2:.4f}')
            print(f'  R\u00b2 adj:          {_c_r2_adj:.4f}')
            print(f'  R\u00b2 pred (PRESS): {_c_r2_pred_str}  (leave-one-out; gap vs R\u00b2 indicates overfitting)')
            print(f'  SNR (\u0177\u0305/SD\u03b5):    {_c_snr:.4f}')
            print(f'  Shapiro-Wilk:     W={_c_sw_stat:.4f}, p={_c_sw_p:.6f}')
            print(f'  Levene test:      F={_c_lev_stat:.4f}, p={_c_lev_p:.6f}')
            if _c_cooks_d is not None:
                print(f'  Cook\'s D thresh:  {_c_cooks_thresh:.4f}  (4/(n\u2212p))')
                print(f'  Leverage thresh:  {_c_lev_thresh:.4f}  (2p/n)')
                _c_flag_strs = [f'{m}(S{s})' for m, s in
                                zip(_c_mice_labels[_c_high_cooks], _c_sess_nums_lbl[_c_high_cooks])]
                print(f'  High-influence obs (Cook\'s D > thresh): {len(_c_high_cooks)}'
                      + (f'  [{", ".join(_c_flag_strs)}]' if len(_c_high_cooks) else ''))
                _c_lev_strs = [f'{m}(S{s})' for m, s in
                               zip(_c_mice_labels[_c_high_lev], _c_sess_nums_lbl[_c_high_lev])]
                print(f'  High-leverage obs (h\u1d62\u1d62 > thresh):      {len(_c_high_lev)}'
                      + (f'  [{", ".join(_c_lev_strs)}]' if len(_c_high_lev) else ''))
            if _c_pg_text:
                print(_c_pg_text)

        except ImportError as _e:
            print(f'[expl_cap_rm_anova_resid] Missing dependency: {_e}')
            print('  Install statsmodels: conda install statsmodels')
        except Exception as _e:
            print(f'[expl_cap_rm_anova_resid] Error: {_e}')

    # ── Exploratory: capacitive sensor value distribution fit ─────────────────
    # DV = per-session mean raw capacitive value (avg_cap, pooled across all mice).
    # Tests Normal, Log-normal, and Gamma families via MLE + AIC.
    # Figure layout: 3 rows × 2 cols
    #   [0, :] Histogram of raw avg_cap values with fitted PDFs + KDE
    #   [1, 0] Normal Q-Q  |  [1, 1] Log-normal Q-Q
    #   [2, 0] Gamma Q-Q   |  [2, 1] AIC / fit-statistic comparison bar chart
    expl_cap_distfit_fig = None
    if 'expl_cap_distfit' in selected_plots:
        try:
            from scipy.stats import (norm as _norm_cd, probplot as _probplot_cd,
                                     shapiro as _sw_cd, beta as _beta_cd)
            from matplotlib.gridspec import GridSpec as _GridSpec_cd
            import warnings as _warnings_cd

            # ── One mean per mouse, grouped by starting_condition ─────────────
            _cd_groups = {}
            for _r in all_results:
                _cv = pd.to_numeric(
                    _r['df'].get('avg_cap', pd.Series(dtype=float)), errors='coerce'
                ).dropna().values
                if len(_cv) > 0:
                    _cond_cd = _r.get('starting_condition', 'Unknown')
                    _cd_groups.setdefault(_cond_cd, []).append(float(np.mean(_cv)))

            _cd_cond_names = sorted(_cd_groups.keys())
            _n_cg_cd = len(_cd_cond_names)
            if _n_cg_cd == 0:
                raise ValueError('No valid capacitance data found')

            # ── Shapiro-Wilk per cohort ────────────────────────────────────────
            _cd_sw = {}
            for _cn in _cd_cond_names:
                _arr = np.array(_cd_groups[_cn])
                if len(_arr) >= 3:
                    _w, _p = _sw_cd(_arr)
                    _cd_sw[_cn] = (float(_w), float(_p), len(_arr))
                else:
                    _cd_sw[_cn] = (np.nan, np.nan, len(_arr))

            # ── Figure layout ─────────────────────────────────────────────────
            # Row 0 (full width): overlaid histograms + Normal fit + rug marks
            # Rows 1.._n_qq_rows: per-cohort Normal Q-Q plots (2-column grid)
            # Last row (full width): summary normality table
            _n_qq_rows_cd = (_n_cg_cd + 1) // 2
            _n_rows_cd = 1 + 1 + _n_qq_rows_cd + 1  # +1 row for box plot
            _coh_pal_cd = [plt.cm.tab10(i / max(_n_cg_cd - 1, 1))
                           for i in range(_n_cg_cd)]

            expl_cap_distfit_fig = plt.figure(figsize=(13, 4.5 * _n_rows_cd))
            _gs_cd = _GridSpec_cd(_n_rows_cd, 2, figure=expl_cap_distfit_fig,
                                  hspace=0.60, wspace=0.38)
            ax_cdhist = expl_cap_distfit_fig.add_subplot(_gs_cd[0, :])

            # [row 0] Overlaid histograms + Normal fit + rug marks
            _cd_rng = np.random.default_rng(42)
            for _ci, _cn in enumerate(_cd_cond_names):
                _arr = np.array(_cd_groups[_cn])
                _col_cd = _coh_pal_cd[_ci]
                _sw_lbl = (f'SW p={_cd_sw[_cn][1]:.3f}'
                           if not np.isnan(_cd_sw[_cn][1]) else 'SW: n<3')
                _nbins_cd = max(4, len(_arr) // 2 + 1)
                ax_cdhist.hist(_arr, bins=_nbins_cd, alpha=0.40, color=_col_cd,
                               density=True, edgecolor='white', linewidth=0.5,
                               label=f'{_cn}  (n={len(_arr)}, {_sw_lbl})')
                if len(_arr) >= 3:
                    _mn_cd = float(np.mean(_arr))
                    _sdv_cd = float(np.std(_arr, ddof=1))
                    _xfit_cd = np.linspace(_mn_cd - 4 * _sdv_cd,
                                           _mn_cd + 4 * _sdv_cd, 300)
                    ax_cdhist.plot(_xfit_cd, _norm_cd.pdf(_xfit_cd, _mn_cd, _sdv_cd),
                                   color=_col_cd, linewidth=2.0, linestyle='--')
                _jit_cd = _cd_rng.uniform(-0.003, 0.003, len(_arr))
                ax_cdhist.plot(_arr, _jit_cd, '|', color=_col_cd,
                               markersize=14, markeredgewidth=2.5, alpha=0.75)

            ax_cdhist.set_xlabel('Per-mouse mean z-scored capacitive sensor value', fontsize=10)
            ax_cdhist.set_ylabel('Density', fontsize=10)
            ax_cdhist.set_title(
                'Per-mouse mean capacitive sensor value — distribution by cohort\n'
                '(each point = one mouse; dashed lines = Normal fit; '
                'tick marks = individual mice)',
                fontsize=10,
            )
            ax_cdhist.legend(fontsize=8)
            ax_cdhist.spines['top'].set_visible(False)
            ax_cdhist.spines['right'].set_visible(False)
            ax_cdhist.tick_params(axis='both', direction='in')

            # [row 1] Box-and-whisker with 1.5×IQR outlier highlighting
            ax_cdbox = expl_cap_distfit_fig.add_subplot(_gs_cd[1, :])
            _bp_data_cd = [np.array(_cd_groups[_cn]) for _cn in _cd_cond_names]
            _bplot_cd = ax_cdbox.boxplot(
                _bp_data_cd, labels=_cd_cond_names, patch_artist=True,
                showfliers=False, widths=0.50,
                boxprops=dict(linewidth=1.4),
                whiskerprops=dict(linewidth=1.2, linestyle='--'),
                capprops=dict(linewidth=1.4),
                medianprops=dict(color='black', linewidth=2.0),
            )
            for _bi, _bpatch in enumerate(_bplot_cd['boxes']):
                _bc = _coh_pal_cd[_bi]
                _bpatch.set_facecolor([*_bc[:3], 0.35])
                _bpatch.set_edgecolor(_bc)
            _rng_box_cd = np.random.default_rng(1)
            _outlier_legend_added_cd = False
            for _bi, _cn in enumerate(_cd_cond_names):
                _arr_bx = np.array(_cd_groups[_cn])
                _q1_cd  = float(np.percentile(_arr_bx, 25))
                _q3_cd  = float(np.percentile(_arr_bx, 75))
                _iqr_cd = _q3_cd - _q1_cd
                _lo_cd  = _q1_cd - 1.5 * _iqr_cd
                _hi_cd  = _q3_cd + 1.5 * _iqr_cd
                _out_cd = (_arr_bx < _lo_cd) | (_arr_bx > _hi_cd)
                _jit_cd = _rng_box_cd.uniform(-0.15, 0.15, len(_arr_bx))
                ax_cdbox.scatter(
                    np.full(int(np.sum(~_out_cd)), _bi + 1) + _jit_cd[~_out_cd],
                    _arr_bx[~_out_cd], color=_coh_pal_cd[_bi],
                    s=40, alpha=0.70, zorder=3,
                )
                if _out_cd.any():
                    ax_cdbox.scatter(
                        np.full(int(np.sum(_out_cd)), _bi + 1) + _jit_cd[_out_cd],
                        _arr_bx[_out_cd], color='red', s=80, alpha=0.90,
                        zorder=4, edgecolors='darkred', linewidths=1.5, marker='D',
                        label='Outlier (1.5\u00d7IQR)' if not _outlier_legend_added_cd else '',
                    )
                    _outlier_legend_added_cd = True
                    for _ov in _arr_bx[_out_cd]:
                        ax_cdbox.annotate(
                            f'{_ov:.3g}', xy=(_bi + 1, _ov),
                            xytext=(8, 0), textcoords='offset points',
                            fontsize=7, color='darkred', va='center',
                        )
            ax_cdbox.set_xlabel('Cohort', fontsize=10)
            ax_cdbox.set_ylabel('Per-mouse mean (z-score)', fontsize=10)
            ax_cdbox.set_title(
                'Box-and-whisker \u2014 per-mouse mean z-scored capacitive value by cohort\n'
                '(whiskers extend to 1.5\u00d7IQR; \u25c6 red diamonds = outliers beyond fence)',
                fontsize=10,
            )
            _handles_cd, _labels_cd = ax_cdbox.get_legend_handles_labels()
            if _handles_cd:
                ax_cdbox.legend(fontsize=8)
            ax_cdbox.spines['top'].set_visible(False)
            ax_cdbox.spines['right'].set_visible(False)
            ax_cdbox.tick_params(axis='both', direction='in')

            # [rows 2+] Per-cohort Normal Q-Q with 95% CI
            for _ci, _cn in enumerate(_cd_cond_names):
                _arr = np.array(_cd_groups[_cn])
                _row_qq = 2 + _ci // 2
                _col_qq = _ci % 2
                _ax_qq = expl_cap_distfit_fig.add_subplot(_gs_cd[_row_qq, _col_qq])
                _col_cd = _coh_pal_cd[_ci]
                if len(_arr) >= 3:
                    (_osm_cd, _osr_cd), (_sl_cd, _int_cd, _) = _probplot_cd(
                        _arr, dist='norm',
                    )
                    _n_qq_cd = len(_arr)
                    with _warnings_cd.catch_warnings():
                        _warnings_cd.simplefilter('ignore')
                        _ci_lo_cd = np.array([
                            _norm_cd.ppf(_beta_cd.ppf(0.025, _i + 1, _n_qq_cd - _i))
                            for _i in range(_n_qq_cd)
                        ])
                        _ci_hi_cd = np.array([
                            _norm_cd.ppf(_beta_cd.ppf(0.975, _i + 1, _n_qq_cd - _i))
                            for _i in range(_n_qq_cd)
                        ])
                    _ax_qq.fill_between(
                        _osm_cd,
                        _sl_cd * _ci_lo_cd + _int_cd,
                        _sl_cd * _ci_hi_cd + _int_cd,
                        color=_col_cd, alpha=0.18, label='95% CI',
                    )
                    _ax_qq.plot(_osm_cd, _osr_cd, 'o', color=_col_cd,
                                markersize=7, alpha=0.85, label='Mouse mean')
                    _ax_qq.plot(
                        [_osm_cd[0], _osm_cd[-1]],
                        [_sl_cd * _osm_cd[0] + _int_cd,
                         _sl_cd * _osm_cd[-1] + _int_cd],
                        'k-', linewidth=1.4, label='Reference line',
                    )
                    _w_cd, _p_cd = _cd_sw[_cn][:2]
                    _verdict_cd = ('Normal (p>0.05) \u2014 t-test OK \u2713'
                                   if _p_cd > 0.05
                                   else 'Non-normal (p\u22640.05) \u2014 consider Mann-Whitney')
                    _ax_qq.text(
                        0.04, 0.96,
                        f'SW: W={_w_cd:.4f}, p={_p_cd:.4f}\n{_verdict_cd}',
                        transform=_ax_qq.transAxes, fontsize=8, va='top',
                        bbox=dict(boxstyle='round,pad=0.3',
                                  facecolor='#e6ffe6' if _p_cd > 0.05 else '#ffe6e6',
                                  edgecolor=_col_cd, alpha=0.90),
                    )
                else:
                    _ax_qq.text(
                        0.5, 0.5,
                        f'n={len(_arr)} \u2014 need \u22653\nfor Shapiro-Wilk / Q-Q',
                        transform=_ax_qq.transAxes,
                        ha='center', va='center', fontsize=10,
                    )
                _ax_qq.set_title(f'{_cn} \u2014 Normal Q-Q  (n={len(_arr)} mice)', fontsize=9)
                _ax_qq.set_xlabel('Theoretical quantiles (Normal)', fontsize=8)
                _ax_qq.set_ylabel('Observed quantiles', fontsize=8)
                _ax_qq.legend(fontsize=7)
                _ax_qq.spines['top'].set_visible(False)
                _ax_qq.spines['right'].set_visible(False)
                _ax_qq.tick_params(axis='both', direction='in')

            # Hide unused Q-Q cell when odd cohort count
            if _n_cg_cd % 2 == 1:
                _ax_empty_cd = expl_cap_distfit_fig.add_subplot(
                    _gs_cd[2 + (_n_cg_cd - 1) // 2, 1])
                _ax_empty_cd.axis('off')

            # [last row] Summary table
            ax_cdtbl = expl_cap_distfit_fig.add_subplot(_gs_cd[_n_rows_cd - 1, :])
            ax_cdtbl.axis('off')
            _tbl_cols_cd = ['Cohort', 'N mice', 'Mean (z-score)', 'SD',
                            'SW W', 'SW p', 'T-test normality OK?']
            _tbl_rows_cd = []
            for _cn in _cd_cond_names:
                _arr = np.array(_cd_groups[_cn])
                _w_cd, _p_cd, _nn_cd = _cd_sw[_cn]
                _mn_cd = float(np.mean(_arr))
                _sdv_str = (f'{np.std(_arr, ddof=1):.4g}'
                            if _nn_cd >= 2 else 'n/a')
                _ok_cd = ('Yes' if (not np.isnan(_p_cd) and _p_cd > 0.05)
                          else ('Insufficient n (need \u22653)'
                                if np.isnan(_p_cd)
                                else 'No \u2014 consider Mann-Whitney U'))
                _tbl_rows_cd.append([
                    _cn, str(_nn_cd), f'{_mn_cd:.4g}', _sdv_str,
                    f'{_w_cd:.4f}' if not np.isnan(_w_cd) else 'n/a',
                    f'{_p_cd:.4f}' if not np.isnan(_p_cd) else 'n/a',
                    _ok_cd,
                ])
            _tbl_cd = ax_cdtbl.table(
                cellText=_tbl_rows_cd, colLabels=_tbl_cols_cd,
                cellLoc='center', loc='center', bbox=[0, 0, 1, 1],
            )
            _tbl_cd.auto_set_font_size(False)
            _tbl_cd.set_fontsize(9)
            for (_ri, _ci2), _cell in _tbl_cd.get_celld().items():
                if _ri == 0:
                    _cell.set_facecolor('#d0d8e8')
                    _cell.set_text_props(fontweight='bold')
                elif _ci2 == 6 and _ri > 0:
                    _txt2 = _tbl_rows_cd[_ri - 1][6]
                    _cell.set_facecolor(
                        '#e6ffe6' if 'Yes' in _txt2
                        else '#ffe6e6' if 'No' in _txt2
                        else '#fffff0'
                    )
            ax_cdtbl.set_title(
                'Normality summary \u2014 one mean per mouse per cohort '
                '(Shapiro-Wilk, \u03b1=0.05)',
                fontsize=9, pad=6,
            )

            expl_cap_distfit_fig.suptitle(
                'Z-Scored Capacitive Sensor Value \u2014 Per-Mouse Means by Cohort\n'
                '(N=1 value per mouse = mean across all sessions; '
                'goal: assess normality for between-cohort t-test)',
                fontsize=11, y=1.01,
            )
            expl_cap_distfit_fig.tight_layout()

            print('\n\u2500\u2500 Capacitive Sensor Value \u2014 Per-Mouse Means by Cohort \u2500\u2500')
            for _cn in _cd_cond_names:
                _arr = np.array(_cd_groups[_cn])
                _w_cd, _p_cd, _nn_cd = _cd_sw[_cn]
                _sdv_str = f'{np.std(_arr, ddof=1):.4g}' if _nn_cd >= 2 else 'n/a'
                _w_str   = f'{_w_cd:.4f}' if not np.isnan(_w_cd) else 'n/a'
                _p_str   = f'{_p_cd:.4f}' if not np.isnan(_p_cd) else 'n/a'
                print(f'  {_cn}: n={_nn_cd}, mean={np.mean(_arr):.4g}, '
                      f'SD={_sdv_str}, SW W={_w_str}, p={_p_str}')

        except Exception as _e:
            import traceback as _tb_cd
            print(f'[expl_cap_distfit] Error: {_e}')
            _tb_cd.print_exc()
            expl_cap_distfit_fig = None

    # ── Exploratory: lick count Poisson vs Negative Binomial distribution fit ──
    # Step 1 before any ANOVA: determine whether the count DV is Poisson-
    # distributed (mean ≈ variance) or overdispersed (variance >> mean, → NB).
    # The figure contains four panels:
    #   [0,0] Hanging rootogram — Poisson fit overlaid on count histogram
    #   [0,1] Hanging rootogram — NB fit overlaid on count histogram
    #   [1,0] Mean–variance plot (per-mouse means vs per-mouse variances)
    #   [1,1] Q-Q plot of counts against Poisson quantiles (pooled data)
    # Summary text reports: mean, variance, dispersion index (V/M), p-value
    # from a chi-squared goodness-of-fit for both Poisson and NB, and AIC.
    expl_lick_distfit_fig = None
    if 'expl_lick_distfit' in selected_plots:
        try:
            from scipy.stats import poisson as _poisson_dist, chi2 as _chi2_dist
            from scipy.stats import nbinom as _nbinom_dist
            import warnings as _warnings_df

            # ── Collect integer lick counts ───────────────────────────────────
            _df_lk_counts = []
            _df_per_mouse_means = []
            _df_per_mouse_vars  = []
            for _r in all_results:
                _lc = pd.to_numeric(_r['df']['lick_count'], errors='coerce').dropna()
                _lc_int = _lc[_lc >= 0].values.astype(float)
                _df_lk_counts.extend(_lc_int.tolist())
                if len(_lc_int) >= 2:
                    _df_per_mouse_means.append(float(np.mean(_lc_int)))
                    _df_per_mouse_vars.append(float(np.var(_lc_int, ddof=1)))

            _df_counts = np.array(_df_lk_counts, dtype=float)
            _df_n      = len(_df_counts)
            _df_mean   = float(np.mean(_df_counts))
            _df_var    = float(np.var(_df_counts, ddof=1))
            _df_disp   = _df_var / _df_mean if _df_mean > 0 else np.nan  # dispersion index

            # ── MLE for Poisson: λ = sample mean ─────────────────────────────
            _pois_lam = _df_mean

            # ── MLE for Negative Binomial via statsmodels ─────────────────────
            # NB parameterisation: mean=μ, dispersion=α → var = μ + α·μ²
            # statsmodels NB2 gives (params['const'], alpha)
            _nb_mu    = np.nan
            _nb_alpha = np.nan
            _nb_r     = np.nan   # r = 1/alpha  (scipy nbinom uses n=r, p=r/(r+mu))
            _nb_p_sc  = np.nan
            try:
                import statsmodels.api as _sm_nb
                _nb_endog = _df_counts.astype(int)
                _nb_exog  = np.ones(len(_nb_endog))
                with _warnings_df.catch_warnings():
                    _warnings_df.simplefilter('ignore')
                    _nb_res = _sm_nb.NegativeBinomial(_nb_endog, _nb_exog).fit(
                        disp=False, method='nm', maxiter=500,
                    )
                _nb_mu    = float(np.exp(_nb_res.params[0]))
                _nb_alpha = float(_nb_res.params[-1])         # overdispersion α
                _nb_r     = float(1.0 / _nb_alpha) if _nb_alpha > 0 else np.inf
                _nb_p_sc  = float(_nb_r / (_nb_r + _nb_mu))  # scipy p parameter
            except Exception:
                pass  # NB fit unavailable; panels will show Poisson only

            # ── Bin counts for rootogram & chi-squared GoF ────────────────────
            _max_bin  = int(np.percentile(_df_counts, 99)) + 1
            _bins     = np.arange(0, _max_bin + 2)
            _obs_freq, _ = np.histogram(_df_counts, bins=_bins)
            _bin_vals    = _bins[:-1]   # 0, 1, 2, …, _max_bin

            # Expected frequencies — Poisson
            _pois_pmf  = _poisson_dist.pmf(_bin_vals, _pois_lam)
            _pois_pmf[-1] += (1.0 - _poisson_dist.cdf(_max_bin, _pois_lam))  # right tail
            _pois_exp  = _pois_pmf * _df_n

            # Expected frequencies — NB
            if not np.isnan(_nb_r):
                _nb_pmf    = _nbinom_dist.pmf(_bin_vals, _nb_r, _nb_p_sc)
                _nb_pmf[-1] += (1.0 - _nbinom_dist.cdf(_max_bin, _nb_r, _nb_p_sc))
                _nb_exp    = _nb_pmf * _df_n
            else:
                _nb_exp = None

            # ── Chi-squared GoF (pool bins with expected < 5) ─────────────────
            def _chisq_gof(obs, exp):
                """Pool tail bins until all expected >= 5, return (chi2, df, p)."""
                obs, exp = np.array(obs, dtype=float), np.array(exp, dtype=float)
                # Pool from right
                while len(obs) > 1 and exp[-1] < 5:
                    obs[-2] += obs[-1];  exp[-2] += exp[-1]
                    obs = obs[:-1];      exp = exp[:-1]
                # Pool from left
                while len(obs) > 1 and exp[0] < 5:
                    obs[1] += obs[0];   exp[1] += exp[0]
                    obs = obs[1:];      exp = exp[1:]
                if len(obs) < 2:
                    return np.nan, np.nan, np.nan
                stat = float(np.sum((obs - exp) ** 2 / np.where(exp > 0, exp, np.nan)))
                df   = len(obs) - 2   # -1 for constraint, -1 for estimated param
                df   = max(df, 1)
                p    = float(1.0 - _chi2_dist.cdf(stat, df))
                return stat, df, p

            _pois_chi2, _pois_df, _pois_p = _chisq_gof(_obs_freq, _pois_exp)
            if _nb_exp is not None:
                _nb_chi2, _nb_df, _nb_p = _chisq_gof(_obs_freq, _nb_exp)
            else:
                _nb_chi2 = _nb_df = _nb_p = np.nan

            # ── AIC (log-likelihood based) ────────────────────────────────────
            # Poisson log-likelihood
            _pois_ll = float(np.sum(_poisson_dist.logpmf(_df_counts.astype(int), _pois_lam)))
            _pois_aic = 2 * 1 - 2 * _pois_ll   # 1 free param (λ)
            # NB log-likelihood
            if not np.isnan(_nb_r):
                _nb_ll  = float(np.sum(_nbinom_dist.logpmf(
                    _df_counts.astype(int), _nb_r, _nb_p_sc)))
                _nb_aic = 2 * 2 - 2 * _nb_ll   # 2 free params (μ, α)
                _delta_aic = _nb_aic - _pois_aic  # negative = NB preferred
            else:
                _nb_ll = _nb_aic = _delta_aic = np.nan

            # ── Build figure: 3 rows × 2 cols (top row = histogram, spans both cols) ──
            from matplotlib.gridspec import GridSpec as _GridSpec
            expl_lick_distfit_fig = plt.figure(figsize=(14, 16))
            _gs = _GridSpec(3, 2, figure=expl_lick_distfit_fig,
                            hspace=0.42, wspace=0.35)
            ax_dfhist = expl_lick_distfit_fig.add_subplot(_gs[0, :])   # row 0, both cols
            ax_dfp    = expl_lick_distfit_fig.add_subplot(_gs[1, 0])   # row 1, col 0
            ax_dfn    = expl_lick_distfit_fig.add_subplot(_gs[1, 1])   # row 1, col 1
            ax_dfmv   = expl_lick_distfit_fig.add_subplot(_gs[2, 0])   # row 2, col 0
            ax_dfqq   = expl_lick_distfit_fig.add_subplot(_gs[2, 1])   # row 2, col 1

            # ── [0, :] Raw lick count histogram with PMF overlays ────────────
            from scipy.stats import gaussian_kde as _gkde_df
            ax_dfhist.hist(_df_counts, bins='auto', color='mediumpurple', alpha=0.55,
                           edgecolor='white', linewidth=0.5, density=True,
                           label='Observed (density)')
            # KDE of raw data
            if len(_df_counts) >= 3:
                try:
                    _hist_kde = _gkde_df(_df_counts)
                    _hist_x = np.linspace(max(0, _df_counts.min()), _df_counts.max(), 500)
                    ax_dfhist.plot(_hist_x, _hist_kde(_hist_x), color='indigo',
                                   linewidth=2, label='KDE (observed)')
                except Exception:
                    pass
            # Poisson PMF as stem plot (scaled to density)
            _hist_pmf_x = np.arange(0, _max_bin + 2)
            _hist_pois_pmf = _poisson_dist.pmf(_hist_pmf_x, _pois_lam)
            ax_dfhist.vlines(_hist_pmf_x, 0, _hist_pois_pmf,
                             color='navy', linewidth=0.8, alpha=0.6)
            ax_dfhist.plot(_hist_pmf_x, _hist_pois_pmf, 'o', color='navy',
                           markersize=3, alpha=0.7, label=f'Poisson PMF (λ={_pois_lam:.1f})')
            # NB PMF
            if not np.isnan(_nb_r):
                _hist_nb_pmf = _nbinom_dist.pmf(_hist_pmf_x, _nb_r, _nb_p_sc)
                ax_dfhist.vlines(_hist_pmf_x, 0, _hist_nb_pmf,
                                 color='darkgreen', linewidth=0.8, alpha=0.6)
                ax_dfhist.plot(_hist_pmf_x, _hist_nb_pmf, '^', color='darkgreen',
                               markersize=3, alpha=0.7,
                               label=f'NB PMF (μ={_nb_mu:.1f}, α={_nb_alpha:.3f})')
            ax_dfhist.set_title(
                f'Raw Lick Count Histogram — all sessions (n={_df_n})  |  '
                f'mean={_df_mean:.1f}  var={_df_var:.1f}  V/M={_df_disp:.2f}',
                fontsize=9,
            )
            ax_dfhist.set_xlabel('Lick count per session')
            ax_dfhist.set_ylabel('Density')
            ax_dfhist.legend(fontsize=8)
            ax_dfhist.tick_params(axis='both', direction='in')
            ax_dfhist.spines['top'].set_visible(False)
            ax_dfhist.spines['right'].set_visible(False)

            # ── [1,0] Hanging rootogram — Poisson ─────────────────────────────
            # Rootogram: bars show sqrt(observed); curve shows sqrt(expected);
            # bars are "hung" from the curve so deviations are visible at the base.
            _bw = 0.8
            _sqrt_obs_p = np.sqrt(_obs_freq.astype(float))
            _sqrt_exp_p = np.sqrt(_pois_exp)
            for _i, (_xv, _so, _se) in enumerate(zip(_bin_vals, _sqrt_obs_p, _sqrt_exp_p)):
                _bottom = _se - _so
                _color  = 'crimson' if _bottom < -0.5 else ('gold' if abs(_bottom) < 0.5 else 'steelblue')
                ax_dfp.bar(_xv, _so, bottom=_bottom, width=_bw,
                           color=_color, alpha=0.75, edgecolor='white', linewidth=0.4)
            ax_dfp.axhline(0, color='black', linewidth=1.0, linestyle='-')
            ax_dfp.step(np.append(_bin_vals, _bin_vals[-1] + 1) - 0.5,
                        np.append(_sqrt_exp_p, _sqrt_exp_p[-1]),
                        where='post', color='navy', linewidth=1.8,
                        label=f'Poisson(\u03bb={_pois_lam:.1f})')
            _pois_label = (f'\u03c7\u00b2({_pois_df})={_pois_chi2:.2f}, p={_pois_p:.4f}\n'
                           f'AIC={_pois_aic:.1f}')
            ax_dfp.text(0.97, 0.97, _pois_label, transform=ax_dfp.transAxes,
                        fontsize=8, va='top', ha='right',
                        bbox=dict(boxstyle='round,pad=0.3', facecolor='lightyellow',
                                  edgecolor='gray', alpha=0.85))
            ax_dfp.set_title('Hanging Rootogram — Poisson fit\n'
                             '(crimson=underpredicted, blue=overpredicted, gold=close)')
            ax_dfp.set_xlabel('Lick count per session')
            ax_dfp.set_ylabel('\u221aObserved  (hung from \u221aExpected)')
            ax_dfp.legend(fontsize=8)
            ax_dfp.tick_params(axis='both', direction='in')
            ax_dfp.spines['top'].set_visible(False)
            ax_dfp.spines['right'].set_visible(False)

            # ── [1,1] Hanging rootogram — Negative Binomial ───────────────────
            if _nb_exp is not None:
                _sqrt_obs_n = np.sqrt(_obs_freq.astype(float))
                _sqrt_exp_n = np.sqrt(_nb_exp)
                for _i, (_xv, _so, _se) in enumerate(zip(_bin_vals, _sqrt_obs_n, _sqrt_exp_n)):
                    _bottom = _se - _so
                    _color  = 'crimson' if _bottom < -0.5 else ('gold' if abs(_bottom) < 0.5 else 'steelblue')
                    ax_dfn.bar(_xv, _so, bottom=_bottom, width=_bw,
                               color=_color, alpha=0.75, edgecolor='white', linewidth=0.4)
                ax_dfn.axhline(0, color='black', linewidth=1.0, linestyle='-')
                ax_dfn.step(np.append(_bin_vals, _bin_vals[-1] + 1) - 0.5,
                            np.append(_sqrt_exp_n, _sqrt_exp_n[-1]),
                            where='post', color='darkgreen', linewidth=1.8,
                            label=f'NB(\u03bc={_nb_mu:.1f}, \u03b1={_nb_alpha:.3f})')
                _nb_label = (f'\u03c7\u00b2({_nb_df})={_nb_chi2:.2f}, p={_nb_p:.4f}\n'
                             f'AIC={_nb_aic:.1f}  (\u0394AIC vs Poisson={_delta_aic:+.1f})')
                ax_dfn.text(0.97, 0.97, _nb_label, transform=ax_dfn.transAxes,
                            fontsize=8, va='top', ha='right',
                            bbox=dict(boxstyle='round,pad=0.3', facecolor='lightyellow',
                                      edgecolor='gray', alpha=0.85))
                ax_dfn.set_title('Hanging Rootogram — Negative Binomial fit\n'
                                 '(\u0394AIC < 0 = NB preferred over Poisson)')
                ax_dfn.set_xlabel('Lick count per session')
                ax_dfn.set_ylabel('\u221aObserved  (hung from \u221aExpected)')
                ax_dfn.legend(fontsize=8)
                ax_dfn.tick_params(axis='both', direction='in')
                ax_dfn.spines['top'].set_visible(False)
                ax_dfn.spines['right'].set_visible(False)
            else:
                ax_dfn.text(0.5, 0.5, 'NB fit unavailable\n(install statsmodels)',
                            transform=ax_dfn.transAxes, ha='center', va='center', fontsize=10)
                ax_dfn.axis('off')

            # ── [2,0] Mean–variance plot ──────────────────────────────────────
            # Each point = one mouse; Poisson expectation: var = mean (slope 1 line).
            # NB expectation: var = mean + α·mean² (quadratic curve).
            _mv_means = np.array(_df_per_mouse_means, dtype=float)
            _mv_vars  = np.array(_df_per_mouse_vars,  dtype=float)
            if len(_mv_means) > 0:
                ax_dfmv.scatter(_mv_means, _mv_vars, color='mediumpurple', s=50,
                                alpha=0.8, zorder=3, label='Per-mouse')
                _mv_x = np.linspace(0, max(_mv_means) * 1.1 + 1, 200)
                ax_dfmv.plot(_mv_x, _mv_x, 'b--', linewidth=1.5,
                             label='Poisson (var = mean)')
                if not np.isnan(_nb_alpha):
                    ax_dfmv.plot(_mv_x, _mv_x + _nb_alpha * _mv_x ** 2, 'g-',
                                 linewidth=1.5,
                                 label=f'NB (var = mean + {_nb_alpha:.3f}\u00b7mean\u00b2)')
                ax_dfmv.set_title('Mean\u2013Variance Relationship\n(per-mouse; Poisson: var=mean)')
                ax_dfmv.set_xlabel('Per-mouse mean lick count')
                ax_dfmv.set_ylabel('Per-mouse variance (lick count)')
                ax_dfmv.legend(fontsize=8)
                ax_dfmv.tick_params(axis='both', direction='in')
                ax_dfmv.spines['top'].set_visible(False)
                ax_dfmv.spines['right'].set_visible(False)
                # Annotate with dispersion index
                _disp_text = (f'Dispersion index (pooled)\n'
                              f'V/M = {_df_disp:.3f}  '
                              f'(=1 Poisson; >1 overdispersed)')
                ax_dfmv.text(0.03, 0.97, _disp_text, transform=ax_dfmv.transAxes,
                             fontsize=8, va='top', ha='left',
                             bbox=dict(boxstyle='round,pad=0.3', facecolor='#f0f4ff',
                                       edgecolor='steelblue', alpha=0.88))
            else:
                ax_dfmv.text(0.5, 0.5, 'Insufficient data',
                             transform=ax_dfmv.transAxes, ha='center', va='center')
                ax_dfmv.axis('off')

            # ── [2,1] Poisson Q-Q plot (pooled counts) ────────────────────────
            # Sort observed counts; compare to quantiles of fitted Poisson.
            _qq_obs_sorted = np.sort(_df_counts)
            _n_pts = len(_qq_obs_sorted)
            _probs  = (np.arange(1, _n_pts + 1) - 0.5) / _n_pts
            _pois_qq_theoretical = _poisson_dist.ppf(_probs, _pois_lam)
            ax_dfqq.scatter(_pois_qq_theoretical, _qq_obs_sorted,
                            color='steelblue', s=10, alpha=0.5,
                            label='Poisson Q-Q')
            if not np.isnan(_nb_r):
                _nb_qq_theoretical = _nbinom_dist.ppf(_probs, _nb_r, _nb_p_sc)
                ax_dfqq.scatter(_nb_qq_theoretical, _qq_obs_sorted,
                                color='darkgreen', s=10, alpha=0.5,
                                marker='^', label='NB Q-Q')
            _diag_max = max(float(_pois_qq_theoretical.max()), float(_qq_obs_sorted.max())) * 1.05
            ax_dfqq.plot([0, _diag_max], [0, _diag_max], 'k--',
                         linewidth=1.5, label='Perfect fit')
            ax_dfqq.set_title('Count Q-Q Plot\n(Poisson = blue circles, NB = green triangles)')
            ax_dfqq.set_xlabel('Theoretical quantiles')
            ax_dfqq.set_ylabel('Observed quantiles')
            ax_dfqq.legend(fontsize=8)
            ax_dfqq.tick_params(axis='both', direction='in')
            ax_dfqq.spines['top'].set_visible(False)
            ax_dfqq.spines['right'].set_visible(False)

            # ── Suptitle with key statistics ──────────────────────────────────
            _preferred = ('Negative Binomial' if (not np.isnan(_delta_aic) and _delta_aic < -2)
                          else ('Poisson' if (not np.isnan(_delta_aic) and _delta_aic > 2)
                                else 'Inconclusive (|ΔAIC| ≤ 2)'))
            _df_summary = (
                f'Lick Count Distribution Fit  |  n={_df_n} sessions, '
                f'{len(_mv_means)} mice  |  '
                f'Mean={_df_mean:.1f}  Var={_df_var:.1f}  '
                f'Dispersion V/M={_df_disp:.3f}  |  '
                f'Poisson AIC={_pois_aic:.1f}  '
                f'NB AIC={_nb_aic:.1f}  \u0394AIC={_delta_aic:+.1f}  '
                f'\u2192 Preferred: {_preferred}'
            )
            expl_lick_distfit_fig.suptitle(_df_summary, fontsize=9, y=1.01, wrap=True)

            # ── Console summary ───────────────────────────────────────────────
            print('\n\u2500\u2500 Lick Count Distribution Fit \u2500\u2500')
            print(f'  N sessions:        {_df_n}')
            print(f'  N mice:            {len(_mv_means)}')
            print(f'  Mean:              {_df_mean:.4f}')
            print(f'  Variance:          {_df_var:.4f}')
            print(f'  Dispersion (V/M):  {_df_disp:.4f}  (1 = Poisson; >1 = overdispersed)')
            print(f'  Poisson  \u03bb={_pois_lam:.4f}  \u03c7\u00b2({_pois_df})={_pois_chi2:.4f}  p={_pois_p:.6f}  AIC={_pois_aic:.2f}')
            if not np.isnan(_nb_r):
                print(f'  NB       \u03bc={_nb_mu:.4f}  \u03b1={_nb_alpha:.4f}  r={_nb_r:.4f}  '
                      f'\u03c7\u00b2({_nb_df})={_nb_chi2:.4f}  p={_nb_p:.6f}  AIC={_nb_aic:.2f}')
                print(f'  \u0394AIC (NB \u2212 Poisson) = {_delta_aic:+.4f}  '
                      f'\u2192 Preferred: {_preferred}')
            else:
                print('  NB fit unavailable')

        except ImportError as _e:
            print(f'[expl_lick_distfit] Missing dependency: {_e}')
            print('  Install statsmodels: conda install statsmodels')
        except Exception as _e:
            import traceback as _tb
            print(f'[expl_lick_distfit] Error: {_e}')
            _tb.print_exc()

    # ── Exploratory: raw lick count box-and-whisker ───────────────────────────
    expl_lick_boxplot_fig = None
    if 'expl_lick_boxplot' in selected_plots:
        _lkbx_names = []
        _lkbx_vals  = []
        _lkbx_all   = []
        for _r in all_results:
            _lc = pd.to_numeric(_r['df']['lick_count'], errors='coerce').dropna()
            _lc_valid = _lc[_lc >= 0].values.tolist()
            if _lc_valid:
                _lkbx_names.append(_r['mouse'])
                _lkbx_vals.append(_lc_valid)
                _lkbx_all.extend(_lc_valid)

        n_mice_lkbx = len(_lkbx_names)
        expl_lick_boxplot_fig, (ax_lkbx1, ax_lkbx2) = plt.subplots(
            1, 2, figsize=(max(10, n_mice_lkbx * 0.8 + 3), 6),
            gridspec_kw={'width_ratios': [max(3, n_mice_lkbx), 1]},
        )
        ax_lkbx1.boxplot(_lkbx_vals, labels=_lkbx_names,
                         patch_artist=True,
                         boxprops=dict(facecolor='mediumpurple', alpha=0.6),
                         medianprops=dict(color='indigo', linewidth=2),
                         whiskerprops=dict(color='mediumpurple'),
                         capprops=dict(color='mediumpurple'),
                         flierprops=dict(marker='o', markerfacecolor='mediumpurple',
                                         markersize=4, alpha=0.5, linestyle='none'))
        ax_lkbx1.set_title('Raw lick count distribution per mouse\n(each session = one data point)')
        ax_lkbx1.set_xlabel('Mouse')
        ax_lkbx1.set_ylabel('Raw lick count per session')
        ax_lkbx1.tick_params(axis='x', rotation=45)
        ax_lkbx1.tick_params(axis='both', direction='in')
        ax_lkbx1.spines['top'].set_visible(False)
        ax_lkbx1.spines['right'].set_visible(False)

        ax_lkbx2.boxplot([_lkbx_all], labels=['All mice'],
                         patch_artist=True,
                         boxprops=dict(facecolor='plum', alpha=0.6),
                         medianprops=dict(color='purple', linewidth=2),
                         whiskerprops=dict(color='plum'),
                         capprops=dict(color='plum'),
                         flierprops=dict(marker='o', markerfacecolor='plum',
                                         markersize=4, alpha=0.5, linestyle='none'))
        ax_lkbx2.set_title(f'Overall\n(n={len(_lkbx_all)} sessions)')
        ax_lkbx2.set_ylabel('Raw lick count per session')
        ax_lkbx2.tick_params(axis='both', direction='in')
        ax_lkbx2.spines['top'].set_visible(False)
        ax_lkbx2.spines['right'].set_visible(False)

        expl_lick_boxplot_fig.suptitle('Raw Lick Count Box-and-Whisker — Exploratory', fontsize=11)
        expl_lick_boxplot_fig.tight_layout()

    # ── Exploratory: raw lick count RM ANOVA residual diagnostics ─────────────
    # DV = log(1 + lick_count) — log(1+x) linearises count data for OLS.
    # The +1 offset avoids log(0) on sessions with zero licks.
    expl_lick_rm_anova_resid_fig = None
    if 'expl_lick_rm_anova_resid' in selected_plots:
        import warnings as _warnings_lk
        try:
            import statsmodels.formula.api as _smf_lk
            from scipy.stats import shapiro as _shapiro_lkr, levene as _levene_lkr
            from scipy.stats import probplot as _probplot_lkr

            # ── Build long-format DataFrame (log(1+lick_count)) ──────────────
            _lk_rows = []
            for _r in all_results:
                _df_r = _r['df'].copy()
                _df_r = _df_r.reset_index(drop=True)
                _df_r['session_num'] = np.arange(1, len(_df_r) + 1, dtype=float)
                _df_r['mouse']       = _r['mouse']
                _df_r['condition']   = _r['starting_condition']
                _raw_lk = pd.to_numeric(_df_r['lick_count'], errors='coerce')
                # log(1+x): defined for x >= 0; negative / NaN → NaN
                _df_r['log1p_lick'] = np.where(
                    (_raw_lk >= 0) & _raw_lk.notna(),
                    np.log1p(_raw_lk.values.astype(float)),
                    np.nan,
                )
                _lk_rows.append(_df_r[['mouse', 'condition', 'session_num', 'log1p_lick']])
            _df_lk_long = pd.concat(_lk_rows, ignore_index=True).dropna(subset=['log1p_lick'])

            # ── Fit OLS with mouse as fixed-effect blocking factor ────────────
            _lk_formula = 'log1p_lick ~ C(condition) + session_num + C(condition):session_num + C(mouse)'
            with _warnings_lk.catch_warnings():
                _warnings_lk.simplefilter('ignore')
                _lk_ols = _smf_lk.ols(_lk_formula, data=_df_lk_long).fit()

            _lk_resid  = _lk_ols.resid.values
            _lk_fitted = _lk_ols.fittedvalues.values
            _lk_conds  = _df_lk_long['condition'].values
            _lk_resid_mean = float(np.mean(_lk_resid))
            _lk_resid_var  = float(_lk_ols.mse_resid)
            _lk_resid_sd   = float(np.sqrt(_lk_resid_var))
            _lk_r2         = float(_lk_ols.rsquared)
            _lk_r2_adj     = float(_lk_ols.rsquared_adj)
            _lk_y_var      = float(_df_lk_long['log1p_lick'].var(ddof=1))
            _lk_y_mean     = float(_df_lk_long['log1p_lick'].mean())
            _lk_unexplained = 1.0 - _lk_r2
            _lk_snr         = float(np.mean(_lk_fitted)) / _lk_resid_sd if _lk_resid_sd > 0 else np.nan

            # PRESS / predicted R²
            try:
                _lk_influence = _lk_ols.get_influence()
                _lk_hat       = _lk_influence.hat_matrix_diag
                _lk_press_r   = _lk_resid / (1.0 - np.clip(_lk_hat, None, 0.9999))
                _lk_press     = float(np.sum(_lk_press_r ** 2))
                _lk_ss_total  = float(np.sum((_df_lk_long['log1p_lick'].values - _lk_y_mean) ** 2))
                _lk_r2_pred   = float(1.0 - _lk_press / _lk_ss_total) if _lk_ss_total > 0 else np.nan
            except Exception:
                _lk_press, _lk_r2_pred, _lk_hat = np.nan, np.nan, None
                _lk_influence = None

            # ── Shapiro-Wilk on residuals ─────────────────────────────────────
            _lk_sw_stat, _lk_sw_p = (_shapiro_lkr(_lk_resid)
                                      if len(_lk_resid) >= 3 else (np.nan, np.nan))

            # ── Levene's test across condition groups ─────────────────────────
            _lk_cond_groups = [_lk_resid[_lk_conds == c] for c in np.unique(_lk_conds)
                               if np.sum(_lk_conds == c) >= 2]
            if len(_lk_cond_groups) >= 2:
                _lk_lev_stat, _lk_lev_p = _levene_lkr(*_lk_cond_groups)
            else:
                _lk_lev_stat, _lk_lev_p = np.nan, np.nan

            # ── pingouin mixed ANOVA table (optional) ─────────────────────────
            _lk_pg_text = ''
            try:
                import pingouin as _pg_lk
                _lk_pg_result = _pg_lk.mixed_anova(
                    data=_df_lk_long, dv='log1p_lick', within='session_num',
                    between='condition', subject='mouse',
                )
                _lk_pg_lines = ['Mixed ANOVA (pingouin):']
                for _, _row in _lk_pg_result.iterrows():
                    _src  = _row.get('Source', '')
                    _f    = _row.get('F', np.nan)
                    _pval = _row.get('p-unc', np.nan)
                    _eta  = _row.get('np2', np.nan)
                    _lk_pg_lines.append(
                        f"  {_src:<28} F={_f:.3f}  p={_pval:.4f}  \u03b7\u00b2={_eta:.3f}"
                    )
                _lk_pg_text = '\n'.join(_lk_pg_lines)
            except Exception:
                _lk_pg_text = 'pingouin not available \u2014 ANOVA table omitted'

            # ── Cook's D and leverage ─────────────────────────────────────────
            _lk_mice_labels   = _df_lk_long['mouse'].values
            _lk_sess_nums_lbl = _df_lk_long['session_num'].values.astype(int)
            _lk_obs_labels    = np.array([f'{m}\nS{s}' for m, s in
                                          zip(_lk_mice_labels, _lk_sess_nums_lbl)])
            _lk_obs_index     = np.arange(len(_lk_resid))
            try:
                _lk_cooks_d  = _lk_influence.cooks_distance[0]
                _lk_lev_diag = _lk_hat.copy()
                _lk_n_params = len(_lk_ols.params)
                _lk_cooks_thresh = 4.0 / max(len(_lk_resid) - _lk_n_params, 1)
                _lk_lev_thresh   = 2.0 * _lk_n_params / max(len(_lk_resid), 1)
                _lk_high_cooks   = np.where(_lk_cooks_d > _lk_cooks_thresh)[0]
                _lk_high_lev     = np.where(_lk_lev_diag > _lk_lev_thresh)[0]
            except Exception:
                _lk_cooks_d = _lk_lev_diag = None
                _lk_cooks_thresh = _lk_lev_thresh = np.nan
                _lk_high_cooks = _lk_high_lev = np.array([], dtype=int)

            # ── Build figure: 3 rows × 2 cols ────────────────────────────────
            expl_lick_rm_anova_resid_fig, _lk_axes = plt.subplots(3, 2, figsize=(13, 15))
            (ax_lkqq, ax_lkrh), (ax_lkrf, ax_lkrs), (ax_lkcd, ax_lklv) = _lk_axes

            # ── Q-Q plot ──────────────────────────────────────────────────────
            (_lk_qq_osm, _lk_qq_osr), (_lk_qq_slope, _lk_qq_int, _) = _probplot_lkr(_lk_resid, dist='norm')
            ax_lkqq.plot(_lk_qq_osm, _lk_qq_osr, 'o', color='mediumpurple',
                         markersize=4, alpha=0.7, label='Residuals')
            ax_lkqq.plot(
                [_lk_qq_osm[0], _lk_qq_osm[-1]],
                [_lk_qq_slope * _lk_qq_osm[0] + _lk_qq_int,
                 _lk_qq_slope * _lk_qq_osm[-1] + _lk_qq_int],
                'r-', linewidth=1.5, label='Normal line',
            )
            ax_lkqq.set_title('Normal Q-Q Plot of Residuals')
            ax_lkqq.set_xlabel('Theoretical quantiles')
            ax_lkqq.set_ylabel('Sample quantiles')
            ax_lkqq.tick_params(axis='both', direction='in')
            ax_lkqq.spines['top'].set_visible(False)
            ax_lkqq.spines['right'].set_visible(False)
            ax_lkqq.legend(fontsize=8)
            _lk_sw_label = (f'Shapiro-Wilk: W={_lk_sw_stat:.3f}, p={_lk_sw_p:.4f}\n'
                            f'{"Normal (p>0.05)" if _lk_sw_p > 0.05 else "Non-normal (p\u22640.05)"}')
            ax_lkqq.text(0.03, 0.97, _lk_sw_label, transform=ax_lkqq.transAxes,
                         fontsize=8, va='top', ha='left',
                         bbox=dict(boxstyle='round,pad=0.3', facecolor='lightyellow',
                                   edgecolor='gray', alpha=0.8))
            _lk_formula_label = (
                'Error terms: \u03b5\u1d62 = y\u1d62 \u2212 \u0177\u1d62\n'
                'y\u1d62 = log(1 + lick_count, session i)\n'
                '\u0177\u1d62 = model fitted value\n'
                'Model: log(1+lick) ~ condition\n'
                '       + session + condition\u00d7session\n'
                '       + mouse  (blocking factor)'
            )
            ax_lkqq.text(0.97, 0.03, _lk_formula_label, transform=ax_lkqq.transAxes,
                         fontsize=7.5, va='bottom', ha='right', family='monospace',
                         bbox=dict(boxstyle='round,pad=0.35', facecolor='#f5f5f5',
                                   edgecolor='gray', alpha=0.88))

            # ── Residual histogram ────────────────────────────────────────────
            from scipy.stats import gaussian_kde as _gkde_lkr
            ax_lkrh.hist(_lk_resid, bins='auto', color='mediumpurple', alpha=0.65,
                         edgecolor='white', linewidth=0.5, density=True)
            if len(_lk_resid) >= 3:
                _lk_kde_x = np.linspace(_lk_resid.min(), _lk_resid.max(), 300)
                try:
                    _lk_kde2 = _gkde_lkr(_lk_resid)
                    ax_lkrh.plot(_lk_kde_x, _lk_kde2(_lk_kde_x), color='indigo',
                                 linewidth=2, label='KDE')
                except Exception:
                    pass
            ax_lkrh.axvline(0, color='red', linewidth=1.5, linestyle='--', label='Zero')
            ax_lkrh.axvline(_lk_resid_mean, color='darkorange', linewidth=1.5,
                            linestyle=':', label=f'Mean={_lk_resid_mean:.3f}')
            ax_lkrh.set_title('Histogram of Residuals')
            ax_lkrh.set_xlabel('Residual \u03b5\u1d62  [log(1+count) units]')
            ax_lkrh.set_ylabel('Density')
            ax_lkrh.tick_params(axis='both', direction='in')
            ax_lkrh.spines['top'].set_visible(False)
            ax_lkrh.spines['right'].set_visible(False)
            ax_lkrh.legend(fontsize=8)
            _lk_r2_pred_str = f'{_lk_r2_pred:.4f}' if not np.isnan(_lk_r2_pred) else 'n/a'
            _lk_stats_box = (
                f'E[\u03b5]         = {_lk_resid_mean:+.5f}  (\u22480 by OLS)\n'
                f'Var(\u03b5)  = MSE = {_lk_resid_var:.5f}\n'
                f'SD(\u03b5)   = RMSE = {_lk_resid_sd:.5f}\n'
                f'\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\u2015\n'
                f'Var(Y)         = {_lk_y_var:.5f}\n'
                f'Var(\u03b5)/Var(Y) = {_lk_unexplained:.4f}  (= 1\u2212R\u00b2)\n'
                f'R\u00b2             = {_lk_r2:.4f}\n'
                f'R\u00b2 adj         = {_lk_r2_adj:.4f}\n'
                f'R\u00b2 pred (PRESS) = {_lk_r2_pred_str}\n'
                f'SNR (fitted/SD\u03b5) = {_lk_snr:.3f}'
            )
            ax_lkrh.text(0.98, 0.98, _lk_stats_box, transform=ax_lkrh.transAxes,
                         fontsize=7.5, va='top', ha='right', family='monospace',
                         bbox=dict(boxstyle='round,pad=0.35', facecolor='#f0f4ff',
                                   edgecolor='steelblue', alpha=0.92))

            # ── Residuals vs Fitted ───────────────────────────────────────────
            _lk_uniq_conds = np.unique(_lk_conds)
            _lk_cond_color = {c: _condition_to_color(c) for c in _lk_uniq_conds}
            for _lkc in _lk_uniq_conds:
                _lkmask = _lk_conds == _lkc
                ax_lkrf.scatter(_lk_fitted[_lkmask], _lk_resid[_lkmask],
                                color=_lk_cond_color[_lkc], alpha=0.5, s=18, label=str(_lkc))
            ax_lkrf.axhline(0, color='red', linewidth=1.5, linestyle='--')
            try:
                from statsmodels.nonparametric.smoothers_lowess import lowess as _lk_lowess
                _lk_lw = _lk_lowess(_lk_resid, _lk_fitted, frac=0.4)
                ax_lkrf.plot(_lk_lw[:, 0], _lk_lw[:, 1], color='black',
                             linewidth=1.5, linestyle='-', label='Lowess')
            except Exception:
                pass
            ax_lkrf.set_title('Residuals vs Fitted Values')
            ax_lkrf.set_xlabel('Fitted values [log(1+count) units]')
            ax_lkrf.set_ylabel('Residuals [log(1+count) units]')
            ax_lkrf.tick_params(axis='both', direction='in')
            ax_lkrf.spines['top'].set_visible(False)
            ax_lkrf.spines['right'].set_visible(False)
            ax_lkrf.legend(fontsize=8, title='Condition')

            # ── Residuals vs session number ───────────────────────────────────
            _lk_sess_nums = _df_lk_long['session_num'].values
            ax_lkrs.scatter(_lk_sess_nums, _lk_resid,
                            color='mediumpurple', alpha=0.4, s=18)
            ax_lkrs.axhline(0, color='red', linewidth=1.5, linestyle='--')
            try:
                _lk_lw2 = _lk_lowess(_lk_resid, _lk_sess_nums, frac=0.4)
                ax_lkrs.plot(_lk_lw2[:, 0], _lk_lw2[:, 1], color='black',
                             linewidth=1.5, linestyle='-', label='Lowess')
                ax_lkrs.legend(fontsize=8)
            except Exception:
                pass
            ax_lkrs.set_title('Residuals vs Session Number\n(check for time-trend)')
            ax_lkrs.set_xlabel('Session number')
            ax_lkrs.set_ylabel('Residuals [log(1+count) units]')
            ax_lkrs.tick_params(axis='both', direction='in')
            ax_lkrs.spines['top'].set_visible(False)
            ax_lkrs.spines['right'].set_visible(False)

            # ── Cook's D bar chart ────────────────────────────────────────────
            if _lk_cooks_d is not None:
                _lk_cd_colors = np.where(_lk_cooks_d > _lk_cooks_thresh, 'crimson', 'mediumpurple')
                ax_lkcd.bar(_lk_obs_index, _lk_cooks_d, color=_lk_cd_colors, alpha=0.75, width=0.8)
                ax_lkcd.axhline(_lk_cooks_thresh, color='crimson', linewidth=1.5,
                                linestyle='--',
                                label=f'Threshold 4/(n\u2212p)={_lk_cooks_thresh:.3f}')
                for _idx in _lk_high_cooks:
                    ax_lkcd.text(_idx, _lk_cooks_d[_idx] * 1.04, _lk_obs_labels[_idx],
                                 fontsize=6, ha='center', va='bottom', color='crimson',
                                 linespacing=1.2)
                ax_lkcd.set_title(
                    f"Cook's D per Observation\n"
                    f"({len(_lk_high_cooks)} flagged > threshold, shown in red)"
                )
                ax_lkcd.set_xlabel('Observation index')
                ax_lkcd.set_ylabel("Cook's D")
                ax_lkcd.tick_params(axis='both', direction='in')
                ax_lkcd.spines['top'].set_visible(False)
                ax_lkcd.spines['right'].set_visible(False)
                ax_lkcd.legend(fontsize=8)
            else:
                ax_lkcd.text(0.5, 0.5, "Cook's D unavailable",
                             transform=ax_lkcd.transAxes, ha='center', va='center')
                ax_lkcd.axis('off')

            # ── Leverage vs Cook's D (influence plot) ─────────────────────────
            if _lk_cooks_d is not None:
                for _lkc in _lk_uniq_conds:
                    _lkmask = _lk_conds == _lkc
                    ax_lklv.scatter(_lk_lev_diag[_lkmask], _lk_cooks_d[_lkmask],
                                    color=_lk_cond_color[_lkc], alpha=0.6,
                                    s=20, label=str(_lkc))
                ax_lklv.axvline(_lk_lev_thresh, color='darkorange', linewidth=1.5,
                                linestyle='--',
                                label=f'Leverage 2p/n={_lk_lev_thresh:.3f}')
                ax_lklv.axhline(_lk_cooks_thresh, color='crimson', linewidth=1.5,
                                linestyle='--',
                                label=f"Cook's D 4/(n\u2212p)={_lk_cooks_thresh:.3f}")
                _lk_both_flag = np.where((_lk_cooks_d > _lk_cooks_thresh) &
                                         (_lk_lev_diag > _lk_lev_thresh))[0]
                for _idx in _lk_both_flag:
                    ax_lklv.annotate(_lk_obs_labels[_idx],
                                     (_lk_lev_diag[_idx], _lk_cooks_d[_idx]),
                                     textcoords='offset points', xytext=(4, 4),
                                     fontsize=6, color='crimson')
                ax_lklv.set_title(
                    f'Leverage vs Cook\u2019s D (Influence Plot)\n'
                    f'High-leverage: {len(_lk_high_lev)} obs  |  '
                    f'High-influence: {len(_lk_high_cooks)} obs  |  '
                    f'Both: {len(_lk_both_flag)} obs'
                )
                ax_lklv.set_xlabel('Leverage h\u1d62\u1d62')
                ax_lklv.set_ylabel("Cook's D")
                ax_lklv.tick_params(axis='both', direction='in')
                ax_lklv.spines['top'].set_visible(False)
                ax_lklv.spines['right'].set_visible(False)
                ax_lklv.legend(fontsize=7, title='Condition', ncol=2)
            else:
                ax_lklv.text(0.5, 0.5, 'Leverage unavailable',
                             transform=ax_lklv.transAxes, ha='center', va='center')
                ax_lklv.axis('off')

            # ── Suptitle: summary stats ───────────────────────────────────────
            _lk_n_mice   = _df_lk_long['mouse'].nunique()
            _lk_n_obs    = len(_df_lk_long)
            _lk_lev_str  = (f'Levene (across conditions): F={_lk_lev_stat:.3f}, p={_lk_lev_p:.4f}'
                            if not np.isnan(_lk_lev_p)
                            else 'Levene: insufficient groups')
            _lk_summary  = (
                f'Raw Lick Count RM ANOVA residual diagnostics  |  '
                f'DV = log(1+lick_count)  |  '
                f'n={_lk_n_mice} mice, {_lk_n_obs} obs  |  '
                f'E[\u03b5]={_lk_resid_mean:+.4f} (\u22480)  '
                f'Var(\u03b5)={_lk_resid_var:.4f}  SD(\u03b5)={_lk_resid_sd:.4f}  '
                f'Var(\u03b5)/Var(Y)={_lk_unexplained:.3f} (=1\u2212R\u00b2)  '
                f'R\u00b2={_lk_r2:.3f}  |  {_lk_lev_str}'
            )
            expl_lick_rm_anova_resid_fig.suptitle(_lk_summary, fontsize=8.5, y=1.01, wrap=True)

            if _lk_pg_text:
                expl_lick_rm_anova_resid_fig.text(
                    0.01, -0.02, _lk_pg_text,
                    fontsize=7.5, family='monospace', va='top',
                    transform=expl_lick_rm_anova_resid_fig.transFigure,
                    bbox=dict(boxstyle='round,pad=0.4', facecolor='#f5f5f5',
                              edgecolor='gray', alpha=0.9),
                )

            expl_lick_rm_anova_resid_fig.tight_layout()
            print('\n\u2500\u2500 Raw Lick Count RM ANOVA Residual Diagnostics \u2500\u2500')
            print(f'  Formula:          \u03b5\u1d62 = y\u1d62 \u2212 \u0177\u1d62  (observed \u2212 model fitted value)')
            print(f'  Model DV:         log(1 + lick_count)  [log(1+count) units]')
            print(f'  N mice:           {_lk_n_mice}')
            print(f'  N observations:   {_lk_n_obs}')
            print(f'  E[\u03b5]  (mean):    {_lk_resid_mean:+.6f}  (\u22480 by OLS construction)')
            print(f'  Var(\u03b5) (MSE):    {_lk_resid_var:.6f}')
            print(f'  SD(\u03b5)  (RMSE):   {_lk_resid_sd:.6f}')
            print(f'  Var(Y):           {_lk_y_var:.6f}')
            print(f'  Var(\u03b5)/Var(Y):   {_lk_unexplained:.4f}  (= 1 \u2212 R\u00b2 = unexplained variance fraction)')
            print(f'  R\u00b2:              {_lk_r2:.4f}')
            print(f'  R\u00b2 adj:          {_lk_r2_adj:.4f}')
            print(f'  R\u00b2 pred (PRESS): {_lk_r2_pred_str}  (leave-one-out; gap vs R\u00b2 indicates overfitting)')
            print(f'  SNR (\u0177\u0305/SD\u03b5):    {_lk_snr:.4f}')
            print(f'  Shapiro-Wilk:     W={_lk_sw_stat:.4f}, p={_lk_sw_p:.6f}')
            print(f'  Levene test:      F={_lk_lev_stat:.4f}, p={_lk_lev_p:.6f}')
            if _lk_cooks_d is not None:
                print(f'  Cook\'s D thresh:  {_lk_cooks_thresh:.4f}  (4/(n\u2212p))')
                print(f'  Leverage thresh:  {_lk_lev_thresh:.4f}  (2p/n)')
                _lk_flag_strs = [f'{m}(S{s})' for m, s in
                                 zip(_lk_mice_labels[_lk_high_cooks], _lk_sess_nums_lbl[_lk_high_cooks])]
                print(f'  High-influence obs (Cook\'s D > thresh): {len(_lk_high_cooks)}'
                      + (f'  [{", ".join(_lk_flag_strs)}]' if len(_lk_high_cooks) else ''))
                _lk_lev_strs = [f'{m}(S{s})' for m, s in
                                zip(_lk_mice_labels[_lk_high_lev], _lk_sess_nums_lbl[_lk_high_lev])]
                print(f'  High-leverage obs (h\u1d62\u1d62 > thresh):      {len(_lk_high_lev)}'
                      + (f'  [{", ".join(_lk_lev_strs)}]' if len(_lk_high_lev) else ''))
            if _lk_pg_text:
                print(_lk_pg_text)

        except ImportError as _e:
            print(f'[expl_lick_rm_anova_resid] Missing dependency: {_e}')
            print('  Install statsmodels: conda install statsmodels')
        except Exception as _e:
            print(f'[expl_lick_rm_anova_resid] Error: {_e}')

    # ── Exploratory: average lick rate distribution fit (per-mouse means by cohort) ─
    # One mean per mouse; grouped by starting_condition.
    # Goal: assess normality per cohort to decide whether a t-test is appropriate.
    expl_lick_rate_distfit_fig = None
    if 'expl_lick_rate_distfit' in selected_plots:
        try:
            from scipy.stats import (norm as _norm_lr, probplot as _probplot_lr,
                                     shapiro as _sw_lr, beta as _beta_lr)
            from matplotlib.gridspec import GridSpec as _GridSpec_lr
            import warnings as _warnings_lr

            # ── One mean per mouse, grouped by starting_condition ─────────────
            _lr_groups = {}
            for _r in all_results:
                _df_lr = _r['df']
                _lc = pd.to_numeric(_df_lr['lick_count'], errors='coerce')
                _sl = pd.to_numeric(_df_lr['session_length'], errors='coerce')
                _valid_lr = (_lc >= 0) & (_sl > 0) & _lc.notna() & _sl.notna()
                _lrv = (_lc[_valid_lr] / _sl[_valid_lr]).values
                if len(_lrv) > 0:
                    _cond_lr = _r.get('starting_condition', 'Unknown')
                    _lr_groups.setdefault(_cond_lr, []).append(float(np.mean(_lrv)))

            _lr_cond_names = sorted(_lr_groups.keys())
            _n_cg_lr = len(_lr_cond_names)
            if _n_cg_lr == 0:
                raise ValueError('No valid lick rate data found')

            # ── Shapiro-Wilk per cohort ────────────────────────────────────────
            _lr_sw = {}
            for _cn in _lr_cond_names:
                _arr = np.array(_lr_groups[_cn])
                if len(_arr) >= 3:
                    _w, _p = _sw_lr(_arr)
                    _lr_sw[_cn] = (float(_w), float(_p), len(_arr))
                else:
                    _lr_sw[_cn] = (np.nan, np.nan, len(_arr))

            # ── Figure layout ─────────────────────────────────────────────────
            _n_qq_rows_lr = (_n_cg_lr + 1) // 2
            _n_rows_lr = 1 + 1 + _n_qq_rows_lr + 1  # +1 row for box plot
            _coh_pal_lr = [plt.cm.tab10(i / max(_n_cg_lr - 1, 1))
                           for i in range(_n_cg_lr)]

            expl_lick_rate_distfit_fig = plt.figure(figsize=(13, 4.5 * _n_rows_lr))
            _gs_lr = _GridSpec_lr(_n_rows_lr, 2, figure=expl_lick_rate_distfit_fig,
                                  hspace=0.60, wspace=0.38)
            ax_lrhist = expl_lick_rate_distfit_fig.add_subplot(_gs_lr[0, :])

            # [row 0] Overlaid histograms + Normal fit + rug marks
            _lr_rng = np.random.default_rng(42)
            for _ci, _cn in enumerate(_lr_cond_names):
                _arr = np.array(_lr_groups[_cn])
                _col_lr = _coh_pal_lr[_ci]
                _sw_lbl_lr = (f'SW p={_lr_sw[_cn][1]:.3f}'
                              if not np.isnan(_lr_sw[_cn][1]) else 'SW: n<3')
                _nbins_lr = max(4, len(_arr) // 2 + 1)
                ax_lrhist.hist(_arr, bins=_nbins_lr, alpha=0.40, color=_col_lr,
                               density=True, edgecolor='white', linewidth=0.5,
                               label=f'{_cn}  (n={len(_arr)}, {_sw_lbl_lr})')
                if len(_arr) >= 3:
                    _mn_lr = float(np.mean(_arr))
                    _sdv_lr = float(np.std(_arr, ddof=1))
                    _xfit_lr = np.linspace(_mn_lr - 4 * _sdv_lr,
                                           _mn_lr + 4 * _sdv_lr, 300)
                    ax_lrhist.plot(_xfit_lr, _norm_lr.pdf(_xfit_lr, _mn_lr, _sdv_lr),
                                   color=_col_lr, linewidth=2.0, linestyle='--')
                _jit_lr = _lr_rng.uniform(-0.003, 0.003, len(_arr))
                ax_lrhist.plot(_arr, _jit_lr, '|', color=_col_lr,
                               markersize=14, markeredgewidth=2.5, alpha=0.75)

            ax_lrhist.set_xlabel('Per-mouse mean lick rate (licks/min)', fontsize=10)
            ax_lrhist.set_ylabel('Density', fontsize=10)
            ax_lrhist.set_title(
                'Per-mouse mean lick rate — distribution by cohort\n'
                '(each point = one mouse; dashed lines = Normal fit; '
                'tick marks = individual mice)',
                fontsize=10,
            )
            ax_lrhist.legend(fontsize=8)
            ax_lrhist.spines['top'].set_visible(False)
            ax_lrhist.spines['right'].set_visible(False)
            ax_lrhist.tick_params(axis='both', direction='in')

            # [row 1] Box-and-whisker with 1.5×IQR outlier highlighting
            ax_lrbox = expl_lick_rate_distfit_fig.add_subplot(_gs_lr[1, :])
            _bp_data_lr = [np.array(_lr_groups[_cn]) for _cn in _lr_cond_names]
            _bplot_lr = ax_lrbox.boxplot(
                _bp_data_lr, labels=_lr_cond_names, patch_artist=True,
                showfliers=False, widths=0.50,
                boxprops=dict(linewidth=1.4),
                whiskerprops=dict(linewidth=1.2, linestyle='--'),
                capprops=dict(linewidth=1.4),
                medianprops=dict(color='black', linewidth=2.0),
            )
            for _bi, _bpatch in enumerate(_bplot_lr['boxes']):
                _bc = _coh_pal_lr[_bi]
                _bpatch.set_facecolor([*_bc[:3], 0.35])
                _bpatch.set_edgecolor(_bc)
            _rng_box_lr = np.random.default_rng(1)
            _outlier_legend_added_lr = False
            for _bi, _cn in enumerate(_lr_cond_names):
                _arr_bx = np.array(_lr_groups[_cn])
                _q1_lr  = float(np.percentile(_arr_bx, 25))
                _q3_lr  = float(np.percentile(_arr_bx, 75))
                _iqr_lr = _q3_lr - _q1_lr
                _lo_lr  = _q1_lr - 1.5 * _iqr_lr
                _hi_lr  = _q3_lr + 1.5 * _iqr_lr
                _out_lr = (_arr_bx < _lo_lr) | (_arr_bx > _hi_lr)
                _jit_lr = _rng_box_lr.uniform(-0.15, 0.15, len(_arr_bx))
                ax_lrbox.scatter(
                    np.full(int(np.sum(~_out_lr)), _bi + 1) + _jit_lr[~_out_lr],
                    _arr_bx[~_out_lr], color=_coh_pal_lr[_bi],
                    s=40, alpha=0.70, zorder=3,
                )
                if _out_lr.any():
                    ax_lrbox.scatter(
                        np.full(int(np.sum(_out_lr)), _bi + 1) + _jit_lr[_out_lr],
                        _arr_bx[_out_lr], color='red', s=80, alpha=0.90,
                        zorder=4, edgecolors='darkred', linewidths=1.5, marker='D',
                        label='Outlier (1.5\u00d7IQR)' if not _outlier_legend_added_lr else '',
                    )
                    _outlier_legend_added_lr = True
                    for _ov in _arr_bx[_out_lr]:
                        ax_lrbox.annotate(
                            f'{_ov:.3g}', xy=(_bi + 1, _ov),
                            xytext=(8, 0), textcoords='offset points',
                            fontsize=7, color='darkred', va='center',
                        )
            ax_lrbox.set_xlabel('Cohort', fontsize=10)
            ax_lrbox.set_ylabel('Per-mouse mean (licks/min)', fontsize=10)
            ax_lrbox.set_title(
                'Box-and-whisker \u2014 per-mouse mean lick rate by cohort\n'
                '(whiskers extend to 1.5\u00d7IQR; \u25c6 red diamonds = outliers beyond fence)',
                fontsize=10,
            )
            _handles_lr, _labels_lr = ax_lrbox.get_legend_handles_labels()
            if _handles_lr:
                ax_lrbox.legend(fontsize=8)
            ax_lrbox.spines['top'].set_visible(False)
            ax_lrbox.spines['right'].set_visible(False)
            ax_lrbox.tick_params(axis='both', direction='in')

            # [rows 2+] Per-cohort Normal Q-Q with 95% CI
            for _ci, _cn in enumerate(_lr_cond_names):
                _arr = np.array(_lr_groups[_cn])
                _row_qq_lr = 2 + _ci // 2
                _col_qq_lr = _ci % 2
                _ax_qq_lr = expl_lick_rate_distfit_fig.add_subplot(
                    _gs_lr[_row_qq_lr, _col_qq_lr])
                _col_lr = _coh_pal_lr[_ci]
                if len(_arr) >= 3:
                    (_osm_lr, _osr_lr), (_sl_lr, _int_lr, _) = _probplot_lr(
                        _arr, dist='norm',
                    )
                    _n_qq_lr = len(_arr)
                    with _warnings_lr.catch_warnings():
                        _warnings_lr.simplefilter('ignore')
                        _ci_lo_lr = np.array([
                            _norm_lr.ppf(_beta_lr.ppf(0.025, _i + 1, _n_qq_lr - _i))
                            for _i in range(_n_qq_lr)
                        ])
                        _ci_hi_lr = np.array([
                            _norm_lr.ppf(_beta_lr.ppf(0.975, _i + 1, _n_qq_lr - _i))
                            for _i in range(_n_qq_lr)
                        ])
                    _ax_qq_lr.fill_between(
                        _osm_lr,
                        _sl_lr * _ci_lo_lr + _int_lr,
                        _sl_lr * _ci_hi_lr + _int_lr,
                        color=_col_lr, alpha=0.18, label='95% CI',
                    )
                    _ax_qq_lr.plot(_osm_lr, _osr_lr, 'o', color=_col_lr,
                                   markersize=7, alpha=0.85, label='Mouse mean')
                    _ax_qq_lr.plot(
                        [_osm_lr[0], _osm_lr[-1]],
                        [_sl_lr * _osm_lr[0] + _int_lr,
                         _sl_lr * _osm_lr[-1] + _int_lr],
                        'k-', linewidth=1.4, label='Reference line',
                    )
                    _w_lr, _p_lr = _lr_sw[_cn][:2]
                    _verdict_lr = ('Normal (p>0.05) — t-test OK \u2713'
                                   if _p_lr > 0.05
                                   else 'Non-normal (p\u22640.05) — consider Mann-Whitney')
                    _ax_qq_lr.text(
                        0.04, 0.96,
                        f'SW: W={_w_lr:.4f}, p={_p_lr:.4f}\n{_verdict_lr}',
                        transform=_ax_qq_lr.transAxes, fontsize=8, va='top',
                        bbox=dict(boxstyle='round,pad=0.3',
                                  facecolor='#e6ffe6' if _p_lr > 0.05 else '#ffe6e6',
                                  edgecolor=_col_lr, alpha=0.90),
                    )
                else:
                    _ax_qq_lr.text(
                        0.5, 0.5,
                        f'n={len(_arr)} — need \u22653\nfor Shapiro-Wilk / Q-Q',
                        transform=_ax_qq_lr.transAxes,
                        ha='center', va='center', fontsize=10,
                    )
                _ax_qq_lr.set_title(f'{_cn} — Normal Q-Q  (n={len(_arr)} mice)', fontsize=9)
                _ax_qq_lr.set_xlabel('Theoretical quantiles (Normal)', fontsize=8)
                _ax_qq_lr.set_ylabel('Observed quantiles', fontsize=8)
                _ax_qq_lr.legend(fontsize=7)
                _ax_qq_lr.spines['top'].set_visible(False)
                _ax_qq_lr.spines['right'].set_visible(False)
                _ax_qq_lr.tick_params(axis='both', direction='in')

            # Hide unused Q-Q cell when odd cohort count
            if _n_cg_lr % 2 == 1:
                _ax_empty_lr = expl_lick_rate_distfit_fig.add_subplot(
                    _gs_lr[2 + (_n_cg_lr - 1) // 2, 1])
                _ax_empty_lr.axis('off')

            # [last row] Summary table
            ax_lrtbl = expl_lick_rate_distfit_fig.add_subplot(_gs_lr[_n_rows_lr - 1, :])
            ax_lrtbl.axis('off')
            _tbl_cols_lr = ['Cohort', 'N mice', 'Mean (licks/min)', 'SD', 'SW W', 'SW p',
                            'T-test normality OK?']
            _tbl_rows_lr = []
            for _cn in _lr_cond_names:
                _arr = np.array(_lr_groups[_cn])
                _w_lr, _p_lr, _nn_lr = _lr_sw[_cn]
                _mn_lr = float(np.mean(_arr))
                _sdv_str_lr = (f'{np.std(_arr, ddof=1):.4g}'
                               if _nn_lr >= 2 else 'n/a')
                _ok_lr = ('Yes' if (not np.isnan(_p_lr) and _p_lr > 0.05)
                          else ('Insufficient n (need \u22653)'
                                if np.isnan(_p_lr)
                                else 'No \u2014 consider Mann-Whitney U'))
                _tbl_rows_lr.append([
                    _cn, str(_nn_lr), f'{_mn_lr:.4g}', _sdv_str_lr,
                    f'{_w_lr:.4f}' if not np.isnan(_w_lr) else 'n/a',
                    f'{_p_lr:.4f}' if not np.isnan(_p_lr) else 'n/a',
                    _ok_lr,
                ])
            _tbl_lr = ax_lrtbl.table(
                cellText=_tbl_rows_lr, colLabels=_tbl_cols_lr,
                cellLoc='center', loc='center', bbox=[0, 0, 1, 1],
            )
            _tbl_lr.auto_set_font_size(False)
            _tbl_lr.set_fontsize(9)
            for (_ri, _ci2), _cell in _tbl_lr.get_celld().items():
                if _ri == 0:
                    _cell.set_facecolor('#d0d8e8')
                    _cell.set_text_props(fontweight='bold')
                elif _ci2 == 6 and _ri > 0:
                    _txt2 = _tbl_rows_lr[_ri - 1][6]
                    _cell.set_facecolor(
                        '#e6ffe6' if 'Yes' in _txt2
                        else '#ffe6e6' if 'No' in _txt2
                        else '#fffff0'
                    )
            ax_lrtbl.set_title(
                'Normality summary — one mean per mouse per cohort '
                '(Shapiro-Wilk, \u03b1=0.05)',
                fontsize=9, pad=6,
            )

            expl_lick_rate_distfit_fig.suptitle(
                'Average Lick Rate — Per-Mouse Means by Cohort\n'
                '(N=1 value per mouse = mean across all sessions; '
                'goal: assess normality for between-cohort t-test)',
                fontsize=11, y=1.01,
            )
            expl_lick_rate_distfit_fig.tight_layout()

            print('\n\u2500\u2500 Average Lick Rate \u2014 Per-Mouse Means by Cohort \u2500\u2500')
            for _cn in _lr_cond_names:
                _arr = np.array(_lr_groups[_cn])
                _w_lr, _p_lr, _nn_lr = _lr_sw[_cn]
                _sdv_str_lr = f'{np.std(_arr, ddof=1):.4g}' if _nn_lr >= 2 else 'n/a'
                _w_str_lr   = f'{_w_lr:.4f}' if not np.isnan(_w_lr) else 'n/a'
                _p_str_lr   = f'{_p_lr:.4f}' if not np.isnan(_p_lr) else 'n/a'
                print(f'  {_cn}: n={_nn_lr}, mean={np.mean(_arr):.4g}, '
                      f'SD={_sdv_str_lr}, SW W={_w_str_lr}, p={_p_str_lr}')

        except Exception as _e:
            import traceback as _tb_lr
            print(f'[expl_lick_rate_distfit] Error: {_e}')
            _tb_lr.print_exc()
            expl_lick_rate_distfit_fig = None

    # ── Exploratory: lick/reward ratio distribution — KDE + normality ────────
    expl_lick_reward_ratio_distfit_fig = None
    if 'expl_lick_reward_ratio_distfit' in selected_plots:
        try:
            from scipy.stats import (norm as _norm_lrr, probplot as _probplot_lrr,
                                     shapiro as _sw_lrr, beta as _beta_lrr)
            from matplotlib.gridspec import GridSpec as _GridSpec_lrr
            import warnings as _warnings_lrr

            # ── One mean per mouse, grouped by starting_condition ─────────────
            _lrr_groups = {}
            for _r in all_results:
                _df_r = _r['df']
                _sess_ratios = []
                for _, _row in _df_r.iterrows():
                    _h = pd.to_numeric(_row.get('hits_gap_aware', np.nan), errors='coerce')
                    _l = pd.to_numeric(_row.get('lick_count',      np.nan), errors='coerce')
                    if pd.notna(_h) and _h > 0 and pd.notna(_l):
                        _sess_ratios.append(_l / _h)
                if _sess_ratios:
                    _cond_lrr = _r.get('starting_condition', 'Unknown')
                    _lrr_groups.setdefault(_cond_lrr, []).append(float(np.mean(_sess_ratios)))

            _lrr_cond_names = sorted(_lrr_groups.keys())
            _n_cg_lrr = len(_lrr_cond_names)
            if _n_cg_lrr == 0:
                raise ValueError('No valid lick/reward ratio data found')

            # ── Shapiro-Wilk per cohort ────────────────────────────────────────
            _lrr_sw = {}
            for _cn in _lrr_cond_names:
                _arr = np.array(_lrr_groups[_cn])
                if len(_arr) >= 3:
                    _w, _p = _sw_lrr(_arr)
                    _lrr_sw[_cn] = (float(_w), float(_p), len(_arr))
                else:
                    _lrr_sw[_cn] = (np.nan, np.nan, len(_arr))

            # ── Figure layout ─────────────────────────────────────────────────
            # Row 0 (full width): overlaid histograms + Normal fit + rug marks
            # Rows 1.._n_qq_rows: per-cohort Normal Q-Q plots (2-column grid)
            # Last row (full width): summary normality table
            _n_qq_rows_lrr = (_n_cg_lrr + 1) // 2
            _n_rows_lrr = 1 + 1 + _n_qq_rows_lrr + 1  # +1 row for box plot
            _coh_pal_lrr = [plt.cm.tab10(i / max(_n_cg_lrr - 1, 1))
                            for i in range(_n_cg_lrr)]

            expl_lick_reward_ratio_distfit_fig = plt.figure(figsize=(13, 4.5 * _n_rows_lrr))
            _gs_lrr = _GridSpec_lrr(_n_rows_lrr, 2, figure=expl_lick_reward_ratio_distfit_fig,
                                    hspace=0.60, wspace=0.38)
            ax_lrrhist = expl_lick_reward_ratio_distfit_fig.add_subplot(_gs_lrr[0, :])

            # [row 0] Overlaid histograms + Normal fit + rug marks
            _lrr_rng = np.random.default_rng(42)
            for _ci, _cn in enumerate(_lrr_cond_names):
                _arr = np.array(_lrr_groups[_cn])
                _col_lrr = _coh_pal_lrr[_ci]
                _sw_lbl_lrr = (f'SW p={_lrr_sw[_cn][1]:.3f}'
                               if not np.isnan(_lrr_sw[_cn][1]) else 'SW: n<3')
                _nbins_lrr = max(4, len(_arr) // 2 + 1)
                ax_lrrhist.hist(_arr, bins=_nbins_lrr, alpha=0.40, color=_col_lrr,
                                density=True, edgecolor='white', linewidth=0.5,
                                label=f'{_cn}  (n={len(_arr)}, {_sw_lbl_lrr})')
                if len(_arr) >= 3:
                    _mn_lrr = float(np.mean(_arr))
                    _sdv_lrr = float(np.std(_arr, ddof=1))
                    _xfit_lrr = np.linspace(_mn_lrr - 4 * _sdv_lrr,
                                            _mn_lrr + 4 * _sdv_lrr, 300)
                    ax_lrrhist.plot(_xfit_lrr, _norm_lrr.pdf(_xfit_lrr, _mn_lrr, _sdv_lrr),
                                    color=_col_lrr, linewidth=2.0, linestyle='--')
                _jit_lrr = _lrr_rng.uniform(-0.003, 0.003, len(_arr))
                ax_lrrhist.plot(_arr, _jit_lrr, '|', color=_col_lrr,
                                markersize=14, markeredgewidth=2.5, alpha=0.75)

            ax_lrrhist.set_xlabel('Per-mouse mean lick / reward ratio', fontsize=10)
            ax_lrrhist.set_ylabel('Density', fontsize=10)
            ax_lrrhist.set_title(
                'Per-mouse mean lick/reward ratio — distribution by cohort\n'
                '(each point = one mouse; dashed lines = Normal fit; '
                'tick marks = individual mice)',
                fontsize=10,
            )
            ax_lrrhist.legend(fontsize=8)
            ax_lrrhist.spines['top'].set_visible(False)
            ax_lrrhist.spines['right'].set_visible(False)
            ax_lrrhist.tick_params(axis='both', direction='in')

            # [row 1] Box-and-whisker with 1.5×IQR outlier highlighting
            ax_lrrbox = expl_lick_reward_ratio_distfit_fig.add_subplot(_gs_lrr[1, :])
            _bp_data_lrr = [np.array(_lrr_groups[_cn]) for _cn in _lrr_cond_names]
            _bplot_lrr = ax_lrrbox.boxplot(
                _bp_data_lrr, labels=_lrr_cond_names, patch_artist=True,
                showfliers=False, widths=0.50,
                boxprops=dict(linewidth=1.4),
                whiskerprops=dict(linewidth=1.2, linestyle='--'),
                capprops=dict(linewidth=1.4),
                medianprops=dict(color='black', linewidth=2.0),
            )
            for _bi, _bpatch in enumerate(_bplot_lrr['boxes']):
                _bc = _coh_pal_lrr[_bi]
                _bpatch.set_facecolor([*_bc[:3], 0.35])
                _bpatch.set_edgecolor(_bc)
            _rng_box_lrr = np.random.default_rng(1)
            _outlier_legend_added_lrr = False
            for _bi, _cn in enumerate(_lrr_cond_names):
                _arr_bx = np.array(_lrr_groups[_cn])
                _q1_lrr  = float(np.percentile(_arr_bx, 25))
                _q3_lrr  = float(np.percentile(_arr_bx, 75))
                _iqr_lrr = _q3_lrr - _q1_lrr
                _lo_lrr  = _q1_lrr - 1.5 * _iqr_lrr
                _hi_lrr  = _q3_lrr + 1.5 * _iqr_lrr
                _out_lrr = (_arr_bx < _lo_lrr) | (_arr_bx > _hi_lrr)
                _jit_lrr = _rng_box_lrr.uniform(-0.15, 0.15, len(_arr_bx))
                ax_lrrbox.scatter(
                    np.full(int(np.sum(~_out_lrr)), _bi + 1) + _jit_lrr[~_out_lrr],
                    _arr_bx[~_out_lrr], color=_coh_pal_lrr[_bi],
                    s=40, alpha=0.70, zorder=3,
                )
                if _out_lrr.any():
                    ax_lrrbox.scatter(
                        np.full(int(np.sum(_out_lrr)), _bi + 1) + _jit_lrr[_out_lrr],
                        _arr_bx[_out_lrr], color='red', s=80, alpha=0.90,
                        zorder=4, edgecolors='darkred', linewidths=1.5, marker='D',
                        label='Outlier (1.5\u00d7IQR)' if not _outlier_legend_added_lrr else '',
                    )
                    _outlier_legend_added_lrr = True
                    for _ov in _arr_bx[_out_lrr]:
                        ax_lrrbox.annotate(
                            f'{_ov:.3g}', xy=(_bi + 1, _ov),
                            xytext=(8, 0), textcoords='offset points',
                            fontsize=7, color='darkred', va='center',
                        )
            ax_lrrbox.set_xlabel('Cohort', fontsize=10)
            ax_lrrbox.set_ylabel('Per-mouse mean (licks/reward)', fontsize=10)
            ax_lrrbox.set_title(
                'Box-and-whisker \u2014 per-mouse mean lick/reward ratio by cohort\n'
                '(whiskers extend to 1.5\u00d7IQR; \u25c6 red diamonds = outliers beyond fence)',
                fontsize=10,
            )
            _handles_lrr, _labels_lrr = ax_lrrbox.get_legend_handles_labels()
            if _handles_lrr:
                ax_lrrbox.legend(fontsize=8)
            ax_lrrbox.spines['top'].set_visible(False)
            ax_lrrbox.spines['right'].set_visible(False)
            ax_lrrbox.tick_params(axis='both', direction='in')

            # [rows 2+] Per-cohort Normal Q-Q with 95% CI
            for _ci, _cn in enumerate(_lrr_cond_names):
                _arr = np.array(_lrr_groups[_cn])
                _row_qq = 2 + _ci // 2
                _col_qq = _ci % 2
                _ax_qq = expl_lick_reward_ratio_distfit_fig.add_subplot(_gs_lrr[_row_qq, _col_qq])
                _col_lrr = _coh_pal_lrr[_ci]
                if len(_arr) >= 3:
                    (_osm_lrr, _osr_lrr), (_sl_lrr, _int_lrr, _) = _probplot_lrr(
                        _arr, dist='norm',
                    )
                    _n_qq_lrr = len(_arr)
                    with _warnings_lrr.catch_warnings():
                        _warnings_lrr.simplefilter('ignore')
                        _ci_lo_lrr = np.array([
                            _norm_lrr.ppf(_beta_lrr.ppf(0.025, _i + 1, _n_qq_lrr - _i))
                            for _i in range(_n_qq_lrr)
                        ])
                        _ci_hi_lrr = np.array([
                            _norm_lrr.ppf(_beta_lrr.ppf(0.975, _i + 1, _n_qq_lrr - _i))
                            for _i in range(_n_qq_lrr)
                        ])
                    _ax_qq.fill_between(
                        _osm_lrr,
                        _sl_lrr * _ci_lo_lrr + _int_lrr,
                        _sl_lrr * _ci_hi_lrr + _int_lrr,
                        color=_col_lrr, alpha=0.18, label='95% CI',
                    )
                    _ax_qq.plot(_osm_lrr, _osr_lrr, 'o', color=_col_lrr,
                                markersize=7, alpha=0.85, label='Mouse mean')
                    _ax_qq.plot(
                        [_osm_lrr[0], _osm_lrr[-1]],
                        [_sl_lrr * _osm_lrr[0] + _int_lrr,
                         _sl_lrr * _osm_lrr[-1] + _int_lrr],
                        'k-', linewidth=1.4, label='Reference line',
                    )
                    _w_lrr, _p_lrr = _lrr_sw[_cn][:2]
                    _verdict_lrr = ('Normal (p>0.05) \u2014 t-test OK \u2713'
                                    if _p_lrr > 0.05
                                    else 'Non-normal (p\u22640.05) \u2014 consider Mann-Whitney')
                    _ax_qq.text(
                        0.04, 0.96,
                        f'SW: W={_w_lrr:.4f}, p={_p_lrr:.4f}\n{_verdict_lrr}',
                        transform=_ax_qq.transAxes, fontsize=8, va='top',
                        bbox=dict(boxstyle='round,pad=0.3',
                                  facecolor='#e6ffe6' if _p_lrr > 0.05 else '#ffe6e6',
                                  edgecolor=_col_lrr, alpha=0.90),
                    )
                else:
                    _ax_qq.text(
                        0.5, 0.5,
                        f'n={len(_arr)} \u2014 need \u22653\nfor Shapiro-Wilk / Q-Q',
                        transform=_ax_qq.transAxes,
                        ha='center', va='center', fontsize=10,
                    )
                _ax_qq.set_title(f'{_cn} \u2014 Normal Q-Q  (n={len(_arr)} mice)', fontsize=9)
                _ax_qq.set_xlabel('Theoretical quantiles (Normal)', fontsize=8)
                _ax_qq.set_ylabel('Observed quantiles', fontsize=8)
                _ax_qq.legend(fontsize=7)
                _ax_qq.spines['top'].set_visible(False)
                _ax_qq.spines['right'].set_visible(False)
                _ax_qq.tick_params(axis='both', direction='in')

            # Hide unused Q-Q cell when odd cohort count
            if _n_cg_lrr % 2 == 1:
                _ax_empty_lrr = expl_lick_reward_ratio_distfit_fig.add_subplot(
                    _gs_lrr[2 + (_n_cg_lrr - 1) // 2, 1])
                _ax_empty_lrr.axis('off')

            # [last row] Summary table
            ax_lrrtbl = expl_lick_reward_ratio_distfit_fig.add_subplot(_gs_lrr[_n_rows_lrr - 1, :])
            ax_lrrtbl.axis('off')
            _tbl_cols_lrr = ['Cohort', 'N mice', 'Mean (licks/reward)', 'SD',
                             'SW W', 'SW p', 'T-test normality OK?']
            _tbl_rows_lrr = []
            for _cn in _lrr_cond_names:
                _arr = np.array(_lrr_groups[_cn])
                _w_lrr, _p_lrr, _nn_lrr = _lrr_sw[_cn]
                _mn_lrr = float(np.mean(_arr))
                _sdv_str = (f'{np.std(_arr, ddof=1):.4g}'
                            if _nn_lrr >= 2 else 'n/a')
                _ok_lrr = ('Yes' if (not np.isnan(_p_lrr) and _p_lrr > 0.05)
                           else ('Insufficient n (need \u22653)'
                                 if np.isnan(_p_lrr)
                                 else 'No \u2014 consider Mann-Whitney U'))
                _tbl_rows_lrr.append([
                    _cn, str(_nn_lrr), f'{_mn_lrr:.4g}', _sdv_str,
                    f'{_w_lrr:.4f}' if not np.isnan(_w_lrr) else 'n/a',
                    f'{_p_lrr:.4f}' if not np.isnan(_p_lrr) else 'n/a',
                    _ok_lrr,
                ])
            _tbl_lrr = ax_lrrtbl.table(
                cellText=_tbl_rows_lrr, colLabels=_tbl_cols_lrr,
                cellLoc='center', loc='center', bbox=[0, 0, 1, 1],
            )
            _tbl_lrr.auto_set_font_size(False)
            _tbl_lrr.set_fontsize(9)
            for (_ri, _ci2), _cell in _tbl_lrr.get_celld().items():
                if _ri == 0:
                    _cell.set_facecolor('#d0d8e8')
                    _cell.set_text_props(fontweight='bold')
                elif _ci2 == 6 and _ri > 0:
                    _txt2 = _tbl_rows_lrr[_ri - 1][6]
                    _cell.set_facecolor(
                        '#e6ffe6' if 'Yes' in _txt2
                        else '#ffe6e6' if 'No' in _txt2
                        else '#fffff0'
                    )
            ax_lrrtbl.set_title(
                'Normality summary \u2014 one mean per mouse per cohort '
                '(Shapiro-Wilk, \u03b1=0.05)',
                fontsize=9, pad=6,
            )

            expl_lick_reward_ratio_distfit_fig.suptitle(
                'Lick / Reward Ratio \u2014 Per-Mouse Means by Cohort\n'
                '(N=1 value per mouse = mean across all sessions with rewards > 0; '
                'goal: assess normality for between-cohort t-test)',
                fontsize=11, y=1.01,
            )
            expl_lick_reward_ratio_distfit_fig.tight_layout()

            print('\n\u2500\u2500 Lick/Reward Ratio \u2014 Per-Mouse Means by Cohort \u2500\u2500')
            for _cn in _lrr_cond_names:
                _arr = np.array(_lrr_groups[_cn])
                _w_lrr, _p_lrr, _nn_lrr = _lrr_sw[_cn]
                _sdv_str = f'{np.std(_arr, ddof=1):.4g}' if _nn_lrr >= 2 else 'n/a'
                _w_str   = f'{_w_lrr:.4f}' if not np.isnan(_w_lrr) else 'n/a'
                _p_str   = f'{_p_lrr:.4f}' if not np.isnan(_p_lrr) else 'n/a'
                print(f'  {_cn}: n={_nn_lrr}, mean={np.mean(_arr):.4g}, '
                      f'SD={_sdv_str}, SW W={_w_str}, p={_p_str}')

        except Exception as _e:
            import traceback as _tb_lrr
            print(f'[expl_lick_reward_ratio_distfit] Error: {_e}')
            _tb_lrr.print_exc()
            expl_lick_reward_ratio_distfit_fig = None

    # Create collapsed condition bar plot for bout count
    condition_bout_count_bar_fig = None
    if 'condition_bout_count_bar' in selected_plots:
        condition_bout_count_bar_fig, ax_bcbar = plt.subplots(figsize=(8, 6))

        condition_mouse_bcs: dict[str, list] = {}
        for result in all_results:
            condition = result['starting_condition']
            df_r = result['df']
            session_bcs = pd.to_numeric(df_r['bout_count'], errors='coerce').dropna().tolist()
            if session_bcs:
                if condition not in condition_mouse_bcs:
                    condition_mouse_bcs[condition] = []
                condition_mouse_bcs[condition].append((result['mouse'], float(np.mean(session_bcs))))

        conditions_sorted_bcbar = sorted(condition_mouse_bcs.keys())
        x_pos_bcbar = np.arange(len(conditions_sorted_bcbar))

        rng_bcbar = np.random.default_rng(seed=42)
        for ci, condition in enumerate(conditions_sorted_bcbar):
            entries = condition_mouse_bcs[condition]
            mouse_bcs = [v for _, v in entries]
            mean_bc = float(np.mean(mouse_bcs))
            sem_bc  = float(np.std(mouse_bcs, ddof=1) / np.sqrt(len(mouse_bcs))) if len(mouse_bcs) > 1 else 0.0
            color = condition_color_map[condition]
            ax_bcbar.bar(ci, mean_bc, width=0.5, color=color, alpha=0.8,
                         yerr=sem_bc, capsize=7, error_kw={'elinewidth': 1.5, 'capthick': 1.5})
            jitter = (rng_bcbar.random(len(mouse_bcs)) - 0.5) * 0.22
            for j, (_, bc_val) in enumerate(entries):
                ax_bcbar.plot(ci + jitter[j], bc_val, 'o',
                              color='white', markeredgecolor=color,
                              markeredgewidth=1.8, markersize=7, zorder=3)

        ax_bcbar.set_xticks(x_pos_bcbar)
        ax_bcbar.set_xticklabels(conditions_sorted_bcbar)
        ax_bcbar.set_title('Average Bout Count by Starting Condition\n(collapsed across all sessions)')
        ax_bcbar.set_xlabel('Starting Condition')
        ax_bcbar.set_ylabel('Bout Count (Mean \u00b1 SEM)')
        ax_bcbar.set_ylim(bottom=0)
        ax_bcbar.tick_params(axis='both', direction='in')
        ax_bcbar.spines['top'].set_visible(False)
        ax_bcbar.spines['right'].set_visible(False)
        condition_bout_count_bar_fig.tight_layout()

    # Create collapsed condition bar plot for average speed per bout
    condition_bout_avg_speed_bar_fig = None
    if 'condition_bout_avg_speed_bar' in selected_plots:
        condition_bout_avg_speed_bar_fig, ax_basbar = plt.subplots(figsize=(8, 6))

        condition_mouse_bas: dict[str, list] = {}
        for result in all_results:
            condition = result['starting_condition']
            df_r = result['df']
            vals = pd.to_numeric(df_r['avg_speed_per_bout'], errors='coerce').dropna().tolist()
            if vals:
                if condition not in condition_mouse_bas:
                    condition_mouse_bas[condition] = []
                condition_mouse_bas[condition].append((result['mouse'], float(np.mean(vals))))

        conditions_sorted_basbar = sorted(condition_mouse_bas.keys())
        x_pos_basbar = np.arange(len(conditions_sorted_basbar))
        rng_basbar = np.random.default_rng(seed=42)
        for ci, condition in enumerate(conditions_sorted_basbar):
            entries = condition_mouse_bas[condition]
            mouse_vals = [v for _, v in entries]
            mean_v = float(np.mean(mouse_vals))
            sem_v  = float(np.std(mouse_vals, ddof=1) / np.sqrt(len(mouse_vals))) if len(mouse_vals) > 1 else 0.0
            color = condition_color_map[condition]
            ax_basbar.bar(ci, mean_v, width=0.5, color=color, alpha=0.8,
                          yerr=sem_v, capsize=7, error_kw={'elinewidth': 1.5, 'capthick': 1.5})
            jitter = (rng_basbar.random(len(mouse_vals)) - 0.5) * 0.22
            for j, (_, val) in enumerate(entries):
                ax_basbar.plot(ci + jitter[j], val, 'o',
                               color='white', markeredgecolor=color,
                               markeredgewidth=1.8, markersize=7, zorder=3)
        ax_basbar.set_xticks(x_pos_basbar)
        ax_basbar.set_xticklabels(conditions_sorted_basbar)
        ax_basbar.set_title('Average Speed per Locomotion Bout by Starting Condition\n(collapsed across all sessions)')
        ax_basbar.set_xlabel('Starting Condition')
        ax_basbar.set_ylabel('Speed per Bout (cm/s, Mean \u00b1 SEM)')
        ax_basbar.set_ylim(bottom=0)
        ax_basbar.tick_params(axis='both', direction='in')
        ax_basbar.spines['top'].set_visible(False)
        ax_basbar.spines['right'].set_visible(False)
        condition_bout_avg_speed_bar_fig.tight_layout()

    # Create collapsed condition bar plot for average distance per bout
    condition_bout_avg_dist_bar_fig = None
    if 'condition_bout_avg_dist_bar' in selected_plots:
        condition_bout_avg_dist_bar_fig, ax_badbar = plt.subplots(figsize=(8, 6))

        condition_mouse_bad: dict[str, list] = {}
        for result in all_results:
            condition = result['starting_condition']
            df_r = result['df']
            vals = pd.to_numeric(df_r['avg_dist_per_bout'], errors='coerce').dropna().tolist()
            if vals:
                if condition not in condition_mouse_bad:
                    condition_mouse_bad[condition] = []
                condition_mouse_bad[condition].append(
                    (result['mouse'], float(np.mean(vals)) / 1000.0))  # mm → m

        conditions_sorted_badbar = sorted(condition_mouse_bad.keys())
        x_pos_badbar = np.arange(len(conditions_sorted_badbar))
        rng_badbar = np.random.default_rng(seed=42)
        for ci, condition in enumerate(conditions_sorted_badbar):
            entries = condition_mouse_bad[condition]
            mouse_vals = [v for _, v in entries]
            mean_v = float(np.mean(mouse_vals))
            sem_v  = float(np.std(mouse_vals, ddof=1) / np.sqrt(len(mouse_vals))) if len(mouse_vals) > 1 else 0.0
            color = condition_color_map[condition]
            ax_badbar.bar(ci, mean_v, width=0.5, color=color, alpha=0.8,
                          yerr=sem_v, capsize=7, error_kw={'elinewidth': 1.5, 'capthick': 1.5})
            jitter = (rng_badbar.random(len(mouse_vals)) - 0.5) * 0.22
            for j, (_, val) in enumerate(entries):
                ax_badbar.plot(ci + jitter[j], val, 'o',
                               color='white', markeredgecolor=color,
                               markeredgewidth=1.8, markersize=7, zorder=3)
        ax_badbar.set_xticks(x_pos_badbar)
        ax_badbar.set_xticklabels(conditions_sorted_badbar)
        ax_badbar.set_title('Average Distance per Locomotion Bout by Starting Condition\n(collapsed across all sessions)')
        ax_badbar.set_xlabel('Starting Condition')
        ax_badbar.set_ylabel('Distance per Bout (m, Mean \u00b1 SEM)')
        ax_badbar.set_ylim(bottom=0)
        ax_badbar.tick_params(axis='both', direction='in')
        ax_badbar.spines['top'].set_visible(False)
        ax_badbar.spines['right'].set_visible(False)
        condition_bout_avg_dist_bar_fig.tight_layout()

    # Create collapsed condition bar plot for lick rate
    condition_lick_bar_fig = None
    if 'condition_lick_bar' in selected_plots:
        condition_lick_bar_fig, ax_lbar = plt.subplots(figsize=(8, 6))

        condition_mouse_lpms: dict[str, list] = {}
        for result in all_results:
            condition = result['starting_condition']
            df_r = result['df']
            session_lpms = []
            for _, row in df_r.iterrows():
                if pd.notna(row['session_length']) and row['session_length'] > 0 and pd.notna(row['lick_count']):
                    session_lpms.append(row['lick_count'] / row['session_length'])
            if session_lpms:
                if condition not in condition_mouse_lpms:
                    condition_mouse_lpms[condition] = []
                condition_mouse_lpms[condition].append((result['mouse'], float(np.mean(session_lpms))))

        conditions_sorted_lbar = sorted(condition_mouse_lpms.keys())
        x_pos_lbar = np.arange(len(conditions_sorted_lbar))

        rng_lbar = np.random.default_rng(seed=42)
        for ci, condition in enumerate(conditions_sorted_lbar):
            entries = condition_mouse_lpms[condition]
            mouse_lpms = [v for _, v in entries]
            mean_lpm = float(np.mean(mouse_lpms))
            sem_lpm  = float(np.std(mouse_lpms, ddof=1) / np.sqrt(len(mouse_lpms))) if len(mouse_lpms) > 1 else 0.0
            color = condition_color_map[condition]
            ax_lbar.bar(ci, mean_lpm, width=0.5, color=color, alpha=0.8,
                        yerr=sem_lpm, capsize=7, error_kw={'elinewidth': 1.5, 'capthick': 1.5})
            jitter = (rng_lbar.random(len(mouse_lpms)) - 0.5) * 0.22
            for j, (mouse_name_lbar, lpm_val) in enumerate(entries):
                ax_lbar.plot(ci + jitter[j], lpm_val, 'o',
                             color='white', markeredgecolor=color,
                             markeredgewidth=1.8, markersize=7, zorder=3)

        ax_lbar.set_xticks(x_pos_lbar)
        ax_lbar.set_xticklabels(conditions_sorted_lbar)
        ax_lbar.set_title('Average Lick Rate by Starting Condition\n(collapsed across all sessions)')
        ax_lbar.set_xlabel('Starting Condition')
        ax_lbar.set_ylabel('Licks per Minute (Mean \u00b1 SEM)')
        ax_lbar.set_ylim(bottom=0)
        ax_lbar.tick_params(axis='both', direction='in')
        ax_lbar.spines['top'].set_visible(False)
        ax_lbar.spines['right'].set_visible(False)
        condition_lick_bar_fig.tight_layout()

    # Create collapsed condition bar plot for average distance per session (mm → m)
    condition_distance_bar_fig = None
    if 'condition_distance_bar' in selected_plots:
        condition_distance_bar_fig, ax_dbar = plt.subplots(figsize=(8, 6))

        condition_mouse_dist_avg: dict[str, list] = {}
        for result in all_results:
            condition = result['starting_condition']
            df_r = result['df']
            session_dists = pd.to_numeric(df_r['total_distance'], errors='coerce').dropna().tolist()
            if session_dists:
                if condition not in condition_mouse_dist_avg:
                    condition_mouse_dist_avg[condition] = []
                condition_mouse_dist_avg[condition].append(
                    (result['mouse'], float(np.mean(session_dists)) / 1000.0))  # mm → m

        conditions_sorted_dbar = sorted(condition_mouse_dist_avg.keys())
        x_pos_dbar = np.arange(len(conditions_sorted_dbar))

        rng_dbar = np.random.default_rng(seed=42)
        for ci, condition in enumerate(conditions_sorted_dbar):
            entries = condition_mouse_dist_avg[condition]
            mouse_dists = [v for _, v in entries]
            mean_d = float(np.mean(mouse_dists))
            sem_d  = float(np.std(mouse_dists, ddof=1) / np.sqrt(len(mouse_dists))) if len(mouse_dists) > 1 else 0.0
            color = condition_color_map[condition]
            ax_dbar.bar(ci, mean_d, width=0.5, color=color, alpha=0.8,
                        yerr=sem_d, capsize=7, error_kw={'elinewidth': 1.5, 'capthick': 1.5})
            jitter = (rng_dbar.random(len(mouse_dists)) - 0.5) * 0.22
            for j, (_, dist_val) in enumerate(entries):
                ax_dbar.plot(ci + jitter[j], dist_val, 'o',
                             color='white', markeredgecolor=color,
                             markeredgewidth=1.8, markersize=7, zorder=3)

        ax_dbar.set_xticks(x_pos_dbar)
        ax_dbar.set_xticklabels(conditions_sorted_dbar)
        ax_dbar.set_title('Average Distance Per Session by Starting Condition\n(collapsed across all sessions)')
        ax_dbar.set_xlabel('Starting Condition')
        ax_dbar.set_ylabel('Average Distance per Session (m, Mean \u00b1 SEM)')
        ax_dbar.set_ylim(bottom=0)
        ax_dbar.tick_params(axis='both', direction='in')
        ax_dbar.spines['top'].set_visible(False)
        ax_dbar.spines['right'].set_visible(False)
        condition_distance_bar_fig.tight_layout()

    # Create collapsed condition bar plot for total distance (sum across all sessions, mm → m)
    total_distance_bar_fig = None
    if 'total_distance_bar' in selected_plots:
        total_distance_bar_fig, ax_tbar = plt.subplots(figsize=(8, 6))

        condition_mouse_dist_total: dict[str, list] = {}
        for result in all_results:
            condition = result['starting_condition']
            df_r = result['df']
            session_dists = pd.to_numeric(df_r['total_distance'], errors='coerce').dropna().tolist()
            if session_dists:
                if condition not in condition_mouse_dist_total:
                    condition_mouse_dist_total[condition] = []
                condition_mouse_dist_total[condition].append(
                    (result['mouse'], float(np.sum(session_dists)) / 1000.0))  # mm → m

        conditions_sorted_tbar = sorted(condition_mouse_dist_total.keys())
        x_pos_tbar = np.arange(len(conditions_sorted_tbar))

        rng_tbar = np.random.default_rng(seed=42)
        for ci, condition in enumerate(conditions_sorted_tbar):
            entries = condition_mouse_dist_total[condition]
            mouse_totals = [v for _, v in entries]
            mean_t = float(np.mean(mouse_totals))
            sem_t  = float(np.std(mouse_totals, ddof=1) / np.sqrt(len(mouse_totals))) if len(mouse_totals) > 1 else 0.0
            color = condition_color_map[condition]
            ax_tbar.bar(ci, mean_t, width=0.5, color=color, alpha=0.8,
                        yerr=sem_t, capsize=7, error_kw={'elinewidth': 1.5, 'capthick': 1.5})
            jitter = (rng_tbar.random(len(mouse_totals)) - 0.5) * 0.22
            for j, (_, total_val) in enumerate(entries):
                ax_tbar.plot(ci + jitter[j], total_val, 'o',
                             color='white', markeredgecolor=color,
                             markeredgewidth=1.8, markersize=7, zorder=3)

        ax_tbar.set_xticks(x_pos_tbar)
        ax_tbar.set_xticklabels(conditions_sorted_tbar)
        ax_tbar.set_title('Total Distance Traveled per Mouse by Starting Condition\n(summed across all sessions)')
        ax_tbar.set_xlabel('Starting Condition')
        ax_tbar.set_ylabel('Total Distance (m, Mean \u00b1 SEM)')
        ax_tbar.set_ylim(bottom=0)
        ax_tbar.tick_params(axis='both', direction='in')
        ax_tbar.spines['top'].set_visible(False)
        ax_tbar.spines['right'].set_visible(False)
        total_distance_bar_fig.tight_layout()

    # ── Behavioral epoch plots ────────────────────────────────────────────────
    epoch_speed_per_mouse_fig            = epoch_speed_cond_fig            = None
    epoch_cap_per_mouse_fig              = epoch_cap_cond_fig              = None
    epoch_speed_sess_per_mouse_fig       = epoch_speed_sess_cond_fig       = None
    epoch_cap_sess_per_mouse_fig         = epoch_cap_sess_cond_fig         = None
    epoch_speed_early_per_mouse_fig    = epoch_speed_late_per_mouse_fig    = None
    epoch_speed_early_cond_fig         = epoch_speed_late_cond_fig         = None
    epoch_cap_early_per_mouse_fig      = epoch_cap_late_per_mouse_fig      = None
    epoch_cap_early_cond_fig           = epoch_cap_late_cond_fig           = None
    epoch_speed_sess_cond_clean_fig    = None
    epoch_cap_sess_cond_clean_fig      = None
    epoch_speed_early_cond_clean_fig   = epoch_speed_late_cond_clean_fig   = None
    epoch_cap_early_cond_clean_fig     = epoch_cap_late_cond_clean_fig     = None
    epoch_speed_early_ev_per_mouse_fig = epoch_speed_late_ev_per_mouse_fig = None
    epoch_speed_early_ev_cond_fig      = epoch_speed_late_ev_cond_fig      = None
    epoch_cap_early_ev_per_mouse_fig   = epoch_cap_late_ev_per_mouse_fig   = None
    epoch_cap_early_ev_cond_fig        = epoch_cap_late_ev_cond_fig        = None
    epoch_speed_early_ev_cond_clean_fig = epoch_speed_late_ev_cond_clean_fig = None
    epoch_cap_early_ev_cond_clean_fig   = epoch_cap_late_ev_cond_clean_fig   = None
    punish_speed_per_mouse_fig         = punish_speed_cond_fig         = None
    punish_cap_per_mouse_fig           = punish_cap_cond_fig           = None
    punish_speed_sess_per_mouse_fig    = punish_speed_sess_cond_fig    = None
    punish_cap_sess_per_mouse_fig      = punish_cap_sess_cond_fig      = None
    punish_speed_sess_cond_clean_fig   = None
    punish_cap_sess_cond_clean_fig     = None
    epoch_speed_sess_sex_per_mouse_fig  = epoch_speed_sess_sex_fig  = None
    epoch_cap_sess_sex_per_mouse_fig    = epoch_cap_sess_sex_fig    = None
    epoch_punish_speed_sess_sex_per_mouse_fig = epoch_punish_speed_sess_sex_fig = None
    epoch_punish_cap_sess_sex_per_mouse_fig   = epoch_punish_cap_sess_sex_fig   = None
    epoch_reward_speed_pre_post_fig = None
    epoch_reward_speed_diff_fig = None
    epoch_reward_cap_pre_post_fig = None
    epoch_reward_cap_diff_fig = None
    epoch_reward_speed_pre_post_entry_fig = None
    epoch_reward_speed_diff_entry_fig = None
    epoch_reward_speed_pre_post_entry_1s_fig = None
    epoch_reward_speed_diff_entry_1s_fig = None
    epoch_reward_lick_count_sess_per_mouse_fig = epoch_reward_lick_count_sess_cond_fig = None
    epoch_reward_lick_count_sess_cond_clean_fig = None
    epoch_punish_lick_count_sess_per_mouse_fig = epoch_punish_lick_count_sess_cond_fig = None
    epoch_punish_speed_pre_post_fig = None
    epoch_punish_speed_diff_fig = None
    epoch_punish_speed_pre_post_entry_fig = None
    epoch_punish_speed_diff_entry_fig = None
    epoch_punish_cap_pre_post_fig = None
    epoch_punish_cap_diff_fig = None
    epoch_punish_cap_pre_post_entry_fig = None
    epoch_punish_cap_diff_entry_fig = None
    _epoch_keys = {'epoch_reward_speed', 'epoch_reward_cap',
                   'epoch_reward_speed_sess', 'epoch_reward_cap_sess',
                   'epoch_reward_speed_early_late', 'epoch_reward_cap_early_late',
                   'epoch_reward_speed_early_late_ev', 'epoch_reward_cap_early_late_ev',
                   'epoch_reward_speed_early_late_ev_clean', 'epoch_reward_cap_early_late_ev_clean',
                   'epoch_reward_speed_sess_clean', 'epoch_reward_cap_sess_clean',
                   'epoch_reward_speed_early_late_clean', 'epoch_reward_cap_early_late_clean',
                   'epoch_punish_speed', 'epoch_punish_cap',
                   'epoch_punish_speed_sess', 'epoch_punish_cap_sess',
                   'epoch_punish_speed_sess_clean', 'epoch_punish_cap_sess_clean',
                   'epoch_reward_speed_sess_sex', 'epoch_reward_cap_sess_sex',
                   'epoch_punish_speed_sess_sex', 'epoch_punish_cap_sess_sex',
                   'epoch_reward_speed_pre_post',
                   'epoch_reward_speed_diff',
                   'epoch_reward_cap_pre_post',
                   'epoch_reward_cap_diff',
                   'epoch_reward_speed_pre_post_entry',
                   'epoch_reward_speed_diff_entry',
                   'epoch_reward_speed_pre_post_entry_1s',
                   'epoch_reward_speed_diff_entry_1s',
                   'epoch_reward_lick_count_sess',
                   'epoch_reward_lick_count_sess_clean',
                   'epoch_punish_lick_count_sess',
                   'epoch_punish_speed_pre_post',
                   'epoch_punish_speed_diff',
                   'epoch_punish_speed_pre_post_entry',
                   'epoch_punish_speed_diff_entry',
                   'epoch_punish_cap_pre_post',
                   'epoch_punish_cap_diff',
                   'epoch_punish_cap_pre_post_entry',
                   'epoch_punish_cap_diff_entry'}
    if _epoch_keys & set(selected_plots):
        _any_speed = any(r.get('speed_epoch_matrix') is not None for r in all_results)
        _any_cap   = any(r.get('cap_epoch_matrix')   is not None for r in all_results)

        if 'epoch_reward_speed' in selected_plots and _any_speed:
            epoch_speed_per_mouse_fig, epoch_speed_cond_fig = _plot_epoch_panels(
                all_results, 'speed_epoch_matrix',
                ylabel='Treadmill Speed (cm/s)',
                title_prefix='Speed Aligned to Reward Zone Entry',
                condition_color_map=condition_color_map,
                hierarchy='event',
                use_sd=True,
            )

        if 'epoch_reward_cap' in selected_plots and _any_cap:
            epoch_cap_per_mouse_fig, epoch_cap_cond_fig = _plot_epoch_panels(
                all_results, 'cap_epoch_matrix',
                ylabel='Capacitive Value (z-score)',
                title_prefix='Capacitive Value Aligned to Reward Zone Entry',
                condition_color_map=condition_color_map,
                hierarchy='event',
                use_sd=True,
            )

        if 'epoch_reward_speed_sess' in selected_plots and _any_speed:
            epoch_speed_sess_per_mouse_fig, epoch_speed_sess_cond_fig = _plot_epoch_panels(
                all_results, 'speed_epoch_session_means',
                ylabel='Treadmill Speed (cm/s)',
                title_prefix='Speed Aligned to Reward Zone Entry',
                condition_color_map=condition_color_map,
                hierarchy='session',
            )
        if 'epoch_reward_speed_sess_clean' in selected_plots and _any_speed:
            _, epoch_speed_sess_cond_clean_fig = _plot_epoch_panels(
                all_results, 'speed_epoch_session_means',
                ylabel='Treadmill Speed (cm/s)',
                title_prefix='Speed Aligned to Reward Zone Entry',
                condition_color_map=condition_color_map,
                hierarchy='session',
                show_individual_traces=False,
            )

        if 'epoch_reward_cap_sess' in selected_plots and _any_cap:
            epoch_cap_sess_per_mouse_fig, epoch_cap_sess_cond_fig = _plot_epoch_panels(
                all_results, 'cap_epoch_session_means',
                ylabel='Capacitive Value (z-score)',
                title_prefix='Capacitive Value Aligned to Reward Zone Entry',
                condition_color_map=condition_color_map,
                hierarchy='session',
            )
        if 'epoch_reward_cap_sess_clean' in selected_plots and _any_cap:
            _, epoch_cap_sess_cond_clean_fig = _plot_epoch_panels(
                all_results, 'cap_epoch_session_means',
                ylabel='Capacitive Value (z-score)',
                title_prefix='Capacitive Value Aligned to Reward Zone Entry',
                condition_color_map=condition_color_map,
                hierarchy='session',
                show_individual_traces=False,
            )

        if 'epoch_reward_speed_early_late' in selected_plots and _any_speed:
            (epoch_speed_early_per_mouse_fig, epoch_speed_late_per_mouse_fig,
             epoch_speed_early_cond_fig,      epoch_speed_late_cond_fig) = \
                _plot_epoch_early_late_panels(
                    all_results, 'speed_epoch_session_means',
                    ylabel='Treadmill Speed (cm/s)',
                    title_prefix='Speed Aligned to Reward Zone Entry',
                    condition_color_map=condition_color_map,
                )
        if 'epoch_reward_speed_early_late_clean' in selected_plots and _any_speed:
            (_, _,
             epoch_speed_early_cond_clean_fig, epoch_speed_late_cond_clean_fig) = \
                _plot_epoch_early_late_panels(
                    all_results, 'speed_epoch_session_means',
                    ylabel='Treadmill Speed (cm/s)',
                    title_prefix='Speed Aligned to Reward Zone Entry',
                    condition_color_map=condition_color_map,
                    show_individual_traces=False,
                )

        if 'epoch_reward_cap_early_late' in selected_plots and _any_cap:
            (epoch_cap_early_per_mouse_fig, epoch_cap_late_per_mouse_fig,
             epoch_cap_early_cond_fig,      epoch_cap_late_cond_fig) = \
                _plot_epoch_early_late_panels(
                    all_results, 'cap_epoch_session_means',
                    ylabel='Capacitive Value (z-score)',
                    title_prefix='Capacitive Value Aligned to Reward Zone Entry',
                    condition_color_map=condition_color_map,
                )
        if 'epoch_reward_cap_early_late_clean' in selected_plots and _any_cap:
            (_, _,
             epoch_cap_early_cond_clean_fig, epoch_cap_late_cond_clean_fig) = \
                _plot_epoch_early_late_panels(
                    all_results, 'cap_epoch_session_means',
                    ylabel='Capacitive Value (z-score)',
                    title_prefix='Capacitive Value Aligned to Reward Zone Entry',
                    condition_color_map=condition_color_map,
                    show_individual_traces=False,
                )

        if 'epoch_reward_speed_early_late_ev' in selected_plots and _any_speed:
            (epoch_speed_early_ev_per_mouse_fig, epoch_speed_late_ev_per_mouse_fig,
             epoch_speed_early_ev_cond_fig,      epoch_speed_late_ev_cond_fig) = \
                _plot_epoch_early_late_panels(
                    all_results, 'speed_epoch_matrix',
                    ylabel='Treadmill Speed (cm/s)',
                    title_prefix='Speed Aligned to Reward Zone Entry',
                    condition_color_map=condition_color_map,
                    indices_key='speed_epoch_event_indices',
                    row_unit='events',
                    use_sd=True,
                )

        if 'epoch_reward_cap_early_late_ev' in selected_plots and _any_cap:
            (epoch_cap_early_ev_per_mouse_fig, epoch_cap_late_ev_per_mouse_fig,
             epoch_cap_early_ev_cond_fig,      epoch_cap_late_ev_cond_fig) = \
                _plot_epoch_early_late_panels(
                    all_results, 'cap_epoch_matrix',
                    ylabel='Capacitive Value (z-score)',
                    title_prefix='Capacitive Value Aligned to Reward Zone Entry',
                    condition_color_map=condition_color_map,
                    indices_key='cap_epoch_event_indices',
                    row_unit='events',
                    use_sd=True,
                )
        if 'epoch_reward_speed_early_late_ev_clean' in selected_plots and _any_speed:
            (_, _,
             epoch_speed_early_ev_cond_clean_fig, epoch_speed_late_ev_cond_clean_fig) = \
                _plot_epoch_early_late_panels(
                    all_results, 'speed_epoch_matrix',
                    ylabel='Treadmill Speed (cm/s)',
                    title_prefix='Speed Aligned to Reward Zone Entry',
                    condition_color_map=condition_color_map,
                    indices_key='speed_epoch_event_indices',
                    row_unit='events',
                    use_sd=True,
                    show_individual_traces=False,
                )
        if 'epoch_reward_cap_early_late_ev_clean' in selected_plots and _any_cap:
            (_, _,
             epoch_cap_early_ev_cond_clean_fig, epoch_cap_late_ev_cond_clean_fig) = \
                _plot_epoch_early_late_panels(
                    all_results, 'cap_epoch_matrix',
                    ylabel='Capacitive Value (z-score)',
                    title_prefix='Capacitive Value Aligned to Reward Zone Entry',
                    condition_color_map=condition_color_map,
                    indices_key='cap_epoch_event_indices',
                    row_unit='events',
                    use_sd=True,
                    show_individual_traces=False,
                )

        # ── Reward zone: session-averaged lick count epoch ───────────────────
        _any_reward_lick = any(r.get('lick_epoch_session_means') is not None for r in all_results)
        if 'epoch_reward_lick_count_sess' in selected_plots and _any_reward_lick:
            epoch_reward_lick_count_sess_per_mouse_fig, epoch_reward_lick_count_sess_cond_fig = \
                _plot_epoch_panels(
                    all_results, 'lick_epoch_session_means',
                    ylabel='Lick count (licks per 500 ms bin)',
                    title_prefix='Lick Count Aligned to Reward Zone Entry',
                    condition_color_map=condition_color_map,
                    hierarchy='session',
                    reward_delivery_vline=True,
                    canonical_time=LICK_EPOCH_TIME,
                )
        if 'epoch_reward_lick_count_sess_clean' in selected_plots and _any_reward_lick:
            _, epoch_reward_lick_count_sess_cond_clean_fig = \
                _plot_epoch_panels(
                    all_results, 'lick_epoch_session_means',
                    ylabel='Lick count (licks per 500 ms bin)',
                    title_prefix='Lick Count Aligned to Reward Zone Entry',
                    condition_color_map=condition_color_map,
                    hierarchy='session',
                    reward_delivery_vline=True,
                    show_individual_traces=False,
                    canonical_time=LICK_EPOCH_TIME,
                )

        # ── Punishment zone epoch plots ───────────────────────────────────────
        _any_punish_speed = any(r.get('punish_speed_epoch_matrix') is not None for r in all_results)
        _any_punish_cap   = any(r.get('punish_cap_epoch_matrix')   is not None for r in all_results)

        _n_punish_speed = sum(1 for r in all_results if r.get('punish_speed_epoch_matrix') is not None)
        _n_punish_cap   = sum(1 for r in all_results if r.get('punish_cap_epoch_matrix')   is not None)
        print(f"[Epoch] Punishment zone — speed data: {_n_punish_speed}/{len(all_results)} mice, "
              f"capacitive data: {_n_punish_cap}/{len(all_results)} mice")
        if not _any_punish_speed and not _any_punish_cap:
            print("  [WARN] No punishment zone epoch data found for any mouse — "
                  "check that trial logs contain 'stay_punish_texture_change_time' "
                  "(new hallway format) or 'texture_history' (old format).")
            print("  [WARN] Per-mouse punishment epoch matrices:")
            for r in all_results:
                _has_spd = r.get('punish_speed_epoch_matrix') is not None
                _has_cap = r.get('punish_cap_epoch_matrix') is not None
                print(f"    {r['mouse']}: speed={_has_spd}, cap={_has_cap}")

        if 'epoch_punish_speed' in selected_plots and _any_punish_speed:
            punish_speed_per_mouse_fig, punish_speed_cond_fig = _plot_epoch_panels(
                all_results, 'punish_speed_epoch_matrix',
                ylabel='Treadmill Speed (cm/s)',
                title_prefix='Speed Aligned to Punishment Zone Entry',
                condition_color_map=condition_color_map,
                hierarchy='event',
                use_sd=True,
                reward_delivery_vline=False,
            )

        if 'epoch_punish_cap' in selected_plots and _any_punish_cap:
            punish_cap_per_mouse_fig, punish_cap_cond_fig = _plot_epoch_panels(
                all_results, 'punish_cap_epoch_matrix',
                ylabel='Capacitive Value (z-score)',
                title_prefix='Capacitive Value Aligned to Punishment Zone Entry',
                condition_color_map=condition_color_map,
                hierarchy='event',
                use_sd=True,
                reward_delivery_vline=False,
            )

        if 'epoch_punish_speed_sess' in selected_plots and _any_punish_speed:
            punish_speed_sess_per_mouse_fig, punish_speed_sess_cond_fig = _plot_epoch_panels(
                all_results, 'punish_speed_epoch_session_means',
                ylabel='Treadmill Speed (cm/s)',
                title_prefix='Speed Aligned to Punishment Zone Entry',
                condition_color_map=condition_color_map,
                hierarchy='session',
                reward_delivery_vline=False,
            )
        if 'epoch_punish_speed_sess_clean' in selected_plots and _any_punish_speed:
            _, punish_speed_sess_cond_clean_fig = _plot_epoch_panels(
                all_results, 'punish_speed_epoch_session_means',
                ylabel='Treadmill Speed (cm/s)',
                title_prefix='Speed Aligned to Punishment Zone Entry',
                condition_color_map=condition_color_map,
                hierarchy='session',
                show_individual_traces=False,
                reward_delivery_vline=False,
            )

        if 'epoch_punish_cap_sess' in selected_plots and _any_punish_cap:
            punish_cap_sess_per_mouse_fig, punish_cap_sess_cond_fig = _plot_epoch_panels(
                all_results, 'punish_cap_epoch_session_means',
                ylabel='Capacitive Value (z-score)',
                title_prefix='Capacitive Value Aligned to Punishment Zone Entry',
                condition_color_map=condition_color_map,
                hierarchy='session',
                reward_delivery_vline=False,
            )
        if 'epoch_punish_cap_sess_clean' in selected_plots and _any_punish_cap:
            _, punish_cap_sess_cond_clean_fig = _plot_epoch_panels(
                all_results, 'punish_cap_epoch_session_means',
                ylabel='Capacitive Value (z-score)',
                title_prefix='Capacitive Value Aligned to Punishment Zone Entry',
                condition_color_map=condition_color_map,
                hierarchy='session',
                show_individual_traces=False,
                reward_delivery_vline=False,
            )

        # ── Sex-split session-averaged epoch plots ───────────────────────────────────
        _sex_color_map = {'male': 'green', 'female': 'purple'}

        if 'epoch_reward_speed_sess_sex' in selected_plots and _any_speed:
            epoch_speed_sess_sex_per_mouse_fig, epoch_speed_sess_sex_fig = _plot_epoch_panels(
                all_results, 'speed_epoch_session_means',
                ylabel='Treadmill Speed (cm/s)',
                title_prefix='Speed Aligned to Reward Zone Entry',
                condition_color_map=condition_color_map,
                hierarchy='session',
                show_individual_traces=False,
                group_key='sex',
                group_color_map=_sex_color_map,
                group_label='By Sex',
            )

        if 'epoch_reward_cap_sess_sex' in selected_plots and _any_cap:
            epoch_cap_sess_sex_per_mouse_fig, epoch_cap_sess_sex_fig = _plot_epoch_panels(
                all_results, 'cap_epoch_session_means',
                ylabel='Capacitive Value (z-score)',
                title_prefix='Capacitive Value Aligned to Reward Zone Entry',
                condition_color_map=condition_color_map,
                hierarchy='session',
                show_individual_traces=False,
                group_key='sex',
                group_color_map=_sex_color_map,
                group_label='By Sex',
            )

        if 'epoch_punish_speed_sess_sex' in selected_plots and _any_punish_speed:
            epoch_punish_speed_sess_sex_per_mouse_fig, epoch_punish_speed_sess_sex_fig = _plot_epoch_panels(
                all_results, 'punish_speed_epoch_session_means',
                ylabel='Treadmill Speed (cm/s)',
                title_prefix='Speed Aligned to Punishment Zone Entry',
                condition_color_map=condition_color_map,
                hierarchy='session',
                show_individual_traces=False,
                reward_delivery_vline=False,
                group_key='sex',
                group_color_map=_sex_color_map,
                group_label='By Sex',
            )

        if 'epoch_punish_cap_sess_sex' in selected_plots and _any_punish_cap:
            epoch_punish_cap_sess_sex_per_mouse_fig, epoch_punish_cap_sess_sex_fig = _plot_epoch_panels(
                all_results, 'punish_cap_epoch_session_means',
                ylabel='Capacitive Value (z-score)',
                title_prefix='Capacitive Value Aligned to Punishment Zone Entry',
                condition_color_map=condition_color_map,
                hierarchy='session',
                show_individual_traces=False,
                reward_delivery_vline=False,
                group_key='sex',
                group_color_map=_sex_color_map,
                group_label='By Sex',
            )

        # ── Punishment zone: session-averaged lick count epoch ───────────────
        _any_punish_lick = any(r.get('punish_lick_epoch_session_means') is not None for r in all_results)
        if 'epoch_punish_lick_count_sess' in selected_plots and _any_punish_lick:
            epoch_punish_lick_count_sess_per_mouse_fig, epoch_punish_lick_count_sess_cond_fig = \
                _plot_epoch_panels(
                    all_results, 'punish_lick_epoch_session_means',
                    ylabel='Lick count (licks per 500 ms bin)',
                    title_prefix='Lick Count Aligned to Punishment Zone Entry',
                    condition_color_map=condition_color_map,
                    hierarchy='session',
                    reward_delivery_vline=False,
                    canonical_time=LICK_EPOCH_TIME,
                )

        # ── Punishment zone: pre/post 0.65 s cutoff speed bar chart ──────────
        if 'epoch_punish_speed_pre_post' in selected_plots and _any_punish_speed:
            _pre_mask_pp  = (EPOCH_CANONICAL_TIME >= 0.0)  & (EPOCH_CANONICAL_TIME <= 0.65)
            _post_mask_pp = (EPOCH_CANONICAL_TIME >  0.65) & (EPOCH_CANONICAL_TIME <= 1.30)
            _ppp_by_cond: dict = {}
            for _r in all_results:
                _sess_mat = _r.get('punish_speed_epoch_session_means')
                if _sess_mat is None or _sess_mat.shape[0] == 0:
                    continue
                _cond = _r['starting_condition']
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore', RuntimeWarning)
                    _pre_per_sess  = np.nanmean(_sess_mat[:, _pre_mask_pp],  axis=1)
                    _post_per_sess = np.nanmean(_sess_mat[:, _post_mask_pp], axis=1)
                    _pre_mean  = float(np.nanmean(_pre_per_sess))
                    _post_mean = float(np.nanmean(_post_per_sess))
                _ppp_by_cond.setdefault(_cond, []).append((_r['mouse'], _pre_mean, _post_mean))

            if _ppp_by_cond:
                _conds_ppp   = sorted(_ppp_by_cond.keys())
                _n_conds_ppp = len(_conds_ppp)
                epoch_punish_speed_pre_post_fig, _axs_ppp = plt.subplots(
                    1, _n_conds_ppp,
                    figsize=(4 * _n_conds_ppp + 1, 5),
                    sharey=True, squeeze=False,
                )
                _all_ppp_yvals = []
                for _ci, _cond in enumerate(_conds_ppp):
                    _ax       = _axs_ppp[0, _ci]
                    _color    = condition_color_map.get(_cond, 'steelblue')
                    _entries  = _ppp_by_cond[_cond]
                    _n_ppp    = len(_entries)
                    _pre_vals = [e[1] for e in _entries]
                    _post_vals= [e[2] for e in _entries]
                    with warnings.catch_warnings():
                        warnings.simplefilter('ignore', RuntimeWarning)
                        _mn_pre   = float(np.nanmean(_pre_vals))
                        _mn_post  = float(np.nanmean(_post_vals))
                        _sem_pre  = (float(np.nanstd(_pre_vals,  ddof=1) / np.sqrt(_n_ppp))
                                     if _n_ppp > 1 else 0.0)
                        _sem_post = (float(np.nanstd(_post_vals, ddof=1) / np.sqrt(_n_ppp))
                                     if _n_ppp > 1 else 0.0)
                    _all_ppp_yvals.extend([_mn_pre + _sem_pre, _mn_post + _sem_post,
                                           _mn_pre - _sem_pre, _mn_post - _sem_post])
                    _all_ppp_yvals.extend(_pre_vals + _post_vals)
                    _ax.bar(0, _mn_pre,  width=0.5, color=_color, alpha=0.7,
                            yerr=_sem_pre,  capsize=7,
                            error_kw={'elinewidth': 1.5, 'capthick': 1.5})
                    _ax.bar(1, _mn_post, width=0.5, color=_color, alpha=0.7,
                            yerr=_sem_post, capsize=7,
                            error_kw={'elinewidth': 1.5, 'capthick': 1.5})
                    _rng_ppp = np.random.default_rng(seed=42)
                    _jitter  = (_rng_ppp.random(_n_ppp) - 0.5) * 0.18
                    for _j, (_mname, _pv, _qv) in enumerate(_entries):
                        _xp = 0 + _jitter[_j]
                        _xq = 1 + _jitter[_j]
                        _ax.plot([_xp, _xq], [_pv, _qv], '-',
                                 color=_color, linewidth=0.9, alpha=0.5, zorder=2)
                        _ax.plot(_xp, _pv, 'o', color='white',
                                 markeredgecolor=_color, markeredgewidth=1.5,
                                 markersize=7, zorder=3)
                        _ax.plot(_xq, _qv, 'o', color='white',
                                 markeredgecolor=_color, markeredgewidth=1.5,
                                 markersize=7, zorder=3)
                    _ax.set_xticks([0, 1])
                    _ax.set_xticklabels(['Pre-cutoff\n(0\u20130.65 s)', 'Post-cutoff\n(0.65\u20131.3 s)'],
                                        fontsize=9)
                    _ax.set_title(f'{_cond}\n(n={_n_ppp} mice)', fontsize=10)
                    _ax.set_ylabel('Treadmill Speed (cm/s)' if _ci == 0 else '', fontsize=9)
                    _ax.set_xlim(-0.6, 1.6)
                    _ax.tick_params(axis='both', direction='in')
                    _ax.spines['top'].set_visible(False)
                    _ax.spines['right'].set_visible(False)
                if _all_ppp_yvals:
                    _ymax_ppp = float(np.nanmax(_all_ppp_yvals))
                    _ymin_ppp = float(np.nanmin(_all_ppp_yvals))
                else:
                    _ymax_ppp, _ymin_ppp = 1.0, 0.0
                _bot_ppp = _ymin_ppp * 1.05 if _ymin_ppp < 0 else 0.0
                _axs_ppp[0, 0].set_ylim(_bot_ppp, _ymax_ppp * 1.05)
                epoch_punish_speed_pre_post_fig.suptitle(
                    'Average Speed: Pre- vs Post-Cutoff (Punishment Zone)\n'
                    '(session-averaged punishment zone entry epochs, by condition)',
                    fontsize=12,
                )
                epoch_punish_speed_pre_post_fig.tight_layout()

        # ── Punishment zone: pre-minus-post speed difference bar chart ────────
        if 'epoch_punish_speed_diff' in selected_plots and _any_punish_speed:
            _pre_mask_pd  = (EPOCH_CANONICAL_TIME >= 0.0)  & (EPOCH_CANONICAL_TIME <= 0.65)
            _post_mask_pd = (EPOCH_CANONICAL_TIME >  0.65) & (EPOCH_CANONICAL_TIME <= 1.30)
            _pdiff_by_cond: dict = {}
            for _r in all_results:
                _sess_mat = _r.get('punish_speed_epoch_session_means')
                if _sess_mat is None or _sess_mat.shape[0] == 0:
                    continue
                _cond = _r['starting_condition']
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore', RuntimeWarning)
                    _pre_pd  = float(np.nanmean(np.nanmean(_sess_mat[:, _pre_mask_pd],  axis=1)))
                    _post_pd = float(np.nanmean(np.nanmean(_sess_mat[:, _post_mask_pd], axis=1)))
                _pdiff_by_cond.setdefault(_cond, []).append((_r['mouse'], _post_pd - _pre_pd))

            if _pdiff_by_cond:
                _conds_pd   = sorted(_pdiff_by_cond.keys())
                _n_conds_pd = len(_conds_pd)
                epoch_punish_speed_diff_fig, _ax_pd = plt.subplots(
                    1, 1, figsize=(max(4, _n_conds_pd * 1.4 + 1.5), 5)
                )
                _all_pdiff_vals = []
                _bar_x_pd = np.arange(_n_conds_pd)
                _rng_pd   = np.random.default_rng(seed=42)
                for _ci, _cond in enumerate(_conds_pd):
                    _color    = condition_color_map.get(_cond, 'steelblue')
                    _entries  = _pdiff_by_cond[_cond]
                    _n_pd     = len(_entries)
                    _dvals_pd = [e[1] for e in _entries]
                    with warnings.catch_warnings():
                        warnings.simplefilter('ignore', RuntimeWarning)
                        _mn_pd  = float(np.nanmean(_dvals_pd))
                        _sem_pd = (float(np.nanstd(_dvals_pd, ddof=1) / np.sqrt(_n_pd))
                                   if _n_pd > 1 else 0.0)
                    _all_pdiff_vals.extend(_dvals_pd + [_mn_pd + _sem_pd, _mn_pd - _sem_pd])
                    _ax_pd.bar(_ci, _mn_pd, width=0.55, color=_color, alpha=0.7,
                               yerr=_sem_pd, capsize=7,
                               error_kw={'elinewidth': 1.5, 'capthick': 1.5})
                    _jitter_pd = (_rng_pd.random(_n_pd) - 0.5) * 0.22
                    for _j, (_mname, _dv) in enumerate(_entries):
                        _ax_pd.plot(_ci + _jitter_pd[_j], _dv, 'o',
                                    color='white', markeredgecolor=_color,
                                    markeredgewidth=1.5, markersize=7, zorder=3)
                _ax_pd.axhline(0, color='black', linewidth=0.9, linestyle='--', zorder=1)
                _ax_pd.set_xticks(_bar_x_pd)
                _ax_pd.set_xticklabels(_conds_pd, fontsize=10)
                _ax_pd.set_ylabel('Speed difference (cm/s)\n[post-cutoff \u2212 pre-cutoff]', fontsize=9)
                _ax_pd.set_xlabel('Condition', fontsize=10)
                if _all_pdiff_vals:
                    _ymax_pd = float(np.nanmax(_all_pdiff_vals))
                    _ymin_pd = float(np.nanmin(_all_pdiff_vals))
                    _pad_pd  = max(abs(_ymax_pd), abs(_ymin_pd)) * 0.12 or 0.5
                    _ax_pd.set_ylim(_ymin_pd - _pad_pd, _ymax_pd + _pad_pd)
                _ax_pd.tick_params(axis='both', direction='in')
                _ax_pd.spines['top'].set_visible(False)
                _ax_pd.spines['right'].set_visible(False)

                # ── Mann-Whitney U tests for all condition pairs ──────────────
                import itertools as _itertools_pd
                _pairs_pd = list(_itertools_pd.combinations(range(_n_conds_pd), 2))
                _ylim_cur_pd = list(_ax_pd.get_ylim())
                _bracket_step_pd = (_ylim_cur_pd[1] - _ylim_cur_pd[0]) * 0.14
                _bracket_base_pd = _ylim_cur_pd[1]
                for _pi, (_ia, _ib) in enumerate(_pairs_pd):
                    _vals_a_pd = [e[1] for e in _pdiff_by_cond[_conds_pd[_ia]]]
                    _vals_b_pd = [e[1] for e in _pdiff_by_cond[_conds_pd[_ib]]]
                    if len(_vals_a_pd) < 2 or len(_vals_b_pd) < 2:
                        continue
                    _u_stat_pd, _p_val_pd = mannwhitneyu(_vals_a_pd, _vals_b_pd, alternative='two-sided')
                    if _p_val_pd < 0.001:
                        _sig_str_pd = f'p = {_p_val_pd:.2e}***'
                    elif _p_val_pd < 0.01:
                        _sig_str_pd = f'p = {_p_val_pd:.3f}**'
                    elif _p_val_pd < 0.05:
                        _sig_str_pd = f'p = {_p_val_pd:.3f}*'
                    else:
                        _sig_str_pd = f'p = {_p_val_pd:.3f} (ns)'
                    _bh_pd = _bracket_base_pd + _bracket_step_pd * (_pi + 0.6)
                    _ax_pd.plot([_ia, _ia, _ib, _ib],
                                [_bh_pd - _bracket_step_pd * 0.15,
                                 _bh_pd,
                                 _bh_pd,
                                 _bh_pd - _bracket_step_pd * 0.15],
                                color='black', linewidth=1.0)
                    _ax_pd.text((_ia + _ib) / 2, _bh_pd + _bracket_step_pd * 0.05,
                                _sig_str_pd, ha='center', va='bottom', fontsize=8)
                if _pairs_pd:
                    _new_top_pd = _bracket_base_pd + _bracket_step_pd * (len(_pairs_pd) + 1.5)
                    _ax_pd.set_ylim(_ylim_cur_pd[0], _new_top_pd)

                epoch_punish_speed_diff_fig.suptitle(
                    'Post- vs Pre-Cutoff Speed Difference by Condition (Punishment Zone)\n'
                    '(mean \u00b1 SEM across mice; positive = faster after 0.65 s cutoff)',
                    fontsize=11,
                )
                epoch_punish_speed_diff_fig.tight_layout()

        # ── Punishment zone: 1 s pre- vs 1 s post-zone entry bar chart ────────
        if 'epoch_punish_speed_pre_post_entry' in selected_plots and _any_punish_speed:
            _pre_mask_pe  = (EPOCH_CANONICAL_TIME >= -1.0) & (EPOCH_CANONICAL_TIME <  0.0)
            _post_mask_pe = (EPOCH_CANONICAL_TIME >= 0.0)  & (EPOCH_CANONICAL_TIME <= 1.0)
            _ppe_by_cond: dict = {}
            for _r in all_results:
                _sess_mat = _r.get('punish_speed_epoch_session_means')
                if _sess_mat is None or _sess_mat.shape[0] == 0:
                    continue
                _cond = _r['starting_condition']
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore', RuntimeWarning)
                    _pre_per_sess  = np.nanmean(_sess_mat[:, _pre_mask_pe],  axis=1)
                    _post_per_sess = np.nanmean(_sess_mat[:, _post_mask_pe], axis=1)
                    _pre_mean  = float(np.nanmean(_pre_per_sess))
                    _post_mean = float(np.nanmean(_post_per_sess))
                _ppe_by_cond.setdefault(_cond, []).append((_r['mouse'], _pre_mean, _post_mean))

            if _ppe_by_cond:
                _conds_ppe   = sorted(_ppe_by_cond.keys())
                _n_conds_ppe = len(_conds_ppe)
                epoch_punish_speed_pre_post_entry_fig, _axs_ppe = plt.subplots(
                    1, _n_conds_ppe,
                    figsize=(4 * _n_conds_ppe + 1, 5),
                    sharey=True, squeeze=False,
                )
                _all_ppe_yvals = []
                for _ci, _cond in enumerate(_conds_ppe):
                    _ax       = _axs_ppe[0, _ci]
                    _color    = condition_color_map.get(_cond, 'steelblue')
                    _entries  = _ppe_by_cond[_cond]
                    _n_ppe    = len(_entries)
                    _pre_vals = [e[1] for e in _entries]
                    _post_vals= [e[2] for e in _entries]
                    with warnings.catch_warnings():
                        warnings.simplefilter('ignore', RuntimeWarning)
                        _mn_pre   = float(np.nanmean(_pre_vals))
                        _mn_post  = float(np.nanmean(_post_vals))
                        _sem_pre  = (float(np.nanstd(_pre_vals,  ddof=1) / np.sqrt(_n_ppe))
                                     if _n_ppe > 1 else 0.0)
                        _sem_post = (float(np.nanstd(_post_vals, ddof=1) / np.sqrt(_n_ppe))
                                     if _n_ppe > 1 else 0.0)
                    _all_ppe_yvals.extend([_mn_pre + _sem_pre, _mn_post + _sem_post,
                                           _mn_pre - _sem_pre, _mn_post - _sem_post])
                    _all_ppe_yvals.extend(_pre_vals + _post_vals)
                    _ax.bar(0, _mn_pre,  width=0.5, color=_color, alpha=0.7,
                            yerr=_sem_pre,  capsize=7,
                            error_kw={'elinewidth': 1.5, 'capthick': 1.5})
                    _ax.bar(1, _mn_post, width=0.5, color=_color, alpha=0.7,
                            yerr=_sem_post, capsize=7,
                            error_kw={'elinewidth': 1.5, 'capthick': 1.5})
                    _rng_ppe = np.random.default_rng(seed=42)
                    _jitter  = (_rng_ppe.random(_n_ppe) - 0.5) * 0.18
                    for _j, (_mname, _pv, _qv) in enumerate(_entries):
                        _xp = 0 + _jitter[_j]
                        _xq = 1 + _jitter[_j]
                        _ax.plot([_xp, _xq], [_pv, _qv], '-',
                                 color=_color, linewidth=0.9, alpha=0.5, zorder=2)
                        _ax.plot(_xp, _pv, 'o', color='white',
                                 markeredgecolor=_color, markeredgewidth=1.5,
                                 markersize=7, zorder=3)
                        _ax.plot(_xq, _qv, 'o', color='white',
                                 markeredgecolor=_color, markeredgewidth=1.5,
                                 markersize=7, zorder=3)
                    _ax.set_xticks([0, 1])
                    _ax.set_xticklabels(['Pre-entry\n(−1–0 s)', 'Post-entry\n(0–1 s)'],
                                        fontsize=9)
                    _ax.set_title(f'{_cond}\n(n={_n_ppe} mice)', fontsize=10)
                    _ax.set_ylabel('Treadmill Speed (cm/s)' if _ci == 0 else '', fontsize=9)
                    _ax.set_xlim(-0.6, 1.6)
                    _ax.tick_params(axis='both', direction='in')
                    _ax.spines['top'].set_visible(False)
                    _ax.spines['right'].set_visible(False)
                if _all_ppe_yvals:
                    _ymax_ppe = float(np.nanmax(_all_ppe_yvals))
                    _ymin_ppe = float(np.nanmin(_all_ppe_yvals))
                else:
                    _ymax_ppe, _ymin_ppe = 1.0, 0.0
                _bot_ppe = _ymin_ppe * 1.05 if _ymin_ppe < 0 else 0.0
                _axs_ppe[0, 0].set_ylim(_bot_ppe, _ymax_ppe * 1.05)
                epoch_punish_speed_pre_post_entry_fig.suptitle(
                    'Average Speed: 1 s Pre- vs 1 s Post-Zone Entry (Punishment Zone)\n'
                    '(session-averaged punishment zone entry epochs, by condition)',
                    fontsize=12,
                )
                epoch_punish_speed_pre_post_entry_fig.tight_layout()

        # ── Punishment zone: pre-minus-post zone entry difference bar chart ───
        if 'epoch_punish_speed_diff_entry' in selected_plots and _any_punish_speed:
            _pre_mask_de  = (EPOCH_CANONICAL_TIME >= -1.0) & (EPOCH_CANONICAL_TIME <  0.0)
            _post_mask_de = (EPOCH_CANONICAL_TIME >= 0.0)  & (EPOCH_CANONICAL_TIME <= 1.0)
            _pde_by_cond: dict = {}
            for _r in all_results:
                _sess_mat = _r.get('punish_speed_epoch_session_means')
                if _sess_mat is None or _sess_mat.shape[0] == 0:
                    continue
                _cond = _r['starting_condition']
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore', RuntimeWarning)
                    _pre_de  = float(np.nanmean(np.nanmean(_sess_mat[:, _pre_mask_de],  axis=1)))
                    _post_de = float(np.nanmean(np.nanmean(_sess_mat[:, _post_mask_de], axis=1)))
                _pde_by_cond.setdefault(_cond, []).append((_r['mouse'], _post_de - _pre_de))

            if _pde_by_cond:
                _conds_de   = sorted(_pde_by_cond.keys())
                _n_conds_de = len(_conds_de)
                epoch_punish_speed_diff_entry_fig, _ax_de = plt.subplots(
                    1, 1, figsize=(max(4, _n_conds_de * 1.4 + 1.5), 5)
                )
                _all_pde_vals = []
                _bar_x_de = np.arange(_n_conds_de)
                _rng_de   = np.random.default_rng(seed=42)
                for _ci, _cond in enumerate(_conds_de):
                    _color    = condition_color_map.get(_cond, 'steelblue')
                    _entries  = _pde_by_cond[_cond]
                    _n_de     = len(_entries)
                    _dvals_de = [e[1] for e in _entries]
                    with warnings.catch_warnings():
                        warnings.simplefilter('ignore', RuntimeWarning)
                        _mn_de  = float(np.nanmean(_dvals_de))
                        _sem_de = (float(np.nanstd(_dvals_de, ddof=1) / np.sqrt(_n_de))
                                   if _n_de > 1 else 0.0)
                    _all_pde_vals.extend(_dvals_de + [_mn_de + _sem_de, _mn_de - _sem_de])
                    _ax_de.bar(_ci, _mn_de, width=0.55, color=_color, alpha=0.7,
                               yerr=_sem_de, capsize=7,
                               error_kw={'elinewidth': 1.5, 'capthick': 1.5})
                    _jitter_de = (_rng_de.random(_n_de) - 0.5) * 0.22
                    for _j, (_mname, _dv) in enumerate(_entries):
                        _ax_de.plot(_ci + _jitter_de[_j], _dv, 'o',
                                    color='white', markeredgecolor=_color,
                                    markeredgewidth=1.5, markersize=7, zorder=3)
                _ax_de.axhline(0, color='black', linewidth=0.9, linestyle='--', zorder=1)
                _ax_de.set_xticks(_bar_x_de)
                _ax_de.set_xticklabels(_conds_de, fontsize=10)
                _ax_de.set_ylabel('Speed difference (cm/s)\n[post-entry \u2212 pre-entry]', fontsize=9)
                _ax_de.set_xlabel('Condition', fontsize=10)
                if _all_pde_vals:
                    _ymax_de = float(np.nanmax(_all_pde_vals))
                    _ymin_de = float(np.nanmin(_all_pde_vals))
                    _pad_de  = max(abs(_ymax_de), abs(_ymin_de)) * 0.12 or 0.5
                    _ax_de.set_ylim(_ymin_de - _pad_de, _ymax_de + _pad_de)
                _ax_de.tick_params(axis='both', direction='in')
                _ax_de.spines['top'].set_visible(False)
                _ax_de.spines['right'].set_visible(False)

                # ── Mann-Whitney U tests for all condition pairs ───────────
                import itertools as _itertools_de
                _pairs_de = list(_itertools_de.combinations(range(_n_conds_de), 2))
                _ylim_cur_de = list(_ax_de.get_ylim())
                _bracket_step_de = (_ylim_cur_de[1] - _ylim_cur_de[0]) * 0.14
                _bracket_base_de = _ylim_cur_de[1]
                for _pi, (_ia, _ib) in enumerate(_pairs_de):
                    _vals_a_de = [e[1] for e in _pde_by_cond[_conds_de[_ia]]]
                    _vals_b_de = [e[1] for e in _pde_by_cond[_conds_de[_ib]]]
                    if len(_vals_a_de) < 2 or len(_vals_b_de) < 2:
                        continue
                    _u_stat_de, _p_val_de = mannwhitneyu(_vals_a_de, _vals_b_de, alternative='two-sided')
                    if _p_val_de < 0.001:
                        _sig_str_de = f'p = {_p_val_de:.2e}***'
                    elif _p_val_de < 0.01:
                        _sig_str_de = f'p = {_p_val_de:.3f}**'
                    elif _p_val_de < 0.05:
                        _sig_str_de = f'p = {_p_val_de:.3f}*'
                    else:
                        _sig_str_de = f'p = {_p_val_de:.3f} (ns)'
                    _bh_de = _bracket_base_de + _bracket_step_de * (_pi + 0.6)
                    _ax_de.plot([_ia, _ia, _ib, _ib],
                                [_bh_de - _bracket_step_de * 0.15,
                                 _bh_de,
                                 _bh_de,
                                 _bh_de - _bracket_step_de * 0.15],
                                color='black', linewidth=1.0)
                    _ax_de.text((_ia + _ib) / 2, _bh_de + _bracket_step_de * 0.05,
                                _sig_str_de, ha='center', va='bottom', fontsize=8)
                if _pairs_de:
                    _new_top_de = _bracket_base_de + _bracket_step_de * (len(_pairs_de) + 1.5)
                    _ax_de.set_ylim(_ylim_cur_de[0], _new_top_de)

                epoch_punish_speed_diff_entry_fig.suptitle(
                    'Post- vs Pre-Zone Entry Speed Difference by Condition (Punishment Zone)\n'
                    '(mean ± SEM across mice; positive = faster after zone entry)',
                    fontsize=11,
                )
                epoch_punish_speed_diff_entry_fig.tight_layout()

        # ── Punishment zone: pre/post 0.65 s cutoff capacitive bar chart ─────
        if 'epoch_punish_cap_pre_post' in selected_plots and _any_punish_cap:
            _pcp_pre_mask  = (EPOCH_CANONICAL_TIME >= 0.0)  & (EPOCH_CANONICAL_TIME <= 0.65)
            _pcp_post_mask = (EPOCH_CANONICAL_TIME >  0.65) & (EPOCH_CANONICAL_TIME <= 1.30)
            _pcp_by_cond: dict = {}
            for _r in all_results:
                _sess_mat = _r.get('punish_cap_epoch_session_means')
                if _sess_mat is None or _sess_mat.shape[0] == 0:
                    continue
                _cond = _r['starting_condition']
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore', RuntimeWarning)
                    _pre_per_sess  = np.nanmean(_sess_mat[:, _pcp_pre_mask],  axis=1)
                    _post_per_sess = np.nanmean(_sess_mat[:, _pcp_post_mask], axis=1)
                    _pre_mean  = float(np.nanmean(_pre_per_sess))
                    _post_mean = float(np.nanmean(_post_per_sess))
                _pcp_by_cond.setdefault(_cond, []).append((_r['mouse'], _pre_mean, _post_mean))

            if _pcp_by_cond:
                _conds_pcp   = sorted(_pcp_by_cond.keys())
                _n_conds_pcp = len(_conds_pcp)
                epoch_punish_cap_pre_post_fig, _axs_pcp = plt.subplots(
                    1, _n_conds_pcp,
                    figsize=(4 * _n_conds_pcp + 1, 5),
                    sharey=True, squeeze=False,
                )
                _all_pcp_yvals = []
                for _ci, _cond in enumerate(_conds_pcp):
                    _ax       = _axs_pcp[0, _ci]
                    _color    = condition_color_map.get(_cond, 'steelblue')
                    _entries  = _pcp_by_cond[_cond]
                    _n_pcp    = len(_entries)
                    _pre_vals = [e[1] for e in _entries]
                    _post_vals= [e[2] for e in _entries]
                    with warnings.catch_warnings():
                        warnings.simplefilter('ignore', RuntimeWarning)
                        _mn_pre   = float(np.nanmean(_pre_vals))
                        _mn_post  = float(np.nanmean(_post_vals))
                        _sem_pre  = (float(np.nanstd(_pre_vals,  ddof=1) / np.sqrt(_n_pcp))
                                     if _n_pcp > 1 else 0.0)
                        _sem_post = (float(np.nanstd(_post_vals, ddof=1) / np.sqrt(_n_pcp))
                                     if _n_pcp > 1 else 0.0)
                    _all_pcp_yvals.extend([_mn_pre + _sem_pre, _mn_post + _sem_post,
                                           _mn_pre - _sem_pre, _mn_post - _sem_post])
                    _all_pcp_yvals.extend(_pre_vals + _post_vals)
                    _ax.bar(0, _mn_pre,  width=0.5, color=_color, alpha=0.7,
                            yerr=_sem_pre,  capsize=7,
                            error_kw={'elinewidth': 1.5, 'capthick': 1.5})
                    _ax.bar(1, _mn_post, width=0.5, color=_color, alpha=0.7,
                            yerr=_sem_post, capsize=7,
                            error_kw={'elinewidth': 1.5, 'capthick': 1.5})
                    _rng_pcp = np.random.default_rng(seed=42)
                    _jitter_pcp = (_rng_pcp.random(_n_pcp) - 0.5) * 0.18
                    for _j, (_mname, _pv, _qv) in enumerate(_entries):
                        _xp = 0 + _jitter_pcp[_j]
                        _xq = 1 + _jitter_pcp[_j]
                        _ax.plot([_xp, _xq], [_pv, _qv], '-',
                                 color=_color, linewidth=0.9, alpha=0.5, zorder=2)
                        _ax.plot(_xp, _pv, 'o', color='white',
                                 markeredgecolor=_color, markeredgewidth=1.5,
                                 markersize=7, zorder=3)
                        _ax.plot(_xq, _qv, 'o', color='white',
                                 markeredgecolor=_color, markeredgewidth=1.5,
                                 markersize=7, zorder=3)
                    _ax.set_xticks([0, 1])
                    _ax.set_xticklabels(['Pre-cutoff\n(0\u20130.65 s)', 'Post-cutoff\n(0.65\u20131.3 s)'],
                                        fontsize=9)
                    _ax.set_title(f'{_cond}\n(n={_n_pcp} mice)', fontsize=10)
                    _ax.set_ylabel('Capacitive Sensor (z-score)' if _ci == 0 else '', fontsize=9)
                    _ax.set_xlim(-0.6, 1.6)
                    _ax.axhline(0, color='black', linewidth=0.8, linestyle='--', zorder=1)
                    _ax.tick_params(axis='both', direction='in')
                    _ax.spines['top'].set_visible(False)
                    _ax.spines['right'].set_visible(False)
                if _all_pcp_yvals:
                    _ymax_pcp = float(np.nanmax(_all_pcp_yvals))
                    _ymin_pcp = float(np.nanmin(_all_pcp_yvals))
                else:
                    _ymax_pcp, _ymin_pcp = 1.0, -1.0
                _pad_pcp = max(abs(_ymax_pcp), abs(_ymin_pcp)) * 0.12 or 0.1
                _axs_pcp[0, 0].set_ylim(_ymin_pcp - _pad_pcp, _ymax_pcp + _pad_pcp)
                epoch_punish_cap_pre_post_fig.suptitle(
                    'Average Capacitive Sensor (z-scored): Pre- vs Post-Cutoff (Punishment Zone)\n'
                    '(session-averaged punishment zone entry epochs, by condition)',
                    fontsize=12,
                )
                epoch_punish_cap_pre_post_fig.tight_layout()

        # ── Punishment zone: pre-minus-post 0.65 s capacitive difference ─────
        if 'epoch_punish_cap_diff' in selected_plots and _any_punish_cap:
            _pcd_pre_mask  = (EPOCH_CANONICAL_TIME >= 0.0)  & (EPOCH_CANONICAL_TIME <= 0.65)
            _pcd_post_mask = (EPOCH_CANONICAL_TIME >  0.65) & (EPOCH_CANONICAL_TIME <= 1.30)
            _pcdiff_by_cond: dict = {}
            for _r in all_results:
                _sess_mat = _r.get('punish_cap_epoch_session_means')
                if _sess_mat is None or _sess_mat.shape[0] == 0:
                    continue
                _cond = _r['starting_condition']
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore', RuntimeWarning)
                    _pre_pcd  = float(np.nanmean(np.nanmean(_sess_mat[:, _pcd_pre_mask],  axis=1)))
                    _post_pcd = float(np.nanmean(np.nanmean(_sess_mat[:, _pcd_post_mask], axis=1)))
                _pcdiff_by_cond.setdefault(_cond, []).append((_r['mouse'], _post_pcd - _pre_pcd))

            if _pcdiff_by_cond:
                _conds_pcd   = sorted(_pcdiff_by_cond.keys())
                _n_conds_pcd = len(_conds_pcd)
                epoch_punish_cap_diff_fig, _ax_pcd = plt.subplots(
                    1, 1, figsize=(max(4, _n_conds_pcd * 1.4 + 1.5), 5)
                )
                _all_pcdiff_vals = []
                _bar_x_pcd = np.arange(_n_conds_pcd)
                _rng_pcd   = np.random.default_rng(seed=42)
                for _ci, _cond in enumerate(_conds_pcd):
                    _color    = condition_color_map.get(_cond, 'steelblue')
                    _entries  = _pcdiff_by_cond[_cond]
                    _n_pcd    = len(_entries)
                    _dvals_pcd = [e[1] for e in _entries]
                    with warnings.catch_warnings():
                        warnings.simplefilter('ignore', RuntimeWarning)
                        _mn_pcd  = float(np.nanmean(_dvals_pcd))
                        _sem_pcd = (float(np.nanstd(_dvals_pcd, ddof=1) / np.sqrt(_n_pcd))
                                    if _n_pcd > 1 else 0.0)
                    _all_pcdiff_vals.extend(_dvals_pcd + [_mn_pcd + _sem_pcd, _mn_pcd - _sem_pcd])
                    _ax_pcd.bar(_ci, _mn_pcd, width=0.55, color=_color, alpha=0.7,
                                yerr=_sem_pcd, capsize=7,
                                error_kw={'elinewidth': 1.5, 'capthick': 1.5})
                    _jitter_pcd = (_rng_pcd.random(_n_pcd) - 0.5) * 0.22
                    for _j, (_mname, _dv) in enumerate(_entries):
                        _ax_pcd.plot(_ci + _jitter_pcd[_j], _dv, 'o',
                                     color='white', markeredgecolor=_color,
                                     markeredgewidth=1.5, markersize=7, zorder=3)
                _ax_pcd.axhline(0, color='black', linewidth=0.9, linestyle='--', zorder=1)
                _ax_pcd.set_xticks(_bar_x_pcd)
                _ax_pcd.set_xticklabels(_conds_pcd, fontsize=10)
                _ax_pcd.set_ylabel('Capacitive difference (z-score)\n[post-cutoff \u2212 pre-cutoff]', fontsize=9)
                _ax_pcd.set_xlabel('Condition', fontsize=10)
                if _all_pcdiff_vals:
                    _ymax_pcd = float(np.nanmax(_all_pcdiff_vals))
                    _ymin_pcd = float(np.nanmin(_all_pcdiff_vals))
                    _pad_pcd  = max(abs(_ymax_pcd), abs(_ymin_pcd)) * 0.12 or 0.1
                    _ax_pcd.set_ylim(_ymin_pcd - _pad_pcd, _ymax_pcd + _pad_pcd)
                _ax_pcd.tick_params(axis='both', direction='in')
                _ax_pcd.spines['top'].set_visible(False)
                _ax_pcd.spines['right'].set_visible(False)

                import itertools as _itertools_pcd
                _pairs_pcd = list(_itertools_pcd.combinations(range(_n_conds_pcd), 2))
                _ylim_cur_pcd = list(_ax_pcd.get_ylim())
                _bracket_step_pcd = (_ylim_cur_pcd[1] - _ylim_cur_pcd[0]) * 0.14
                _bracket_base_pcd = _ylim_cur_pcd[1]
                for _pi, (_ia, _ib) in enumerate(_pairs_pcd):
                    _vals_a_pcd = [e[1] for e in _pcdiff_by_cond[_conds_pcd[_ia]]]
                    _vals_b_pcd = [e[1] for e in _pcdiff_by_cond[_conds_pcd[_ib]]]
                    if len(_vals_a_pcd) < 2 or len(_vals_b_pcd) < 2:
                        continue
                    _u_pcd, _p_pcd = mannwhitneyu(_vals_a_pcd, _vals_b_pcd, alternative='two-sided')
                    if _p_pcd < 0.001:
                        _sig_pcd = f'p = {_p_pcd:.2e}***'
                    elif _p_pcd < 0.01:
                        _sig_pcd = f'p = {_p_pcd:.3f}**'
                    elif _p_pcd < 0.05:
                        _sig_pcd = f'p = {_p_pcd:.3f}*'
                    else:
                        _sig_pcd = f'p = {_p_pcd:.3f} (ns)'
                    _bh_pcd = _bracket_base_pcd + _bracket_step_pcd * (_pi + 0.6)
                    _ax_pcd.plot([_ia, _ia, _ib, _ib],
                                 [_bh_pcd - _bracket_step_pcd * 0.15,
                                  _bh_pcd, _bh_pcd,
                                  _bh_pcd - _bracket_step_pcd * 0.15],
                                 color='black', linewidth=1.0)
                    _ax_pcd.text((_ia + _ib) / 2, _bh_pcd + _bracket_step_pcd * 0.05,
                                 _sig_pcd, ha='center', va='bottom', fontsize=8)
                if _pairs_pcd:
                    _new_top_pcd = _bracket_base_pcd + _bracket_step_pcd * (len(_pairs_pcd) + 1.5)
                    _ax_pcd.set_ylim(_ylim_cur_pcd[0], _new_top_pcd)

                epoch_punish_cap_diff_fig.suptitle(
                    'Post- vs Pre-Cutoff Capacitive (z-scored) Difference by Condition (Punishment Zone)\n'
                    '(mean \u00b1 SEM across mice; positive = higher licking after 0.65 s cutoff; Mann-Whitney U)',
                    fontsize=11,
                )
                epoch_punish_cap_diff_fig.tight_layout()

        # ── Punishment zone: 1 s pre- vs 1 s post-zone entry cap bar chart ───
        if 'epoch_punish_cap_pre_post_entry' in selected_plots and _any_punish_cap:
            _pcpe_pre_mask  = (EPOCH_CANONICAL_TIME >= -1.0) & (EPOCH_CANONICAL_TIME <  0.0)
            _pcpe_post_mask = (EPOCH_CANONICAL_TIME >= 0.0)  & (EPOCH_CANONICAL_TIME <= 1.0)
            _pcpe_by_cond: dict = {}
            for _r in all_results:
                _sess_mat = _r.get('punish_cap_epoch_session_means')
                if _sess_mat is None or _sess_mat.shape[0] == 0:
                    continue
                _cond = _r['starting_condition']
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore', RuntimeWarning)
                    _pre_per_sess  = np.nanmean(_sess_mat[:, _pcpe_pre_mask],  axis=1)
                    _post_per_sess = np.nanmean(_sess_mat[:, _pcpe_post_mask], axis=1)
                    _pre_mean  = float(np.nanmean(_pre_per_sess))
                    _post_mean = float(np.nanmean(_post_per_sess))
                _pcpe_by_cond.setdefault(_cond, []).append((_r['mouse'], _pre_mean, _post_mean))

            if _pcpe_by_cond:
                _conds_pcpe   = sorted(_pcpe_by_cond.keys())
                _n_conds_pcpe = len(_conds_pcpe)
                epoch_punish_cap_pre_post_entry_fig, _axs_pcpe = plt.subplots(
                    1, _n_conds_pcpe,
                    figsize=(4 * _n_conds_pcpe + 1, 5),
                    sharey=True, squeeze=False,
                )
                _all_pcpe_yvals = []
                for _ci, _cond in enumerate(_conds_pcpe):
                    _ax       = _axs_pcpe[0, _ci]
                    _color    = condition_color_map.get(_cond, 'steelblue')
                    _entries  = _pcpe_by_cond[_cond]
                    _n_pcpe   = len(_entries)
                    _pre_vals = [e[1] for e in _entries]
                    _post_vals= [e[2] for e in _entries]
                    with warnings.catch_warnings():
                        warnings.simplefilter('ignore', RuntimeWarning)
                        _mn_pre   = float(np.nanmean(_pre_vals))
                        _mn_post  = float(np.nanmean(_post_vals))
                        _sem_pre  = (float(np.nanstd(_pre_vals,  ddof=1) / np.sqrt(_n_pcpe))
                                     if _n_pcpe > 1 else 0.0)
                        _sem_post = (float(np.nanstd(_post_vals, ddof=1) / np.sqrt(_n_pcpe))
                                     if _n_pcpe > 1 else 0.0)
                    _all_pcpe_yvals.extend([_mn_pre + _sem_pre, _mn_post + _sem_post,
                                            _mn_pre - _sem_pre, _mn_post - _sem_post])
                    _all_pcpe_yvals.extend(_pre_vals + _post_vals)
                    _ax.bar(0, _mn_pre,  width=0.5, color=_color, alpha=0.7,
                            yerr=_sem_pre,  capsize=7,
                            error_kw={'elinewidth': 1.5, 'capthick': 1.5})
                    _ax.bar(1, _mn_post, width=0.5, color=_color, alpha=0.7,
                            yerr=_sem_post, capsize=7,
                            error_kw={'elinewidth': 1.5, 'capthick': 1.5})
                    _rng_pcpe = np.random.default_rng(seed=42)
                    _jitter_pcpe = (_rng_pcpe.random(_n_pcpe) - 0.5) * 0.18
                    for _j, (_mname, _pv, _qv) in enumerate(_entries):
                        _xp = 0 + _jitter_pcpe[_j]
                        _xq = 1 + _jitter_pcpe[_j]
                        _ax.plot([_xp, _xq], [_pv, _qv], '-',
                                 color=_color, linewidth=0.9, alpha=0.5, zorder=2)
                        _ax.plot(_xp, _pv, 'o', color='white',
                                 markeredgecolor=_color, markeredgewidth=1.5,
                                 markersize=7, zorder=3)
                        _ax.plot(_xq, _qv, 'o', color='white',
                                 markeredgecolor=_color, markeredgewidth=1.5,
                                 markersize=7, zorder=3)
                    _ax.set_xticks([0, 1])
                    _ax.set_xticklabels(['Pre-entry\n(\u22121\u20130 s)', 'Post-entry\n(0\u20131 s)'],
                                        fontsize=9)
                    _ax.set_title(f'{_cond}\n(n={_n_pcpe} mice)', fontsize=10)
                    _ax.set_ylabel('Capacitive Sensor (z-score)' if _ci == 0 else '', fontsize=9)
                    _ax.set_xlim(-0.6, 1.6)
                    _ax.axhline(0, color='black', linewidth=0.8, linestyle='--', zorder=1)
                    _ax.tick_params(axis='both', direction='in')
                    _ax.spines['top'].set_visible(False)
                    _ax.spines['right'].set_visible(False)
                if _all_pcpe_yvals:
                    _ymax_pcpe = float(np.nanmax(_all_pcpe_yvals))
                    _ymin_pcpe = float(np.nanmin(_all_pcpe_yvals))
                else:
                    _ymax_pcpe, _ymin_pcpe = 1.0, -1.0
                _pad_pcpe = max(abs(_ymax_pcpe), abs(_ymin_pcpe)) * 0.12 or 0.1
                _axs_pcpe[0, 0].set_ylim(_ymin_pcpe - _pad_pcpe, _ymax_pcpe + _pad_pcpe)
                epoch_punish_cap_pre_post_entry_fig.suptitle(
                    'Average Capacitive Sensor (z-scored): 1 s Pre- vs 1 s Post-Zone Entry (Punishment Zone)\n'
                    '(session-averaged punishment zone entry epochs, by condition)',
                    fontsize=12,
                )
                epoch_punish_cap_pre_post_entry_fig.tight_layout()

        # ── Punishment zone: pre-minus-post zone entry cap difference ─────────
        if 'epoch_punish_cap_diff_entry' in selected_plots and _any_punish_cap:
            _pcde_pre_mask  = (EPOCH_CANONICAL_TIME >= -1.0) & (EPOCH_CANONICAL_TIME <  0.0)
            _pcde_post_mask = (EPOCH_CANONICAL_TIME >= 0.0)  & (EPOCH_CANONICAL_TIME <= 1.0)
            _pcdentry_by_cond: dict = {}
            for _r in all_results:
                _sess_mat = _r.get('punish_cap_epoch_session_means')
                if _sess_mat is None or _sess_mat.shape[0] == 0:
                    continue
                _cond = _r['starting_condition']
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore', RuntimeWarning)
                    _pre_pcde  = float(np.nanmean(np.nanmean(_sess_mat[:, _pcde_pre_mask],  axis=1)))
                    _post_pcde = float(np.nanmean(np.nanmean(_sess_mat[:, _pcde_post_mask], axis=1)))
                _pcdentry_by_cond.setdefault(_cond, []).append((_r['mouse'], _post_pcde - _pre_pcde))

            if _pcdentry_by_cond:
                _conds_pcde   = sorted(_pcdentry_by_cond.keys())
                _n_conds_pcde = len(_conds_pcde)
                epoch_punish_cap_diff_entry_fig, _ax_pcde = plt.subplots(
                    1, 1, figsize=(max(4, _n_conds_pcde * 1.4 + 1.5), 5)
                )
                _all_pcde_vals = []
                _bar_x_pcde = np.arange(_n_conds_pcde)
                _rng_pcde   = np.random.default_rng(seed=42)
                for _ci, _cond in enumerate(_conds_pcde):
                    _color    = condition_color_map.get(_cond, 'steelblue')
                    _entries  = _pcdentry_by_cond[_cond]
                    _n_pcde   = len(_entries)
                    _dvals_pcde = [e[1] for e in _entries]
                    with warnings.catch_warnings():
                        warnings.simplefilter('ignore', RuntimeWarning)
                        _mn_pcde  = float(np.nanmean(_dvals_pcde))
                        _sem_pcde = (float(np.nanstd(_dvals_pcde, ddof=1) / np.sqrt(_n_pcde))
                                     if _n_pcde > 1 else 0.0)
                    _all_pcde_vals.extend(_dvals_pcde + [_mn_pcde + _sem_pcde, _mn_pcde - _sem_pcde])
                    _ax_pcde.bar(_ci, _mn_pcde, width=0.55, color=_color, alpha=0.7,
                                 yerr=_sem_pcde, capsize=7,
                                 error_kw={'elinewidth': 1.5, 'capthick': 1.5})
                    _jitter_pcde = (_rng_pcde.random(_n_pcde) - 0.5) * 0.22
                    for _j, (_mname, _dv) in enumerate(_entries):
                        _ax_pcde.plot(_ci + _jitter_pcde[_j], _dv, 'o',
                                      color='white', markeredgecolor=_color,
                                      markeredgewidth=1.5, markersize=7, zorder=3)
                _ax_pcde.axhline(0, color='black', linewidth=0.9, linestyle='--', zorder=1)
                _ax_pcde.set_xticks(_bar_x_pcde)
                _ax_pcde.set_xticklabels(_conds_pcde, fontsize=10)
                _ax_pcde.set_ylabel('Capacitive difference (z-score)\n[post-entry \u2212 pre-entry]', fontsize=9)
                _ax_pcde.set_xlabel('Condition', fontsize=10)
                if _all_pcde_vals:
                    _ymax_pcde = float(np.nanmax(_all_pcde_vals))
                    _ymin_pcde = float(np.nanmin(_all_pcde_vals))
                    _pad_pcde  = max(abs(_ymax_pcde), abs(_ymin_pcde)) * 0.12 or 0.1
                    _ax_pcde.set_ylim(_ymin_pcde - _pad_pcde, _ymax_pcde + _pad_pcde)
                _ax_pcde.tick_params(axis='both', direction='in')
                _ax_pcde.spines['top'].set_visible(False)
                _ax_pcde.spines['right'].set_visible(False)

                import itertools as _itertools_pcde
                _pairs_pcde = list(_itertools_pcde.combinations(range(_n_conds_pcde), 2))
                _ylim_cur_pcde = list(_ax_pcde.get_ylim())
                _bracket_step_pcde = (_ylim_cur_pcde[1] - _ylim_cur_pcde[0]) * 0.14
                _bracket_base_pcde = _ylim_cur_pcde[1]
                for _pi, (_ia, _ib) in enumerate(_pairs_pcde):
                    _vals_a_pcde = [e[1] for e in _pcdentry_by_cond[_conds_pcde[_ia]]]
                    _vals_b_pcde = [e[1] for e in _pcdentry_by_cond[_conds_pcde[_ib]]]
                    if len(_vals_a_pcde) < 2 or len(_vals_b_pcde) < 2:
                        continue
                    _u_pcde, _p_pcde = mannwhitneyu(_vals_a_pcde, _vals_b_pcde, alternative='two-sided')
                    if _p_pcde < 0.001:
                        _sig_pcde = f'p = {_p_pcde:.2e}***'
                    elif _p_pcde < 0.01:
                        _sig_pcde = f'p = {_p_pcde:.3f}**'
                    elif _p_pcde < 0.05:
                        _sig_pcde = f'p = {_p_pcde:.3f}*'
                    else:
                        _sig_pcde = f'p = {_p_pcde:.3f} (ns)'
                    _bh_pcde = _bracket_base_pcde + _bracket_step_pcde * (_pi + 0.6)
                    _ax_pcde.plot([_ia, _ia, _ib, _ib],
                                  [_bh_pcde - _bracket_step_pcde * 0.15,
                                   _bh_pcde, _bh_pcde,
                                   _bh_pcde - _bracket_step_pcde * 0.15],
                                  color='black', linewidth=1.0)
                    _ax_pcde.text((_ia + _ib) / 2, _bh_pcde + _bracket_step_pcde * 0.05,
                                  _sig_pcde, ha='center', va='bottom', fontsize=8)
                if _pairs_pcde:
                    _new_top_pcde = _bracket_base_pcde + _bracket_step_pcde * (len(_pairs_pcde) + 1.5)
                    _ax_pcde.set_ylim(_ylim_cur_pcde[0], _new_top_pcde)

                epoch_punish_cap_diff_entry_fig.suptitle(
                    'Post- vs Pre-Zone Entry Capacitive (z-scored) Difference by Condition (Punishment Zone)\n'
                    '(mean \u00b1 SEM across mice; positive = higher licking after zone entry; Mann-Whitney U)',
                    fontsize=11,
                )
                epoch_punish_cap_diff_entry_fig.tight_layout()

        # ── Pre/post-reward delivery speed bar chart ───────────────────────
        if 'epoch_reward_speed_pre_post' in selected_plots and _any_speed:
            _pre_mask  = (EPOCH_CANONICAL_TIME >= 0.0)  & (EPOCH_CANONICAL_TIME <= 0.65)
            _post_mask = (EPOCH_CANONICAL_TIME >  0.65) & (EPOCH_CANONICAL_TIME <= 1.30)
            # Collect per-mouse session-averaged bin means, grouped by condition
            _pp_by_cond: dict = {}
            for _r in all_results:
                _sess_mat = _r.get('speed_epoch_session_means')
                if _sess_mat is None or _sess_mat.shape[0] == 0:
                    continue
                _cond = _r['starting_condition']
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore', RuntimeWarning)
                    _pre_per_sess  = np.nanmean(_sess_mat[:, _pre_mask],  axis=1)
                    _post_per_sess = np.nanmean(_sess_mat[:, _post_mask], axis=1)
                    _pre_mean  = float(np.nanmean(_pre_per_sess))
                    _post_mean = float(np.nanmean(_post_per_sess))
                _pp_by_cond.setdefault(_cond, []).append((_r['mouse'], _pre_mean, _post_mean))

            if _pp_by_cond:
                _conds_pp   = sorted(_pp_by_cond.keys())
                _n_conds_pp = len(_conds_pp)
                epoch_reward_speed_pre_post_fig, _axs_pp = plt.subplots(
                    1, _n_conds_pp,
                    figsize=(4 * _n_conds_pp + 1, 5),
                    sharey=True, squeeze=False,
                )
                _all_pp_yvals = []
                for _ci, _cond in enumerate(_conds_pp):
                    _ax      = _axs_pp[0, _ci]
                    _color   = condition_color_map.get(_cond, 'steelblue')
                    _entries = _pp_by_cond[_cond]
                    _n_pp    = len(_entries)
                    _pre_vals  = [e[1] for e in _entries]
                    _post_vals = [e[2] for e in _entries]
                    with warnings.catch_warnings():
                        warnings.simplefilter('ignore', RuntimeWarning)
                        _mn_pre   = float(np.nanmean(_pre_vals))
                        _mn_post  = float(np.nanmean(_post_vals))
                        _sem_pre  = (float(np.nanstd(_pre_vals,  ddof=1) / np.sqrt(_n_pp))
                                     if _n_pp > 1 else 0.0)
                        _sem_post = (float(np.nanstd(_post_vals, ddof=1) / np.sqrt(_n_pp))
                                     if _n_pp > 1 else 0.0)
                    _all_pp_yvals.extend([_mn_pre + _sem_pre, _mn_post + _sem_post,
                                          _mn_pre - _sem_pre, _mn_post - _sem_post])
                    _all_pp_yvals.extend(_pre_vals + _post_vals)
                    _ax.bar(0, _mn_pre,  width=0.5, color=_color, alpha=0.7,
                            yerr=_sem_pre,  capsize=7,
                            error_kw={'elinewidth': 1.5, 'capthick': 1.5})
                    _ax.bar(1, _mn_post, width=0.5, color=_color, alpha=0.7,
                            yerr=_sem_post, capsize=7,
                            error_kw={'elinewidth': 1.5, 'capthick': 1.5})
                    _rng_pp = np.random.default_rng(seed=42)
                    _jitter = (_rng_pp.random(_n_pp) - 0.5) * 0.18
                    for _j, (_mname, _pv, _qv) in enumerate(_entries):
                        _xp = 0 + _jitter[_j]
                        _xq = 1 + _jitter[_j]
                        _ax.plot([_xp, _xq], [_pv, _qv], '-',
                                 color=_color, linewidth=0.9, alpha=0.5, zorder=2)
                        _ax.plot(_xp, _pv, 'o', color='white',
                                 markeredgecolor=_color, markeredgewidth=1.5,
                                 markersize=7, zorder=3)
                        _ax.plot(_xq, _qv, 'o', color='white',
                                 markeredgecolor=_color, markeredgewidth=1.5,
                                 markersize=7, zorder=3)
                    _ax.set_xticks([0, 1])
                    _ax.set_xticklabels(['Pre-reward\n(0\u20130.65 s)', 'Post-reward\n(0.65\u20131.3 s)'],
                                        fontsize=9)
                    _ax.set_title(f'{_cond}\n(n={_n_pp} mice)', fontsize=10)
                    _ax.set_ylabel('Treadmill Speed (cm/s)' if _ci == 0 else '', fontsize=9)
                    _ax.set_xlim(-0.6, 1.6)
                    _ax.tick_params(axis='both', direction='in')
                    _ax.spines['top'].set_visible(False)
                    _ax.spines['right'].set_visible(False)
                if _all_pp_yvals:
                    _ymax_pp = float(np.nanmax(_all_pp_yvals))
                    _ymin_pp = float(np.nanmin(_all_pp_yvals))
                else:
                    _ymax_pp, _ymin_pp = 1.0, 0.0
                _bot_pp = _ymin_pp * 1.05 if _ymin_pp < 0 else 0.0
                _axs_pp[0, 0].set_ylim(_bot_pp, _ymax_pp * 1.05)
                epoch_reward_speed_pre_post_fig.suptitle(
                    'Average Speed: Pre- vs Post-Reward Delivery\n'
                    '(session-averaged reward zone entry epochs, by condition)',
                    fontsize=12,
                )
                epoch_reward_speed_pre_post_fig.tight_layout()

        # ── Pre-minus-post reward speed difference bar chart ──────────────────
        if 'epoch_reward_speed_diff' in selected_plots and _any_speed:
            _pre_mask_d  = (EPOCH_CANONICAL_TIME >= 0.0)  & (EPOCH_CANONICAL_TIME <= 0.65)
            _post_mask_d = (EPOCH_CANONICAL_TIME >  0.65) & (EPOCH_CANONICAL_TIME <= 1.30)
            # condition -> list of (mouse_name, diff_value)
            _diff_by_cond: dict = {}
            for _r in all_results:
                _sess_mat = _r.get('speed_epoch_session_means')
                if _sess_mat is None or _sess_mat.shape[0] == 0:
                    continue
                _cond = _r['starting_condition']
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore', RuntimeWarning)
                    _pre_d  = float(np.nanmean(np.nanmean(_sess_mat[:, _pre_mask_d],  axis=1)))
                    _post_d = float(np.nanmean(np.nanmean(_sess_mat[:, _post_mask_d], axis=1)))
                _diff_by_cond.setdefault(_cond, []).append((_r['mouse'], _post_d - _pre_d))

            if _diff_by_cond:
                _conds_d   = sorted(_diff_by_cond.keys())
                _n_conds_d = len(_conds_d)
                epoch_reward_speed_diff_fig, _ax_d = plt.subplots(
                    1, 1, figsize=(max(4, _n_conds_d * 1.4 + 1.5), 5)
                )
                _all_diff_vals = []
                _bar_x = np.arange(_n_conds_d)
                _rng_d = np.random.default_rng(seed=42)
                for _ci, _cond in enumerate(_conds_d):
                    _color   = condition_color_map.get(_cond, 'steelblue')
                    _entries = _diff_by_cond[_cond]
                    _n_d     = len(_entries)
                    _dvals   = [e[1] for e in _entries]
                    with warnings.catch_warnings():
                        warnings.simplefilter('ignore', RuntimeWarning)
                        _mn_d  = float(np.nanmean(_dvals))
                        _sem_d = (float(np.nanstd(_dvals, ddof=1) / np.sqrt(_n_d))
                                  if _n_d > 1 else 0.0)
                    _all_diff_vals.extend(_dvals + [_mn_d + _sem_d, _mn_d - _sem_d])
                    _ax_d.bar(_ci, _mn_d, width=0.55, color=_color, alpha=0.7,
                              yerr=_sem_d, capsize=7,
                              error_kw={'elinewidth': 1.5, 'capthick': 1.5})
                    _jitter_d = (_rng_d.random(_n_d) - 0.5) * 0.22
                    for _j, (_mname, _dv) in enumerate(_entries):
                        _ax_d.plot(_ci + _jitter_d[_j], _dv, 'o',
                                   color='white', markeredgecolor=_color,
                                   markeredgewidth=1.5, markersize=7, zorder=3)
                _ax_d.axhline(0, color='black', linewidth=0.9, linestyle='--', zorder=1)
                _ax_d.set_xticks(_bar_x)
                _ax_d.set_xticklabels(_conds_d, fontsize=10)
                _ax_d.set_ylabel('Speed difference (cm/s)\n[post-reward \u2212 pre-reward]', fontsize=9)
                _ax_d.set_xlabel('Condition', fontsize=10)
                if _all_diff_vals:
                    _ymax_d = float(np.nanmax(_all_diff_vals))
                    _ymin_d = float(np.nanmin(_all_diff_vals))
                    _pad_d  = max(abs(_ymax_d), abs(_ymin_d)) * 0.12 or 0.5
                    _ax_d.set_ylim(_ymin_d - _pad_d, _ymax_d + _pad_d)
                _ax_d.tick_params(axis='both', direction='in')
                _ax_d.spines['top'].set_visible(False)
                _ax_d.spines['right'].set_visible(False)

                # ── Mann-Whitney U tests for all condition pairs ──────────────
                import itertools as _itertools_d
                _pairs_d = list(_itertools_d.combinations(range(_n_conds_d), 2))
                _ylim_cur = list(_ax_d.get_ylim())
                _bracket_step = (_ylim_cur[1] - _ylim_cur[0]) * 0.14
                _bracket_base = _ylim_cur[1]
                for _pi, (_ia, _ib) in enumerate(_pairs_d):
                    _vals_a = [e[1] for e in _diff_by_cond[_conds_d[_ia]]]
                    _vals_b = [e[1] for e in _diff_by_cond[_conds_d[_ib]]]
                    if len(_vals_a) < 2 or len(_vals_b) < 2:
                        continue
                    _u_stat_d, _p_val_d = mannwhitneyu(_vals_a, _vals_b, alternative='two-sided')
                    if _p_val_d < 0.001:
                        _sig_str = f'p = {_p_val_d:.2e}***'
                    elif _p_val_d < 0.01:
                        _sig_str = f'p = {_p_val_d:.3f}**'
                    elif _p_val_d < 0.05:
                        _sig_str = f'p = {_p_val_d:.3f}*'
                    else:
                        _sig_str = f'p = {_p_val_d:.3f} (ns)'
                    _bh = _bracket_base + _bracket_step * (_pi + 0.6)
                    _ax_d.plot([_ia, _ia, _ib, _ib],
                               [_bh - _bracket_step * 0.15,
                                _bh,
                                _bh,
                                _bh - _bracket_step * 0.15],
                               color='black', linewidth=1.0)
                    _ax_d.text((_ia + _ib) / 2, _bh + _bracket_step * 0.05,
                               _sig_str, ha='center', va='bottom', fontsize=8)
                if _pairs_d:
                    _new_top = _bracket_base + _bracket_step * (len(_pairs_d) + 1.5)
                    _ax_d.set_ylim(_ylim_cur[0], _new_top)

                epoch_reward_speed_diff_fig.suptitle(
                    'Post- vs Pre-Reward Speed Difference by Condition\n'
                    '(mean \u00b1 SEM across mice; positive = faster after reward delivery)',
                    fontsize=11,
                )
                epoch_reward_speed_diff_fig.tight_layout()

        # ── Pre/post-reward delivery capacitive (z-scored) bar chart ──────────
        if 'epoch_reward_cap_pre_post' in selected_plots and _any_cap:
            _cap_pre_mask  = (EPOCH_CANONICAL_TIME >= 0.0)  & (EPOCH_CANONICAL_TIME <= 0.65)
            _cap_post_mask = (EPOCH_CANONICAL_TIME >  0.65) & (EPOCH_CANONICAL_TIME <= 1.30)
            _cap_pp_by_cond: dict = {}
            for _r in all_results:
                _sess_mat = _r.get('cap_epoch_session_means')
                if _sess_mat is None or _sess_mat.shape[0] == 0:
                    continue
                _cond = _r['starting_condition']
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore', RuntimeWarning)
                    _pre_per_sess  = np.nanmean(_sess_mat[:, _cap_pre_mask],  axis=1)
                    _post_per_sess = np.nanmean(_sess_mat[:, _cap_post_mask], axis=1)
                    _pre_mean  = float(np.nanmean(_pre_per_sess))
                    _post_mean = float(np.nanmean(_post_per_sess))
                _cap_pp_by_cond.setdefault(_cond, []).append((_r['mouse'], _pre_mean, _post_mean))

            if _cap_pp_by_cond:
                _conds_cap_pp   = sorted(_cap_pp_by_cond.keys())
                _n_conds_cap_pp = len(_conds_cap_pp)
                epoch_reward_cap_pre_post_fig, _axs_cap_pp = plt.subplots(
                    1, _n_conds_cap_pp,
                    figsize=(4 * _n_conds_cap_pp + 1, 5),
                    sharey=True, squeeze=False,
                )
                _all_cap_pp_yvals = []
                for _ci, _cond in enumerate(_conds_cap_pp):
                    _ax      = _axs_cap_pp[0, _ci]
                    _color   = condition_color_map.get(_cond, 'steelblue')
                    _entries = _cap_pp_by_cond[_cond]
                    _n_cap_pp = len(_entries)
                    _pre_vals  = [e[1] for e in _entries]
                    _post_vals = [e[2] for e in _entries]
                    with warnings.catch_warnings():
                        warnings.simplefilter('ignore', RuntimeWarning)
                        _mn_pre   = float(np.nanmean(_pre_vals))
                        _mn_post  = float(np.nanmean(_post_vals))
                        _sem_pre  = (float(np.nanstd(_pre_vals,  ddof=1) / np.sqrt(_n_cap_pp))
                                     if _n_cap_pp > 1 else 0.0)
                        _sem_post = (float(np.nanstd(_post_vals, ddof=1) / np.sqrt(_n_cap_pp))
                                     if _n_cap_pp > 1 else 0.0)
                    _all_cap_pp_yvals.extend([_mn_pre + _sem_pre, _mn_post + _sem_post,
                                              _mn_pre - _sem_pre, _mn_post - _sem_post])
                    _all_cap_pp_yvals.extend(_pre_vals + _post_vals)
                    _ax.bar(0, _mn_pre,  width=0.5, color=_color, alpha=0.7,
                            yerr=_sem_pre,  capsize=7,
                            error_kw={'elinewidth': 1.5, 'capthick': 1.5})
                    _ax.bar(1, _mn_post, width=0.5, color=_color, alpha=0.7,
                            yerr=_sem_post, capsize=7,
                            error_kw={'elinewidth': 1.5, 'capthick': 1.5})
                    _rng_cap_pp = np.random.default_rng(seed=42)
                    _jitter_cap_pp = (_rng_cap_pp.random(_n_cap_pp) - 0.5) * 0.18
                    for _j, (_mname, _pv, _qv) in enumerate(_entries):
                        _xp = 0 + _jitter_cap_pp[_j]
                        _xq = 1 + _jitter_cap_pp[_j]
                        _ax.plot([_xp, _xq], [_pv, _qv], '-',
                                 color=_color, linewidth=0.9, alpha=0.5, zorder=2)
                        _ax.plot(_xp, _pv, 'o', color='white',
                                 markeredgecolor=_color, markeredgewidth=1.5,
                                 markersize=7, zorder=3)
                        _ax.plot(_xq, _qv, 'o', color='white',
                                 markeredgecolor=_color, markeredgewidth=1.5,
                                 markersize=7, zorder=3)
                    _ax.set_xticks([0, 1])
                    _ax.set_xticklabels(['Pre-reward\n(0\u20130.65 s)', 'Post-reward\n(0.65\u20131.3 s)'],
                                        fontsize=9)
                    _ax.set_title(f'{_cond}\n(n={_n_cap_pp} mice)', fontsize=10)
                    _ax.set_ylabel('Capacitive Sensor (z-score)' if _ci == 0 else '', fontsize=9)
                    _ax.set_xlim(-0.6, 1.6)
                    _ax.axhline(0, color='black', linewidth=0.8, linestyle='--', zorder=1)
                    _ax.tick_params(axis='both', direction='in')
                    _ax.spines['top'].set_visible(False)
                    _ax.spines['right'].set_visible(False)
                if _all_cap_pp_yvals:
                    _ymax_cap_pp = float(np.nanmax(_all_cap_pp_yvals))
                    _ymin_cap_pp = float(np.nanmin(_all_cap_pp_yvals))
                else:
                    _ymax_cap_pp, _ymin_cap_pp = 1.0, -1.0
                _pad_cap_pp = max(abs(_ymax_cap_pp), abs(_ymin_cap_pp)) * 0.12 or 0.1
                _axs_cap_pp[0, 0].set_ylim(_ymin_cap_pp - _pad_cap_pp, _ymax_cap_pp + _pad_cap_pp)
                epoch_reward_cap_pre_post_fig.suptitle(
                    'Average Capacitive Sensor (z-scored): Pre- vs Post-Reward Delivery\n'
                    '(session-averaged reward zone entry epochs, by condition)',
                    fontsize=12,
                )
                epoch_reward_cap_pre_post_fig.tight_layout()

        # ── Pre-minus-post reward capacitive difference bar chart (Mann-Whitney U) ──
        if 'epoch_reward_cap_diff' in selected_plots and _any_cap:
            _cap_pre_mask_d  = (EPOCH_CANONICAL_TIME >= 0.0)  & (EPOCH_CANONICAL_TIME <= 0.65)
            _cap_post_mask_d = (EPOCH_CANONICAL_TIME >  0.65) & (EPOCH_CANONICAL_TIME <= 1.30)
            _cap_diff_by_cond: dict = {}
            for _r in all_results:
                _sess_mat = _r.get('cap_epoch_session_means')
                if _sess_mat is None or _sess_mat.shape[0] == 0:
                    continue
                _cond = _r['starting_condition']
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore', RuntimeWarning)
                    _pre_d  = float(np.nanmean(np.nanmean(_sess_mat[:, _cap_pre_mask_d],  axis=1)))
                    _post_d = float(np.nanmean(np.nanmean(_sess_mat[:, _cap_post_mask_d], axis=1)))
                _cap_diff_by_cond.setdefault(_cond, []).append((_r['mouse'], _post_d - _pre_d))

            if _cap_diff_by_cond:
                _conds_cap_d   = sorted(_cap_diff_by_cond.keys())
                _n_conds_cap_d = len(_conds_cap_d)
                epoch_reward_cap_diff_fig, _ax_cap_d = plt.subplots(
                    1, 1, figsize=(max(4, _n_conds_cap_d * 1.4 + 1.5), 5)
                )
                _all_cap_diff_vals = []
                _bar_x_cap = np.arange(_n_conds_cap_d)
                _rng_cap_d = np.random.default_rng(seed=42)
                for _ci, _cond in enumerate(_conds_cap_d):
                    _color   = condition_color_map.get(_cond, 'steelblue')
                    _entries = _cap_diff_by_cond[_cond]
                    _n_cap_d = len(_entries)
                    _dvals   = [e[1] for e in _entries]
                    with warnings.catch_warnings():
                        warnings.simplefilter('ignore', RuntimeWarning)
                        _mn_cap_d  = float(np.nanmean(_dvals))
                        _sem_cap_d = (float(np.nanstd(_dvals, ddof=1) / np.sqrt(_n_cap_d))
                                      if _n_cap_d > 1 else 0.0)
                    _all_cap_diff_vals.extend(_dvals + [_mn_cap_d + _sem_cap_d, _mn_cap_d - _sem_cap_d])
                    _ax_cap_d.bar(_ci, _mn_cap_d, width=0.55, color=_color, alpha=0.7,
                                  yerr=_sem_cap_d, capsize=7,
                                  error_kw={'elinewidth': 1.5, 'capthick': 1.5})
                    _jitter_cap_d = (_rng_cap_d.random(_n_cap_d) - 0.5) * 0.22
                    for _j, (_mname, _dv) in enumerate(_entries):
                        _ax_cap_d.plot(_ci + _jitter_cap_d[_j], _dv, 'o',
                                       color='white', markeredgecolor=_color,
                                       markeredgewidth=1.5, markersize=7, zorder=3)
                _ax_cap_d.axhline(0, color='black', linewidth=0.9, linestyle='--', zorder=1)
                _ax_cap_d.set_xticks(_bar_x_cap)
                _ax_cap_d.set_xticklabels(_conds_cap_d, fontsize=10)
                _ax_cap_d.set_ylabel('Capacitive difference (z-score)\n[post-reward \u2212 pre-reward]', fontsize=9)
                _ax_cap_d.set_xlabel('Condition', fontsize=10)
                if _all_cap_diff_vals:
                    _ymax_cap_d = float(np.nanmax(_all_cap_diff_vals))
                    _ymin_cap_d = float(np.nanmin(_all_cap_diff_vals))
                    _pad_cap_d  = max(abs(_ymax_cap_d), abs(_ymin_cap_d)) * 0.12 or 0.1
                    _ax_cap_d.set_ylim(_ymin_cap_d - _pad_cap_d, _ymax_cap_d + _pad_cap_d)
                _ax_cap_d.tick_params(axis='both', direction='in')
                _ax_cap_d.spines['top'].set_visible(False)
                _ax_cap_d.spines['right'].set_visible(False)

                # ── Mann-Whitney U tests for all condition pairs ──────────────
                import itertools as _itertools_cap_d
                _pairs_cap_d = list(_itertools_cap_d.combinations(range(_n_conds_cap_d), 2))
                _ylim_cur_cap = list(_ax_cap_d.get_ylim())
                _bracket_step_cap = (_ylim_cur_cap[1] - _ylim_cur_cap[0]) * 0.14
                _bracket_base_cap = _ylim_cur_cap[1]
                for _pi, (_ia, _ib) in enumerate(_pairs_cap_d):
                    _vals_a = [e[1] for e in _cap_diff_by_cond[_conds_cap_d[_ia]]]
                    _vals_b = [e[1] for e in _cap_diff_by_cond[_conds_cap_d[_ib]]]
                    if len(_vals_a) < 2 or len(_vals_b) < 2:
                        continue
                    _u_stat_cap, _p_val_cap = mannwhitneyu(_vals_a, _vals_b, alternative='two-sided')
                    if _p_val_cap < 0.001:
                        _sig_str_cap = f'p = {_p_val_cap:.2e}***'
                    elif _p_val_cap < 0.01:
                        _sig_str_cap = f'p = {_p_val_cap:.3f}**'
                    elif _p_val_cap < 0.05:
                        _sig_str_cap = f'p = {_p_val_cap:.3f}*'
                    else:
                        _sig_str_cap = f'p = {_p_val_cap:.3f} (ns)'
                    _bh_cap = _bracket_base_cap + _bracket_step_cap * (_pi + 0.6)
                    _ax_cap_d.plot([_ia, _ia, _ib, _ib],
                                   [_bh_cap - _bracket_step_cap * 0.15,
                                    _bh_cap,
                                    _bh_cap,
                                    _bh_cap - _bracket_step_cap * 0.15],
                                   color='black', linewidth=1.0)
                    _ax_cap_d.text((_ia + _ib) / 2, _bh_cap + _bracket_step_cap * 0.05,
                                   _sig_str_cap, ha='center', va='bottom', fontsize=8)
                if _pairs_cap_d:
                    _new_top_cap = _bracket_base_cap + _bracket_step_cap * (len(_pairs_cap_d) + 1.5)
                    _ax_cap_d.set_ylim(_ylim_cur_cap[0], _new_top_cap)

                epoch_reward_cap_diff_fig.suptitle(
                    'Post- vs Pre-Reward Capacitive (z-scored) Difference by Condition\n'
                    '(mean \u00b1 SEM across mice; positive = higher licking after reward delivery)',
                    fontsize=11,
                )
                epoch_reward_cap_diff_fig.tight_layout()

        # ── Reward zone: 0.65 s pre- vs 0.65 s post-zone entry bar chart ─────
        if 'epoch_reward_speed_pre_post_entry' in selected_plots and _any_speed:
            _pre_mask_re  = (EPOCH_CANONICAL_TIME >= -0.65) & (EPOCH_CANONICAL_TIME <  0.0)
            _post_mask_re = (EPOCH_CANONICAL_TIME >= 0.0)   & (EPOCH_CANONICAL_TIME <= 0.65)
            _rpe_by_cond: dict = {}
            for _r in all_results:
                _sess_mat = _r.get('speed_epoch_session_means')
                if _sess_mat is None or _sess_mat.shape[0] == 0:
                    continue
                _cond = _r['starting_condition']
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore', RuntimeWarning)
                    _pre_per_sess  = np.nanmean(_sess_mat[:, _pre_mask_re],  axis=1)
                    _post_per_sess = np.nanmean(_sess_mat[:, _post_mask_re], axis=1)
                    _pre_mean  = float(np.nanmean(_pre_per_sess))
                    _post_mean = float(np.nanmean(_post_per_sess))
                _rpe_by_cond.setdefault(_cond, []).append((_r['mouse'], _pre_mean, _post_mean))

            if _rpe_by_cond:
                _conds_rpe   = sorted(_rpe_by_cond.keys())
                _n_conds_rpe = len(_conds_rpe)
                epoch_reward_speed_pre_post_entry_fig, _axs_rpe = plt.subplots(
                    1, _n_conds_rpe,
                    figsize=(4 * _n_conds_rpe + 1, 5),
                    sharey=True, squeeze=False,
                )
                _all_rpe_yvals = []
                for _ci, _cond in enumerate(_conds_rpe):
                    _ax       = _axs_rpe[0, _ci]
                    _color    = condition_color_map.get(_cond, 'steelblue')
                    _entries  = _rpe_by_cond[_cond]
                    _n_rpe    = len(_entries)
                    _pre_vals = [e[1] for e in _entries]
                    _post_vals= [e[2] for e in _entries]
                    with warnings.catch_warnings():
                        warnings.simplefilter('ignore', RuntimeWarning)
                        _mn_pre   = float(np.nanmean(_pre_vals))
                        _mn_post  = float(np.nanmean(_post_vals))
                        _sem_pre  = (float(np.nanstd(_pre_vals,  ddof=1) / np.sqrt(_n_rpe))
                                     if _n_rpe > 1 else 0.0)
                        _sem_post = (float(np.nanstd(_post_vals, ddof=1) / np.sqrt(_n_rpe))
                                     if _n_rpe > 1 else 0.0)
                    _all_rpe_yvals.extend([_mn_pre + _sem_pre, _mn_post + _sem_post,
                                           _mn_pre - _sem_pre, _mn_post - _sem_post])
                    _all_rpe_yvals.extend(_pre_vals + _post_vals)
                    _ax.bar(0, _mn_pre,  width=0.5, color=_color, alpha=0.7,
                            yerr=_sem_pre,  capsize=7,
                            error_kw={'elinewidth': 1.5, 'capthick': 1.5})
                    _ax.bar(1, _mn_post, width=0.5, color=_color, alpha=0.7,
                            yerr=_sem_post, capsize=7,
                            error_kw={'elinewidth': 1.5, 'capthick': 1.5})
                    _rng_rpe = np.random.default_rng(seed=42)
                    _jitter  = (_rng_rpe.random(_n_rpe) - 0.5) * 0.18
                    for _j, (_mname, _pv, _qv) in enumerate(_entries):
                        _xp = 0 + _jitter[_j]
                        _xq = 1 + _jitter[_j]
                        _ax.plot([_xp, _xq], [_pv, _qv], '-',
                                 color=_color, linewidth=0.9, alpha=0.5, zorder=2)
                        _ax.plot(_xp, _pv, 'o', color='white',
                                 markeredgecolor=_color, markeredgewidth=1.5,
                                 markersize=7, zorder=3)
                        _ax.plot(_xq, _qv, 'o', color='white',
                                 markeredgecolor=_color, markeredgewidth=1.5,
                                 markersize=7, zorder=3)
                    _ax.set_xticks([0, 1])
                    _ax.set_xticklabels(['Pre-entry\n(−0.65–0 s)', 'Post-entry\n(0–0.65 s)'],
                                        fontsize=9)
                    _ax.set_title(f'{_cond}\n(n={_n_rpe} mice)', fontsize=10)
                    _ax.set_ylabel('Treadmill Speed (cm/s)' if _ci == 0 else '', fontsize=9)
                    _ax.set_xlim(-0.6, 1.6)
                    _ax.tick_params(axis='both', direction='in')
                    _ax.spines['top'].set_visible(False)
                    _ax.spines['right'].set_visible(False)
                if _all_rpe_yvals:
                    _ymax_rpe = float(np.nanmax(_all_rpe_yvals))
                    _ymin_rpe = float(np.nanmin(_all_rpe_yvals))
                else:
                    _ymax_rpe, _ymin_rpe = 1.0, 0.0
                _bot_rpe = _ymin_rpe * 1.05 if _ymin_rpe < 0 else 0.0
                _axs_rpe[0, 0].set_ylim(_bot_rpe, _ymax_rpe * 1.05)
                epoch_reward_speed_pre_post_entry_fig.suptitle(
                    'Average Speed: 0.65 s Pre- vs 0.65 s Post-Zone Entry (Reward Zone)\n'
                    '(session-averaged reward zone entry epochs, by condition)',
                    fontsize=12,
                )
                epoch_reward_speed_pre_post_entry_fig.tight_layout()

        # ── Reward zone: pre-minus-post zone entry difference bar chart ───────
        if 'epoch_reward_speed_diff_entry' in selected_plots and _any_speed:
            _pre_mask_rde  = (EPOCH_CANONICAL_TIME >= -0.65) & (EPOCH_CANONICAL_TIME <  0.0)
            _post_mask_rde = (EPOCH_CANONICAL_TIME >= 0.0)   & (EPOCH_CANONICAL_TIME <= 0.65)
            _rde_by_cond: dict = {}
            for _r in all_results:
                _sess_mat = _r.get('speed_epoch_session_means')
                if _sess_mat is None or _sess_mat.shape[0] == 0:
                    continue
                _cond = _r['starting_condition']
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore', RuntimeWarning)
                    _pre_rde  = float(np.nanmean(np.nanmean(_sess_mat[:, _pre_mask_rde],  axis=1)))
                    _post_rde = float(np.nanmean(np.nanmean(_sess_mat[:, _post_mask_rde], axis=1)))
                _rde_by_cond.setdefault(_cond, []).append((_r['mouse'], _post_rde - _pre_rde))

            if _rde_by_cond:
                _conds_rde   = sorted(_rde_by_cond.keys())
                _n_conds_rde = len(_conds_rde)
                epoch_reward_speed_diff_entry_fig, _ax_rde = plt.subplots(
                    1, 1, figsize=(max(4, _n_conds_rde * 1.4 + 1.5), 5)
                )
                _all_rde_vals = []
                _bar_x_rde = np.arange(_n_conds_rde)
                _rng_rde   = np.random.default_rng(seed=42)
                for _ci, _cond in enumerate(_conds_rde):
                    _color    = condition_color_map.get(_cond, 'steelblue')
                    _entries  = _rde_by_cond[_cond]
                    _n_rde    = len(_entries)
                    _dvals_rde = [e[1] for e in _entries]
                    with warnings.catch_warnings():
                        warnings.simplefilter('ignore', RuntimeWarning)
                        _mn_rde  = float(np.nanmean(_dvals_rde))
                        _sem_rde = (float(np.nanstd(_dvals_rde, ddof=1) / np.sqrt(_n_rde))
                                    if _n_rde > 1 else 0.0)
                    _all_rde_vals.extend(_dvals_rde + [_mn_rde + _sem_rde, _mn_rde - _sem_rde])
                    _ax_rde.bar(_ci, _mn_rde, width=0.55, color=_color, alpha=0.7,
                                yerr=_sem_rde, capsize=7,
                                error_kw={'elinewidth': 1.5, 'capthick': 1.5})
                    _jitter_rde = (_rng_rde.random(_n_rde) - 0.5) * 0.22
                    for _j, (_mname, _dv) in enumerate(_entries):
                        _ax_rde.plot(_ci + _jitter_rde[_j], _dv, 'o',
                                     color='white', markeredgecolor=_color,
                                     markeredgewidth=1.5, markersize=7, zorder=3)
                _ax_rde.axhline(0, color='black', linewidth=0.9, linestyle='--', zorder=1)
                _ax_rde.set_xticks(_bar_x_rde)
                _ax_rde.set_xticklabels(_conds_rde, fontsize=10)
                _ax_rde.set_ylabel('Speed difference (cm/s)\n[post-entry \u2212 pre-entry]', fontsize=9)
                _ax_rde.set_xlabel('Condition', fontsize=10)
                if _all_rde_vals:
                    _ymax_rde = float(np.nanmax(_all_rde_vals))
                    _ymin_rde = float(np.nanmin(_all_rde_vals))
                    _pad_rde  = max(abs(_ymax_rde), abs(_ymin_rde)) * 0.12 or 0.5
                    _ax_rde.set_ylim(_ymin_rde - _pad_rde, _ymax_rde + _pad_rde)
                _ax_rde.tick_params(axis='both', direction='in')
                _ax_rde.spines['top'].set_visible(False)
                _ax_rde.spines['right'].set_visible(False)

                # ── Mann-Whitney U tests for all condition pairs ───────────
                import itertools as _itertools_rde
                _pairs_rde = list(_itertools_rde.combinations(range(_n_conds_rde), 2))
                _ylim_cur_rde = list(_ax_rde.get_ylim())
                _bracket_step_rde = (_ylim_cur_rde[1] - _ylim_cur_rde[0]) * 0.14
                _bracket_base_rde = _ylim_cur_rde[1]
                for _pi, (_ia, _ib) in enumerate(_pairs_rde):
                    _vals_a_rde = [e[1] for e in _rde_by_cond[_conds_rde[_ia]]]
                    _vals_b_rde = [e[1] for e in _rde_by_cond[_conds_rde[_ib]]]
                    if len(_vals_a_rde) < 2 or len(_vals_b_rde) < 2:
                        continue
                    _u_stat_rde, _p_val_rde = mannwhitneyu(_vals_a_rde, _vals_b_rde, alternative='two-sided')
                    if _p_val_rde < 0.001:
                        _sig_str_rde = f'p = {_p_val_rde:.2e}***'
                    elif _p_val_rde < 0.01:
                        _sig_str_rde = f'p = {_p_val_rde:.3f}**'
                    elif _p_val_rde < 0.05:
                        _sig_str_rde = f'p = {_p_val_rde:.3f}*'
                    else:
                        _sig_str_rde = f'p = {_p_val_rde:.3f} (ns)'
                    _bh_rde = _bracket_base_rde + _bracket_step_rde * (_pi + 0.6)
                    _ax_rde.plot([_ia, _ia, _ib, _ib],
                                 [_bh_rde - _bracket_step_rde * 0.15,
                                  _bh_rde,
                                  _bh_rde,
                                  _bh_rde - _bracket_step_rde * 0.15],
                                 color='black', linewidth=1.0)
                    _ax_rde.text((_ia + _ib) / 2, _bh_rde + _bracket_step_rde * 0.05,
                                 _sig_str_rde, ha='center', va='bottom', fontsize=8)
                if _pairs_rde:
                    _new_top_rde = _bracket_base_rde + _bracket_step_rde * (len(_pairs_rde) + 1.5)
                    _ax_rde.set_ylim(_ylim_cur_rde[0], _new_top_rde)

                epoch_reward_speed_diff_entry_fig.suptitle(
                    'Post- vs Pre-Zone Entry Speed Difference by Condition (Reward Zone)\n'
                    '(mean ± SEM across mice; positive = faster after zone entry)',
                    fontsize=11,
                )
                epoch_reward_speed_diff_entry_fig.tight_layout()

        # ── Reward zone: 1 s pre- vs 1 s post-zone entry bar chart ──────────
        if 'epoch_reward_speed_pre_post_entry_1s' in selected_plots and _any_speed:
            _pre_mask_re1  = (EPOCH_CANONICAL_TIME >= -1.0) & (EPOCH_CANONICAL_TIME <  0.0)
            _post_mask_re1 = (EPOCH_CANONICAL_TIME >= 0.0)  & (EPOCH_CANONICAL_TIME <= 1.0)
            _re1_by_cond: dict = {}
            for _r in all_results:
                _sess_mat = _r.get('speed_epoch_session_means')
                if _sess_mat is None or _sess_mat.shape[0] == 0:
                    continue
                _cond = _r['starting_condition']
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore', RuntimeWarning)
                    _pre_per_sess  = np.nanmean(_sess_mat[:, _pre_mask_re1],  axis=1)
                    _post_per_sess = np.nanmean(_sess_mat[:, _post_mask_re1], axis=1)
                    _pre_mean  = float(np.nanmean(_pre_per_sess))
                    _post_mean = float(np.nanmean(_post_per_sess))
                _re1_by_cond.setdefault(_cond, []).append((_r['mouse'], _pre_mean, _post_mean))

            if _re1_by_cond:
                _conds_re1   = sorted(_re1_by_cond.keys())
                _n_conds_re1 = len(_conds_re1)
                epoch_reward_speed_pre_post_entry_1s_fig, _axs_re1 = plt.subplots(
                    1, _n_conds_re1,
                    figsize=(4 * _n_conds_re1 + 1, 5),
                    sharey=True, squeeze=False,
                )
                _all_re1_yvals = []
                for _ci, _cond in enumerate(_conds_re1):
                    _ax       = _axs_re1[0, _ci]
                    _color    = condition_color_map.get(_cond, 'steelblue')
                    _entries  = _re1_by_cond[_cond]
                    _n_re1    = len(_entries)
                    _pre_vals = [e[1] for e in _entries]
                    _post_vals= [e[2] for e in _entries]
                    with warnings.catch_warnings():
                        warnings.simplefilter('ignore', RuntimeWarning)
                        _mn_pre   = float(np.nanmean(_pre_vals))
                        _mn_post  = float(np.nanmean(_post_vals))
                        _sem_pre  = (float(np.nanstd(_pre_vals,  ddof=1) / np.sqrt(_n_re1))
                                     if _n_re1 > 1 else 0.0)
                        _sem_post = (float(np.nanstd(_post_vals, ddof=1) / np.sqrt(_n_re1))
                                     if _n_re1 > 1 else 0.0)
                    _all_re1_yvals.extend([_mn_pre + _sem_pre, _mn_post + _sem_post,
                                           _mn_pre - _sem_pre, _mn_post - _sem_post])
                    _all_re1_yvals.extend(_pre_vals + _post_vals)
                    _ax.bar(0, _mn_pre,  width=0.5, color=_color, alpha=0.7,
                            yerr=_sem_pre,  capsize=7,
                            error_kw={'elinewidth': 1.5, 'capthick': 1.5})
                    _ax.bar(1, _mn_post, width=0.5, color=_color, alpha=0.7,
                            yerr=_sem_post, capsize=7,
                            error_kw={'elinewidth': 1.5, 'capthick': 1.5})
                    _rng_re1 = np.random.default_rng(seed=42)
                    _jitter  = (_rng_re1.random(_n_re1) - 0.5) * 0.18
                    for _j, (_mname, _pv, _qv) in enumerate(_entries):
                        _xp = 0 + _jitter[_j]
                        _xq = 1 + _jitter[_j]
                        _ax.plot([_xp, _xq], [_pv, _qv], '-',
                                 color=_color, linewidth=0.9, alpha=0.5, zorder=2)
                        _ax.plot(_xp, _pv, 'o', color='white',
                                 markeredgecolor=_color, markeredgewidth=1.5,
                                 markersize=7, zorder=3)
                        _ax.plot(_xq, _qv, 'o', color='white',
                                 markeredgecolor=_color, markeredgewidth=1.5,
                                 markersize=7, zorder=3)
                    _ax.set_xticks([0, 1])
                    _ax.set_xticklabels(['Pre-entry\n(\u22121\u20130 s)', 'Post-entry\n(0\u20131 s)'],
                                        fontsize=9)
                    _ax.set_title(f'{_cond}\n(n={_n_re1} mice)', fontsize=10)
                    _ax.set_ylabel('Treadmill Speed (cm/s)' if _ci == 0 else '', fontsize=9)
                    _ax.set_xlim(-0.6, 1.6)
                    _ax.tick_params(axis='both', direction='in')
                    _ax.spines['top'].set_visible(False)
                    _ax.spines['right'].set_visible(False)
                if _all_re1_yvals:
                    _ymax_re1 = float(np.nanmax(_all_re1_yvals))
                    _ymin_re1 = float(np.nanmin(_all_re1_yvals))
                else:
                    _ymax_re1, _ymin_re1 = 1.0, 0.0
                _bot_re1 = _ymin_re1 * 1.05 if _ymin_re1 < 0 else 0.0
                _axs_re1[0, 0].set_ylim(_bot_re1, _ymax_re1 * 1.05)
                epoch_reward_speed_pre_post_entry_1s_fig.suptitle(
                    'Average Speed: 1 s Pre- vs 1 s Post-Zone Entry (Reward Zone)\n'
                    '(session-averaged reward zone entry epochs, by condition)',
                    fontsize=12,
                )
                epoch_reward_speed_pre_post_entry_1s_fig.tight_layout()

        # ── Reward zone: 1 s pre-minus-post zone entry difference bar chart ──
        if 'epoch_reward_speed_diff_entry_1s' in selected_plots and _any_speed:
            _pre_mask_rde1  = (EPOCH_CANONICAL_TIME >= -1.0) & (EPOCH_CANONICAL_TIME <  0.0)
            _post_mask_rde1 = (EPOCH_CANONICAL_TIME >= 0.0)  & (EPOCH_CANONICAL_TIME <= 1.0)
            _rde1_by_cond: dict = {}
            for _r in all_results:
                _sess_mat = _r.get('speed_epoch_session_means')
                if _sess_mat is None or _sess_mat.shape[0] == 0:
                    continue
                _cond = _r['starting_condition']
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore', RuntimeWarning)
                    _pre_rde1  = float(np.nanmean(np.nanmean(_sess_mat[:, _pre_mask_rde1],  axis=1)))
                    _post_rde1 = float(np.nanmean(np.nanmean(_sess_mat[:, _post_mask_rde1], axis=1)))
                _rde1_by_cond.setdefault(_cond, []).append((_r['mouse'], _post_rde1 - _pre_rde1))

            if _rde1_by_cond:
                _conds_rde1   = sorted(_rde1_by_cond.keys())
                _n_conds_rde1 = len(_conds_rde1)
                epoch_reward_speed_diff_entry_1s_fig, _ax_rde1 = plt.subplots(
                    1, 1, figsize=(max(4, _n_conds_rde1 * 1.4 + 1.5), 5)
                )
                _all_rde1_vals = []
                _bar_x_rde1 = np.arange(_n_conds_rde1)
                _rng_rde1   = np.random.default_rng(seed=42)
                for _ci, _cond in enumerate(_conds_rde1):
                    _color     = condition_color_map.get(_cond, 'steelblue')
                    _entries   = _rde1_by_cond[_cond]
                    _n_rde1    = len(_entries)
                    _dvals_rde1 = [e[1] for e in _entries]
                    with warnings.catch_warnings():
                        warnings.simplefilter('ignore', RuntimeWarning)
                        _mn_rde1  = float(np.nanmean(_dvals_rde1))
                        _sem_rde1 = (float(np.nanstd(_dvals_rde1, ddof=1) / np.sqrt(_n_rde1))
                                     if _n_rde1 > 1 else 0.0)
                    _all_rde1_vals.extend(_dvals_rde1 + [_mn_rde1 + _sem_rde1, _mn_rde1 - _sem_rde1])
                    _ax_rde1.bar(_ci, _mn_rde1, width=0.55, color=_color, alpha=0.7,
                                 yerr=_sem_rde1, capsize=7,
                                 error_kw={'elinewidth': 1.5, 'capthick': 1.5})
                    _jitter_rde1 = (_rng_rde1.random(_n_rde1) - 0.5) * 0.22
                    for _j, (_mname, _dv) in enumerate(_entries):
                        _ax_rde1.plot(_ci + _jitter_rde1[_j], _dv, 'o',
                                      color='white', markeredgecolor=_color,
                                      markeredgewidth=1.5, markersize=7, zorder=3)
                _ax_rde1.axhline(0, color='black', linewidth=0.9, linestyle='--', zorder=1)
                _ax_rde1.set_xticks(_bar_x_rde1)
                _ax_rde1.set_xticklabels(_conds_rde1, fontsize=10)
                _ax_rde1.set_ylabel('Speed difference (cm/s)\n[post-entry \u2212 pre-entry]', fontsize=9)
                _ax_rde1.set_xlabel('Condition', fontsize=10)
                if _all_rde1_vals:
                    _ymax_rde1 = float(np.nanmax(_all_rde1_vals))
                    _ymin_rde1 = float(np.nanmin(_all_rde1_vals))
                    _pad_rde1  = max(abs(_ymax_rde1), abs(_ymin_rde1)) * 0.12 or 0.5
                    _ax_rde1.set_ylim(_ymin_rde1 - _pad_rde1, _ymax_rde1 + _pad_rde1)
                _ax_rde1.tick_params(axis='both', direction='in')
                _ax_rde1.spines['top'].set_visible(False)
                _ax_rde1.spines['right'].set_visible(False)

                # ── Mann-Whitney U tests for all condition pairs ───────────────
                import itertools as _itertools_rde1
                _pairs_rde1 = list(_itertools_rde1.combinations(range(_n_conds_rde1), 2))
                _ylim_cur_rde1 = list(_ax_rde1.get_ylim())
                _bracket_step_rde1 = (_ylim_cur_rde1[1] - _ylim_cur_rde1[0]) * 0.14
                _bracket_base_rde1 = _ylim_cur_rde1[1]
                for _pi, (_ia, _ib) in enumerate(_pairs_rde1):
                    _vals_a_rde1 = [e[1] for e in _rde1_by_cond[_conds_rde1[_ia]]]
                    _vals_b_rde1 = [e[1] for e in _rde1_by_cond[_conds_rde1[_ib]]]
                    if len(_vals_a_rde1) < 2 or len(_vals_b_rde1) < 2:
                        continue
                    _u_stat_rde1, _p_val_rde1 = mannwhitneyu(_vals_a_rde1, _vals_b_rde1, alternative='two-sided')
                    if _p_val_rde1 < 0.001:
                        _sig_str_rde1 = f'p = {_p_val_rde1:.2e}***'
                    elif _p_val_rde1 < 0.01:
                        _sig_str_rde1 = f'p = {_p_val_rde1:.3f}**'
                    elif _p_val_rde1 < 0.05:
                        _sig_str_rde1 = f'p = {_p_val_rde1:.3f}*'
                    else:
                        _sig_str_rde1 = f'p = {_p_val_rde1:.3f} (ns)'
                    _bh_rde1 = _bracket_base_rde1 + _bracket_step_rde1 * (_pi + 0.6)
                    _ax_rde1.plot([_ia, _ia, _ib, _ib],
                                  [_bh_rde1 - _bracket_step_rde1 * 0.15,
                                   _bh_rde1,
                                   _bh_rde1,
                                   _bh_rde1 - _bracket_step_rde1 * 0.15],
                                  color='black', linewidth=1.0)
                    _ax_rde1.text((_ia + _ib) / 2, _bh_rde1 + _bracket_step_rde1 * 0.05,
                                  _sig_str_rde1, ha='center', va='bottom', fontsize=8)
                if _pairs_rde1:
                    _new_top_rde1 = _bracket_base_rde1 + _bracket_step_rde1 * (len(_pairs_rde1) + 1.5)
                    _ax_rde1.set_ylim(_ylim_cur_rde1[0], _new_top_rde1)

                epoch_reward_speed_diff_entry_1s_fig.suptitle(
                    'Post- vs Pre-Zone Entry Speed Difference by Condition (Reward Zone, 1 s windows)\n'
                    '(mean \u00b1 SEM across mice; positive = faster after zone entry)',
                    fontsize=11,
                )
                epoch_reward_speed_diff_entry_1s_fig.tight_layout()

    # Create the level-based analysis plots
    level_reward_fig = level_speed_collapsed_fig = level_speed_condition_fig = None
    level_lick_collapsed_fig = level_lick_condition_fig = None
    level_dist_collapsed_fig = level_dist_condition_fig = level_dist_condition_excl_last_fig = None
    level_bout_collapsed_fig = level_bout_condition_fig = None
    level_bout_avg_speed_collapsed_fig = level_bout_avg_speed_condition_fig = None
    level_bout_avg_dist_collapsed_fig = level_bout_avg_dist_condition_fig = None
    last_level_bar_fig = None
    level_survivor_fig = None
    time_to_level2_fig = None
    _level_stats_data = None
    if transitions_csv_path or any(k in selected_plots for k in ('levels', 'level_speed', 'level_speed_condition',
                                                                   'level_lick', 'level_lick_condition',
                                                                   'level_dist', 'level_dist_condition',
                                                                   'level_dist_condition_excl_last',
                                                                   'level_bout', 'level_bout_condition',
                                                                   'level_bout_avg_speed', 'level_bout_avg_speed_condition',
                                                                   'level_bout_avg_dist', 'level_bout_avg_dist_condition',
                                                                   'last_level_bar', 'level_survivor', 'time_to_level2')):
        level_reward_fig, level_speed_collapsed_fig, level_speed_condition_fig, \
        level_lick_collapsed_fig, level_lick_condition_fig, \
        level_dist_collapsed_fig, level_dist_condition_fig, \
        level_dist_condition_excl_last_fig, \
        level_bout_collapsed_fig, level_bout_condition_fig, \
        level_bout_avg_speed_collapsed_fig, level_bout_avg_speed_condition_fig, \
        level_bout_avg_dist_collapsed_fig, level_bout_avg_dist_condition_fig, \
        last_level_bar_fig, level_survivor_fig, time_to_level2_fig, \
        _level_stats_data = analyze_levels(
            data_files, transitions_csv_path, animal_conditions=conditions,
            selected_plots=selected_plots,
        )

    # ── Missing data report ───────────────────────────────────────────────────
    all_session_dates = set(
        d.date() for result in all_results for d in result['df']['date']
    )

    report_lines = []
    report_lines.append("=" * 70)
    report_lines.append("MISSING DATA REPORT")
    report_lines.append(f"Global date range: {global_start}  →  {global_end}  ({total_days} calendar days)")
    report_lines.append(f"Session dates across all mice: {len(all_session_dates)}")
    report_lines.append("=" * 70)

    for result in all_results:
        mouse_name  = result['mouse']
        mouse_dates = set(d.date() for d in result['df']['date'])
        file_errors = result.get('session_file_errors', {})
        error_dates = set(datetime.strptime(ds, '%Y-%m-%d').date() for ds in file_errors)
        fully_missing = sorted(all_session_dates - mouse_dates - error_dates)
        present       = sorted(mouse_dates)

        report_lines.append(f"\nMouse: {mouse_name}")
        report_lines.append(f"  Sessions present   : {len(present)}")
        for d in present:
            day_offset = (d - global_start).days
            report_lines.append(f"    Day {day_offset:>3}  {d}")

        if file_errors:
            report_lines.append(f"  Incomplete sessions (missing file): {len(file_errors)}")
            for date_str, missing_types in sorted(file_errors.items()):
                d = datetime.strptime(date_str, '%Y-%m-%d').date()
                day_offset = (d - global_start).days
                report_lines.append(f"    Day {day_offset:>3}  {date_str}  <-- MISSING FILE(S): {', '.join(missing_types)}")

        if fully_missing:
            report_lines.append(f"  Fully absent dates : {len(fully_missing)}")
            for d in fully_missing:
                day_offset = (d - global_start).days
                report_lines.append(f"    Day {day_offset:>3}  {d}  <-- NO SESSION")

        if not file_errors and not fully_missing:
            report_lines.append(f"  Missing sessions   : none")

    report_lines.append("\n" + "=" * 70)
    report_text = "\n".join(report_lines)

    print("\n" + report_text)

    report_dir = output_dir if output_dir else os.path.dirname(os.path.abspath(data_files[0]))
    timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = os.path.join(report_dir, f"missing_data_report_{timestamp_str}.txt")
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(report_text + "\n")
    print(f"\nMissing data report saved to: {report_path}")

    # Print weekday alignment table so the user can verify Mon→Tue→Thu→Fri cycle
    print_session_weekday_alignment(all_results)

    return speed_fig, sensitivity_fig, lick_fig, reward_fig, lick_reward_ratio_fig, false_alarm_fig, correct_rejection_fig, specificity_fig, dprime_fig, avg_reward_fig, sex_reward_fig, avg_sex_speed_fig, distance_fig, bout_count_fig, avg_bout_count_fig, rewards_per_bout_fig, first_lick_latency_fig, condition_rewards_per_bout_fig, condition_rewards_per_bout_bar_fig, condition_first_lick_latency_fig, condition_first_lick_latency_bar_fig, condition_lick_after_reward_prop_fig, condition_lick_after_reward_prop_bar_fig, weekday_reward_bar_fig, weekday_reward_bar_condition_fig, bout_avg_speed_fig, bout_avg_dist_fig, sex_distance_fig, condition_distance_fig, condition_distance_bar_fig, total_distance_bar_fig, avg_lick_rate_fig, sex_lick_rate_fig, condition_reward_fig, condition_speed_fig, condition_bout_count_fig, condition_bout_avg_speed_fig, condition_bout_avg_dist_fig, condition_lick_fig, condition_lick_rate_fig, condition_lick_reward_ratio_fig, condition_lick_reward_ratio_bar_fig, condition_punish_zone_pct_bar_fig, condition_bar_fig, condition_speed_bar_fig, condition_bout_count_bar_fig, condition_bout_avg_speed_bar_fig, condition_bout_avg_dist_bar_fig, condition_lick_bar_fig, level_reward_fig, level_speed_collapsed_fig, level_speed_condition_fig, level_lick_collapsed_fig, level_lick_condition_fig, level_dist_collapsed_fig, level_dist_condition_fig, level_dist_condition_excl_last_fig, level_bout_collapsed_fig, level_bout_condition_fig, level_bout_avg_speed_collapsed_fig, level_bout_avg_speed_condition_fig, level_bout_avg_dist_collapsed_fig, level_bout_avg_dist_condition_fig, last_level_bar_fig, level_survivor_fig, time_to_level2_fig, epoch_speed_per_mouse_fig, epoch_speed_cond_fig, epoch_cap_per_mouse_fig, epoch_cap_cond_fig, epoch_speed_sess_per_mouse_fig, epoch_speed_sess_cond_fig, epoch_cap_sess_per_mouse_fig, epoch_cap_sess_cond_fig, epoch_speed_early_per_mouse_fig, epoch_speed_late_per_mouse_fig, epoch_speed_early_cond_fig, epoch_speed_late_cond_fig, epoch_cap_early_per_mouse_fig, epoch_cap_late_per_mouse_fig, epoch_cap_early_cond_fig, epoch_cap_late_cond_fig, epoch_speed_early_ev_per_mouse_fig, epoch_speed_late_ev_per_mouse_fig, epoch_speed_early_ev_cond_fig, epoch_speed_late_ev_cond_fig, epoch_cap_early_ev_per_mouse_fig, epoch_cap_late_ev_per_mouse_fig, epoch_cap_early_ev_cond_fig, epoch_cap_late_ev_cond_fig, epoch_speed_early_ev_cond_clean_fig, epoch_speed_late_ev_cond_clean_fig, epoch_cap_early_ev_cond_clean_fig, epoch_cap_late_ev_cond_clean_fig, epoch_speed_sess_cond_clean_fig, epoch_cap_sess_cond_clean_fig, epoch_speed_early_cond_clean_fig, epoch_speed_late_cond_clean_fig, epoch_cap_early_cond_clean_fig, epoch_cap_late_cond_clean_fig, punish_speed_per_mouse_fig, punish_speed_cond_fig, punish_cap_per_mouse_fig, punish_cap_cond_fig, punish_speed_sess_per_mouse_fig, punish_speed_sess_cond_fig, punish_cap_sess_per_mouse_fig, punish_cap_sess_cond_fig, punish_speed_sess_cond_clean_fig, punish_cap_sess_cond_clean_fig, sex_speed_fig, sex_distance_indiv_fig, sex_reward_indiv_fig, epoch_speed_sess_sex_per_mouse_fig, epoch_speed_sess_sex_fig, epoch_cap_sess_sex_per_mouse_fig, epoch_cap_sess_sex_fig, epoch_punish_speed_sess_sex_per_mouse_fig, epoch_punish_speed_sess_sex_fig, epoch_punish_cap_sess_sex_per_mouse_fig, epoch_punish_cap_sess_sex_fig, epoch_reward_speed_pre_post_fig, epoch_reward_speed_diff_fig, epoch_reward_cap_pre_post_fig, epoch_reward_cap_diff_fig, epoch_reward_speed_pre_post_entry_fig, epoch_reward_speed_diff_entry_fig, epoch_reward_speed_pre_post_entry_1s_fig, epoch_reward_speed_diff_entry_1s_fig, epoch_reward_lick_count_sess_per_mouse_fig, epoch_reward_lick_count_sess_cond_fig, epoch_reward_lick_count_sess_cond_clean_fig, epoch_punish_lick_count_sess_per_mouse_fig, epoch_punish_lick_count_sess_cond_fig, epoch_punish_speed_pre_post_fig, epoch_punish_speed_diff_fig, epoch_punish_speed_pre_post_entry_fig, epoch_punish_speed_diff_entry_fig, epoch_punish_cap_pre_post_fig, epoch_punish_cap_diff_fig, epoch_punish_cap_pre_post_entry_fig, epoch_punish_cap_diff_entry_fig, expl_speed_histogram_fig, expl_speed_distfit_fig, expl_speed_boxplot_fig, expl_speed_rm_anova_resid_fig, expl_cap_histogram_fig, expl_cap_boxplot_fig, expl_cap_rm_anova_resid_fig, expl_cap_distfit_fig, expl_lick_distfit_fig, expl_lick_boxplot_fig, expl_lick_rm_anova_resid_fig, expl_lick_rate_distfit_fig, expl_lick_reward_ratio_distfit_fig, all_results, _level_stats_data

def _run_weight_correlations(root, file_paths, animal_info):
    """Load a weight CSV, match to session data by mouse ID + date, and produce
    two correlation plots (Total Change vs reward count):
      1. All mice on one axes, each mouse its own colour.
      2. Two subplots split by starting condition.
    """
    from scipy.stats import kendalltau as _kendalltau

    # ── Load weight CSV ───────────────────────────────────────────────────────
    weight_csv_path = filedialog.askopenfilename(
        title='Select weight master CSV (ID, Date, Total Change …)',
        filetypes=[('CSV files', '*.csv'), ('All files', '*.*')],
        initialdir=os.path.dirname(file_paths[0]),
    ) or None
    if not weight_csv_path:
        print("No weight CSV selected — weight correlations cancelled.")
        return

    try:
        wdf = pd.read_csv(weight_csv_path)
        wdf.columns = wdf.columns.str.strip()
        # Normalise column names to lower-case for flexible matching
        wdf_cols_lower = {c.lower().replace(' ', '_'): c for c in wdf.columns}
        id_col    = wdf_cols_lower.get('id',    wdf.columns[0])
        date_col  = wdf_cols_lower.get('date',  None)
        dc_col    = wdf_cols_lower.get('daily_change', wdf_cols_lower.get('daily change', None))
        tc_col    = wdf_cols_lower.get('total_change', wdf_cols_lower.get('total change', None))
        if date_col is None or tc_col is None:
            # Try to find by position (Date=col5, Daily Change=col7, Total Change=col8)
            date_col = wdf.columns[5] if len(wdf.columns) > 5 else wdf.columns[1]
            tc_col   = wdf.columns[8] if len(wdf.columns) > 8 else wdf.columns[2]
        if dc_col is None:
            dc_col = wdf.columns[7] if len(wdf.columns) > 7 else wdf.columns[2]
        wdf[date_col] = pd.to_datetime(wdf[date_col], errors='coerce')
        wdf[tc_col]   = pd.to_numeric(wdf[tc_col],   errors='coerce')
        wdf[dc_col]   = pd.to_numeric(wdf[dc_col],   errors='coerce')
        # Build lookups: (mouse_id, date_normalised) -> total_change / daily_change
        weight_lookup       = {}
        daily_change_lookup = {}
        for _, row in wdf.iterrows():
            mid = str(row[id_col]).strip()
            dt  = row[date_col]
            if not pd.notna(dt):
                continue
            d = dt.normalize()
            tc = row[tc_col]
            dc = row[dc_col]
            if pd.notna(tc):
                weight_lookup[(mid, d)] = float(tc)
            if pd.notna(dc):
                daily_change_lookup[(mid, d)] = float(dc)
        print(f"Weight CSV loaded: {len(weight_lookup)} total-change entries, "
              f"{len(daily_change_lookup)} daily-change entries across "
              f"{len(wdf[id_col].unique())} mice.")
    except Exception as e:
        print(f"[ERROR] Cannot read weight CSV: {e}")
        return

    # ── Load session data ─────────────────────────────────────────────────────
    *_, all_results, _ = analyze_mouse_data(
        file_paths,
        ['s' if animal_info[os.path.basename(fp).split('_')[0]]['sex'] == 'male' else 'o'
         for fp in file_paths
         if os.path.basename(fp).split('_')[0] in animal_info],
        [animal_info[os.path.basename(fp).split('_')[0]]['starting_condition']
         for fp in file_paths
         if os.path.basename(fp).split('_')[0] in animal_info],
        transitions_csv_path=None,
        selected_plots=frozenset(),
    )

    # ── Build merged dataframe ────────────────────────────────────────────────
    _one_day = pd.Timedelta(days=1)
    records = []
    for result in all_results:
        mouse     = result['mouse']
        condition = result['starting_condition']
        df_r      = result['df']
        for _, row in df_r.iterrows():
            dt      = pd.Timestamp(row['date']).normalize()
            hits    = pd.to_numeric(row.get('hits', np.nan), errors='coerce')
            tc      = weight_lookup.get((mouse, dt), np.nan)
            next_dc = daily_change_lookup.get((mouse, dt + _one_day), np.nan)
            if pd.notna(hits) and (pd.notna(tc) or pd.notna(next_dc)):
                records.append({
                    'mouse':        mouse,
                    'condition':    condition,
                    'date':         dt,
                    'reward_count': float(hits),
                    'total_change': float(tc)      if pd.notna(tc)      else np.nan,
                    'next_day_dc':  float(next_dc) if pd.notna(next_dc) else np.nan,
                })

    if not records:
        print("[WARN] No overlapping dates found between weight CSV and session data.")
        return

    merged = pd.DataFrame(records)
    mice_sorted   = sorted(merged['mouse'].unique())
    cmap          = plt.get_cmap('tab20')
    mouse_colors  = {m: cmap(i / max(len(mice_sorted) - 1, 1))
                     for i, m in enumerate(mice_sorted)}
    conditions_sorted = sorted(merged['condition'].unique())

    def _add_corr_annotation(ax, x, y):
        """Add Kendall's tau and p-value annotation to axes."""
        mask = np.isfinite(x) & np.isfinite(y)
        if mask.sum() < 3:
            return
        tau, p = _kendalltau(x[mask], y[mask])
        sig = ('***' if p < 0.001 else '**' if p < 0.01
               else '*' if p < 0.05 else 'ns')
        ax.text(0.97, 0.05,
                f'\u03c4 = {tau:.3f}\np = {p:.3f}  {sig}',
                transform=ax.transAxes, ha='right', va='bottom',
                fontsize=8.5,
                bbox=dict(boxstyle='round,pad=0.3', facecolor='white',
                          edgecolor='lightgray', alpha=0.8))

    def _assumptions_figure(sub_df, title,
                             x_col='total_change', y_col='reward_count',
                             x_label='Total Weight Change (%)',
                             y_label='Reward Count per Session'):
        """Return a 2×2 assumption-check figure for a weight-correlation subset.

        Panels:
          Top-left  — Box plot of X variable with 1.5×IQR whiskers
          Top-right — Box plot of Y variable with 1.5×IQR whiskers
          Bot-left  — Scatter with linear fit + LOWESS (linearity check)
          Bot-right — χ² Q-Q plot of Mahalanobis D² (bivariate normality)
                      + marginal Shapiro-Wilk results annotated
        """
        from scipy.stats import chi2 as _chi2, shapiro as _shapiro

        both_df   = sub_df[[x_col, y_col, 'mouse']].dropna(subset=[x_col, y_col])
        xy        = both_df[[x_col, y_col]].values.astype(float)
        fin       = np.isfinite(xy[:, 0]) & np.isfinite(xy[:, 1])
        x_both    = xy[fin, 0]
        y_both    = xy[fin, 1]
        mouse_arr = both_df['mouse'].values[fin]
        n = len(x_both)

        fig, axes = plt.subplots(2, 2, figsize=(11, 8))
        fig.suptitle(f'Correlation Assumption Checks — {title}',
                     fontsize=13, fontweight='bold')

        # ── Panel 1 (top-left): Box plot — Total Weight Change (1.5×IQR) ─────
        ax1 = axes[0, 0]
        ax1.boxplot(x_both, vert=True, patch_artist=True,
                    boxprops=dict(facecolor='#aec6e8', color='#2c5f82'),
                    medianprops=dict(color='#c0392b', linewidth=2),
                    whiskerprops=dict(color='#2c5f82', linewidth=1.2),
                    capprops=dict(color='#2c5f82', linewidth=1.2),
                    flierprops=dict(marker='o', color='#e74c3c', markersize=5.5,
                                    linestyle='none', markerfacecolor='#e74c3c',
                                    alpha=0.75),
                    whis=1.5)
        ax1.set_xticks([1])
        ax1.set_xticklabels([x_label])
        ax1.set_ylabel('Value')
        ax1.set_title(f'Outlier Check: {x_label}\n(Whiskers = 1.5 × IQR)')
        ax1.spines['top'].set_visible(False)
        ax1.spines['right'].set_visible(False)
        if n > 0:
            q1x, q3x = np.percentile(x_both, [25, 75])
            iqr_x    = q3x - q1x
            lo_x, hi_x = q1x - 1.5 * iqr_x, q3x + 1.5 * iqr_x
            n_out_x  = int(np.sum((x_both < lo_x) | (x_both > hi_x)))
            ax1.text(0.97, 0.97,
                     f'n = {n}\nIQR = {iqr_x:.2f}\n'
                     f'Fences: [{lo_x:.2f}, {hi_x:.2f}]\nOutliers: {n_out_x}',
                     transform=ax1.transAxes, ha='right', va='top', fontsize=7.5,
                     bbox=dict(boxstyle='round,pad=0.3', facecolor='white',
                               edgecolor='lightgray', alpha=0.8))

        # ── Panel 2 (top-right): Box plot — Reward Count (1.5×IQR) ──────────
        ax2 = axes[0, 1]
        ax2.boxplot(y_both, vert=True, patch_artist=True,
                    boxprops=dict(facecolor='#a8e8c0', color='#2c7a4b'),
                    medianprops=dict(color='#c0392b', linewidth=2),
                    whiskerprops=dict(color='#2c7a4b', linewidth=1.2),
                    capprops=dict(color='#2c7a4b', linewidth=1.2),
                    flierprops=dict(marker='o', color='#e74c3c', markersize=5.5,
                                    linestyle='none', markerfacecolor='#e74c3c',
                                    alpha=0.75),
                    whis=1.5)
        ax2.set_xticks([1])
        ax2.set_xticklabels([y_label])
        ax2.set_ylabel('Value')
        ax2.set_title(f'Outlier Check: {y_label}\n(Whiskers = 1.5 × IQR)')
        ax2.spines['top'].set_visible(False)
        ax2.spines['right'].set_visible(False)
        if n > 0:
            q1y, q3y = np.percentile(y_both, [25, 75])
            iqr_y    = q3y - q1y
            lo_y, hi_y = q1y - 1.5 * iqr_y, q3y + 1.5 * iqr_y
            n_out_y  = int(np.sum((y_both < lo_y) | (y_both > hi_y)))
            ax2.text(0.97, 0.97,
                     f'n = {n}\nIQR = {iqr_y:.2f}\n'
                     f'Fences: [{lo_y:.2f}, {hi_y:.2f}]\nOutliers: {n_out_y}',
                     transform=ax2.transAxes, ha='right', va='top', fontsize=7.5,
                     bbox=dict(boxstyle='round,pad=0.3', facecolor='white',
                               edgecolor='lightgray', alpha=0.8))

        # ── Panel 3 (bottom-left): Scatter — Linearity check ─────────────────
        ax3 = axes[1, 0]
        for m in np.unique(mouse_arr):
            idx = mouse_arr == m
            ax3.scatter(x_both[idx], y_both[idx],
                        color=mouse_colors.get(m, 'gray'),
                        s=30, alpha=0.75, edgecolors='none', label=m)
        if n >= 2:
            coef1  = np.polyfit(x_both, y_both, 1)
            x_line = np.linspace(x_both.min(), x_both.max(), 200)
            ax3.plot(x_line, np.polyval(coef1, x_line),
                     color='black', linewidth=1.5, linestyle='--',
                     label='Linear fit', zorder=5)
        try:
            from statsmodels.nonparametric.smoothers_lowess import lowess as _lowess
            if n >= 5:
                order    = np.argsort(x_both)
                smoothed = _lowess(y_both[order], x_both[order],
                                   frac=0.5, return_sorted=True)
                ax3.plot(smoothed[:, 0], smoothed[:, 1],
                         color='#e74c3c', linewidth=1.5, linestyle='-',
                         label='LOWESS', zorder=6)
        except ImportError:
            if n >= 3:
                coef2 = np.polyfit(x_both, y_both, 2)
                ax3.plot(x_line, np.polyval(coef2, x_line),
                         color='#e74c3c', linewidth=1.5, linestyle='-',
                         label='Quadratic fit', zorder=6)
        ax3.set_xlabel(x_label)
        ax3.set_ylabel(y_label)
        ax3.set_title('Linearity Check\n(Dashed = linear fit,  Red = LOWESS / quadratic)')
        ax3.spines['top'].set_visible(False)
        ax3.spines['right'].set_visible(False)
        ax3.legend(fontsize=7, loc='best', framealpha=0.7)

        # ── Panel 4 (bottom-right): Bivariate normality — χ² Q-Q plot ────────
        ax4 = axes[1, 1]
        if n >= 4:
            X_mat    = np.column_stack([x_both, y_both])
            mean_vec = X_mat.mean(axis=0)
            cov_mat  = np.cov(X_mat.T)
            try:
                cov_inv   = np.linalg.inv(cov_mat)
                diff      = X_mat - mean_vec
                d2        = np.einsum('ij,jk,ik->i', diff, cov_inv, diff)
                d2_sorted = np.sort(d2)
                probs     = (np.arange(1, n + 1) - 0.5) / n
                chi2_q    = _chi2.ppf(probs, df=2)
                ax4.scatter(chi2_q, d2_sorted, s=22, alpha=0.75,
                            color='#5b7fce', edgecolors='none',
                            label='Observed D²')
                ref_max = max(float(chi2_q.max()), float(d2_sorted.max()))
                ax4.plot([0, ref_max], [0, ref_max], 'r--',
                         linewidth=1.2, label='Reference (normal)')
                sw_x = _shapiro(x_both)
                sw_y = _shapiro(y_both)
                sig_x = '*' if sw_x.pvalue < 0.05 else 'ns'
                sig_y = '*' if sw_y.pvalue < 0.05 else 'ns'
                ax4.text(0.03, 0.97,
                         'Shapiro-Wilk (marginals):\n'
                         f'  {x_label[:15]}: W={sw_x.statistic:.3f}, '
                         f'p={sw_x.pvalue:.3f} {sig_x}\n'
                         f'  {y_label[:15]}: W={sw_y.statistic:.3f}, '
                         f'p={sw_y.pvalue:.3f} {sig_y}',
                         transform=ax4.transAxes, ha='left', va='top',
                         fontsize=7.5,
                         bbox=dict(boxstyle='round,pad=0.3', facecolor='white',
                                   edgecolor='lightgray', alpha=0.8))
                ax4.legend(fontsize=7.5, loc='lower right')
            except np.linalg.LinAlgError:
                ax4.text(0.5, 0.5,
                         'Singular covariance matrix\n(insufficient data variance)',
                         ha='center', va='center',
                         transform=ax4.transAxes, fontsize=9)
        else:
            ax4.text(0.5, 0.5, f'Insufficient data (n = {n} < 4)',
                     ha='center', va='center',
                     transform=ax4.transAxes, fontsize=9)
        ax4.set_xlabel('Theoretical χ² Quantiles (df = 2)')
        ax4.set_ylabel('Mahalanobis Distance²')
        ax4.set_title('Bivariate Normality Check\n(χ² Q-Q of Mahalanobis D²)')
        ax4.spines['top'].set_visible(False)
        ax4.spines['right'].set_visible(False)

        fig.tight_layout()
        return fig

    # ── Assumption checks (one figure per group) — total change ──────────────
    fig_assump_all   = _assumptions_figure(merged.dropna(subset=['total_change']),
                                            'All Mice Combined')
    fig_assump_conds = {
        cond: _assumptions_figure(
            merged[merged['condition'] == cond].dropna(subset=['total_change']),
            f'Condition: {cond}'
        )
        for cond in conditions_sorted
    }

    # ── Assumption checks — reward count vs next-day daily change ─────────────
    _nd_x_label = 'Reward Count (Day D)'
    _nd_y_label = 'Daily Weight Change Day D+1 (%)'
    merged_nd = merged.dropna(subset=['reward_count', 'next_day_dc']).copy()
    if not merged_nd.empty:
        fig_assump_nd_all = _assumptions_figure(
            merged_nd, 'All Mice Combined \u2014 Next-Day Daily Change',
            x_col='reward_count', y_col='next_day_dc',
            x_label=_nd_x_label, y_label=_nd_y_label,
        )
        fig_assump_nd_conds = {
            cond: _assumptions_figure(
                merged_nd[merged_nd['condition'] == cond],
                f'Condition: {cond} \u2014 Next-Day Daily Change',
                x_col='reward_count', y_col='next_day_dc',
                x_label=_nd_x_label, y_label=_nd_y_label,
            )
            for cond in conditions_sorted
            if not merged_nd[merged_nd['condition'] == cond].empty
        }
    else:
        fig_assump_nd_all  = None
        fig_assump_nd_conds = {}
        print("[WARN] No next-day daily change data found \u2014 skipping next-day correlation plots.")

    # ── Plot 1: all mice combined ─────────────────────────────────────────────
    fig_all, ax_all = plt.subplots(figsize=(8, 6))
    for mouse in mice_sorted:
        sub = merged[merged['mouse'] == mouse]
        ax_all.scatter(sub['total_change'], sub['reward_count'],
                       color=mouse_colors[mouse], label=mouse,
                       s=38, alpha=0.82, edgecolors='none')
    # Overall regression line
    all_x = merged['total_change'].values
    all_y = merged['reward_count'].values
    mask  = np.isfinite(all_x) & np.isfinite(all_y)
    if mask.sum() >= 2:
        coef = np.polyfit(all_x[mask], all_y[mask], 1)
        x_line = np.linspace(all_x[mask].min(), all_x[mask].max(), 200)
        ax_all.plot(x_line, np.polyval(coef, x_line),
                    color='black', linewidth=1.5, linestyle='--', zorder=5)
    _add_corr_annotation(ax_all, all_x, all_y)
    ax_all.set_xlabel('Total Body Weight Change (%)')
    ax_all.set_ylabel('Reward Count per Session')
    ax_all.set_title('Total Weight Change vs Reward Count\n(all mice, all sessions — Kendall\'s \u03c4)')
    ax_all.spines['top'].set_visible(False)
    ax_all.spines['right'].set_visible(False)
    ax_all.legend(title='Mouse', bbox_to_anchor=(1.02, 1), loc='upper left',
                  fontsize=7.5, title_fontsize=8)
    fig_all.tight_layout()

    # ── Plot 2: subplots by starting condition ────────────────────────────────
    n_conds = len(conditions_sorted)
    fig_cond, axes_cond = plt.subplots(1, n_conds,
                                        figsize=(6 * n_conds, 6),
                                        sharey=False)
    if n_conds == 1:
        axes_cond = [axes_cond]
    for ax_c, cond in zip(axes_cond, conditions_sorted):
        sub_cond = merged[merged['condition'] == cond]
        for mouse in sorted(sub_cond['mouse'].unique()):
            sub_m = sub_cond[sub_cond['mouse'] == mouse]
            ax_c.scatter(sub_m['total_change'], sub_m['reward_count'],
                         color=mouse_colors[mouse], label=mouse,
                         s=38, alpha=0.82, edgecolors='none')
        cx = sub_cond['total_change'].values
        cy = sub_cond['reward_count'].values
        cmask = np.isfinite(cx) & np.isfinite(cy)
        if cmask.sum() >= 2:
            coef_c = np.polyfit(cx[cmask], cy[cmask], 1)
            xc_line = np.linspace(cx[cmask].min(), cx[cmask].max(), 200)
            ax_c.plot(xc_line, np.polyval(coef_c, xc_line),
                      color='black', linewidth=1.5, linestyle='--', zorder=5)
        _add_corr_annotation(ax_c, cx, cy)
        ax_c.set_xlabel('Total Body Weight Change (%)')
        ax_c.set_ylabel('Reward Count per Session')
        ax_c.set_title(f'Condition: {cond}')
        ax_c.spines['top'].set_visible(False)
        ax_c.spines['right'].set_visible(False)
        ax_c.legend(title='Mouse', fontsize=7.5, title_fontsize=8)
    fig_cond.suptitle('Total Weight Change vs Reward Count by Starting Condition (Kendall\'s \u03c4)',
                      fontsize=12, fontweight='bold')
    fig_cond.tight_layout()

    # ── Plot 3: next-day daily change — all mice combined ─────────────────────
    if merged_nd.empty:
        fig_nd_all = fig_nd_cond = None
    else:
        fig_nd_all, ax_nd_all = plt.subplots(figsize=(8, 6))
        for mouse in mice_sorted:
            sub = merged_nd[merged_nd['mouse'] == mouse]
            if sub.empty:
                continue
            ax_nd_all.scatter(sub['reward_count'], sub['next_day_dc'],
                              color=mouse_colors[mouse], label=mouse,
                              s=38, alpha=0.82, edgecolors='none')
        nd_x  = merged_nd['reward_count'].values
        nd_y  = merged_nd['next_day_dc'].values
        nd_fin = np.isfinite(nd_x) & np.isfinite(nd_y)
        if nd_fin.sum() >= 2:
            coef_nd = np.polyfit(nd_x[nd_fin], nd_y[nd_fin], 1)
            xl_nd   = np.linspace(nd_x[nd_fin].min(), nd_x[nd_fin].max(), 200)
            ax_nd_all.plot(xl_nd, np.polyval(coef_nd, xl_nd),
                           color='black', linewidth=1.5, linestyle='--', zorder=5)
        _add_corr_annotation(ax_nd_all, nd_x, nd_y)
        ax_nd_all.set_xlabel(_nd_x_label)
        ax_nd_all.set_ylabel(_nd_y_label)
        ax_nd_all.set_title(
            'Reward Count (Day D) vs Next-Day Daily Weight Change\n'
            '(all mice, all sessions \u2014 Kendall\'s \u03c4)')
        ax_nd_all.spines['top'].set_visible(False)
        ax_nd_all.spines['right'].set_visible(False)
        ax_nd_all.legend(title='Mouse', bbox_to_anchor=(1.02, 1), loc='upper left',
                         fontsize=7.5, title_fontsize=8)
        fig_nd_all.tight_layout()

        # ── Plot 4: next-day daily change — by starting condition ──────────────
        fig_nd_cond, axes_nd_cond = plt.subplots(1, n_conds,
                                                  figsize=(6 * n_conds, 6),
                                                  sharey=False)
        if n_conds == 1:
            axes_nd_cond = [axes_nd_cond]
        for ax_ndc, cond in zip(axes_nd_cond, conditions_sorted):
            sub_cond_nd = merged_nd[merged_nd['condition'] == cond]
            for mouse in sorted(sub_cond_nd['mouse'].unique()):
                sub_m = sub_cond_nd[sub_cond_nd['mouse'] == mouse]
                ax_ndc.scatter(sub_m['reward_count'], sub_m['next_day_dc'],
                               color=mouse_colors[mouse], label=mouse,
                               s=38, alpha=0.82, edgecolors='none')
            cndx  = sub_cond_nd['reward_count'].values
            cndy  = sub_cond_nd['next_day_dc'].values
            cnd_fin = np.isfinite(cndx) & np.isfinite(cndy)
            if cnd_fin.sum() >= 2:
                coef_ndc = np.polyfit(cndx[cnd_fin], cndy[cnd_fin], 1)
                xl_ndc   = np.linspace(cndx[cnd_fin].min(), cndx[cnd_fin].max(), 200)
                ax_ndc.plot(xl_ndc, np.polyval(coef_ndc, xl_ndc),
                            color='black', linewidth=1.5, linestyle='--', zorder=5)
            _add_corr_annotation(ax_ndc, cndx, cndy)
            ax_ndc.set_xlabel(_nd_x_label)
            ax_ndc.set_ylabel(_nd_y_label)
            ax_ndc.set_title(f'Condition: {cond}')
            ax_ndc.spines['top'].set_visible(False)
            ax_ndc.spines['right'].set_visible(False)
            ax_ndc.legend(title='Mouse', fontsize=7.5, title_fontsize=8)
        fig_nd_cond.suptitle(
            'Reward Count (Day D) vs Next-Day Daily Weight Change by Condition'
            ' (Kendall\'s \u03c4)',
            fontsize=12, fontweight='bold')
        fig_nd_cond.tight_layout()

    # ── Display and optionally save ───────────────────────────────────────────
    for fig_obj in ([fig_assump_all]
                    + list(fig_assump_conds.values())
                    + [fig_all, fig_cond]
                    + ([fig_assump_nd_all] if fig_assump_nd_all else [])
                    + list(fig_assump_nd_conds.values())
                    + ([fig_nd_all, fig_nd_cond] if fig_nd_all else [])):
        fig_obj.show()
    plt.show()

    plt.rcParams['font.family']    = 'sans-serif'
    plt.rcParams['font.sans-serif'] = ['Arial']
    plt.rcParams['svg.fonttype']   = 'none'
    figs_to_save = (
        [(fig_assump_all, 'weight_assumptions_all')]
        + [(fig_a, f'weight_assumptions_cond_{cond}')
           for cond, fig_a in fig_assump_conds.items()]
        + [(fig_all,  'weight_corr_all_mice'),
           (fig_cond, 'weight_corr_by_condition')]
        + ([(fig_assump_nd_all, 'weight_nd_assumptions_all')]
           if fig_assump_nd_all else [])
        + [(fig_a, f'weight_nd_assumptions_cond_{cond}')
           for cond, fig_a in fig_assump_nd_conds.items()]
        + ([(fig_nd_all,  'weight_nd_corr_all_mice'),
            (fig_nd_cond, 'weight_nd_corr_by_condition')]
           if fig_nd_all else [])
    )
    for fig_obj, fname in figs_to_save:
        save_path = filedialog.asksaveasfilename(
            defaultextension='.svg',
            filetypes=[('SVG files', '*.svg'), ('All files', '*.*')],
            title=f'Save {fname} as',
            initialfile=f'{fname}_{len(mice_sorted)}mice.svg',
        )
        if save_path:
            fig_obj.savefig(save_path, bbox_inches='tight', format='svg')
            print(f"Saved: {save_path}")


def _ask_mode(root):
    """Ask whether to generate plots or run the descriptive stats report.
    Returns 'plots', 'stats', or None if the dialog is dismissed."""
    result = [None]
    dialog = tk.Toplevel(root)
    dialog.title('Select Mode')
    dialog.resizable(False, False)
    dialog.grab_set()

    tk.Label(dialog, text='What would you like to do?',
             font=('Arial', 11, 'bold')).pack(padx=20, pady=(16, 8))

    def _choose(mode):
        result[0] = mode
        dialog.destroy()

    btn_frame = tk.Frame(dialog)
    btn_frame.pack(padx=20, pady=(4, 16))
    tk.Button(btn_frame, text='Generate Plots', width=20,
              command=lambda: _choose('plots')).pack(side='left', padx=6)
    tk.Button(btn_frame, text='Exploratory / Distribution Plots', width=32,
              command=lambda: _choose('expl')).pack(side='left', padx=6)
    tk.Button(btn_frame, text='Descriptive Stats Report', width=24,
              command=lambda: _choose('stats')).pack(side='left', padx=6)
    tk.Button(btn_frame, text='Weight Correlations', width=22,
              command=lambda: _choose('weight_corr')).pack(side='left', padx=6)

    dialog.update_idletasks()
    dialog.geometry(
        f"+{root.winfo_screenwidth() // 2 - dialog.winfo_reqwidth() // 2}"
        f"+{root.winfo_screenheight() // 2 - dialog.winfo_reqheight() // 2}"
    )
    root.wait_window(dialog)
    return result[0]


def main():
    # Create and hide the root window
    root = tk.Tk()
    root.withdraw()

    # Open file dialog to select the master CSV file
    master_csv_path = filedialog.askopenfilename(
        title='Select master CSV file (containing animal_id, sex, starting_condition)',
        filetypes=[('CSV files', '*.csv')],
        initialdir=os.getcwd()
    )
    
    if not master_csv_path:
        print("No master CSV file selected. Exiting...")
        return
    
    # Read the master CSV file
    try:
        master_df = pd.read_csv(master_csv_path)
        # Strip whitespace from column names and values
        master_df.columns = master_df.columns.str.strip()
        master_df['animal_id'] = master_df['animal_id'].str.strip()
        master_df['sex'] = master_df['sex'].str.strip().str.lower()
        master_df['starting_condition'] = master_df['starting_condition'].str.strip()
        
        # Create a dictionary mapping animal_id to sex and starting_condition
        animal_info = {}
        for _, row in master_df.iterrows():
            animal_info[row['animal_id']] = {
                'sex': row['sex'],
                'starting_condition': row['starting_condition']
            }
        
        print(f"Loaded master CSV with {len(animal_info)} animals")
        
    except Exception as e:
        print(f"Error reading master CSV file: {str(e)}")
        return

    # Open file dialog to select multiple data files
    file_paths = filedialog.askopenfilenames(
        title='Select mouse data files',
        filetypes=[('CSV files', '*.csv')],
        initialdir=os.getcwd()  # Start in current directory
    )
    
    if not file_paths:
        print("No file selected. Exiting...")
        return

    # ── Mode selection ────────────────────────────────────────────────────────
    mode = _ask_mode(root)
    if mode is None:
        print("No mode selected. Exiting...")
        return

    # ── Extract markers and starting conditions (shared by both modes) ────────
    markers = []
    starting_conditions = []
    for file_path in file_paths:
        mouse_name = os.path.basename(file_path).split("_")[0]
        if mouse_name in animal_info:
            sex = animal_info[mouse_name]['sex']
            marker = 's' if sex == 'male' else 'o'
            markers.append(marker)
            starting_conditions.append(animal_info[mouse_name]['starting_condition'])
            print(f"{mouse_name}: sex={sex}, marker={marker}, condition={animal_info[mouse_name]['starting_condition']}")
        else:
            print(f"Warning: {mouse_name} not found in master CSV file. Skipping...")
            continue

    # ── STATS mode ────────────────────────────────────────────────────────────
    if mode == 'stats':
        transitions_csv_path = filedialog.askopenfilename(
            title='Select transitions CSV for level stats — cancel to skip level sheets',
            filetypes=[('CSV files', '*.csv'), ('All files', '*.*')],
            initialdir=os.path.dirname(file_paths[0]),
        ) or None
        if transitions_csv_path:
            print(f"Transitions CSV: {os.path.basename(transitions_csv_path)}")
        else:
            print("No transitions CSV selected — level sheets will be omitted.")

        output_dir = filedialog.askdirectory(
            title='Select output folder for the stats report',
            initialdir=os.path.dirname(file_paths[0]),
        ) or None

        *_, all_results, _level_stats_data = analyze_mouse_data(
            file_paths, markers, starting_conditions,
            transitions_csv_path=transitions_csv_path,
            selected_plots=frozenset(),
        )
        generate_descriptive_stats_report(all_results, _level_stats_data, output_dir=output_dir)
        return

    # ── WEIGHT CORRELATIONS mode ──────────────────────────────────────────────
    if mode == 'weight_corr':
        _run_weight_correlations(root, file_paths, animal_info)
        return

    # ── PLOTS mode ────────────────────────────────────────────────────────────
    if mode == 'expl':
        selected_plots = _ask_plot_selection(
            root, labels=_EXPL_PLOT_LABELS,
            title='Select Exploratory / Distribution Plots',
        )
    else:
        selected_plots = _ask_plot_selection(root)
    if not selected_plots:
        print("No plots selected. Exiting...")
        return

    # Select transitions CSV only if a level plot was requested
    transitions_csv_path = None
    if any(k in selected_plots for k in ('levels', 'level_speed', 'level_speed_condition',
                                          'level_lick', 'level_lick_condition',
                                          'level_dist', 'level_dist_condition',
                                          'level_dist_condition_excl_last',
                                          'level_bout', 'level_bout_condition',
                                          'level_bout_avg_speed', 'level_bout_avg_speed_condition',
                                          'level_bout_avg_dist', 'level_bout_avg_dist_condition',
                                          'last_level_bar', 'level_survivor',
                                          'time_to_level2')):
        transitions_csv_path = filedialog.askopenfilename(
            title='Select transitions CSV (from level_sorter.py) — cancel to skip level plot',
            filetypes=[('CSV files', '*.csv'), ('All files', '*.*')],
            initialdir=os.path.dirname(file_paths[0]),
        ) or None
        if transitions_csv_path:
            print(f"Transitions CSV: {os.path.basename(transitions_csv_path)}")
        else:
            print("No transitions CSV selected — level plot will be empty.")

    # Analyze data and plot results
    speed_fig, sensitivity_fig, lick_fig, reward_fig, lick_reward_ratio_fig, false_alarm_fig, correct_rejection_fig, specificity_fig, dprime_fig, avg_reward_fig, sex_reward_fig, avg_sex_speed_fig, distance_fig, bout_count_fig, avg_bout_count_fig, rewards_per_bout_fig, first_lick_latency_fig, condition_rewards_per_bout_fig, condition_rewards_per_bout_bar_fig, condition_first_lick_latency_fig, condition_first_lick_latency_bar_fig, condition_lick_after_reward_prop_fig, condition_lick_after_reward_prop_bar_fig, weekday_reward_bar_fig, weekday_reward_bar_condition_fig, bout_avg_speed_fig, bout_avg_dist_fig, sex_distance_fig, condition_distance_fig, condition_distance_bar_fig, total_distance_bar_fig, avg_lick_rate_fig, sex_lick_rate_fig, condition_reward_fig, condition_speed_fig, condition_bout_count_fig, condition_bout_avg_speed_fig, condition_bout_avg_dist_fig, condition_lick_fig, condition_lick_rate_fig, condition_lick_reward_ratio_fig, condition_lick_reward_ratio_bar_fig, condition_punish_zone_pct_bar_fig, condition_bar_fig, condition_speed_bar_fig, condition_bout_count_bar_fig, condition_bout_avg_speed_bar_fig, condition_bout_avg_dist_bar_fig, condition_lick_bar_fig, level_reward_fig, level_speed_collapsed_fig, level_speed_condition_fig, level_lick_collapsed_fig, level_lick_condition_fig, level_dist_collapsed_fig, level_dist_condition_fig, level_dist_condition_excl_last_fig, level_bout_collapsed_fig, level_bout_condition_fig, level_bout_avg_speed_collapsed_fig, level_bout_avg_speed_condition_fig, level_bout_avg_dist_collapsed_fig, level_bout_avg_dist_condition_fig, last_level_bar_fig, level_survivor_fig, time_to_level2_fig, epoch_speed_per_mouse_fig, epoch_speed_cond_fig, epoch_cap_per_mouse_fig, epoch_cap_cond_fig, epoch_speed_sess_per_mouse_fig, epoch_speed_sess_cond_fig, epoch_cap_sess_per_mouse_fig, epoch_cap_sess_cond_fig, epoch_speed_early_per_mouse_fig, epoch_speed_late_per_mouse_fig, epoch_speed_early_cond_fig, epoch_speed_late_cond_fig, epoch_cap_early_per_mouse_fig, epoch_cap_late_per_mouse_fig, epoch_cap_early_cond_fig, epoch_cap_late_cond_fig, epoch_speed_early_ev_per_mouse_fig, epoch_speed_late_ev_per_mouse_fig, epoch_speed_early_ev_cond_fig, epoch_speed_late_ev_cond_fig, epoch_cap_early_ev_per_mouse_fig, epoch_cap_late_ev_per_mouse_fig, epoch_cap_early_ev_cond_fig, epoch_cap_late_ev_cond_fig, epoch_speed_early_ev_cond_clean_fig, epoch_speed_late_ev_cond_clean_fig, epoch_cap_early_ev_cond_clean_fig, epoch_cap_late_ev_cond_clean_fig, epoch_speed_sess_cond_clean_fig, epoch_cap_sess_cond_clean_fig, epoch_speed_early_cond_clean_fig, epoch_speed_late_cond_clean_fig, epoch_cap_early_cond_clean_fig, epoch_cap_late_cond_clean_fig, punish_speed_per_mouse_fig, punish_speed_cond_fig, punish_cap_per_mouse_fig, punish_cap_cond_fig, punish_speed_sess_per_mouse_fig, punish_speed_sess_cond_fig, punish_cap_sess_per_mouse_fig, punish_cap_sess_cond_fig, punish_speed_sess_cond_clean_fig, punish_cap_sess_cond_clean_fig, sex_speed_fig, sex_distance_indiv_fig, sex_reward_indiv_fig, epoch_speed_sess_sex_per_mouse_fig, epoch_speed_sess_sex_fig, epoch_cap_sess_sex_per_mouse_fig, epoch_cap_sess_sex_fig, epoch_punish_speed_sess_sex_per_mouse_fig, epoch_punish_speed_sess_sex_fig, epoch_punish_cap_sess_sex_per_mouse_fig, epoch_punish_cap_sess_sex_fig, epoch_reward_speed_pre_post_fig, epoch_reward_speed_diff_fig, epoch_reward_cap_pre_post_fig, epoch_reward_cap_diff_fig, epoch_reward_speed_pre_post_entry_fig, epoch_reward_speed_diff_entry_fig, epoch_reward_speed_pre_post_entry_1s_fig, epoch_reward_speed_diff_entry_1s_fig, epoch_reward_lick_count_sess_per_mouse_fig, epoch_reward_lick_count_sess_cond_fig, epoch_reward_lick_count_sess_cond_clean_fig, epoch_punish_lick_count_sess_per_mouse_fig, epoch_punish_lick_count_sess_cond_fig, epoch_punish_speed_pre_post_fig, epoch_punish_speed_diff_fig, epoch_punish_speed_pre_post_entry_fig, epoch_punish_speed_diff_entry_fig, epoch_punish_cap_pre_post_fig, epoch_punish_cap_diff_fig, epoch_punish_cap_pre_post_entry_fig, epoch_punish_cap_diff_entry_fig, expl_speed_histogram_fig, expl_speed_distfit_fig, expl_speed_boxplot_fig, expl_speed_rm_anova_resid_fig, expl_cap_histogram_fig, expl_cap_boxplot_fig, expl_cap_rm_anova_resid_fig, expl_cap_distfit_fig, expl_lick_distfit_fig, expl_lick_boxplot_fig, expl_lick_rm_anova_resid_fig, expl_lick_rate_distfit_fig, expl_lick_reward_ratio_distfit_fig, all_results, _level_stats_data = analyze_mouse_data(
        file_paths, markers, starting_conditions,
        transitions_csv_path=transitions_csv_path,
        selected_plots=selected_plots,
    )

    # All generated figures (None entries are skipped)
    all_figs = [f for f in [
        speed_fig, sensitivity_fig, lick_fig, reward_fig, lick_reward_ratio_fig,
        false_alarm_fig, correct_rejection_fig, specificity_fig, dprime_fig,
        avg_reward_fig, sex_reward_fig, avg_sex_speed_fig,
        distance_fig, bout_count_fig, avg_bout_count_fig, rewards_per_bout_fig, first_lick_latency_fig, condition_rewards_per_bout_fig, condition_rewards_per_bout_bar_fig, condition_first_lick_latency_fig, condition_first_lick_latency_bar_fig, condition_lick_after_reward_prop_fig, condition_lick_after_reward_prop_bar_fig, weekday_reward_bar_fig, weekday_reward_bar_condition_fig, bout_avg_speed_fig, bout_avg_dist_fig, sex_distance_fig, condition_distance_fig,
        condition_distance_bar_fig, total_distance_bar_fig,
        avg_lick_rate_fig, sex_lick_rate_fig,
        condition_reward_fig, condition_speed_fig, condition_bout_count_fig, condition_bout_avg_speed_fig, condition_bout_avg_dist_fig, condition_lick_fig, condition_lick_rate_fig, condition_lick_reward_ratio_fig, condition_lick_reward_ratio_bar_fig,
        condition_punish_zone_pct_bar_fig,
        condition_bar_fig, condition_speed_bar_fig, condition_bout_count_bar_fig, condition_bout_avg_speed_bar_fig, condition_bout_avg_dist_bar_fig, condition_lick_bar_fig,
        level_reward_fig, level_speed_collapsed_fig, level_speed_condition_fig,
        level_lick_collapsed_fig, level_lick_condition_fig,
        level_dist_collapsed_fig, level_dist_condition_fig,
        level_dist_condition_excl_last_fig,
        level_bout_collapsed_fig, level_bout_condition_fig,
        level_bout_avg_speed_collapsed_fig, level_bout_avg_speed_condition_fig,
        level_bout_avg_dist_collapsed_fig, level_bout_avg_dist_condition_fig,
        last_level_bar_fig, level_survivor_fig, time_to_level2_fig,
        epoch_speed_per_mouse_fig, epoch_speed_cond_fig,
        epoch_cap_per_mouse_fig, epoch_cap_cond_fig,
        epoch_speed_sess_per_mouse_fig, epoch_speed_sess_cond_fig,
        epoch_cap_sess_per_mouse_fig, epoch_cap_sess_cond_fig,
        epoch_speed_early_per_mouse_fig, epoch_speed_late_per_mouse_fig,
        epoch_speed_early_cond_fig, epoch_speed_late_cond_fig,
        epoch_cap_early_per_mouse_fig, epoch_cap_late_per_mouse_fig,
        epoch_cap_early_cond_fig, epoch_cap_late_cond_fig,
        epoch_speed_sess_cond_clean_fig, epoch_cap_sess_cond_clean_fig,
        epoch_speed_early_cond_clean_fig, epoch_speed_late_cond_clean_fig,
        epoch_cap_early_cond_clean_fig, epoch_cap_late_cond_clean_fig,
        epoch_speed_early_ev_per_mouse_fig, epoch_speed_late_ev_per_mouse_fig,
        epoch_speed_early_ev_cond_fig, epoch_speed_late_ev_cond_fig,
        epoch_cap_early_ev_per_mouse_fig, epoch_cap_late_ev_per_mouse_fig,
        epoch_cap_early_ev_cond_fig, epoch_cap_late_ev_cond_fig,
        epoch_speed_early_ev_cond_clean_fig, epoch_speed_late_ev_cond_clean_fig,
        epoch_cap_early_ev_cond_clean_fig, epoch_cap_late_ev_cond_clean_fig,
        punish_speed_per_mouse_fig, punish_speed_cond_fig,
        punish_cap_per_mouse_fig, punish_cap_cond_fig,
        punish_speed_sess_per_mouse_fig, punish_speed_sess_cond_fig,
        punish_cap_sess_per_mouse_fig, punish_cap_sess_cond_fig,
        punish_speed_sess_cond_clean_fig, punish_cap_sess_cond_clean_fig,
        sex_speed_fig, sex_distance_indiv_fig, sex_reward_indiv_fig,
        epoch_speed_sess_sex_per_mouse_fig, epoch_speed_sess_sex_fig,
        epoch_cap_sess_sex_per_mouse_fig, epoch_cap_sess_sex_fig,
        epoch_punish_speed_sess_sex_per_mouse_fig, epoch_punish_speed_sess_sex_fig,
        epoch_punish_cap_sess_sex_per_mouse_fig, epoch_punish_cap_sess_sex_fig,
        epoch_reward_speed_pre_post_fig,
        epoch_reward_speed_diff_fig,
        epoch_reward_cap_pre_post_fig,
        epoch_reward_cap_diff_fig,
        epoch_reward_speed_pre_post_entry_fig,
        epoch_reward_speed_diff_entry_fig,
        epoch_reward_speed_pre_post_entry_1s_fig,
        epoch_reward_speed_diff_entry_1s_fig,
        epoch_reward_lick_count_sess_per_mouse_fig,
        epoch_reward_lick_count_sess_cond_fig,
        epoch_reward_lick_count_sess_cond_clean_fig,
        epoch_punish_lick_count_sess_per_mouse_fig,
        epoch_punish_lick_count_sess_cond_fig,
        epoch_punish_speed_pre_post_fig,
        epoch_punish_speed_diff_fig,
        epoch_punish_speed_pre_post_entry_fig,
        epoch_punish_speed_diff_entry_fig,
        epoch_punish_cap_pre_post_fig,
        epoch_punish_cap_diff_fig,
        epoch_punish_cap_pre_post_entry_fig,
        epoch_punish_cap_diff_entry_fig,
        expl_speed_histogram_fig,
        expl_speed_distfit_fig,
        expl_speed_boxplot_fig,
        expl_speed_rm_anova_resid_fig,
        expl_cap_histogram_fig,
        expl_cap_boxplot_fig,
        expl_cap_rm_anova_resid_fig,
        expl_lick_distfit_fig,
        expl_lick_boxplot_fig,
        expl_lick_rm_anova_resid_fig,
        expl_lick_rate_distfit_fig,
        expl_lick_reward_ratio_distfit_fig,
    ] if f is not None]

    # Configure all figures (add legend only when labeled artists exist, then tight layout)
    for fig in all_figs:
        plt.figure(fig.number)
        handles, labels = plt.gca().get_legend_handles_labels()
        if labels:
            if len(file_paths) > 10:
                plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
                plt.subplots_adjust(right=0.85)
            else:
                plt.legend()
        plt.tight_layout()

    # Display all plots
    for fig in all_figs:
        fig.show()
    plt.show()

    # Save all selected plots automatically
    plt.rcParams['font.family'] = 'sans-serif'
    plt.rcParams['font.sans-serif'] = ['Arial']
    plt.rcParams['svg.fonttype'] = 'none'

    # Plot configurations to save (skip None figures)
    plot_configs = [
        (speed_fig,             'speed',               'Speed plot'),
        (sensitivity_fig,       'sensitivity',         'Sensitivity plot'),
        (lick_fig,              'lick_count',          'Lick count plot'),
        (reward_fig,            'reward_count',        'Reward count plot'),
        (lick_reward_ratio_fig, 'lick_reward_ratio',   'Lick count / Reward count ratio plot'),
        (false_alarm_fig,       'false_alarms',        'False alarms plot'),
        (correct_rejection_fig, 'correct_rejections',  'Correct rejections plot'),
        (specificity_fig,       'specificity',         'Specificity plot'),
        (dprime_fig,            'dprime',              "d' plot"),
        (avg_reward_fig,        'avg_reward',          'Average rewards plot'),
        (sex_reward_fig,        'sex_reward',          'Sex-specific average rewards plot'),
        (avg_sex_speed_fig,     'avg_sex_speed',       'Sex-specific average speed plot'),
        (distance_fig,          'distance',            'Distance per session plot'),
        (bout_count_fig,         'bout_count',          'Locomotion bout count plot'),
        (avg_bout_count_fig,     'avg_bout_count',      'Average bout count across all mice plot'),
        (rewards_per_bout_fig,            'rewards_per_bout',            'Average rewards per locomotion bout per session plot'),
        (first_lick_latency_fig,              'first_lick_latency',              'Average first-lick latency after reward delivery per session plot'),
        (condition_rewards_per_bout_fig,  'condition_rewards_per_bout',  'Condition-based rewards per bout plot'),
        (condition_rewards_per_bout_bar_fig, 'condition_rewards_per_bout_bar', 'Condition rewards per bout collapsed bar chart'),
        (condition_first_lick_latency_fig,     'condition_first_lick_latency',     'Condition-based first-lick latency line plot'),
        (condition_first_lick_latency_bar_fig, 'condition_first_lick_latency_bar', 'Condition first-lick latency collapsed bar chart'),
        (condition_lick_after_reward_prop_fig,     'lick_after_reward_prop',     'Condition: proportion of reward deliveries with licks — line plot'),
        (condition_lick_after_reward_prop_bar_fig, 'lick_after_reward_prop_bar', 'Condition: proportion of reward deliveries with licks — collapsed bar chart'),
        (weekday_reward_bar_fig,             'weekday_reward_bar',             'Weekday reward count bar chart (all mice pooled)'),
        (weekday_reward_bar_condition_fig,   'weekday_reward_bar_condition',   'Weekday reward count bar chart (by condition)'),
        (bout_avg_speed_fig,     'bout_avg_speed',      'Average speed per locomotion bout plot'),
        (bout_avg_dist_fig,      'bout_avg_dist',       'Average distance per locomotion bout plot'),
        (sex_distance_fig,      'sex_distance',        'Sex-specific distance per session plot'),
        (condition_distance_fig,'condition_distance',  'Condition-based distance per session plot'),
        (condition_distance_bar_fig, 'condition_distance_bar', 'Condition distance collapsed bar chart'),
        (total_distance_bar_fig,     'total_distance_bar',     'Total distance per mouse collapsed bar chart'),
        (avg_lick_rate_fig,     'avg_lick_rate',       'Average lick rate plot'),
        (sex_lick_rate_fig,     'sex_lick_rate',       'Sex-specific lick rate plot'),
        (condition_reward_fig,  'condition_reward',    'Condition-based average rewards plot'),
        (condition_speed_fig,   'condition_speed',     'Condition-based average speed plot'),
        (condition_bout_count_fig, 'condition_bout_count', 'Condition-based bout count plot'),
        (condition_bout_avg_speed_fig, 'condition_bout_avg_speed', 'Condition-based avg speed per bout plot'),
        (condition_bout_avg_dist_fig,  'condition_bout_avg_dist',  'Condition-based avg distance per bout plot'),
        (condition_lick_fig,              'condition_lick',              'Condition-based average lick count plot'),
        (condition_lick_rate_fig,         'condition_lick_rate',         'Condition-based lick rate plot'),
        (condition_lick_reward_ratio_fig,     'condition_lick_reward_ratio',     'Condition-based lick/reward ratio plot'),
        (condition_lick_reward_ratio_bar_fig,  'condition_lick_reward_ratio_bar', 'Condition lick/reward ratio collapsed bar chart'),
        (condition_punish_zone_pct_bar_fig,    'condition_punish_zone_pct_bar',   'Condition % punishment zones collapsed bar chart'),
        (condition_bar_fig,      'condition_bar',       'Condition collapsed bar chart'),
        (condition_speed_bar_fig,'condition_speed_bar', 'Condition speed collapsed bar chart'),
        (condition_bout_count_bar_fig, 'condition_bout_count_bar', 'Condition bout count collapsed bar chart'),
        (condition_bout_avg_speed_bar_fig, 'condition_bout_avg_speed_bar', 'Condition avg speed per bout bar chart'),
        (condition_bout_avg_dist_bar_fig,  'condition_bout_avg_dist_bar',  'Condition avg distance per bout bar chart'),
        (condition_lick_bar_fig, 'condition_lick_bar',  'Condition lick rate collapsed bar chart'),
        (level_reward_fig,              'level_reward',          'Level-based average rewards plot'),
        (level_speed_collapsed_fig,     'level_speed',           'Level-based average speed — collapsed'),
        (level_speed_condition_fig,     'level_speed_condition', 'Level-based average speed — by condition'),
        (level_lick_collapsed_fig,      'level_lick',            'Level-based average lick rate — collapsed'),
        (level_lick_condition_fig,      'level_lick_condition',  'Level-based average lick rate — by condition'),
        (level_dist_collapsed_fig,      'level_dist',            'Level-based distance — collapsed'),
        (level_dist_condition_fig,      'level_dist_condition',  'Level-based distance — by condition'),
        (level_dist_condition_excl_last_fig, 'level_dist_condition_excl_last', 'Level-based distance — by condition, last level excluded'),
        (level_bout_collapsed_fig,   'level_bout',           'Level-based bout count — collapsed'),
        (level_bout_condition_fig,   'level_bout_condition', 'Level-based bout count — by condition'),
        (level_bout_avg_speed_collapsed_fig,  'level_bout_avg_speed',           'Level-based avg speed per bout — collapsed'),
        (level_bout_avg_speed_condition_fig,  'level_bout_avg_speed_condition', 'Level-based avg speed per bout — by condition'),
        (level_bout_avg_dist_collapsed_fig,   'level_bout_avg_dist',            'Level-based avg distance per bout — collapsed'),
        (level_bout_avg_dist_condition_fig,   'level_bout_avg_dist_condition',  'Level-based avg distance per bout — by condition'),
        (last_level_bar_fig,                  'last_level_bar',                 'Final level reached per mouse bar chart (by condition)'),
        (level_survivor_fig,                  'level_survivor',                 'Level attainment survivor plot (proportion of mice per level, by condition)'),
        (time_to_level2_fig,                  'time_to_level2',                 'Cumulative time to first level 1→2 transition — bar chart by condition'),
        (epoch_speed_per_mouse_fig,      'epoch_reward_speed_per_mouse',           'Speed epoch (event) — per mouse'),
        (epoch_speed_cond_fig,           'epoch_reward_speed_condition',           'Speed epoch (event) — by condition'),
        (epoch_cap_per_mouse_fig,        'epoch_reward_cap_per_mouse',             'Capacitive epoch (event) — per mouse'),
        (epoch_cap_cond_fig,             'epoch_reward_cap_condition',             'Capacitive epoch (event) — by condition'),
        (epoch_speed_sess_per_mouse_fig, 'epoch_reward_speed_sess_per_mouse',      'Speed epoch (session) — per mouse'),
        (epoch_speed_sess_cond_fig,      'epoch_reward_speed_sess_condition',      'Speed epoch (session) — by condition'),
        (epoch_cap_sess_per_mouse_fig,   'epoch_reward_cap_sess_per_mouse',        'Capacitive epoch (session) — per mouse'),
        (epoch_cap_sess_cond_fig,        'epoch_reward_cap_sess_condition',        'Capacitive epoch (session) — by condition'),
        (epoch_speed_early_per_mouse_fig, 'epoch_reward_speed_early_per_mouse',     'Speed epoch (early sessions) — per mouse'),
        (epoch_speed_early_cond_fig,      'epoch_reward_speed_early_condition',     'Speed epoch (early sessions) — by condition'),
        (epoch_speed_late_per_mouse_fig,  'epoch_reward_speed_late_per_mouse',      'Speed epoch (late sessions) — per mouse'),
        (epoch_speed_late_cond_fig,       'epoch_reward_speed_late_condition',      'Speed epoch (late sessions) — by condition'),
        (epoch_cap_early_per_mouse_fig,   'epoch_reward_cap_early_per_mouse',       'Capacitive epoch (early sessions) — per mouse'),
        (epoch_cap_early_cond_fig,        'epoch_reward_cap_early_condition',       'Capacitive epoch (early sessions) — by condition'),
        (epoch_cap_late_per_mouse_fig,    'epoch_reward_cap_late_per_mouse',        'Capacitive epoch (late sessions) — per mouse'),
        (epoch_cap_late_cond_fig,         'epoch_reward_cap_late_condition',        'Capacitive epoch (late sessions) — by condition'),
        (epoch_speed_sess_cond_clean_fig,       'epoch_reward_speed_sess_condition_clean',       'Speed epoch (session) — by condition, no individual traces'),
        (epoch_cap_sess_cond_clean_fig,         'epoch_reward_cap_sess_condition_clean',         'Capacitive epoch (session) — by condition, no individual traces'),
        (epoch_speed_early_cond_clean_fig,      'epoch_reward_speed_early_condition_clean',      'Speed epoch (early sessions) — by condition, no individual traces'),
        (epoch_speed_late_cond_clean_fig,       'epoch_reward_speed_late_condition_clean',       'Speed epoch (late sessions) — by condition, no individual traces'),
        (epoch_cap_early_cond_clean_fig,        'epoch_reward_cap_early_condition_clean',        'Capacitive epoch (early sessions) — by condition, no individual traces'),
        (epoch_cap_late_cond_clean_fig,         'epoch_reward_cap_late_condition_clean',         'Capacitive epoch (late sessions) — by condition, no individual traces'),
        (epoch_speed_early_ev_per_mouse_fig, 'epoch_reward_speed_early_ev_per_mouse', 'Speed epoch ev (early sessions) — per mouse'),
        (epoch_speed_early_ev_cond_fig,      'epoch_reward_speed_early_ev_condition', 'Speed epoch ev (early sessions) — by condition'),
        (epoch_speed_late_ev_per_mouse_fig,  'epoch_reward_speed_late_ev_per_mouse',  'Speed epoch ev (late sessions) — per mouse'),
        (epoch_speed_late_ev_cond_fig,       'epoch_reward_speed_late_ev_condition',  'Speed epoch ev (late sessions) — by condition'),
        (epoch_cap_early_ev_per_mouse_fig,   'epoch_reward_cap_early_ev_per_mouse',   'Capacitive epoch ev (early sessions) — per mouse'),
        (epoch_cap_early_ev_cond_fig,        'epoch_reward_cap_early_ev_condition',   'Capacitive epoch ev (early sessions) — by condition'),
        (epoch_cap_late_ev_per_mouse_fig,    'epoch_reward_cap_late_ev_per_mouse',    'Capacitive epoch ev (late sessions) — per mouse'),
        (epoch_cap_late_ev_cond_fig,         'epoch_reward_cap_late_ev_condition',    'Capacitive epoch ev (late sessions) — by condition'),
        (epoch_speed_early_ev_cond_clean_fig,  'epoch_reward_speed_early_ev_condition_clean', 'Speed epoch ev (early sessions) — by condition, no individual traces'),
        (epoch_speed_late_ev_cond_clean_fig,   'epoch_reward_speed_late_ev_condition_clean',  'Speed epoch ev (late sessions) — by condition, no individual traces'),
        (epoch_cap_early_ev_cond_clean_fig,    'epoch_reward_cap_early_ev_condition_clean',   'Capacitive epoch ev (early sessions) — by condition, no individual traces'),
        (epoch_cap_late_ev_cond_clean_fig,     'epoch_reward_cap_late_ev_condition_clean',    'Capacitive epoch ev (late sessions) — by condition, no individual traces'),
        (punish_speed_per_mouse_fig,          'epoch_punish_speed_per_mouse',          'Speed epoch (punish, event) — per mouse'),
        (punish_speed_cond_fig,               'epoch_punish_speed_condition',          'Speed epoch (punish, event) — by condition'),
        (punish_cap_per_mouse_fig,            'epoch_punish_cap_per_mouse',            'Capacitive epoch (punish, event) — per mouse'),
        (punish_cap_cond_fig,                 'epoch_punish_cap_condition',            'Capacitive epoch (punish, event) — by condition'),
        (punish_speed_sess_per_mouse_fig,     'epoch_punish_speed_sess_per_mouse',     'Speed epoch (punish, session) — per mouse'),
        (punish_speed_sess_cond_fig,          'epoch_punish_speed_sess_condition',     'Speed epoch (punish, session) — by condition'),
        (punish_cap_sess_per_mouse_fig,       'epoch_punish_cap_sess_per_mouse',       'Capacitive epoch (punish, session) — per mouse'),
        (punish_cap_sess_cond_fig,            'epoch_punish_cap_sess_condition',       'Capacitive epoch (punish, session) — by condition'),
        (punish_speed_sess_cond_clean_fig,    'epoch_punish_speed_sess_condition_clean','Speed epoch (punish, session) — by condition, no individual traces'),
        (punish_cap_sess_cond_clean_fig,      'epoch_punish_cap_sess_condition_clean', 'Capacitive epoch (punish, session) — by condition, no individual traces'),
        (sex_speed_fig,          'sex_speed',          'Individual speed plot — by sex'),
        (sex_distance_indiv_fig, 'sex_distance_indiv', 'Individual distance plot — by sex'),
        (sex_reward_indiv_fig,   'sex_reward_indiv',   'Individual reward rate plot — by sex'),
        (epoch_speed_sess_sex_per_mouse_fig, 'epoch_reward_speed_sess_sex_per_mouse', 'Speed epoch (session, reward zone) — per mouse (sex coloring)'),
        (epoch_speed_sess_sex_fig,           'epoch_reward_speed_sess_sex',           'Speed epoch (session, reward zone) — by sex'),
        (epoch_cap_sess_sex_per_mouse_fig,   'epoch_reward_cap_sess_sex_per_mouse',   'Capacitive epoch (session, reward zone) — per mouse (sex coloring)'),
        (epoch_cap_sess_sex_fig,             'epoch_reward_cap_sess_sex',             'Capacitive epoch (session, reward zone) — by sex'),
        (epoch_punish_speed_sess_sex_per_mouse_fig, 'epoch_punish_speed_sess_sex_per_mouse', 'Speed epoch (session, punish zone) — per mouse (sex coloring)'),
        (epoch_punish_speed_sess_sex_fig,           'epoch_punish_speed_sess_sex',           'Speed epoch (session, punish zone) — by sex'),
        (epoch_punish_cap_sess_sex_per_mouse_fig,   'epoch_punish_cap_sess_sex_per_mouse',   'Capacitive epoch (session, punish zone) — per mouse (sex coloring)'),
        (epoch_punish_cap_sess_sex_fig,             'epoch_punish_cap_sess_sex',             'Capacitive epoch (session, punish zone) — by sex'),
        (epoch_reward_speed_pre_post_fig,            'epoch_reward_speed_pre_post',            'Pre/post-reward speed bar chart by condition'),
        (epoch_reward_speed_diff_fig,                'epoch_reward_speed_diff',                'Pre-minus-post-reward speed difference by condition'),
        (epoch_reward_cap_pre_post_fig,              'epoch_reward_cap_pre_post',              'Pre/post-reward capacitive (z-scored) bar chart by condition'),
        (epoch_reward_cap_diff_fig,                  'epoch_reward_cap_diff',                  'Pre-minus-post-reward capacitive difference by condition (Mann-Whitney U)'),
        (epoch_reward_speed_pre_post_entry_fig,      'epoch_reward_speed_pre_post_entry',      'Pre/post-entry reward zone speed bar chart by condition (0.65 s windows)'),
        (epoch_reward_speed_diff_entry_fig,          'epoch_reward_speed_diff_entry',          'Pre-minus-post-entry reward zone speed difference by condition (0.65 s windows)'),
        (epoch_reward_speed_pre_post_entry_1s_fig,   'epoch_reward_speed_pre_post_entry_1s',   'Pre/post-entry reward zone speed bar chart by condition (1 s windows)'),
        (epoch_reward_speed_diff_entry_1s_fig,       'epoch_reward_speed_diff_entry_1s',       'Pre-minus-post-entry reward zone speed difference by condition (1 s windows)'),
        (epoch_reward_lick_count_sess_per_mouse_fig, 'epoch_reward_lick_count_sess_per_mouse', 'Lick count epoch per mouse (reward zone, session-averaged)'),
        (epoch_reward_lick_count_sess_cond_fig,      'epoch_reward_lick_count_sess_cond',      'Lick count epoch by condition (reward zone, session-averaged)'),
        (epoch_reward_lick_count_sess_cond_clean_fig,'epoch_reward_lick_count_sess_cond_clean', 'Lick count epoch by condition only — no individual traces (reward zone, session-averaged)'),
        (epoch_punish_lick_count_sess_per_mouse_fig, 'epoch_punish_lick_count_sess_per_mouse', 'Lick count epoch per mouse (punishment zone, session-averaged)'),
        (epoch_punish_lick_count_sess_cond_fig,      'epoch_punish_lick_count_sess_cond',      'Lick count epoch by condition (punishment zone, session-averaged)'),
        (epoch_punish_speed_pre_post_fig,            'epoch_punish_speed_pre_post',            'Pre/post-cutoff punishment zone speed bar chart by condition'),
        (epoch_punish_speed_diff_fig,                'epoch_punish_speed_diff',                'Pre-minus-post-cutoff punishment zone speed difference by condition'),
        (epoch_punish_speed_pre_post_entry_fig,      'epoch_punish_speed_pre_post_entry',      'Pre/post-entry punishment zone speed bar chart by condition (1 s windows)'),
        (epoch_punish_speed_diff_entry_fig,          'epoch_punish_speed_diff_entry',          'Pre-minus-post-entry punishment zone speed difference by condition (1 s windows)'),
        (epoch_punish_cap_pre_post_fig,              'epoch_punish_cap_pre_post',              'Pre/post-cutoff punishment zone capacitive (z-scored) bar chart by condition (0.65 s windows)'),
        (epoch_punish_cap_diff_fig,                  'epoch_punish_cap_diff',                  'Pre-minus-post-cutoff punishment zone capacitive difference by condition (Mann-Whitney U)'),
        (epoch_punish_cap_pre_post_entry_fig,        'epoch_punish_cap_pre_post_entry',        'Pre/post-entry punishment zone capacitive (z-scored) bar chart by condition (1 s windows)'),
        (epoch_punish_cap_diff_entry_fig,            'epoch_punish_cap_diff_entry',            'Pre-minus-post-entry punishment zone capacitive difference by condition (Mann-Whitney U)'),
        (expl_speed_histogram_fig,                   'expl_speed_histogram',                   'Exploratory: speed histogram (all sessions + per-mouse means)'),
        (expl_speed_distfit_fig,                     'expl_speed_distfit',                     'Exploratory: Average speed distribution fit (Normal, Log-normal, Gamma)'),
        (expl_speed_boxplot_fig,                     'expl_speed_boxplot',                     'Exploratory: speed box-and-whisker (per-mouse + overall)'),
        (expl_speed_rm_anova_resid_fig,              'expl_speed_rm_anova_resid',               'Exploratory: RM ANOVA residual diagnostics'),
        (expl_cap_histogram_fig,                     'expl_cap_histogram',                     'Exploratory: z-scored mean cap value histogram'),
        (expl_cap_boxplot_fig,                       'expl_cap_boxplot',                       'Exploratory: z-scored mean cap value box-and-whisker'),
        (expl_cap_rm_anova_resid_fig,                'expl_cap_rm_anova_resid',                'Exploratory: Capacitive RM ANOVA residual diagnostics'),
        (expl_cap_distfit_fig,                       'expl_cap_distfit',                       'Exploratory: Capacitive sensor value distribution fit (Normal, Log-normal, Gamma)'),
        (expl_lick_distfit_fig,                      'expl_lick_distfit',                      'Exploratory: Lick count Poisson vs NB distribution fit'),
        (expl_lick_boxplot_fig,                      'expl_lick_boxplot',                      'Exploratory: raw lick count box-and-whisker'),
        (expl_lick_rm_anova_resid_fig,               'expl_lick_rm_anova_resid',               'Exploratory: Raw lick count RM ANOVA residual diagnostics'),
        (expl_lick_rate_distfit_fig,                 'expl_lick_rate_distfit',                 'Exploratory: Average lick rate distribution fit (Normal, Log-normal, Gamma)'),
        (expl_lick_reward_ratio_distfit_fig,          'expl_lick_reward_ratio_distfit',          'Exploratory: Lick/reward ratio distribution KDE + normality test'),
    ]

    for fig, name, title in plot_configs:
        if fig is None:
            continue
        save_path = filedialog.asksaveasfilename(
            defaultextension=".svg",
            filetypes=[("SVG files", "*.svg"), ("All files", "*.*")],
            title=f"Save {title} as",
            initialfile=f"mouse_{name}_comparison_{len(file_paths)}mice.svg"
        )
        if save_path:
            fig.savefig(save_path, bbox_inches='tight', format='svg')
            print(f"{title} saved to: {save_path}")
if __name__ == "__main__":
    main()
